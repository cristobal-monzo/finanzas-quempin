# -*- coding: utf-8 -*-
"""
analisis_financiero.py -- Consolidador de costos reales por proyecto para
QUEMPIN SpA. Lee Centro de Costos.xlsx (SOLO LECTURA, nunca lo escribe) y
mantiene Análisis de Proyectos.xlsx (3 hojas: Proyectos, Detalle Costos
Reales, Indicadores). Ver docs/superpowers/specs/2026-07-20-analisis-
financiero-design.md para el diseño completo.
"""

from pathlib import Path

import openpyxl

# ── CONFIGURACIÓN ────────────────────────────────────────────────────────────

RAIZ = Path(__file__).resolve().parent
RAIZ_MODULO = RAIZ.parent
RUTA_EXCEL = RAIZ_MODULO / "Análisis de Proyectos.xlsx"
RAIZ_RESPALDOS = RAIZ_MODULO / "Respaldos"

RAIZ_CENTRO_COSTOS = RAIZ_MODULO.parent / "Centro de Costos"
RUTA_EXCEL_CENTRO_COSTOS = RAIZ_CENTRO_COSTOS / "Excel" / "Centro de Costos.xlsx"
RAIZ_FACTURAS_CENTRO_COSTOS = (
    RAIZ_CENTRO_COSTOS / "Sitio de comunicación - Centro de Costos 1" / "Facturas y Boletas"
)

HOJA_PROYECTOS = "Proyectos"
HOJA_DETALLE_COSTOS_REALES = "Detalle Costos Reales"
HOJA_INDICADORES = "Indicadores"

HEADERS_PROYECTOS = [
    "TAG proyecto", "Nombre del proyecto", "Estado", "Fecha de inicio",
    "Fecha de cierre", "Monto de Venta (sin IVA)",
    "Costos Materiales Proyectados", "Costos Equipos Proyectados",
    "Mano de Obra Proyectada", "Otros Costos Proyectados",
    "Costos Materiales Reales", "Costos Equipos Reales",
    "Otros Costos Reales", "Mano de Obra Real", "Total Proyectado",
    "Total Real", "Margen Proyectado", "Margen Real",
    "Desviación % (Real vs Proyectado)",
]
HEADERS_DETALLE_COSTOS_REALES = ["TAG proyecto", "Subcategoría", "Bucket", "Total sin IVA"]
HEADERS_INDICADORES = [
    "TAG proyecto", "Nombre del proyecto", "Rentabilidad sobre costo",
    "Margen neto %", "Productividad Materiales", "Productividad Equipos",
    "Productividad MO", "Productividad Otros", "Costo Materiales % de venta",
    "Costo Equipos % de venta", "Costo MO % de venta", "Costo Otros % de venta",
    "Desviación % Materiales", "Desviación % Equipos", "Desviación % MO",
    "Desviación % Otros",
]


def asegurar_estructura_workbook(ruta_excel: Path) -> openpyxl.Workbook:
    """Abre ruta_excel si existe, o crea un libro nuevo. Garantiza que las 3
    hojas existan con encabezados en la fila 1 -- si una hoja ya existe, no
    la toca (regla de oro: no reescribir datos ya presentes). Elimina hojas
    default vacías ("Hoja1"/"Sheet") si quedaron de un libro recién creado."""
    if ruta_excel.exists():
        wb = openpyxl.load_workbook(ruta_excel)
    else:
        wb = openpyxl.Workbook()

    for nombre_hoja, headers in (
        (HOJA_PROYECTOS, HEADERS_PROYECTOS),
        (HOJA_DETALLE_COSTOS_REALES, HEADERS_DETALLE_COSTOS_REALES),
        (HOJA_INDICADORES, HEADERS_INDICADORES),
    ):
        if nombre_hoja not in wb.sheetnames:
            ws = wb.create_sheet(nombre_hoja)
            for col, encabezado in enumerate(headers, start=1):
                ws.cell(row=1, column=col, value=encabezado)

    for nombre_default in ("Hoja1", "Sheet"):
        if nombre_default in wb.sheetnames:
            ws_default = wb[nombre_default]
            esta_vacia = all(
                celda.value is None
                for fila in ws_default.iter_rows()
                for celda in fila
            )
            if esta_vacia:
                del wb[nombre_default]

    return wb


# ── MAPEO DE CATEGORÍAS ──────────────────────────────────────────────────────

MAPEO_CATEGORIA_BUCKET = {
    "Materiales": "Materiales",
    "Consumibles": "Materiales",
    "Equipos-Herramientas": "Equipos",
}


def mapear_categoria_a_bucket(categoria_item: str | None) -> tuple[str, bool]:
    """Devuelve (bucket, es_mapeo_explicito). Cualquier categoria_item que no
    esté en MAPEO_CATEGORIA_BUCKET (incluyendo None) cae en "Otros" con
    es_mapeo_explicito=False, para poder avisar sin perder el monto."""
    if categoria_item in MAPEO_CATEGORIA_BUCKET:
        return MAPEO_CATEGORIA_BUCKET[categoria_item], True
    return "Otros", False


# ── LECTURA DE CENTRO DE COSTOS (SOLO LECTURA) ───────────────────────────────

def prefijo_de_n_ref(n_ref: str) -> str:
    """'UMAG-001' -> 'UMAG'. Mismo prefijo que PREFIJOS_PROYECTO en
    Centro de Costos/Sistema/auditor_centro_costos.py."""
    return n_ref.split("-")[0]


def leer_detalle_centro_costos(ruta_excel_cc: Path) -> list[dict]:
    """Lee la hoja 'Detalle' de Centro de Costos.xlsx -- SOLO LECTURA, este
    módulo nunca escribe ese archivo. Filas sin N° Ref. o sin Total sin IVA
    se ignoran (no se puede agrupar ni sumar sin esos dos datos)."""
    wb = openpyxl.load_workbook(ruta_excel_cc, data_only=True)
    ws = wb["Detalle"]
    encabezados = [celda.value for celda in ws[1]]
    col_n_ref = encabezados.index("N° Ref.") + 1
    col_categoria = encabezados.index("Categoría Ítem") + 1
    col_total_sin_iva = encabezados.index("Total sin IVA (CLP)") + 1

    items = []
    for fila in ws.iter_rows(min_row=2):
        n_ref = fila[col_n_ref - 1].value
        total = fila[col_total_sin_iva - 1].value
        if n_ref is None or total is None:
            continue
        categoria = fila[col_categoria - 1].value
        items.append({"n_ref": n_ref, "categoria_item": categoria, "total_sin_iva": float(total)})
    return items


def agrupar_por_proyecto_y_subcategoria(items_detalle: list[dict]) -> dict[tuple[str, str], float]:
    """Suma total_sin_iva agrupado por (prefijo de proyecto, categoria_item
    original -- sin colapsar a bucket todavía, eso lo hace la hoja 'Detalle
    Costos Reales' al escribir, para no perder la subcategoría real)."""
    agrupado: dict[tuple[str, str], float] = {}
    for item in items_detalle:
        prefijo = prefijo_de_n_ref(item["n_ref"])
        clave = (prefijo, item["categoria_item"])
        agrupado[clave] = agrupado.get(clave, 0.0) + item["total_sin_iva"]
    return agrupado


# ── LECTURA DE LA HOJA "PROYECTOS" ───────────────────────────────────────────

def leer_filas_proyectos(ws_proyectos) -> tuple[list[dict], list[str]]:
    """Recorre la hoja 'Proyectos' desde la fila 2. Filas sin TAG o sin
    Nombre se saltan con aviso. TAG duplicado: se queda con la primera
    fila, avisa de las siguientes."""
    filas_validas = []
    avisos = []
    tags_vistos = set()

    for fila_idx in range(2, ws_proyectos.max_row + 1):
        tag = ws_proyectos.cell(row=fila_idx, column=1).value
        nombre = ws_proyectos.cell(row=fila_idx, column=2).value

        if not tag or not nombre:
            if tag or nombre:
                avisos.append(f"Fila {fila_idx}: falta TAG o Nombre, se salta.")
            continue

        if tag in tags_vistos:
            avisos.append(f"Fila {fila_idx}: TAG '{tag}' duplicado, se usa la primera fila.")
            continue

        tags_vistos.add(tag)
        filas_validas.append({"fila": fila_idx, "tag": tag, "nombre": nombre})

    return filas_validas, avisos
