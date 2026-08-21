import importlib.util
import sys
from datetime import date, datetime
from pathlib import Path

import openpyxl

# Ver nota en Centro de Costos/Visualizador Web/tests/test_build_visualizador.py:
# los 3 modulos tienen un "build_visualizador.py" y sys.modules cachea por
# nombre, asi que hay que cargarlo por ruta bajo un nombre unico.
_RUTA_BV = Path(__file__).resolve().parent.parent / "build_visualizador.py"
_spec = importlib.util.spec_from_file_location("build_visualizador_ch", _RUTA_BV)
bv = importlib.util.module_from_spec(_spec)
sys.modules["build_visualizador_ch"] = bv
_spec.loader.exec_module(bv)

HEADERS_MASTER = ["N° Ref.", "Fecha", "Proyecto", "Proveedor"]
HEADERS_DETALLE = [
    "N° Ref.", "Nombre Ítem", "Descripción", "Categoría Ítem",
    "P. Unitario sin IVA", "Total sin IVA (CLP)", "Total con IVA (CLP)",
]


def _wb_con_dos_items(tmp_path):
    wb = openpyxl.Workbook()
    ws_master = wb.active
    ws_master.title = "Master"
    for c, h in enumerate(HEADERS_MASTER, 1):
        ws_master.cell(row=1, column=c, value=h)
    filas_master = [
        ("UMAG-014", datetime(2026, 1, 1), "UMAG", "Ferretería XYZ"),
        ("UMAG-020", datetime(2026, 1, 1), "UMAG", "Ferretería XYZ"),
    ]
    for r, fila in enumerate(filas_master, 2):
        for c, v in enumerate(fila, 1):
            ws_master.cell(row=r, column=c, value=v)

    ws_detalle = wb.create_sheet("Detalle")
    for c, h in enumerate(HEADERS_DETALLE, 1):
        ws_detalle.cell(row=1, column=c, value=h)
    filas_detalle = [
        ("UMAG-014", "Bomba", "Bomba centrífuga Pedrollo 1.5HP", "Equipos-Herramientas", 90000, 90000, 107100),
        ("UMAG-020", "Brocha", "Brocha 4 pulgadas", "Materiales", 3000, 3000, 3570),
    ]
    for r, fila in enumerate(filas_detalle, 2):
        for c, v in enumerate(fila, 1):
            ws_detalle.cell(row=r, column=c, value=v)

    ruta = tmp_path / "Centro de Costos.xlsx"
    wb.save(str(ruta))
    return ruta


def test_extraer_indice_saneado_incluye_todos_los_items(tmp_path, monkeypatch):
    ruta = _wb_con_dos_items(tmp_path)
    monkeypatch.setattr(bv.ch, "consultar_uf_api", lambda fecha: 36000.0)
    monkeypatch.setattr(bv.ch, "RUTA_CACHE_UF", tmp_path / "uf_cache.json")

    data = bv.extraer_indice_saneado(ruta, fecha_hoy=date(2026, 7, 20))

    assert len(data["items"]) == 2
    nombres = sorted(it["nombre_item"] for it in data["items"])
    assert nombres == ["Bomba", "Brocha"]
    assert data["uf_hoy"] == 36000.0
    assert data["excluidos_count"] == 0
    assert data["sin_uf_count"] == 0


def test_extraer_indice_saneado_marca_fuente_mindicador_cuando_responde(tmp_path, monkeypatch):
    ruta = _wb_con_dos_items(tmp_path)
    monkeypatch.setattr(bv.ch, "consultar_uf_api", lambda fecha: 36000.0)
    monkeypatch.setattr(bv.ch, "RUTA_CACHE_UF", tmp_path / "uf_cache.json")

    data = bv.extraer_indice_saneado(ruta, fecha_hoy=date(2026, 7, 20))

    assert data["uf_fuente"] == "mindicador.cl"


def test_extraer_indice_saneado_usa_uf_manual_si_mindicador_falla(tmp_path, monkeypatch):
    ruta = _wb_con_dos_items(tmp_path)

    def _falla(fecha):
        raise bv.ch.UFNoDisponibleError("timeout")
    monkeypatch.setattr(bv.ch, "consultar_uf_api", _falla)
    monkeypatch.setattr(bv.ch, "RUTA_CACHE_UF", tmp_path / "uf_cache.json")

    data = bv.extraer_indice_saneado(
        ruta, fecha_hoy=date(2026, 7, 20),
        uf_manual=39200.5, fuente_manual="Banco Central de Chile, 20-07-2026",
    )

    assert data["uf_hoy"] == 39200.5
    assert data["uf_fuente"] == "Banco Central de Chile, 20-07-2026"


def test_extraer_indice_saneado_conserva_categoria_proyecto_proveedor(tmp_path, monkeypatch):
    ruta = _wb_con_dos_items(tmp_path)
    monkeypatch.setattr(bv.ch, "consultar_uf_api", lambda fecha: 36000.0)
    monkeypatch.setattr(bv.ch, "RUTA_CACHE_UF", tmp_path / "uf_cache.json")

    data = bv.extraer_indice_saneado(ruta, fecha_hoy=date(2026, 7, 20))

    bomba = next(it for it in data["items"] if it["nombre_item"] == "Bomba")
    assert bomba["categoria_item"] == "Equipos-Herramientas"
    assert bomba["proyecto"] == "UMAG"
    assert bomba["proveedor_tag"] == "Ferretería XYZ"
    assert bomba["precio_reajustado_hoy"] == 90000  # UF sin cambio (36000 -> 36000)


def test_build_escribe_snapshot_json_y_html(tmp_path, monkeypatch):
    ruta_excel = _wb_con_dos_items(tmp_path)
    monkeypatch.setattr(bv.ch, "consultar_uf_api", lambda fecha: 36000.0)
    monkeypatch.setattr(bv.ch, "RUTA_CACHE_UF", tmp_path / "uf_cache.json")
    monkeypatch.setattr(bv, "RUTA_EXCEL", ruta_excel)
    ruta_data = tmp_path / "data" / "cotizador-historico.json"
    ruta_build = tmp_path / "build" / "index.html"
    monkeypatch.setattr(bv, "RUTA_DATA_JSON", ruta_data)
    monkeypatch.setattr(bv, "RUTA_BUILD_HTML", ruta_build)

    codigo = bv.build()

    assert codigo == 0
    assert ruta_data.exists()
    assert ruta_build.exists()
    html = ruta_build.read_text(encoding="utf-8")
    assert "__CH_DATA_B64__" not in html  # el placeholder se reemplazo
