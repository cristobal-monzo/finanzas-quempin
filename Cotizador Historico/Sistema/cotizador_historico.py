# -*- coding: utf-8 -*-
"""
cotizador_historico.py — estima el costo actual de un item a partir de sus
compras historicas en Centro de Costos, reajustando cada precio por UF
(fecha de compra -> fecha de la consulta).

Modulo 100% de solo lectura sobre Centro de Costos.xlsx: nunca lo abre en
modo escritura ni lo modifica. Ver ../docs/superpowers/specs/
2026-07-17-cotizador-historico-design.md para el diseno completo.
"""

from datetime import date, datetime
from pathlib import Path
import unicodedata
from difflib import SequenceMatcher
import json
import urllib.error
import urllib.request
import zipfile

import openpyxl

RAIZ_MODULO = Path(__file__).resolve().parent.parent
RUTA_EXCEL_CENTRO_COSTOS = RAIZ_MODULO.parent / "Centro de Costos" / "Excel" / "Centro de Costos.xlsx"
RUTA_CACHE_UF = Path(__file__).resolve().parent / "uf_cache.json"


class ExcelNoDisponibleError(Exception):
    """El archivo Centro de Costos.xlsx no existe o no se pudo abrir para lectura."""


def mapear_encabezados(hoja):
    """dict {texto_encabezado: numero_columna (1-based)} leyendo la fila 1."""
    fila = next(hoja.iter_rows(min_row=1, max_row=1))
    return {celda.value: celda.column for celda in fila if celda.value}


def _fechas_por_ref(ws_master):
    cols = mapear_encabezados(ws_master)
    col_ref = cols["N° Ref."]
    col_fecha = cols["Fecha"]
    fechas = {}
    for fila in ws_master.iter_rows(min_row=2):
        n_ref = fila[col_ref - 1].value
        if n_ref:
            fechas[n_ref] = fila[col_fecha - 1].value
    return fechas


def _proyecto_proveedor_por_ref(ws_master):
    """dict {n_ref: (proyecto, proveedor_tag)} -- ambas columnas son
    opcionales (None si Master no las tiene), a diferencia de N Ref./Fecha
    que _fechas_por_ref exige porque son estructurales."""
    cols = mapear_encabezados(ws_master)
    col_ref = cols["N° Ref."]
    col_proyecto = cols.get("Proyecto")
    col_proveedor = cols.get("Proveedor")
    meta = {}
    for fila in ws_master.iter_rows(min_row=2):
        n_ref = fila[col_ref - 1].value
        if n_ref:
            meta[n_ref] = (
                fila[col_proyecto - 1].value if col_proyecto else None,
                fila[col_proveedor - 1].value if col_proveedor else None,
            )
    return meta


def cargar_items_detalle(ruta_excel=None):
    """Lee Detalle+Master de Centro de Costos.xlsx (solo lectura) y devuelve
    una lista de dicts, uno por item de linea de Detalle, con su fecha ya
    resuelta via Master (cruce por N Ref.). Incluye "total_sin_iva"/
    "total_con_iva" de esa misma fila (permiten derivar la tasa real de IVA
    del documento, ver tasa_iva_real, sin asumir 19% fijo).

    Items cuyo N Ref. no tiene fila en Master, cuya Fecha en Master no es un
    datetime valido, cuyo precio unitario no es un numero, o cuyo precio
    unitario es negativo, quedan con excluido_motivo poblado ("sin_master",
    "fecha_invalida", "precio_invalido" o "precio_negativo") y fecha=None --
    no deben entrar a ninguna busqueda ni agregacion posterior.

    "precio_negativo" es la defensa contra Notas de Credito (devoluciones):
    sus items de Detalle vienen con P. Unitario sin IVA negativo (ver
    UMAG-025, ej. real), y promediarlos junto a compras reales produce un
    "costo actual" mas barato que cualquier compra real -- no sirve para
    evaluar costos futuros. Se filtra por signo del precio, no por "Tipo
    Documento" de Master (que este modulo no lee), porque cargar_items_detalle
    opera a nivel de Detalle y una devolucion siempre es negativa
    independiente de como haya quedado tipificado el documento. Pedido
    explicito del usuario 2026-07-28 -- no reintroducir Notas de Credito al
    indice de este modulo."""
    ruta = Path(ruta_excel) if ruta_excel is not None else RUTA_EXCEL_CENTRO_COSTOS
    try:
        wb = openpyxl.load_workbook(str(ruta), data_only=True, read_only=True)
    except FileNotFoundError as exc:
        raise ExcelNoDisponibleError(f"No existe {ruta}") from exc
    except PermissionError as exc:
        raise ExcelNoDisponibleError(f"No se pudo abrir {ruta} para lectura: {exc}") from exc
    except (zipfile.BadZipFile, KeyError) as exc:
        # zipfile.BadZipFile: el archivo no es un .xlsx valido (ej. quedo a
        # medio escribir por un corte de OneDrive/energia durante un guardado
        # de Centro de Costos). KeyError: es un zip valido pero le faltan las
        # partes internas que openpyxl espera de un xlsx (mismo tipo de
        # corrupcion parcial). Antes de este fix, cualquiera de las dos
        # tumbaba con una traza cruda de zipfile/openpyxl en vez de un error
        # claro y accionable, unico caso de este modulo sin ese tratamiento.
        raise ExcelNoDisponibleError(f"{ruta} existe pero no es un .xlsx valido/legible: {exc}") from exc

    try:
        try:
            ws_detalle = wb["Detalle"]
            ws_master = wb["Master"]
            fechas = _fechas_por_ref(ws_master)
            meta = _proyecto_proveedor_por_ref(ws_master)
            cols = mapear_encabezados(ws_detalle)
            col_ref = cols["N° Ref."]
            col_nombre = cols["Nombre Ítem"]
            col_desc = cols["Descripción"]
            col_precio = cols["P. Unitario sin IVA"]
            col_total_sin_iva = cols["Total sin IVA (CLP)"]
            col_total_con_iva = cols["Total con IVA (CLP)"]
            col_categoria = cols.get("Categoría Ítem")
        except KeyError as exc:
            raise ExcelNoDisponibleError(
                f"Estructura inesperada en {ruta}: falta hoja o columna {exc}"
            ) from exc

        items = []
        for fila in ws_detalle.iter_rows(min_row=2):
            n_ref = fila[col_ref - 1].value
            if not n_ref:
                continue
            fecha = fechas.get(n_ref)
            if n_ref not in fechas:
                excluido_motivo = "sin_master"
            elif not isinstance(fecha, datetime):
                excluido_motivo = "fecha_invalida"
            else:
                excluido_motivo = None

            precio = fila[col_precio - 1].value
            if excluido_motivo is None and not isinstance(precio, (int, float)):
                excluido_motivo = "precio_invalido"
            elif excluido_motivo is None and precio < 0:
                excluido_motivo = "precio_negativo"

            proyecto, proveedor_tag = meta.get(n_ref, (None, None))
            items.append({
                "n_ref": n_ref,
                "nombre_item": fila[col_nombre - 1].value or "",
                "descripcion": fila[col_desc - 1].value or "",
                "categoria_item": fila[col_categoria - 1].value if col_categoria else None,
                "proyecto": proyecto,
                "proveedor_tag": proveedor_tag,
                "precio_unitario_sin_iva": precio,
                "total_sin_iva": fila[col_total_sin_iva - 1].value,
                "total_con_iva": fila[col_total_con_iva - 1].value,
                "fecha": fecha if excluido_motivo is None else None,
                "excluido_motivo": excluido_motivo,
            })
        return items
    finally:
        wb.close()


UMBRAL_SIMILITUD = 0.6
UMBRAL_SUGERENCIA = 0.4
MAX_SUGERENCIAS = 5


def normalizar_texto(texto):
    texto = (texto or "").strip().lower()
    texto = unicodedata.normalize("NFKD", texto)
    return "".join(c for c in texto if not unicodedata.combining(c))


LONGITUD_MINIMA_PALABRA_SIGNIFICATIVA = 4


def similitud(a, b):
    """1.0 si uno es substring del otro. Tambien 1.0 si alguna palabra
    "significativa" (>=4 caracteres, para no engancharse con codigos/medidas
    cortas como "1/2" o marcas de 3 letras) de b es substring de a o
    viceversa -- cubre el caso de un Nombre Item real que no vino
    simplificado (ver Centro de Costos/CLAUDE.md) donde el primer termino
    calza con la consulta pero el string completo no (ej. consulta
    "guantes" contra nombre_item "Guante de trabajo cuero spandex"). Si
    nada de eso aplica, ratio de difflib para tolerar typos/variantes."""
    if not a or not b:
        return 0.0
    if a in b or b in a:
        return 1.0
    for palabra in b.split():
        if len(palabra) >= LONGITUD_MINIMA_PALABRA_SIGNIFICATIVA and (palabra in a or a in palabra):
            return 1.0
    return SequenceMatcher(None, a, b).ratio()


def buscar_items(items, texto_busqueda, umbral=UMBRAL_SIMILITUD, umbral_sugerencia=UMBRAL_SUGERENCIA):
    """Busqueda difusa de texto_busqueda contra Nombre Item/Descripcion.
    Devuelve (coincidencias, sugerencias): coincidencias son items (dicts
    sin modificar) con similitud >= umbral, ordenados de mayor a menor;
    sugerencias son hasta MAX_SUGERENCIAS nombre_item distintos con
    similitud en [umbral_sugerencia, umbral), para cuando no hay match
    directo. Items con excluido_motivo != None se ignoran siempre."""
    consulta = normalizar_texto(texto_busqueda)
    puntuadas = []
    for item in items:
        if item["excluido_motivo"] is not None:
            continue
        s = max(
            similitud(consulta, normalizar_texto(item["nombre_item"])),
            similitud(consulta, normalizar_texto(item["descripcion"])),
        )
        puntuadas.append((s, item))
    puntuadas.sort(key=lambda par: -par[0])

    coincidencias = [item for s, item in puntuadas if s >= umbral]

    sugerencias = []
    for s, item in puntuadas:
        if umbral_sugerencia <= s < umbral and item["nombre_item"] not in sugerencias:
            sugerencias.append(item["nombre_item"])
        if len(sugerencias) >= MAX_SUGERENCIAS:
            break
    return coincidencias, sugerencias


class UFNoDisponibleError(Exception):
    """No se pudo obtener el valor de la UF para una fecha desde mindicador.cl."""


URL_MINDICADOR_UF = "https://mindicador.cl/api/uf/{fecha}"


def consultar_uf_api(fecha):
    """Llama a mindicador.cl y devuelve el valor UF (float) para 'fecha'
    (date o datetime). Lanza UFNoDisponibleError si falla la conexion, la
    respuesta no es JSON valido, o no trae serie de datos. Nunca cachea en
    disco -- eso lo hace el llamador via obtener_valor_uf/guardar_cache_uf."""
    url = URL_MINDICADOR_UF.format(fecha=fecha.strftime("%d-%m-%Y"))
    try:
        with urllib.request.urlopen(url, timeout=10) as respuesta:
            datos = json.loads(respuesta.read().decode("utf-8"))
    except (urllib.error.URLError, OSError, ValueError) as exc:
        raise UFNoDisponibleError(f"No se pudo consultar mindicador.cl para {fecha}: {exc}") from exc

    serie = datos.get("serie") or []
    if not serie:
        raise UFNoDisponibleError(f"mindicador.cl no tiene valor de UF para {fecha}")
    try:
        return serie[0]["valor"]
    except (KeyError, TypeError) as exc:
        raise UFNoDisponibleError(f"Respuesta inesperada de mindicador.cl para {fecha}: {exc}") from exc


def obtener_uf_hoy(fecha, uf_manual=None, fuente_manual=None):
    """UF de 'hoy' con fallback manual: intenta mindicador.cl primero
    (consultar_uf_api); si falla y se paso un valor manual -- buscado por
    el agente en una fuente confiable cuando mindicador.cl no responde --
    lo usa en su lugar. Devuelve (valor, fuente). Sin valor manual, relanza
    UFNoDisponibleError igual que antes: nunca se inventa un valor de UF sin
    que alguien lo haya provisto explicitamente."""
    try:
        return consultar_uf_api(fecha), "mindicador.cl"
    except UFNoDisponibleError:
        if uf_manual is None:
            raise
        return uf_manual, fuente_manual or "fuente manual (sin especificar)"


def cargar_cache_uf(ruta_cache=None):
    ruta = Path(ruta_cache) if ruta_cache is not None else RUTA_CACHE_UF
    if not ruta.exists():
        return {}
    with open(ruta, "r", encoding="utf-8") as f:
        return json.load(f)


def guardar_cache_uf(cache, ruta_cache=None):
    ruta = Path(ruta_cache) if ruta_cache is not None else RUTA_CACHE_UF
    with open(ruta, "w", encoding="utf-8") as f:
        json.dump(cache, f, ensure_ascii=False, indent=2, sort_keys=True)


def obtener_valor_uf(fecha, cache_uf):
    """Valor UF para una fecha HISTORICA (compra pasada), usando cache_uf
    (dict fecha_iso->valor, mutado in-place) para no repetir llamadas a la
    API. El llamador decide si persiste cache_uf con guardar_cache_uf. No
    usar esta funcion para la UF de "hoy" -- ver consultar_item (Task 4),
    que llama a consultar_uf_api directo para hoy, sin pasar por el cache
    de archivo."""
    fecha_iso = fecha.strftime("%Y-%m-%d")
    if fecha_iso in cache_uf:
        return cache_uf[fecha_iso]
    valor = consultar_uf_api(fecha)
    cache_uf[fecha_iso] = valor
    return valor


def calcular_precio_reajustado(precio_original, uf_fecha_compra, uf_hoy):
    factor = uf_hoy / uf_fecha_compra
    return round(precio_original * factor)


def tasa_iva_real(total_sin_iva, total_con_iva):
    """Tasa real de IVA del documento (Total con IVA / Total sin IVA de esa
    fila de Detalle) -- igual que hace Centro de Costos, nunca asume 19%
    fijo, para que tambien sea correcta en documentos exentos o de Zona
    Franca. Si cualquiera de los dos totales no es un numero utilizable, o
    el total sin IVA es 0 (ej. items que figuran en $0 en la factura),
    devuelve 1.0 (sin IVA adicional) como respaldo seguro."""
    if not isinstance(total_sin_iva, (int, float)) or not total_sin_iva:
        return 1.0
    if not isinstance(total_con_iva, (int, float)):
        return 1.0
    return total_con_iva / total_sin_iva


def reajustar_item(item, uf_hoy, cache_uf):
    """Reajusta un item de cargar_items_detalle a la UF de hoy. Devuelve el
    dict de compra reajustada, o None si no se pudo obtener la UF de la
    fecha de compra (UFNoDisponibleError) -- el llamador decide como contar
    ese caso (ver consultar_item y reajustar_todos). 'fecha' queda en ISO
    'YYYY-MM-DD' a proposito (no DD-MM-AAAA): el visualizador web filtra/
    compara este campo como string (item.fecha < desde en template.html),
    lo que solo funciona en orden cronologico correcto en formato ISO;
    cada consumidor (driver.py, visualizador) formatea a DD-MM-AAAA solo
    para mostrarlo, nunca se reformatea en el origen."""
    try:
        uf_compra = obtener_valor_uf(item["fecha"], cache_uf)
    except UFNoDisponibleError:
        return None
    precio_reajustado = calcular_precio_reajustado(item["precio_unitario_sin_iva"], uf_compra, uf_hoy)
    tasa_iva = tasa_iva_real(item.get("total_sin_iva"), item.get("total_con_iva"))
    return {
        "n_ref": item["n_ref"],
        "fecha": item["fecha"].strftime("%Y-%m-%d"),
        "precio_original_sin_iva": item["precio_unitario_sin_iva"],
        "precio_reajustado_hoy": precio_reajustado,
        "precio_reajustado_hoy_con_iva": round(precio_reajustado * tasa_iva),
    }


def reajustar_todos(items, uf_hoy, cache_uf=None):
    """Reajusta TODOS los items indexables (excluido_motivo is None) a la
    UF de hoy, sin filtrar por texto de busqueda -- lo usa el visualizador
    web, que necesita el indice completo, no solo los resultados de una
    consulta puntual. Devuelve (reajustados, sin_uf_count); cada dict de
    reajustados trae ademas nombre_item/descripcion/categoria_item/
    proyecto/proveedor_tag del item original, para no tener que volver a
    cruzarlos despues.

    Si cache_uf es None, carga y persiste el cache de disco el mismo que
    usa consultar_item; si se pasa un dict ya cargado (tests, o un
    llamador que quiere controlar el I/O), se muta in-place y no se
    persiste aqui -- mismo contrato que obtener_valor_uf."""
    propio_cache = cache_uf is None
    if propio_cache:
        cache_uf = cargar_cache_uf()

    reajustados = []
    sin_uf_count = 0
    for item in items:
        if item["excluido_motivo"] is not None:
            continue
        compra = reajustar_item(item, uf_hoy, cache_uf)
        if compra is None:
            sin_uf_count += 1
            continue
        compra["nombre_item"] = item["nombre_item"]
        compra["descripcion"] = item["descripcion"]
        compra["categoria_item"] = item.get("categoria_item")
        compra["proyecto"] = item.get("proyecto")
        compra["proveedor_tag"] = item.get("proveedor_tag")
        reajustados.append(compra)

    if propio_cache:
        guardar_cache_uf(cache_uf)
    return reajustados, sin_uf_count


def consultar_item(texto_busqueda, ruta_excel=None, fecha_hoy=None, uf_manual=None, fuente_manual=None):
    """Orquesta una consulta completa: carga Detalle, busca por texto,
    reajusta cada compra encontrada por UF, y agrega promedio/rango.

    Si la UF de una compra puntual no esta disponible (sin internet, o
    mindicador.cl sin dato para esa fecha), esa compra se excluye del
    resultado (contada en "sin_uf_count") sin abortar la consulta completa
    -- solo si NINGUNA compra pudo reajustarse el resultado queda como "no
    encontrado". La UF de "hoy" (unica para toda la consulta) se pide via
    obtener_uf_hoy: mindicador.cl primero, y si falla y se paso un valor
    manual (buscado en una fuente confiable), lo usa en su lugar -- sin
    valor manual, sigue abortando la consulta completa (UFNoDisponibleError)
    porque sin ella no se puede reajustar nada.

    fecha_hoy es inyectable para tests (default: date.today())."""
    hoy = fecha_hoy or date.today()
    items = cargar_items_detalle(ruta_excel)
    excluidos_count = sum(1 for it in items if it["excluido_motivo"] is not None)

    coincidencias, sugerencias = buscar_items(items, texto_busqueda)
    if not coincidencias:
        return {
            "encontrado": False,
            "compras": [],
            "promedio_reajustado": None,
            "promedio_reajustado_con_iva": None,
            "rango_minimo": None,
            "rango_maximo": None,
            "excluidos_count": excluidos_count,
            "sugerencias": sugerencias,
            "sin_uf_count": 0,
            "uf_fuente": None,
        }

    uf_hoy, uf_fuente = obtener_uf_hoy(hoy, uf_manual=uf_manual, fuente_manual=fuente_manual)
    cache_uf = cargar_cache_uf()
    compras = []
    sin_uf_count = 0
    for item in coincidencias:
        compra = reajustar_item(item, uf_hoy, cache_uf)
        if compra is None:
            sin_uf_count += 1
            continue
        compras.append(compra)
    guardar_cache_uf(cache_uf)

    if not compras:
        return {
            "encontrado": False,
            "compras": [],
            "promedio_reajustado": None,
            "promedio_reajustado_con_iva": None,
            "rango_minimo": None,
            "rango_maximo": None,
            "excluidos_count": excluidos_count,
            "sugerencias": [],
            "sin_uf_count": sin_uf_count,
            "uf_fuente": uf_fuente,
        }

    reajustados = [c["precio_reajustado_hoy"] for c in compras]
    reajustados_con_iva = [c["precio_reajustado_hoy_con_iva"] for c in compras]
    return {
        "encontrado": True,
        "compras": compras,
        "promedio_reajustado": round(sum(reajustados) / len(reajustados)),
        "promedio_reajustado_con_iva": round(sum(reajustados_con_iva) / len(reajustados_con_iva)),
        "rango_minimo": min(reajustados),
        "rango_maximo": max(reajustados),
        "excluidos_count": excluidos_count,
        "sugerencias": [],
        "sin_uf_count": sin_uf_count,
        "uf_fuente": uf_fuente,
    }
