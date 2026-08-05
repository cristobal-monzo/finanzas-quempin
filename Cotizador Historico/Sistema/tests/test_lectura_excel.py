from datetime import datetime

import openpyxl
import pytest

import cotizador_historico as ch


def _crear_excel_prueba(tmp_path, filas_detalle, filas_master):
    """filas_detalle: lista de tuplas (n_ref, nombre_item, descripcion,
    precio_unitario_sin_iva, total_sin_iva, total_con_iva)
    filas_master: lista de tuplas (n_ref, fecha)"""
    wb = openpyxl.Workbook()
    ws_detalle = wb.active
    ws_detalle.title = "Detalle"
    encabezados = [
        "N° Ref.", "Nombre Ítem", "Descripción", "P. Unitario sin IVA",
        "Total sin IVA (CLP)", "Total con IVA (CLP)",
    ]
    for c, h in enumerate(encabezados, 1):
        ws_detalle.cell(row=1, column=c, value=h)
    for r, fila in enumerate(filas_detalle, 2):
        for c, valor in enumerate(fila, 1):
            ws_detalle.cell(row=r, column=c, value=valor)

    ws_master = wb.create_sheet("Master")
    for c, h in enumerate(["N° Ref.", "Fecha"], 1):
        ws_master.cell(row=1, column=c, value=h)
    for r, fila in enumerate(filas_master, 2):
        for c, valor in enumerate(fila, 1):
            ws_master.cell(row=r, column=c, value=valor)

    ruta = tmp_path / "Centro de Costos.xlsx"
    wb.save(ruta)
    return ruta


def test_mapear_encabezados_lee_fila_1(tmp_path):
    ruta = _crear_excel_prueba(tmp_path, filas_detalle=[], filas_master=[])
    wb = openpyxl.load_workbook(ruta, read_only=True)
    cols = ch.mapear_encabezados(wb["Detalle"])
    assert cols["N° Ref."] == 1
    assert cols["Nombre Ítem"] == 2
    assert cols["Descripción"] == 3
    assert cols["P. Unitario sin IVA"] == 4
    assert cols["Total sin IVA (CLP)"] == 5
    assert cols["Total con IVA (CLP)"] == 6


def test_cargar_items_detalle_resuelve_fecha_via_master(tmp_path):
    ruta = _crear_excel_prueba(
        tmp_path,
        filas_detalle=[("UMAG-001", "Taladro", "Taladro percutor 20V", 90000, 90000, 107100)],
        filas_master=[("UMAG-001", datetime(2026, 1, 15))],
    )
    items = ch.cargar_items_detalle(ruta)
    assert len(items) == 1
    item = items[0]
    assert item["n_ref"] == "UMAG-001"
    assert item["nombre_item"] == "Taladro"
    assert item["descripcion"] == "Taladro percutor 20V"
    assert item["precio_unitario_sin_iva"] == 90000
    assert item["total_sin_iva"] == 90000
    assert item["total_con_iva"] == 107100
    assert item["fecha"] == datetime(2026, 1, 15)
    assert item["excluido_motivo"] is None


def test_cargar_items_detalle_excluye_item_sin_master(tmp_path):
    ruta = _crear_excel_prueba(
        tmp_path,
        filas_detalle=[("UMAG-002", "Cemento", "Saco 25kg", 5000, 5000, 5950)],
        filas_master=[],
    )
    items = ch.cargar_items_detalle(ruta)
    assert items[0]["excluido_motivo"] == "sin_master"
    assert items[0]["fecha"] is None


def test_cargar_items_detalle_excluye_fecha_no_parseable(tmp_path):
    ruta = _crear_excel_prueba(
        tmp_path,
        filas_detalle=[("UMAG-003", "Cable", "Cable 10m", 3000, 3000, 3570)],
        filas_master=[("UMAG-003", "sin fecha")],
    )
    items = ch.cargar_items_detalle(ruta)
    assert items[0]["excluido_motivo"] == "fecha_invalida"
    assert items[0]["fecha"] is None


def test_cargar_items_detalle_ignora_filas_sin_n_ref(tmp_path):
    ruta = _crear_excel_prueba(
        tmp_path,
        filas_detalle=[
            ("UMAG-004", "Pintura", "Balde 1 galón", 15000, 15000, 17850),
            (None, None, None, None, None, None),
        ],
        filas_master=[("UMAG-004", datetime(2026, 2, 1))],
    )
    items = ch.cargar_items_detalle(ruta)
    assert len(items) == 1


def test_cargar_items_detalle_archivo_inexistente_lanza_error(tmp_path):
    ruta = tmp_path / "no existe.xlsx"
    with pytest.raises(ch.ExcelNoDisponibleError):
        ch.cargar_items_detalle(ruta)


def test_cargar_items_detalle_archivo_corrupto_lanza_error_claro(tmp_path):
    # No es un .xlsx (zip) valido -- ej. quedo a medio escribir por un corte
    # de OneDrive/energia durante un guardado de Centro de Costos.
    ruta = tmp_path / "corrupto.xlsx"
    ruta.write_bytes(b"esto no es un zip ni un xlsx valido")
    with pytest.raises(ch.ExcelNoDisponibleError):
        ch.cargar_items_detalle(ruta)


def test_cargar_items_detalle_zip_valido_pero_no_es_xlsx_lanza_error_claro(tmp_path):
    # Un zip real, pero sin las partes internas que openpyxl espera de un
    # xlsx (ej. un .zip cualquiera renombrado a .xlsx por error).
    import zipfile
    ruta = tmp_path / "no_es_xlsx.xlsx"
    with zipfile.ZipFile(ruta, "w") as zf:
        zf.writestr("no_es_un_xlsx_real.txt", "contenido")
    with pytest.raises(ch.ExcelNoDisponibleError):
        ch.cargar_items_detalle(ruta)


def test_cargar_items_detalle_excluye_precio_invalido(tmp_path):
    ruta = _crear_excel_prueba(
        tmp_path,
        filas_detalle=[("UMAG-005", "Guante", "Guante de cuero", None, None, None)],
        filas_master=[("UMAG-005", datetime(2026, 2, 1))],
    )
    items = ch.cargar_items_detalle(ruta)
    assert items[0]["excluido_motivo"] == "precio_invalido"
    assert items[0]["fecha"] is None


def test_cargar_items_detalle_excluye_precio_negativo(tmp_path):
    # Precio unitario negativo = devolucion/abono, nunca un costo real.
    # Filtro por signo del precio en Detalle, no por "Tipo Documento" de
    # Master (columna que existe en Centro de Costos pero que este modulo
    # no lee) -- asi funciona igual si una devolucion llega sin tipificar
    # como Nota de Credito en ese campo.
    ruta = _crear_excel_prueba(
        tmp_path,
        filas_detalle=[("UMAG-006", "Válvula", "Devolución - Válvula 1/4", -52101, -260505, -310002)],
        filas_master=[("UMAG-006", datetime(2026, 3, 10))],
    )
    items = ch.cargar_items_detalle(ruta)
    assert items[0]["excluido_motivo"] == "precio_negativo"
    assert items[0]["fecha"] is None


def test_cargar_items_detalle_no_excluye_precio_cero(tmp_path):
    # $0 si es un caso legitimo (items que la factura lista en $0), a
    # diferencia de un precio negativo -- no confundir ambos.
    ruta = _crear_excel_prueba(
        tmp_path,
        filas_detalle=[("UMAG-007", "Flete", "Despacho sin costo", 0, 0, 0)],
        filas_master=[("UMAG-007", datetime(2026, 3, 10))],
    )
    items = ch.cargar_items_detalle(ruta)
    assert items[0]["excluido_motivo"] is None


def test_cargar_items_detalle_falta_hoja_master_lanza_error_claro(tmp_path):
    wb = openpyxl.Workbook()
    ws_detalle = wb.active
    ws_detalle.title = "Detalle"
    for c, h in enumerate(["N° Ref.", "Nombre Ítem", "Descripción", "P. Unitario sin IVA"], 1):
        ws_detalle.cell(row=1, column=c, value=h)
    ruta = tmp_path / "sin_master.xlsx"
    wb.save(ruta)

    with pytest.raises(ch.ExcelNoDisponibleError):
        ch.cargar_items_detalle(ruta)


def _crear_excel_prueba_enriquecido(tmp_path):
    """Detalle con 'Categoría Ítem', Master con 'Proyecto'/'Proveedor' —
    columnas que cargar_items_detalle debe leer si existen, sin exigirlas."""
    wb = openpyxl.Workbook()
    ws_detalle = wb.active
    ws_detalle.title = "Detalle"
    encabezados_d = [
        "N° Ref.", "Nombre Ítem", "Descripción", "Categoría Ítem",
        "P. Unitario sin IVA", "Total sin IVA (CLP)", "Total con IVA (CLP)",
    ]
    for c, h in enumerate(encabezados_d, 1):
        ws_detalle.cell(row=1, column=c, value=h)
    fila_d = ("UMAG-010", "Bomba", "Bomba centrífuga 1.5HP", "Equipos-Herramientas", 90000, 90000, 107100)
    for c, v in enumerate(fila_d, 1):
        ws_detalle.cell(row=2, column=c, value=v)

    ws_master = wb.create_sheet("Master")
    encabezados_m = ["N° Ref.", "Fecha", "Proyecto", "Proveedor"]
    for c, h in enumerate(encabezados_m, 1):
        ws_master.cell(row=1, column=c, value=h)
    fila_m = ("UMAG-010", datetime(2026, 3, 10), "UMAG", "Ferretería XYZ")
    for c, v in enumerate(fila_m, 1):
        ws_master.cell(row=2, column=c, value=v)

    ruta = tmp_path / "Centro de Costos.xlsx"
    wb.save(ruta)
    return ruta


def test_cargar_items_detalle_incluye_categoria_proyecto_proveedor_si_existen(tmp_path):
    ruta = _crear_excel_prueba_enriquecido(tmp_path)
    items = ch.cargar_items_detalle(ruta)
    assert items[0]["categoria_item"] == "Equipos-Herramientas"
    assert items[0]["proyecto"] == "UMAG"
    assert items[0]["proveedor_tag"] == "Ferretería XYZ"


def test_cargar_items_detalle_categoria_proyecto_proveedor_none_si_no_existen(tmp_path):
    # _crear_excel_prueba (la funcion original, ya existente en este archivo)
    # NO tiene esas columnas -- deben quedar en None, nunca lanzar KeyError.
    ruta = _crear_excel_prueba(
        tmp_path,
        filas_detalle=[("UMAG-001", "Taladro", "Taladro percutor 20V", 90000, 90000, 107100)],
        filas_master=[("UMAG-001", datetime(2026, 1, 15))],
    )
    items = ch.cargar_items_detalle(ruta)
    assert items[0]["categoria_item"] is None
    assert items[0]["proyecto"] is None
    assert items[0]["proveedor_tag"] is None
