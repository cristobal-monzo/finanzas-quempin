# -*- coding: utf-8 -*-
from datetime import date

import openpyxl

import driver as drv

HEADERS_PROYECTOS_TEST = [
    "TAG proyecto", "Nombre del proyecto", "Cliente", "Estado",
    "Fecha de inicio", "Fecha de cierre", "Monto de Venta (sin IVA)",
    "Costos Materiales Proyectados", "Costos Equipos Proyectados",
    "Mano de Obra Proyectada", "Otros Costos Proyectados",
    "Costos Materiales Reales", "Costos Equipos Reales",
    "Otros Costos Reales", "Mano de Obra Real", "Total Proyectado",
    "Total Real", "Margen Proyectado", "Margen Real",
    "Desviación % (Real vs Proyectado)", "Categoría",
]  # mismo orden real que master (ver Task 4)


def _fila_completa(**overrides) -> dict:
    base = {
        "TAG proyecto": "UMAG", "Nombre del proyecto": "UMAG", "Cliente": "UMAG",
        "Estado": "Activo", "Fecha de inicio": date(2026, 1, 10),
        "Fecha de cierre": None, "Monto de Venta (sin IVA)": 1000000,
        "Costos Materiales Proyectados": 100000, "Costos Equipos Proyectados": 100000,
        "Mano de Obra Proyectada": 100000, "Otros Costos Proyectados": 100000,
        "Costos Materiales Reales": 90000, "Costos Equipos Reales": 90000,
        "Otros Costos Reales": 90000, "Mano de Obra Real": 90000,
        "Total Proyectado": 400000, "Total Real": 360000,
        "Margen Proyectado": 600000, "Margen Real": 640000,
        "Desviación % (Real vs Proyectado)": -0.1, "Categoría": "I+D+i",
    }
    base.update(overrides)
    return base


def _crear_excel_af(tmp_path, filas_proyectos: list[dict]):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Proyectos"
    for col, h in enumerate(HEADERS_PROYECTOS_TEST, start=1):
        ws.cell(row=1, column=col, value=h)
    for fila_idx, fila in enumerate(filas_proyectos, start=2):
        for col, h in enumerate(HEADERS_PROYECTOS_TEST, start=1):
            ws.cell(row=fila_idx, column=col, value=fila.get(h))
    ruta = tmp_path / "Análisis de Proyectos.xlsx"
    wb.save(ruta)
    return ruta


def test_listar_entidades_incluye_proyectos_clientes_y_categorias(tmp_path):
    ruta = _crear_excel_af(tmp_path, [
        _fila_completa(),
        _fila_completa(**{
            "TAG proyecto": "CFLI", "Nombre del proyecto": "Cesfam Limache",
            "Cliente": "Cesfam Limache", "Categoría": "Mantenimiento",
        }),
    ])
    entidades = drv.listar_entidades(ruta)
    assert entidades["proyecto:UMAG"] == ("proyecto", "UMAG")
    assert entidades["cliente:UMAG"] == ("cliente", "UMAG")
    assert entidades["categoria:I+D+i"] == ("categoria", "I+D+i")
    assert entidades["categoria:Mantenimiento"] == ("categoria", "Mantenimiento")


def test_listar_entidades_excluye_proyectos_sin_datos_completos(tmp_path):
    ruta = _crear_excel_af(tmp_path, [
        _fila_completa(),
        _fila_completa(**{
            "TAG proyecto": "CFLI", "Cliente": "Cesfam Limache",
            "Monto de Venta (sin IVA)": None,  # incompleto
        }),
    ])
    entidades = drv.listar_entidades(ruta)
    assert "proyecto:CFLI" not in entidades
    assert "cliente:Cesfam Limache" not in entidades
    assert "proyecto:UMAG" in entidades


def test_calcular_reportes_pendientes_marca_todo_la_primera_vez(tmp_path):
    ruta = _crear_excel_af(tmp_path, [
        _fila_completa(),
        _fila_completa(**{
            "TAG proyecto": "CFLI", "Nombre del proyecto": "Cesfam Limache",
            "Cliente": "Cesfam Limache", "Categoría": "Mantenimiento",
        }),
    ])
    ruta_estado = tmp_path / "estado_reportes.json"
    pendientes = drv.calcular_reportes_pendientes(ruta, ruta_estado)
    assert set(pendientes) == {
        "proyecto:UMAG", "proyecto:CFLI",
        "cliente:UMAG", "cliente:Cesfam Limache",
        "categoria:I+D+i", "categoria:Mantenimiento",
    }
