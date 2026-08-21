# -*- coding: utf-8 -*-
from datetime import date

import openpyxl
import pytest

import datos_reportes as dr

HEADERS_PROYECTOS_TEST = [
    "TAG proyecto", "Nombre del proyecto", "Cliente", "Estado",
    "Fecha de inicio", "Fecha de cierre", "Monto de Venta (sin IVA)",
    "Costos Materiales Proyectados", "Costos Equipos Proyectados",
    "Mano de Obra Proyectada", "Otros Costos Proyectados",
    "Costos Materiales Reales", "Costos Equipos Reales",
    "Otros Costos Reales", "Mano de Obra Real", "Total Proyectado",
    "Total Real", "Margen Proyectado", "Margen Real",
    "Desviación % (Real vs Proyectado)", "Categoría",
]  # mismo orden real que master (Cliente es la 3a columna, Categoria la ultima)

HEADERS_DETALLE_TEST = ["TAG proyecto", "Subcategoría", "Bucket", "Total sin IVA"]


def _fila_proyecto_completa(**overrides) -> dict:
    """Fila 'feliz' con todos los campos MANUALES requeridos cargados. Las
    columnas derivadas (Costos *Reales, Total/Margen/Desviación) se dejan
    SIN poblar a propósito -- así quedan de verdad tras un 'ejecutar()'
    real (fórmulas de Excel que openpyxl no cachea) -- el código bajo
    prueba debe recalcularlas, nunca leerlas."""
    base = {
        "TAG proyecto": "UMAG", "Nombre del proyecto": "UMAG", "Cliente": "UMAG",
        "Estado": "Activo", "Fecha de inicio": date(2026, 1, 10),
        "Fecha de cierre": None, "Monto de Venta (sin IVA)": 1000000,
        "Costos Materiales Proyectados": 100000, "Costos Equipos Proyectados": 100000,
        "Mano de Obra Proyectada": 100000, "Otros Costos Proyectados": 100000,
        "Mano de Obra Real": 90000, "Categoría": "I+D+i",
    }
    base.update(overrides)
    return base


def _filas_detalle(tag: str) -> list[dict]:
    """Detalle Costos Reales 'feliz' para un tag: Materiales=60.000,
    Equipos=70.000, Otros=20.000 -> Total Real=240.000 (+ Mano de Obra Real
    manual), Margen Real=760.000, Desviación=-40% (ahorro, Real<Proyectado),
    Nota=96 (con el fix 2026-07-28: sin ABS(), el componente de desviación
    da el puntaje máximo; margen neto 76% muy por sobre el 25% objetivo ->
    con la curva 2026-08-20 el score de margen se acerca a 100 sin tocarlo,
    93.95, ya no satura de golpe), Evaluación='Excelente' (ver
    kpis_recalculados.py para la verificación a mano completa)."""
    return [
        {"TAG proyecto": tag, "Subcategoría": "Consumibles", "Bucket": "Materiales", "Total sin IVA": 60000},
        {"TAG proyecto": tag, "Subcategoría": "Equipos-Herramientas", "Bucket": "Equipos", "Total sin IVA": 70000},
        {"TAG proyecto": tag, "Subcategoría": "Combustible", "Bucket": "Otros", "Total sin IVA": 20000},
    ]


def _crear_excel_af(tmp_path, filas_proyectos: list[dict], filas_detalle: list[dict] | None = None):
    wb = openpyxl.Workbook()
    ws_p = wb.active
    ws_p.title = "Proyectos"
    for col, h in enumerate(HEADERS_PROYECTOS_TEST, start=1):
        ws_p.cell(row=1, column=col, value=h)
    for fila_idx, fila in enumerate(filas_proyectos, start=2):
        for col, h in enumerate(HEADERS_PROYECTOS_TEST, start=1):
            ws_p.cell(row=fila_idx, column=col, value=fila.get(h))

    ws_d = wb.create_sheet("Detalle Costos Reales")
    for col, h in enumerate(HEADERS_DETALLE_TEST, start=1):
        ws_d.cell(row=1, column=col, value=h)
    for fila_idx, fila in enumerate(filas_detalle or [], start=2):
        for col, h in enumerate(HEADERS_DETALLE_TEST, start=1):
            ws_d.cell(row=fila_idx, column=col, value=fila.get(h))

    ruta = tmp_path / "Análisis de Proyectos.xlsx"
    wb.save(ruta)
    return ruta


def test_paquete_datos_proyecto_incluye_kpis_recalculados_y_en_desarrollo(tmp_path):
    ruta = _crear_excel_af(tmp_path, [_fila_proyecto_completa()], _filas_detalle("UMAG"))
    paquete = dr.paquete_datos_proyecto(ruta, "UMAG")
    assert paquete["tipo"] == "proyecto"
    assert paquete["proyecto"]["Categoría"] == "I+D+i"
    assert paquete["proyecto"]["Total Real"] == 240000
    assert paquete["proyecto"]["Margen Real"] == 760000
    assert paquete["proyecto"]["Desviación % (Real vs Proyectado)"] == pytest.approx(-0.4)
    assert paquete["indicadores"]["Margen neto %"] == pytest.approx(0.76)
    assert paquete["indicadores"]["Nota del Proyecto"] == 96
    assert paquete["indicadores"]["Evaluación"] == "Excelente"
    assert paquete["en_desarrollo"] is True


def test_paquete_datos_proyecto_kpi_none_si_denominador_es_cero(tmp_path):
    ruta = _crear_excel_af(tmp_path, [
        _fila_proyecto_completa(**{"Costos Materiales Proyectados": 0}),
    ], _filas_detalle("UMAG"))
    paquete = dr.paquete_datos_proyecto(ruta, "UMAG")
    assert paquete["indicadores"]["Desviación % Materiales"] is None


def test_paquete_datos_proyecto_en_desarrollo_false_si_fecha_de_cierre_ya_paso(tmp_path):
    ruta = _crear_excel_af(tmp_path, [
        _fila_proyecto_completa(**{"Fecha de cierre": date(2020, 1, 1)}),
    ], _filas_detalle("UMAG"))
    paquete = dr.paquete_datos_proyecto(ruta, "UMAG")
    assert paquete["en_desarrollo"] is False


def test_paquete_datos_proyecto_lanza_valueerror_si_no_existe(tmp_path):
    ruta = _crear_excel_af(tmp_path, [_fila_proyecto_completa()], _filas_detalle("UMAG"))
    with pytest.raises(ValueError):
        dr.paquete_datos_proyecto(ruta, "NOEXISTE")


def test_paquete_datos_proyecto_lanza_datosincompletos_si_falta_un_campo_manual(tmp_path):
    ruta = _crear_excel_af(tmp_path, [
        _fila_proyecto_completa(**{"Mano de Obra Real": None}),
    ], _filas_detalle("UMAG"))
    with pytest.raises(dr.DatosIncompletosError):
        dr.paquete_datos_proyecto(ruta, "UMAG")


def test_paquete_datos_proyecto_no_requiere_fecha_de_cierre_para_estar_completo(tmp_path):
    ruta = _crear_excel_af(tmp_path, [
        _fila_proyecto_completa(**{"Fecha de cierre": None}),
    ], _filas_detalle("UMAG"))
    paquete = dr.paquete_datos_proyecto(ruta, "UMAG")  # no lanza DatosIncompletosError
    assert paquete["en_desarrollo"] is True


def test_paquete_datos_cliente_incluye_cltv_recalculado_y_sus_proyectos(tmp_path):
    ruta = _crear_excel_af(tmp_path, [_fila_proyecto_completa()], _filas_detalle("UMAG"))
    paquete = dr.paquete_datos_cliente(ruta, "UMAG")
    assert paquete["tipo"] == "cliente"
    # 1 solo proyecto -> vida=1, meses_activo piso de 12 (1 año), frecuencia=1.
    assert paquete["cltv"]["CLTV"] == pytest.approx(760000)
    assert paquete["cltv"]["Clasificación"] == "Clientes estratégicos"  # unico cliente valido
    assert len(paquete["proyectos"]) == 1
    assert paquete["proyectos"][0]["TAG proyecto"] == "UMAG"


def test_paquete_datos_cliente_excluye_proyectos_incompletos(tmp_path):
    ruta = _crear_excel_af(
        tmp_path,
        [
            _fila_proyecto_completa(**{"TAG proyecto": "UMAG"}),
            _fila_proyecto_completa(**{
                "TAG proyecto": "UMAG2", "Nombre del proyecto": "UMAG Fase 2",
                "Monto de Venta (sin IVA)": None,
            }),
        ],
        _filas_detalle("UMAG") + _filas_detalle("UMAG2"),
    )
    paquete = dr.paquete_datos_cliente(ruta, "UMAG")
    assert len(paquete["proyectos"]) == 1
    assert paquete["proyectos"][0]["TAG proyecto"] == "UMAG"


def test_paquete_datos_categoria_agrupa_por_categoria(tmp_path):
    ruta = _crear_excel_af(tmp_path, [
        _fila_proyecto_completa(),
        _fila_proyecto_completa(**{
            "TAG proyecto": "CFLI", "Nombre del proyecto": "Cesfam Limache",
            "Cliente": "Cesfam Limache", "Categoría": "Mantenimiento",
        }),
    ], _filas_detalle("UMAG") + _filas_detalle("CFLI"))
    paquete = dr.paquete_datos_categoria(ruta, "Mantenimiento")
    assert paquete["tipo"] == "categoria"
    assert len(paquete["proyectos"]) == 1
    assert paquete["proyectos"][0]["TAG proyecto"] == "CFLI"


def test_paquete_datos_categoria_excluye_proyectos_incompletos(tmp_path):
    ruta = _crear_excel_af(tmp_path, [
        _fila_proyecto_completa(**{
            "TAG proyecto": "CFLI", "Cliente": "Cesfam Limache", "Categoría": "Mantenimiento",
        }),
        _fila_proyecto_completa(**{
            "TAG proyecto": "CCON", "Cliente": "Cesfam Constitución", "Categoría": "Mantenimiento",
            "Monto de Venta (sin IVA)": None,
        }),
    ], _filas_detalle("CFLI") + _filas_detalle("CCON"))
    paquete = dr.paquete_datos_categoria(ruta, "Mantenimiento")
    assert len(paquete["proyectos"]) == 1
    assert paquete["proyectos"][0]["TAG proyecto"] == "CFLI"


def test_paquete_datos_categoria_lanza_valueerror_si_no_hay_proyectos(tmp_path):
    ruta = _crear_excel_af(tmp_path, [_fila_proyecto_completa()], _filas_detalle("UMAG"))
    with pytest.raises(ValueError):
        dr.paquete_datos_categoria(ruta, "Sin Categoria Real")


def test_paquete_datos_comparacion_combina_entidades(tmp_path):
    ruta = _crear_excel_af(tmp_path, [
        _fila_proyecto_completa(),
        _fila_proyecto_completa(**{
            "TAG proyecto": "CFLI", "Cliente": "Cesfam Limache", "Categoría": "Mantenimiento",
        }),
    ], _filas_detalle("UMAG") + _filas_detalle("CFLI"))
    paquete = dr.paquete_datos_comparacion(ruta, [("proyecto", "UMAG"), ("proyecto", "CFLI")])
    assert paquete["tipo"] == "comparacion"
    assert len(paquete["entidades"]) == 2
    assert paquete["entidades"][0]["proyecto"]["TAG proyecto"] == "UMAG"


def test_paquete_datos_comparacion_rechaza_tipo_desconocido(tmp_path):
    ruta = _crear_excel_af(tmp_path, [_fila_proyecto_completa()], _filas_detalle("UMAG"))
    with pytest.raises(ValueError):
        dr.paquete_datos_comparacion(ruta, [("no_existe", "UMAG")])
