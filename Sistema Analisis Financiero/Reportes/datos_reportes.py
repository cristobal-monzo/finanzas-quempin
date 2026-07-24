# -*- coding: utf-8 -*-
"""
datos_reportes.py -- Paquetes de datos de solo lectura (dict) para que el
agente redacte reportes PDF sin leer celdas de Excel a mano. Nunca escribe
Análisis de Proyectos.xlsx.
"""

import sys
from datetime import date, datetime
from pathlib import Path

import openpyxl

RAIZ_REPORTES = Path(__file__).resolve().parent
RAIZ_SISTEMA = RAIZ_REPORTES.parent / "Sistema"
if str(RAIZ_SISTEMA) not in sys.path:
    sys.path.insert(0, str(RAIZ_SISTEMA))

from analisis_financiero import HOJA_DETALLE_COSTOS_REALES, HOJA_PROYECTOS  # noqa: E402

from kpis_recalculados import (  # noqa: E402
    calcular_cltv_clientes, costos_reales_por_proyecto, recalcular_proyecto,
)

# Campos manuales que un proyecto debe tener cargados para generar reporte
# (spec §6). "Fecha de cierre" queda deliberadamente FUERA -- su ausencia (o
# una fecha futura) marca al proyecto como "en desarrollo", no incompleto.
# "Cliente"/"Categoría" tampoco cuentan -- se resuelven automaticamente, no
# son carga manual del usuario.
CAMPOS_MANUALES_REQUERIDOS = [
    "Estado", "Fecha de inicio", "Monto de Venta (sin IVA)",
    "Costos Materiales Proyectados", "Costos Equipos Proyectados",
    "Mano de Obra Proyectada", "Otros Costos Proyectados", "Mano de Obra Real",
]


class DatosIncompletosError(ValueError):
    """Un proyecto no tiene todos sus campos manuales requeridos cargados."""


def proyecto_tiene_datos_completos(proyecto: dict) -> bool:
    return all(
        proyecto.get(campo) not in (None, "") for campo in CAMPOS_MANUALES_REQUERIDOS
    )


def proyecto_esta_en_desarrollo(proyecto: dict, hoy: date | None = None) -> bool:
    """Sin 'Fecha de cierre', o con una posterior a 'hoy' (fecha real actual
    por defecto), el proyecto se considera en desarrollo. Un valor no
    interpretable como fecha se trata igual (conservador: en desarrollo)."""
    hoy = hoy or date.today()
    fecha_cierre = proyecto.get("Fecha de cierre")
    if not fecha_cierre:
        return True
    if isinstance(fecha_cierre, datetime):
        fecha_cierre = fecha_cierre.date()
    if not isinstance(fecha_cierre, date):
        return True
    return fecha_cierre > hoy


def _mapa_encabezados(ws) -> dict[str, int]:
    return {celda.value: idx + 1 for idx, celda in enumerate(ws[1]) if celda.value}


def _todas_las_filas(ws, mapa: dict[str, int]) -> list[dict]:
    """Todas las filas (>=2), como dicts encabezado->valor."""
    return [
        {h: ws.cell(row=fila_idx, column=c).value for h, c in mapa.items()}
        for fila_idx in range(2, ws.max_row + 1)
    ]


def _todos_los_proyectos_recalculados(wb) -> list[dict]:
    """Lee 'Proyectos' + 'Detalle Costos Reales' (ambas data_only=True --
    'Detalle Costos Reales' sí trae valores literales, nunca fórmulas,
    analisis_financiero.py:612-635) y devuelve una entrada por proyecto con
    TAG: {"proyecto": <dict recalculado>, "indicadores": <dict
    recalculado>}. Nunca lee 'Indicadores'/'Clientes' -- openpyxl no
    cachea el resultado de una fórmula que él mismo escribe, así que esas
    hojas quedan en None hasta que alguien abre el Excel a mano (ver
    kpis_recalculados.py)."""
    ws_p = wb[HOJA_PROYECTOS]
    mapa_p = _mapa_encabezados(ws_p)
    ws_d = wb[HOJA_DETALLE_COSTOS_REALES]
    mapa_d = _mapa_encabezados(ws_d)

    costos_reales = costos_reales_por_proyecto(_todas_las_filas(ws_d, mapa_d))

    entradas = []
    for proyecto in _todas_las_filas(ws_p, mapa_p):
        tag = proyecto.get("TAG proyecto")
        if not tag:
            continue
        proyecto_actualizado, indicadores = recalcular_proyecto(
            proyecto, costos_reales.get(tag, {})
        )
        entradas.append({"proyecto": proyecto_actualizado, "indicadores": indicadores})
    return entradas


def paquete_datos_proyecto(ruta_excel: Path, tag: str) -> dict:
    """Datos de 'Proyectos' + KPIs derivados (recalculados en Python, ver
    kpis_recalculados.py) para un proyecto por su TAG. Lanza
    DatosIncompletosError si le faltan campos manuales requeridos."""
    wb = openpyxl.load_workbook(ruta_excel, data_only=True)
    entrada = next(
        (e for e in _todos_los_proyectos_recalculados(wb) if e["proyecto"].get("TAG proyecto") == tag),
        None,
    )
    if entrada is None:
        raise ValueError(f"TAG de proyecto '{tag}' no encontrado en '{HOJA_PROYECTOS}'.")
    proyecto = entrada["proyecto"]
    if not proyecto_tiene_datos_completos(proyecto):
        raise DatosIncompletosError(
            f"Proyecto '{tag}' no tiene todos sus datos manuales cargados -- "
            f"no se genera reporte hasta que se complete."
        )

    return {
        "tipo": "proyecto",
        "tag": tag,
        "proyecto": proyecto,
        "indicadores": entrada["indicadores"],
        "en_desarrollo": proyecto_esta_en_desarrollo(proyecto),
    }


def paquete_datos_cliente(ruta_excel: Path, nombre_cliente: str) -> dict:
    """CLTV (recalculado en Python) + proyectos con datos completos del
    cliente -- los incompletos se excluyen del agregado, no bloquean el
    reporte. El CLTV se calcula sobre TODOS los proyectos completos del
    libro (la Clasificación depende del percentil entre todos los
    clientes), luego se selecciona la entrada de este cliente."""
    wb = openpyxl.load_workbook(ruta_excel, data_only=True)
    completos = [
        e["proyecto"] for e in _todos_los_proyectos_recalculados(wb)
        if proyecto_tiene_datos_completos(e["proyecto"])
    ]
    proyectos_cliente = [p for p in completos if p.get("Cliente") == nombre_cliente]
    cltv = calcular_cltv_clientes(completos).get(nombre_cliente, {})

    if not proyectos_cliente and not cltv:
        raise ValueError(
            f"Cliente '{nombre_cliente}' no encontrado, o ninguno de sus "
            f"proyectos tiene datos completos."
        )

    return {"tipo": "cliente", "cliente": nombre_cliente, "cltv": cltv, "proyectos": proyectos_cliente}


def paquete_datos_categoria(ruta_excel: Path, categoria: str) -> dict:
    """Proyectos con datos completos (recalculados en Python) cuya
    Categoría calza -- los incompletos se excluyen del agregado."""
    wb = openpyxl.load_workbook(ruta_excel, data_only=True)
    proyectos = [
        e["proyecto"] for e in _todos_los_proyectos_recalculados(wb)
        if proyecto_tiene_datos_completos(e["proyecto"]) and e["proyecto"].get("Categoría") == categoria
    ]
    if not proyectos:
        raise ValueError(f"Ningún proyecto con datos completos y Categoría '{categoria}'.")
    return {"tipo": "categoria", "categoria": categoria, "proyectos": proyectos}


_FUNCIONES_POR_TIPO = {
    "proyecto": paquete_datos_proyecto,
    "cliente": paquete_datos_cliente,
    "categoria": paquete_datos_categoria,
}


def paquete_datos_comparacion(ruta_excel: Path, entidades: list[tuple[str, str]]) -> dict:
    """entidades: lista de (tipo, identificador), tipo en 'proyecto'/'cliente'/'categoria'."""
    paquetes = []
    for tipo, identificador in entidades:
        funcion = _FUNCIONES_POR_TIPO.get(tipo)
        if funcion is None:
            raise ValueError(f"Tipo de entidad desconocido: '{tipo}'.")
        paquetes.append(funcion(ruta_excel, identificador))
    return {"tipo": "comparacion", "entidades": paquetes}
