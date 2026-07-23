import json

import openpyxl

import analisis_financiero as af


def _armar_pendiente(tmp_path):
    ruta_excel = tmp_path / "Análisis de Proyectos.xlsx"
    ruta_pendientes = tmp_path / "clientes_pendientes.json"
    wb = af.asegurar_estructura_workbook(ruta_excel)
    ws = wb[af.HOJA_PROYECTOS]
    ws.cell(row=2, column=1, value="AGCI1")
    ws.cell(row=2, column=2, value="AGCID Febrero")
    ws.cell(row=3, column=1, value="AGCI2")
    ws.cell(row=3, column=2, value="AGCID Marzo")
    filas_validas = [
        {"fila": 2, "tag": "AGCI1", "nombre": "AGCID Febrero"},
        {"fila": 3, "tag": "AGCI2", "nombre": "AGCID Marzo"},
    ]
    af.asegurar_columna_cliente(ws, filas_validas, ruta_pendientes)
    wb.save(ruta_excel)
    return ruta_excel, ruta_pendientes


def test_preview_no_toca_el_excel_ni_el_json(tmp_path):
    ruta_excel, ruta_pendientes = _armar_pendiente(tmp_path)
    contenido_antes = ruta_pendientes.read_text(encoding="utf-8")

    resultado = af.confirmar_clientes_pendientes(
        None, ruta_excel=ruta_excel, ruta_pendientes=ruta_pendientes,
        ruta_respaldos=tmp_path / "Respaldos",
    )

    assert len(resultado) == 1
    assert resultado[0]["tag"] == "AGCI2"
    assert ruta_pendientes.read_text(encoding="utf-8") == contenido_antes
    assert not (tmp_path / "Respaldos").exists()


def test_todos_aplica_la_sugerencia_y_recolorea_azul_marino(tmp_path):
    ruta_excel, ruta_pendientes = _armar_pendiente(tmp_path)

    aplicados = af.confirmar_clientes_pendientes(
        "TODOS", ruta_excel=ruta_excel, ruta_pendientes=ruta_pendientes,
        ruta_respaldos=tmp_path / "Respaldos",
    )

    assert len(aplicados) == 1
    wb = openpyxl.load_workbook(ruta_excel)
    ws = wb[af.HOJA_PROYECTOS]
    col_cliente = af.HEADERS_PROYECTOS.index("Cliente") + 1
    celda = ws.cell(row=3, column=col_cliente)
    assert celda.value == "AGCID Febrero"
    assert celda.font.color.rgb == "001F3864"

    guardado = json.loads(ruta_pendientes.read_text(encoding="utf-8"))
    assert guardado[0]["estado"] == "Confirmado"


def test_filtra_por_tag_especifico(tmp_path):
    ruta_excel, ruta_pendientes = _armar_pendiente(tmp_path)

    aplicados = af.confirmar_clientes_pendientes(
        ["NO-EXISTE"], ruta_excel=ruta_excel, ruta_pendientes=ruta_pendientes,
        ruta_respaldos=tmp_path / "Respaldos",
    )

    assert aplicados == []
    guardado = json.loads(ruta_pendientes.read_text(encoding="utf-8"))
    assert guardado[0]["estado"] == "Pendiente"
