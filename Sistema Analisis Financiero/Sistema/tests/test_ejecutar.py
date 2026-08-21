import openpyxl

import analisis_financiero as af


def _crear_excel_cc(tmp_path, filas, con_master=None):
    """con_master: lista opcional de tuplas (n_ref, nombre_proyecto) para
    agregar tambien la hoja 'Master' -- solo la necesitan los tests que
    ejercitan la deteccion de proyectos nuevos (leer_nombres_proyecto_
    centro_costos la requiere; el resto de ejecutar() no la toca)."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Detalle"
    encabezados = [
        "N° Ref.", "Proyecto", "Tipo de Proyecto", "N° Documento", "Nombre Ítem",
        "Descripción", "Categoría Ítem", "Cantidad", "P. Unitario sin IVA",
        "Total sin IVA (CLP)", "Total con IVA (CLP)",
    ]
    for col, encabezado in enumerate(encabezados, start=1):
        ws.cell(row=1, column=col, value=encabezado)
    for fila_idx, (n_ref, categoria_item, total_sin_iva) in enumerate(filas, start=2):
        ws.cell(row=fila_idx, column=1, value=n_ref)
        ws.cell(row=fila_idx, column=7, value=categoria_item)
        ws.cell(row=fila_idx, column=10, value=total_sin_iva)

    if con_master is not None:
        ws_master = wb.create_sheet("Master")
        encabezados_master = [
            "N° Ref.", "Proyecto", "Tipo de Proyecto", "Fecha", "N° Documento",
            "Tipo Documento", "Proveedor", "Proveedor (Razón Social)", "Categoría",
            "Resumen Ítems", "Total sin IVA (CLP)", "IVA 19% (CLP)",
            "Total con IVA (CLP)", "Estado", "Archivo origen", "Fecha modificación",
        ]
        for col, encabezado in enumerate(encabezados_master, start=1):
            ws_master.cell(row=1, column=col, value=encabezado)
        for fila_idx, (n_ref, nombre_proyecto) in enumerate(con_master, start=2):
            ws_master.cell(row=fila_idx, column=1, value=n_ref)
            ws_master.cell(row=fila_idx, column=2, value=nombre_proyecto)

    ruta = tmp_path / "Centro de Costos.xlsx"
    wb.save(ruta)
    return ruta


def _crear_excel_af_con_un_proyecto(tmp_path):
    ruta = tmp_path / "Análisis de Proyectos.xlsx"
    wb = af.asegurar_estructura_workbook(ruta)
    ws = wb[af.HOJA_PROYECTOS]
    ws.cell(row=2, column=1, value="UMAG")
    ws.cell(row=2, column=2, value="UMAG")
    col_venta = af.HEADERS_PROYECTOS.index("Monto de Venta (sin IVA)") + 1
    ws.cell(row=2, column=col_venta, value=1000000)
    wb.save(ruta)
    return ruta


def test_ejecutar_de_punta_a_punta_crea_carpeta_regenera_hojas_y_guarda(tmp_path):
    ruta_af = _crear_excel_af_con_un_proyecto(tmp_path)
    ruta_cc = _crear_excel_cc(tmp_path, [("UMAG-001", "Materiales", 50000.0)])
    raiz_facturas = tmp_path / "Facturas y Boletas"
    raiz_respaldos = tmp_path / "Respaldos"

    resumen = af.ejecutar(ruta_af, ruta_cc, raiz_facturas, raiz_respaldos)

    assert resumen["error"] is None
    assert resumen["carpetas_creadas"] == ["UMAG"]
    assert (raiz_facturas / "UMAG").is_dir()

    wb = openpyxl.load_workbook(ruta_af)
    ws_detalle = wb[af.HOJA_DETALLE_COSTOS_REALES]
    assert ws_detalle.cell(row=2, column=4).value == 50000.0
    ws_proyectos = wb[af.HOJA_PROYECTOS]
    col_mat_real = af.HEADERS_PROYECTOS.index("Costos Materiales Reales") + 1
    assert ws_proyectos.cell(row=2, column=col_mat_real).value.startswith("=SUMIFS(")
    # el archivo AF ya existia en disco (lo creo el fixture) antes de ejecutar(),
    # asi que se respalda antes de sobrescribirlo -- regla de oro del modulo.
    assert len(list((raiz_respaldos).rglob("*.xlsx"))) == 1


def test_ejecutar_dos_veces_es_idempotente(tmp_path):
    ruta_af = _crear_excel_af_con_un_proyecto(tmp_path)
    ruta_cc = _crear_excel_cc(tmp_path, [("UMAG-001", "Materiales", 50000.0)])
    raiz_facturas = tmp_path / "Facturas y Boletas"
    raiz_respaldos = tmp_path / "Respaldos"

    af.ejecutar(ruta_af, ruta_cc, raiz_facturas, raiz_respaldos)
    wb_1 = openpyxl.load_workbook(ruta_af)
    detalle_1 = [
        [wb_1[af.HOJA_DETALLE_COSTOS_REALES].cell(row=r, column=c).value for c in range(1, 5)]
        for r in range(1, wb_1[af.HOJA_DETALLE_COSTOS_REALES].max_row + 1)
    ]

    af.ejecutar(ruta_af, ruta_cc, raiz_facturas, raiz_respaldos)
    wb_2 = openpyxl.load_workbook(ruta_af)
    detalle_2 = [
        [wb_2[af.HOJA_DETALLE_COSTOS_REALES].cell(row=r, column=c).value for c in range(1, 5)]
        for r in range(1, wb_2[af.HOJA_DETALLE_COSTOS_REALES].max_row + 1)
    ]

    assert detalle_1 == detalle_2
    # la segunda corrida ya no crea la carpeta (ya existia)
    resumen_2 = af.ejecutar(ruta_af, ruta_cc, raiz_facturas, raiz_respaldos)
    assert resumen_2["carpetas_creadas"] == []


def test_ejecutar_dry_run_no_escribe_nada(tmp_path):
    ruta_af = _crear_excel_af_con_un_proyecto(tmp_path)
    ruta_cc = _crear_excel_cc(tmp_path, [("UMAG-001", "Combustible", 10000.0)])
    raiz_facturas = tmp_path / "Facturas y Boletas"
    raiz_respaldos = tmp_path / "Respaldos"
    contenido_antes = ruta_af.read_bytes()

    resumen = af.ejecutar(ruta_af, ruta_cc, raiz_facturas, raiz_respaldos, dry_run=True)

    assert ruta_af.read_bytes() == contenido_antes
    assert not (raiz_facturas / "UMAG").exists()
    assert resumen["carpetas_creadas"] == ["UMAG"]  # lo que SE CREARIA, sin crearlo
    assert resumen["categorias_no_mapeadas"] == ["Combustible"]


def test_ejecutar_sin_centro_de_costos_avisa_y_no_falla(tmp_path):
    ruta_af = _crear_excel_af_con_un_proyecto(tmp_path)
    ruta_cc = tmp_path / "no existe.xlsx"
    raiz_facturas = tmp_path / "Facturas y Boletas"
    raiz_respaldos = tmp_path / "Respaldos"

    resumen = af.ejecutar(ruta_af, ruta_cc, raiz_facturas, raiz_respaldos)

    assert resumen["error"] is None
    assert any("no se encontr" in aviso.lower() for aviso in resumen["avisos"])


def test_ejecutar_crea_fila_para_proyecto_nuevo_de_centro_de_costos(tmp_path):
    ruta_af = _crear_excel_af_con_un_proyecto(tmp_path)  # solo tiene UMAG
    ruta_cc = _crear_excel_cc(
        tmp_path,
        [("UMAG-001", "Materiales", 50000.0), ("CFLI-001", "Materiales", 20000.0)],
        con_master=[("UMAG-001", "UMAG"), ("CFLI-001", "Cesfam Limache")],
    )
    raiz_facturas = tmp_path / "Facturas y Boletas"
    raiz_respaldos = tmp_path / "Respaldos"

    resumen = af.ejecutar(ruta_af, ruta_cc, raiz_facturas, raiz_respaldos)

    assert resumen["error"] is None
    assert resumen["proyectos_nuevos"] == ["Cesfam Limache"]

    wb = openpyxl.load_workbook(ruta_af)
    ws_proyectos = wb[af.HOJA_PROYECTOS]
    assert ws_proyectos.cell(row=3, column=1).value == "CFLI"
    assert ws_proyectos.cell(row=3, column=2).value == "Cesfam Limache"
    # Las columnas manuales de la fila nueva quedan en blanco (las llena el
    # usuario a mano); Estado es la primera columna manual tras Categoría.
    col_estado = af.HEADERS_PROYECTOS.index("Estado") + 1
    assert ws_proyectos.cell(row=3, column=col_estado).value is None
    # La fila nueva sí queda enganchada al resto del pipeline: la formula de
    # costos reales se genera igual que para una fila preexistente.
    col_mat_real = af.HEADERS_PROYECTOS.index("Costos Materiales Reales") + 1
    assert ws_proyectos.cell(row=3, column=col_mat_real).value.startswith("=SUMIFS(")


def test_ejecutar_dry_run_previsualiza_proyectos_nuevos_sin_escribir(tmp_path):
    ruta_af = _crear_excel_af_con_un_proyecto(tmp_path)  # solo tiene UMAG
    ruta_cc = _crear_excel_cc(
        tmp_path,
        [("UMAG-001", "Materiales", 50000.0), ("CFLI-001", "Materiales", 20000.0)],
        con_master=[("UMAG-001", "UMAG"), ("CFLI-001", "Cesfam Limache")],
    )
    raiz_facturas = tmp_path / "Facturas y Boletas"
    raiz_respaldos = tmp_path / "Respaldos"
    contenido_antes = ruta_af.read_bytes()

    resumen = af.ejecutar(ruta_af, ruta_cc, raiz_facturas, raiz_respaldos, dry_run=True)

    assert ruta_af.read_bytes() == contenido_antes
    assert resumen["proyectos_nuevos"] == ["Cesfam Limache"]


def test_ejecutar_no_crea_fila_para_proyecto_que_ya_existe(tmp_path):
    ruta_af = _crear_excel_af_con_un_proyecto(tmp_path)  # ya tiene UMAG
    ruta_cc = _crear_excel_cc(
        tmp_path,
        [("UMAG-001", "Materiales", 50000.0)],
        con_master=[("UMAG-001", "UMAG")],
    )
    raiz_facturas = tmp_path / "Facturas y Boletas"
    raiz_respaldos = tmp_path / "Respaldos"

    resumen = af.ejecutar(ruta_af, ruta_cc, raiz_facturas, raiz_respaldos)

    assert resumen["proyectos_nuevos"] == []


def test_ejecutar_no_aborta_si_falla_el_visualizador(tmp_path, monkeypatch):
    def _falla():
        raise RuntimeError("boom")

    monkeypatch.setattr(af, "actualizar_visualizador_af", _falla)

    ruta_excel_af = tmp_path / "Análisis de Proyectos.xlsx"
    ruta_excel_cc = _crear_excel_cc(tmp_path, [("UMAG-001", "Materiales", 50000.0)])

    resumen = af.ejecutar(
        ruta_excel_af=ruta_excel_af, ruta_excel_cc=ruta_excel_cc,
        raiz_facturas_cc=tmp_path / "facturas", raiz_respaldos=tmp_path / "respaldos",
        ruta_clientes_pendientes=tmp_path / "pendientes.json",
    )

    assert resumen["error"] is None
    assert any("visualizador" in aviso.lower() for aviso in resumen["avisos"])
    assert ruta_excel_af.exists()  # el Excel si quedo guardado
