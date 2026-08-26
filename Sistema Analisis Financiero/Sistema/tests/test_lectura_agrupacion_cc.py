# -*- coding: utf-8 -*-
import openpyxl

import analisis_financiero as af


def _crear_excel_cc_minimo(tmp_path, filas):
    """filas: lista de tuplas (n_ref, categoria_item, total_sin_iva)"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Detalle"
    encabezados = [
        "N° Ref.", "Proyecto", "Tipo de Proyecto", "N° Documento", "Nombre Ítem",
        "Descripción", "Categoría Ítem", "Cantidad", "P. Unitario sin IVA",
        "Total sin IVA (CLP)", "Total con IVA (CLP)",
    ]
    for col, encabezado in enumerate(encabezados, start=1):
        ws.cell(row=1, column=col, value=encabezado)
    for fila_idx, (n_ref, categoria_item, total_sin_iva) in enumerate(filas, start=2):
        ws.cell(row=fila_idx, column=1, value=n_ref)
        ws.cell(row=fila_idx, column=7, value=categoria_item)
        ws.cell(row=fila_idx, column=10, value=total_sin_iva)
    ruta = tmp_path / "Centro de Costos.xlsx"
    wb.save(ruta)
    return ruta


def test_prefijo_de_n_ref_toma_lo_anterior_al_guion():
    assert af.prefijo_de_n_ref("UMAG-001") == "UMAG"
    assert af.prefijo_de_n_ref("CFLI-014") == "CFLI"


def test_leer_detalle_centro_costos_lee_las_3_columnas_relevantes(tmp_path):
    ruta = _crear_excel_cc_minimo(tmp_path, [("UMAG-001", "Materiales", 50000)])
    items = af.leer_detalle_centro_costos(ruta)
    assert items == [{"n_ref": "UMAG-001", "categoria_item": "Materiales", "total_sin_iva": 50000.0}]


def _wb_centro_costos_pe(tmp_path, filas_detalle):
    """filas_detalle: lista de tuplas (n_ref, categoria_item, total_sin_igv).
    Headers reales de Peru/Centro de Costos/Excel/Centro de Costos Perú.xlsx
    (confirmados via openpyxl) -- IGV/PEN en vez de IVA/CLP."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Detalle"
    headers = ["N° Ref.", "Categoría Ítem", "Total sin IGV (PEN)"]
    for c, h in enumerate(headers, 1):
        ws.cell(row=1, column=c, value=h)
    for r, fila in enumerate(filas_detalle, 2):
        for c, valor in enumerate(fila, 1):
            ws.cell(row=r, column=c, value=valor)
    ruta = tmp_path / "Centro de Costos Perú.xlsx"
    wb.save(ruta)
    return ruta


def test_leer_detalle_centro_costos_pais_pe_lee_columna_igv_pen(tmp_path):
    ruta = _wb_centro_costos_pe(tmp_path, [("LIMA-001", "Materiales", 300.0)])
    items = af.leer_detalle_centro_costos(ruta, pais="PE")
    assert items == [{"n_ref": "LIMA-001", "categoria_item": "Materiales", "total_sin_iva": 300.0}]


def test_leer_detalle_centro_costos_pais_cl_no_cambia_con_el_parametro_default(tmp_path):
    ruta = _crear_excel_cc_minimo(tmp_path, [("UMAG-001", "Materiales", 90000)])
    items = af.leer_detalle_centro_costos(ruta, pais="CL")
    assert items == [{"n_ref": "UMAG-001", "categoria_item": "Materiales", "total_sin_iva": 90000.0}]


def test_leer_detalle_centro_costos_ignora_filas_sin_total(tmp_path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Detalle"
    ws.cell(row=1, column=1, value="N° Ref.")
    ws.cell(row=1, column=7, value="Categoría Ítem")
    ws.cell(row=1, column=10, value="Total sin IVA (CLP)")
    ws.cell(row=2, column=1, value="UMAG-001")
    ws.cell(row=2, column=7, value="Materiales")
    # sin valor en Total sin IVA (CLP)
    ruta = tmp_path / "Centro de Costos.xlsx"
    wb.save(ruta)
    assert af.leer_detalle_centro_costos(ruta) == []


def test_agrupar_por_proyecto_y_subcategoria_suma_por_prefijo_y_categoria():
    items = [
        {"n_ref": "UMAG-001", "categoria_item": "Materiales", "total_sin_iva": 50000.0},
        {"n_ref": "UMAG-002", "categoria_item": "Materiales", "total_sin_iva": 20000.0},
        {"n_ref": "UMAG-003", "categoria_item": "Equipos-Herramientas", "total_sin_iva": 90000.0},
        {"n_ref": "CFLI-001", "categoria_item": "Materiales", "total_sin_iva": 15000.0},
    ]
    agrupado = af.agrupar_por_proyecto_y_subcategoria(items)
    assert agrupado == {
        ("UMAG", "Materiales"): 70000.0,
        ("UMAG", "Equipos-Herramientas"): 90000.0,
        ("CFLI", "Materiales"): 15000.0,
    }
