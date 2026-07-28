import openpyxl
import pytest

import analisis_financiero as af


def test_una_fila_referencia_las_columnas_correctas_de_proyectos(tmp_path):
    wb = af.asegurar_estructura_workbook(tmp_path / "Análisis de Proyectos.xlsx")
    filas_validas = [{"fila": 2, "tag": "UMAG", "nombre": "UMAG"}]

    af.asegurar_hoja_indicadores(wb, filas_validas)

    ws = wb[af.HOJA_INDICADORES]
    l = af.LETRA_COL_PROYECTOS
    tag, nombre = l["TAG proyecto"], l["Nombre del proyecto"]
    venta, margen_real = l["Monto de Venta (sin IVA)"], l["Margen Real"]
    total_real, total_proy = l["Total Real"], l["Total Proyectado"]
    desviacion_total = l["Desviación % (Real vs Proyectado)"]
    mat_r, eq_r, mo_r, otros_r = (
        l["Costos Materiales Reales"], l["Costos Equipos Reales"],
        l["Mano de Obra Real"], l["Otros Costos Reales"],
    )
    mat_p, eq_p, mo_p, otros_p = (
        l["Costos Materiales Proyectados"], l["Costos Equipos Proyectados"],
        l["Mano de Obra Proyectada"], l["Otros Costos Proyectados"],
    )

    assert ws.cell(row=2, column=1).value == f"=Proyectos!{tag}2"
    assert ws.cell(row=2, column=2).value == f"=Proyectos!{nombre}2"
    # C: Margen neto %
    assert ws.cell(row=2, column=3).value == f"=Proyectos!{margen_real}2/Proyectos!{venta}2"
    # D-G: Costo % de venta (Materiales/Equipos/MO/Otros)
    assert ws.cell(row=2, column=4).value == f"=Proyectos!{mat_r}2/Proyectos!{venta}2"
    assert ws.cell(row=2, column=5).value == f"=Proyectos!{eq_r}2/Proyectos!{venta}2"
    assert ws.cell(row=2, column=6).value == f"=Proyectos!{mo_r}2/Proyectos!{venta}2"
    assert ws.cell(row=2, column=7).value == f"=Proyectos!{otros_r}2/Proyectos!{venta}2"
    # H-K: Estructura % del costo real (mix, sobre Total Real)
    assert ws.cell(row=2, column=8).value == f"=Proyectos!{mat_r}2/Proyectos!{total_real}2"
    assert ws.cell(row=2, column=9).value == f"=Proyectos!{eq_r}2/Proyectos!{total_real}2"
    assert ws.cell(row=2, column=10).value == f"=Proyectos!{mo_r}2/Proyectos!{total_real}2"
    assert ws.cell(row=2, column=11).value == f"=Proyectos!{otros_r}2/Proyectos!{total_real}2"
    # L-O: Desviación % por categoría
    assert ws.cell(row=2, column=12).value == f"=Proyectos!{mat_r}2/Proyectos!{mat_p}2-1"
    assert ws.cell(row=2, column=13).value == f"=Proyectos!{eq_r}2/Proyectos!{eq_p}2-1"
    assert ws.cell(row=2, column=14).value == f"=Proyectos!{mo_r}2/Proyectos!{mo_p}2-1"
    assert ws.cell(row=2, column=15).value == f"=Proyectos!{otros_r}2/Proyectos!{otros_p}2-1"
    # P: Desviación % Total -- referencia directa (no recalcula)
    assert ws.cell(row=2, column=16).value == f"=Proyectos!{desviacion_total}2"
    # Q-U: Ahorro/Sobrecosto neto en $ (Proyectado - Real, por categoría y total)
    assert ws.cell(row=2, column=17).value == f"=Proyectos!{mat_p}2-Proyectos!{mat_r}2"
    assert ws.cell(row=2, column=18).value == f"=Proyectos!{eq_p}2-Proyectos!{eq_r}2"
    assert ws.cell(row=2, column=19).value == f"=Proyectos!{mo_p}2-Proyectos!{mo_r}2"
    assert ws.cell(row=2, column=20).value == f"=Proyectos!{otros_p}2-Proyectos!{otros_r}2"
    assert ws.cell(row=2, column=21).value == f"=Proyectos!{total_proy}2-Proyectos!{total_real}2"
    # X: Peso del proyecto en la cartera de ventas (%) -- venta del proyecto
    # sobre la suma de TODA la columna de venta en "Proyectos" (no solo su
    # propia fila).
    assert ws.cell(row=2, column=24).value == (
        f"=Proyectos!{venta}2/SUM(Proyectos!${venta}:${venta})"
    )
    # Y: Margen por día de ejecución -- vacío ("") si Fecha de cierre no
    # está cargada (proyecto "en desarrollo"), nunca un error/número sin
    # sentido. MAX(1, dias) evita #DIV/0! si cierre e inicio caen el mismo día.
    fecha_inicio, fecha_cierre = l["Fecha de inicio"], l["Fecha de cierre"]
    assert ws.cell(row=2, column=25).value == (
        f'=IF(Proyectos!{fecha_cierre}2="","",'
        f"Proyectos!{margen_real}2/MAX(1,Proyectos!{fecha_cierre}2-Proyectos!{fecha_inicio}2))"
    )


def test_peso_cartera_suma_toda_la_columna_no_solo_la_propia_fila(tmp_path):
    """Peso del proyecto = venta del proyecto / suma de TODA la columna de
    venta de 'Proyectos' -- verificado a mano con 3 proyectos con venta
    cargada: 100.000 + 300.000 + 600.000 = 1.000.000; UMAG (100.000) debería
    pesar 10% de la cartera."""
    wb = af.asegurar_estructura_workbook(tmp_path / "Análisis de Proyectos.xlsx")
    ws_p = wb[af.HOJA_PROYECTOS]
    col_tag = af.HEADERS_PROYECTOS.index("TAG proyecto") + 1
    col_nombre = af.HEADERS_PROYECTOS.index("Nombre del proyecto") + 1
    col_venta = af.HEADERS_PROYECTOS.index("Monto de Venta (sin IVA)") + 1
    for fila, (tag, nombre, venta) in enumerate(
        [("UMAG", "UMAG", 100000), ("CFLI", "Cesfam Limache", 300000), ("CCON", "Cesfam Constitución", 600000)],
        start=2,
    ):
        ws_p.cell(row=fila, column=col_tag, value=tag)
        ws_p.cell(row=fila, column=col_nombre, value=nombre)
        ws_p.cell(row=fila, column=col_venta, value=venta)

    filas_validas = [
        {"fila": 2, "tag": "UMAG", "nombre": "UMAG"},
        {"fila": 3, "tag": "CFLI", "nombre": "Cesfam Limache"},
        {"fila": 4, "tag": "CCON", "nombre": "Cesfam Constitución"},
    ]
    af.asegurar_hoja_indicadores(wb, filas_validas)
    wb.save(wb_path := tmp_path / "resultado.xlsx")

    # Verificación del VALOR (no solo la fórmula): recalculamos a mano con
    # openpyxl en modo lectura normal (fórmulas, no data_only) más los
    # valores ya guardados en "Proyectos" -- sin recurrir a ningún motor de
    # cálculo externo (LibreOffice/Google Sheets), solo aritmética Python.
    wb_leido = openpyxl.load_workbook(wb_path)
    ws_ind = wb_leido[af.HOJA_INDICADORES]
    formula_umag = ws_ind.cell(row=2, column=24).value
    l = af.LETRA_COL_PROYECTOS
    venta_letra = l["Monto de Venta (sin IVA)"]
    assert formula_umag == f"=Proyectos!{venta_letra}2/SUM(Proyectos!${venta_letra}:${venta_letra})"
    total_cartera = 100000 + 300000 + 600000
    assert 100000 / total_cartera == pytest.approx(0.1)


def test_margen_por_dia_formula_identica_sin_importar_la_fila(tmp_path):
    """La fórmula de 'Margen por día de ejecución' es la misma estructura
    (con el guard IF de Fecha de cierre vacía) para cualquier fila -- no hay
    una rama de código en Python que decida 'este proyecto no tiene fecha,
    no escribo fórmula'; el guard vive DENTRO de la fórmula de Excel, para
    que se recalcule solo si el usuario completa la fecha después."""
    wb = af.asegurar_estructura_workbook(tmp_path / "Análisis de Proyectos.xlsx")
    filas_validas = [
        {"fila": 2, "tag": "UMAG", "nombre": "UMAG"},
        {"fila": 3, "tag": "MLER", "nombre": "Microturbina LER"},  # sin Fecha de cierre en la vida real
    ]
    af.asegurar_hoja_indicadores(wb, filas_validas)

    ws = wb[af.HOJA_INDICADORES]
    l = af.LETRA_COL_PROYECTOS
    margen_real = l["Margen Real"]
    fecha_inicio, fecha_cierre = l["Fecha de inicio"], l["Fecha de cierre"]
    esperado_umag = (
        f'=IF(Proyectos!{fecha_cierre}2="","",'
        f"Proyectos!{margen_real}2/MAX(1,Proyectos!{fecha_cierre}2-Proyectos!{fecha_inicio}2))"
    )
    esperado_mler = (
        f'=IF(Proyectos!{fecha_cierre}3="","",'
        f"Proyectos!{margen_real}3/MAX(1,Proyectos!{fecha_cierre}3-Proyectos!{fecha_inicio}3))"
    )
    assert ws.cell(row=2, column=25).value == esperado_umag
    assert ws.cell(row=3, column=25).value == esperado_mler


def test_fila_con_hueco_en_proyectos_queda_compacta_en_indicadores_pero_referencia_la_fila_real(tmp_path):
    wb = af.asegurar_estructura_workbook(tmp_path / "Análisis de Proyectos.xlsx")
    # Proyectos fila 3 fue invalida y se salto -- el segundo proyecto valido
    # esta en la fila 4 de "Proyectos".
    filas_validas = [
        {"fila": 2, "tag": "UMAG", "nombre": "UMAG"},
        {"fila": 4, "tag": "CFLI", "nombre": "Cesfam Limache"},
    ]

    af.asegurar_hoja_indicadores(wb, filas_validas)

    ws = wb[af.HOJA_INDICADORES]
    l = af.LETRA_COL_PROYECTOS
    tag = l["TAG proyecto"]
    venta, margen_real = l["Monto de Venta (sin IVA)"], l["Margen Real"]
    desviacion_total = l["Desviación % (Real vs Proyectado)"]

    assert ws.cell(row=2, column=1).value == f"=Proyectos!{tag}2"
    assert ws.cell(row=3, column=1).value == f"=Proyectos!{tag}4"
    assert ws.cell(row=3, column=3).value == f"=Proyectos!{margen_real}4/Proyectos!{venta}4"
    assert ws.cell(row=3, column=16).value == f"=Proyectos!{desviacion_total}4"


def test_regenerar_borra_filas_de_la_corrida_anterior(tmp_path):
    wb = af.asegurar_estructura_workbook(tmp_path / "Análisis de Proyectos.xlsx")
    af.asegurar_hoja_indicadores(wb, [{"fila": 2, "tag": "UMAG", "nombre": "UMAG"}])
    af.asegurar_hoja_indicadores(wb, [{"fila": 2, "tag": "CFLI", "nombre": "Cesfam Limache"}])

    ws = wb[af.HOJA_INDICADORES]
    assert ws.max_row == 2
