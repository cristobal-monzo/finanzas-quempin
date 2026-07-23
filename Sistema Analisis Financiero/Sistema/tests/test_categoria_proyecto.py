# -*- coding: utf-8 -*-
import openpyxl

import analisis_financiero as af


def _crear_excel_cc_con_tipo_proyecto(tmp_path, filas):
    """filas: lista de tuplas (n_ref, tipo_proyecto)"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Master"
    encabezados = [
        "N° Ref.", "Proyecto", "Tipo de Proyecto", "Fecha", "N° Documento",
        "Tipo Documento", "Proveedor", "Proveedor (Razón Social)", "Categoría",
        "Resumen Ítems", "Total sin IVA (CLP)", "IVA 19% (CLP)",
        "Total con IVA (CLP)", "Estado", "Archivo origen", "Fecha modificación",
    ]
    for col, encabezado in enumerate(encabezados, start=1):
        ws.cell(row=1, column=col, value=encabezado)
    for fila_idx, (n_ref, tipo_proyecto) in enumerate(filas, start=2):
        ws.cell(row=fila_idx, column=1, value=n_ref)
        ws.cell(row=fila_idx, column=3, value=tipo_proyecto)
    ruta = tmp_path / "Centro de Costos.xlsx"
    wb.save(ruta)
    return ruta


def test_leer_tipo_proyecto_lee_un_solo_valor_por_proyecto(tmp_path):
    ruta = _crear_excel_cc_con_tipo_proyecto(tmp_path, [
        ("UMAG-001", "I+D+i"),
        ("UMAG-002", "I+D+i"),
        ("CFLI-001", "Mantenimiento"),
    ])
    assert af.leer_tipo_proyecto_centro_costos(ruta) == {
        "UMAG": "I+D+i",
        "CFLI": "Mantenimiento",
    }


def test_leer_tipo_proyecto_usa_el_mas_frecuente_si_hay_inconsistencia(tmp_path):
    ruta = _crear_excel_cc_con_tipo_proyecto(tmp_path, [
        ("UMAG-001", "I+D+i"),
        ("UMAG-002", "I+D+i"),
        ("UMAG-003", "Mantenimiento"),
    ])
    assert af.leer_tipo_proyecto_centro_costos(ruta) == {"UMAG": "I+D+i"}


def test_leer_tipo_proyecto_ignora_filas_sin_tipo(tmp_path):
    ruta = _crear_excel_cc_con_tipo_proyecto(tmp_path, [
        ("UMAG-001", None),
        ("UMAG-002", "I+D+i"),
    ])
    assert af.leer_tipo_proyecto_centro_costos(ruta) == {"UMAG": "I+D+i"}


def test_asegurar_categoria_proyectos_escribe_valor_y_avisa_si_falta(tmp_path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Proyectos"
    ws.cell(row=1, column=1, value="TAG proyecto")
    ws.cell(row=2, column=1, value="UMAG")
    ws.cell(row=3, column=1, value="MLER")
    col_categoria = 5  # columna arbitraria de prueba, no depende del archivo real
    filas_validas = [
        {"fila": 2, "tag": "UMAG", "nombre": "UMAG"},
        {"fila": 3, "tag": "MLER", "nombre": "Microturbina LER"},
    ]
    avisos = af.asegurar_categoria_proyectos(
        ws, filas_validas, {"UMAG": "I+D+i"}, columna=col_categoria,
    )
    assert ws.cell(row=2, column=col_categoria).value == "I+D+i"
    assert ws.cell(row=3, column=col_categoria).value is None
    assert len(avisos) == 1
    assert "MLER" in avisos[0]


def test_asegurar_categoria_proyectos_limpia_celda_obsoleta(tmp_path):
    """Cuando un proyecto pierde sus documentos en Centro de Costos,
    la celda Categoría debe limpiarse (None), no quedar con valor stale."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Proyectos"
    ws.cell(row=1, column=1, value="TAG proyecto")
    ws.cell(row=2, column=1, value="UMAG")
    col_categoria = 5
    # Simular una ejecución anterior donde UMAG tenía categoría
    ws.cell(row=2, column=col_categoria, value="I+D+i")

    filas_validas = [
        {"fila": 2, "tag": "UMAG", "nombre": "UMAG"},
    ]
    # En esta ejecución, UMAG ya no aparece en categoria_por_prefijo
    # (p. ej., todos sus documentos fueron eliminados o reclasificados)
    avisos = af.asegurar_categoria_proyectos(
        ws, filas_validas, {}, columna=col_categoria,
    )

    # La celda debe ser limpiada, no quedarse con el valor "I+D+i"
    assert ws.cell(row=2, column=col_categoria).value is None
    assert len(avisos) == 1
    assert "UMAG" in avisos[0]
    assert "Categoría queda vacía" in avisos[0]
