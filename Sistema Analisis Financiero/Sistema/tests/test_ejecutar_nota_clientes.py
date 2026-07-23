import json

import openpyxl

import analisis_financiero as af


def _armar_centro_costos(tmp_path):
    """Centro de Costos.xlsx mínimo con una hoja 'Detalle' vacía -- basta
    para que ejecutar() no aborte por archivo faltante."""
    ruta_cc = tmp_path / "Centro de Costos.xlsx"
    wb_cc = openpyxl.Workbook()
    ws = wb_cc.active
    ws.title = "Detalle"
    ws.append(["N° Ref.", "Proyecto", "Tipo de Proyecto", "N° Documento",
               "Nombre Ítem", "Descripción", "Categoría Ítem", "Cantidad",
               "P. Unitario sin IVA", "Total sin IVA (CLP)", "Total con IVA (CLP)"])
    wb_cc.save(ruta_cc)
    return ruta_cc


def test_ejecutar_completa_cliente_y_genera_hojas_clientes_y_glosario(tmp_path):
    ruta_af = tmp_path / "Análisis de Proyectos.xlsx"
    ruta_cc = _armar_centro_costos(tmp_path)
    ruta_pendientes = tmp_path / "clientes_pendientes.json"
    wb = af.asegurar_estructura_workbook(ruta_af)
    ws = wb[af.HOJA_PROYECTOS]
    ws.cell(row=2, column=1, value="AGCI1")
    ws.cell(row=2, column=2, value="AGCID Febrero")
    ws.cell(row=2, column=6, value=1_000_000)
    wb.save(ruta_af)

    resumen = af.ejecutar(
        ruta_excel_af=ruta_af,
        ruta_excel_cc=ruta_cc,
        raiz_facturas_cc=tmp_path / "Facturas y Boletas",
        raiz_respaldos=tmp_path / "Respaldos",
        ruta_clientes_pendientes=ruta_pendientes,
    )

    assert resumen["error"] is None
    wb_final = openpyxl.load_workbook(ruta_af)
    ws_proyectos = wb_final[af.HOJA_PROYECTOS]
    col_cliente = af.HEADERS_PROYECTOS.index("Cliente") + 1
    assert ws_proyectos.cell(row=2, column=col_cliente).value == "AGCID Febrero"

    ws_clientes = wb_final[af.HOJA_CLIENTES]
    assert ws_clientes.cell(row=2, column=1).value == "AGCID Febrero"

    ws_glosario = wb_final[af.HOJA_GLOSARIO_KPIS]
    assert ws_glosario.max_row == 1 + len(af.GLOSARIO_KPIS)


def test_resumen_incluye_clientes_pendientes_nuevos(tmp_path):
    ruta_af = tmp_path / "Análisis de Proyectos.xlsx"
    ruta_cc = _armar_centro_costos(tmp_path)
    ruta_pendientes = tmp_path / "clientes_pendientes.json"
    wb = af.asegurar_estructura_workbook(ruta_af)
    ws = wb[af.HOJA_PROYECTOS]
    ws.cell(row=2, column=1, value="AGCI1")
    ws.cell(row=2, column=2, value="AGCID Febrero")
    ws.cell(row=3, column=1, value="AGCI2")
    ws.cell(row=3, column=2, value="AGCID Marzo")
    wb.save(ruta_af)

    resumen = af.ejecutar(
        ruta_excel_af=ruta_af,
        ruta_excel_cc=ruta_cc,
        raiz_facturas_cc=tmp_path / "Facturas y Boletas",
        raiz_respaldos=tmp_path / "Respaldos",
        ruta_clientes_pendientes=ruta_pendientes,
    )

    assert len(resumen["clientes_pendientes"]) == 1
    assert resumen["clientes_pendientes"][0]["tag"] == "AGCI2"
    assert ruta_pendientes.exists()


def test_dry_run_no_escribe_clientes_pendientes_json(tmp_path):
    ruta_af = tmp_path / "Análisis de Proyectos.xlsx"
    ruta_cc = _armar_centro_costos(tmp_path)
    ruta_pendientes = tmp_path / "clientes_pendientes.json"
    wb = af.asegurar_estructura_workbook(ruta_af)
    ws = wb[af.HOJA_PROYECTOS]
    ws.cell(row=2, column=1, value="AGCI1")
    ws.cell(row=2, column=2, value="AGCID Febrero")
    ws.cell(row=3, column=1, value="AGCI2")
    ws.cell(row=3, column=2, value="AGCID Marzo")
    wb.save(ruta_af)

    af.ejecutar(
        ruta_excel_af=ruta_af,
        ruta_excel_cc=ruta_cc,
        raiz_facturas_cc=tmp_path / "Facturas y Boletas",
        raiz_respaldos=tmp_path / "Respaldos",
        ruta_clientes_pendientes=ruta_pendientes,
        dry_run=True,
    )

    assert not ruta_pendientes.exists()
