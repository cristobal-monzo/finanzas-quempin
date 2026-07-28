import openpyxl

import analisis_financiero as af


def test_crea_las_3_hojas_con_encabezados_si_no_existe_el_archivo(tmp_path):
    ruta = tmp_path / "Análisis de Proyectos.xlsx"
    wb = af.asegurar_estructura_workbook(ruta)

    assert af.HOJA_PROYECTOS in wb.sheetnames
    assert af.HOJA_DETALLE_COSTOS_REALES in wb.sheetnames
    assert af.HOJA_INDICADORES in wb.sheetnames

    ws = wb[af.HOJA_PROYECTOS]
    assert ws.cell(row=1, column=1).value == "TAG proyecto"
    assert ws.cell(row=1, column=2).value == "Nombre del proyecto"
    assert ws.cell(row=1, column=3).value == "Cliente"
    ultima_col = len(af.HEADERS_PROYECTOS)
    assert ws.cell(row=1, column=ultima_col).value == "Desviación % (Real vs Proyectado)"


def test_elimina_hoja1_vacia_al_migrar_un_archivo_existente(tmp_path):
    ruta = tmp_path / "Análisis de Proyectos.xlsx"
    wb_previo = openpyxl.Workbook()
    wb_previo.active.title = "Hoja1"
    wb_previo.save(ruta)

    wb = af.asegurar_estructura_workbook(ruta)

    assert "Hoja1" not in wb.sheetnames
    assert af.HOJA_PROYECTOS in wb.sheetnames


def test_no_toca_una_hoja_proyectos_que_ya_existe(tmp_path):
    ruta = tmp_path / "Análisis de Proyectos.xlsx"
    wb_previo = openpyxl.Workbook()
    ws_previo = wb_previo.active
    ws_previo.title = af.HOJA_PROYECTOS
    ws_previo.cell(row=2, column=1, value="UMAG")
    ws_previo.cell(row=2, column=2, value="UMAG")
    wb_previo.save(ruta)

    wb = af.asegurar_estructura_workbook(ruta)

    ws = wb[af.HOJA_PROYECTOS]
    assert ws.cell(row=2, column=1).value == "UMAG"


def test_agrega_encabezados_nuevos_a_una_hoja_existente_sin_tocar_los_viejos(tmp_path):
    ruta = tmp_path / "Análisis de Proyectos.xlsx"
    wb_previo = openpyxl.Workbook()
    ws_previo = wb_previo.active
    ws_previo.title = af.HOJA_PROYECTOS
    # Simula un archivo real al que le falta el último header del esquema
    # actual -- tiene todos menos el último.
    ultima_col = len(af.HEADERS_PROYECTOS)
    for col, encabezado in enumerate(af.HEADERS_PROYECTOS[:-1], start=1):
        ws_previo.cell(row=1, column=col, value=encabezado)
    ws_previo.cell(row=2, column=1, value="UMAG")
    ws_previo.cell(row=2, column=2, value="UMAG")
    wb_previo.save(ruta)

    wb = af.asegurar_estructura_workbook(ruta)

    ws = wb[af.HOJA_PROYECTOS]
    assert ws.cell(row=1, column=1).value == "TAG proyecto"
    assert ws.cell(row=1, column=ultima_col).value == "Desviación % (Real vs Proyectado)"
    assert ws.cell(row=2, column=1).value == "UMAG"


def test_hoja_indicadores_con_esquema_viejo_se_reescribe_completa_al_esquema_nuevo(tmp_path):
    """Regresión del bug real encontrado 2026-07-28: 'Indicadores' es 100%
    regenerada (datos Y encabezados) -- si el archivo en disco tiene un
    esquema de KPIs viejo (menos columnas, distinto orden), el encabezado
    debe reescribirse completo para no quedar desalineado con las fórmulas
    del esquema nuevo. Simula el Excel real antes del fix: 18 columnas
    viejas (con 'Rentabilidad sobre costo'/'Productividad *' en 3-8, 'Nota
    del Proyecto'/'Evaluación' en 17-18)."""
    ruta = tmp_path / "Análisis de Proyectos.xlsx"
    wb_previo = openpyxl.Workbook()
    ws_previo = wb_previo.active
    ws_previo.title = af.HOJA_INDICADORES
    headers_viejos = [
        "TAG proyecto", "Nombre del proyecto", "Rentabilidad sobre costo",
        "Margen neto %", "Productividad Materiales", "Productividad Equipos",
        "Productividad MO", "Productividad Otros", "Costo Materiales % de venta",
        "Costo Equipos % de venta", "Costo MO % de venta", "Costo Otros % de venta",
        "Desviación % Materiales", "Desviación % Equipos", "Desviación % MO",
        "Desviación % Otros", "Nota del Proyecto", "Evaluación",
    ]
    for col, encabezado in enumerate(headers_viejos, start=1):
        ws_previo.cell(row=1, column=col, value=encabezado)
    wb_previo.save(ruta)

    wb = af.asegurar_estructura_workbook(ruta)

    ws = wb[af.HOJA_INDICADORES]
    nombres = [ws.cell(row=1, column=c).value for c in range(1, len(af.HEADERS_INDICADORES) + 1)]
    assert nombres == af.HEADERS_INDICADORES
    assert "Rentabilidad sobre costo" not in nombres
    # no debe quedar ningún resto del esquema viejo más allá de la última
    # columna del esquema nuevo tampoco (acá el nuevo es más largo, no aplica,
    # pero se verifica que no haya columnas extra con basura vieja).
    assert ws.cell(row=1, column=len(af.HEADERS_INDICADORES) + 1).value is None


def test_hoja_proyectos_incluye_columna_cliente_entre_nombre_y_categoria(tmp_path):
    ruta = tmp_path / "Análisis de Proyectos.xlsx"
    wb = af.asegurar_estructura_workbook(ruta)

    ws = wb[af.HOJA_PROYECTOS]
    assert ws.cell(row=1, column=2).value == "Nombre del proyecto"
    assert ws.cell(row=1, column=3).value == "Cliente"
    assert ws.cell(row=1, column=4).value == "Categoría"
    assert ws.cell(row=1, column=5).value == "Estado"


def test_hoja_indicadores_incluye_nota_y_evaluacion(tmp_path):
    ruta = tmp_path / "Análisis de Proyectos.xlsx"
    wb = af.asegurar_estructura_workbook(ruta)

    ws = wb[af.HOJA_INDICADORES]
    assert ws.cell(row=1, column=22).value == "Nota del Proyecto"
    assert ws.cell(row=1, column=23).value == "Evaluación"


def test_hoja_indicadores_no_incluye_los_5_kpis_redundantes_eliminados(tmp_path):
    ruta = tmp_path / "Análisis de Proyectos.xlsx"
    wb = af.asegurar_estructura_workbook(ruta)

    ws = wb[af.HOJA_INDICADORES]
    nombres = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
    for eliminado in (
        "Rentabilidad sobre costo", "Productividad Materiales",
        "Productividad Equipos", "Productividad MO", "Productividad Otros",
    ):
        assert eliminado not in nombres


def test_hoja_indicadores_incluye_los_4_kpis_nuevos(tmp_path):
    ruta = tmp_path / "Análisis de Proyectos.xlsx"
    wb = af.asegurar_estructura_workbook(ruta)

    ws = wb[af.HOJA_INDICADORES]
    nombres = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
    for nuevo in (
        "Desviación % Total", "Estructura % Materiales",
        "Ahorro/Sobrecosto Materiales", "Ahorro/Sobrecosto Total",
    ):
        assert nuevo in nombres


def test_hoja_indicadores_incluye_los_2_kpis_nuevos_2026_07_28(tmp_path):
    ruta = tmp_path / "Análisis de Proyectos.xlsx"
    wb = af.asegurar_estructura_workbook(ruta)

    ws = wb[af.HOJA_INDICADORES]
    assert ws.cell(row=1, column=24).value == "Peso del proyecto en la cartera de ventas (%)"
    assert ws.cell(row=1, column=25).value == "Margen por día de ejecución"
    assert ws.cell(row=1, column=26).value is None


def test_hoja_detalle_costos_reales_incluye_columna_de_porcentaje(tmp_path):
    ruta = tmp_path / "Análisis de Proyectos.xlsx"
    wb = af.asegurar_estructura_workbook(ruta)

    ws = wb[af.HOJA_DETALLE_COSTOS_REALES]
    assert ws.cell(row=1, column=5).value == "% del Total Real del proyecto"


def test_crea_hoja_clientes_con_encabezados(tmp_path):
    ruta = tmp_path / "Análisis de Proyectos.xlsx"
    wb = af.asegurar_estructura_workbook(ruta)

    assert af.HOJA_CLIENTES in wb.sheetnames
    ws = wb[af.HOJA_CLIENTES]
    assert ws.cell(row=1, column=1).value == "Cliente"
    assert ws.cell(row=1, column=7).value == "CLTV"
    assert ws.cell(row=1, column=8).value == "Clasificación"


def test_crea_hoja_glosario_kpis_con_encabezados(tmp_path):
    ruta = tmp_path / "Análisis de Proyectos.xlsx"
    wb = af.asegurar_estructura_workbook(ruta)

    assert af.HOJA_GLOSARIO_KPIS in wb.sheetnames
    ws = wb[af.HOJA_GLOSARIO_KPIS]
    assert ws.cell(row=1, column=1).value == "KPI"
    assert ws.cell(row=1, column=4).value == "Qué significa el resultado"
