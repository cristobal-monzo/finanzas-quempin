# -*- coding: utf-8 -*-
import openpyxl

import analisis_financiero as af


def _crear_excel_cc_con_proyecto(tmp_path, filas):
    """filas: lista de tuplas (n_ref, nombre_proyecto)"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Master"
    encabezados = [
        "N° Ref.", "Proyecto", "Tipo de Proyecto", "Fecha", "N° Documento",
        "Tipo Documento", "Proveedor", "Proveedor (Razón Social)", "Categoría",
        "Resumen Ítems", "Total sin IVA (CLP)", "IVA 19% (CLP)",
        "Total con IVA (CLP)", "Estado", "Archivo origen", "Fecha modificación",
    ]
    for col, encabezado in enumerate(encabezados, start=1):
        ws.cell(row=1, column=col, value=encabezado)
    for fila_idx, (n_ref, nombre_proyecto) in enumerate(filas, start=2):
        ws.cell(row=fila_idx, column=1, value=n_ref)
        ws.cell(row=fila_idx, column=2, value=nombre_proyecto)
    ruta = tmp_path / "Centro de Costos.xlsx"
    wb.save(ruta)
    return ruta


def test_leer_nombres_proyecto_lee_un_nombre_por_prefijo(tmp_path):
    ruta = _crear_excel_cc_con_proyecto(tmp_path, [
        ("UMAG-001", "UMAG"),
        ("UMAG-002", "UMAG"),
        ("CFLI-001", "Cesfam Limache"),
    ])
    assert af.leer_nombres_proyecto_centro_costos(ruta) == {
        "UMAG": "UMAG",
        "CFLI": "Cesfam Limache",
    }


def test_leer_nombres_proyecto_ignora_filas_sin_nombre(tmp_path):
    ruta = _crear_excel_cc_con_proyecto(tmp_path, [
        ("UMAG-001", None),
        ("UMAG-002", "UMAG"),
    ])
    assert af.leer_nombres_proyecto_centro_costos(ruta) == {"UMAG": "UMAG"}


def test_leer_nombres_proyecto_sin_hoja_master_devuelve_vacio(tmp_path):
    wb = openpyxl.Workbook()
    wb.active.title = "Detalle"
    ruta = tmp_path / "Centro de Costos.xlsx"
    wb.save(ruta)
    assert af.leer_nombres_proyecto_centro_costos(ruta) == {}


def _ws_proyectos_con_filas(tmp_path, filas):
    wb = af.asegurar_estructura_workbook(tmp_path / "Análisis de Proyectos.xlsx")
    ws = wb[af.HOJA_PROYECTOS]
    for idx, (tag, nombre) in enumerate(filas, start=2):
        ws.cell(row=idx, column=1, value=tag)
        ws.cell(row=idx, column=2, value=nombre)
    return ws


def test_crea_fila_para_prefijo_nuevo(tmp_path):
    ws = _ws_proyectos_con_filas(tmp_path, [("UMAG", "UMAG")])
    filas_validas = [{"fila": 2, "tag": "UMAG", "nombre": "UMAG"}]

    nuevas = af.crear_filas_proyectos_nuevos(ws, filas_validas, {"CFLI": "Cesfam Limache"})

    assert nuevas == [{"fila": 3, "tag": "CFLI", "nombre": "Cesfam Limache"}]
    assert ws.cell(row=3, column=1).value == "CFLI"
    assert ws.cell(row=3, column=2).value == "Cesfam Limache"


def test_no_crea_nada_si_el_diccionario_esta_vacio(tmp_path):
    ws = _ws_proyectos_con_filas(tmp_path, [("UMAG", "UMAG")])
    filas_validas = [{"fila": 2, "tag": "UMAG", "nombre": "UMAG"}]

    nuevas = af.crear_filas_proyectos_nuevos(ws, filas_validas, {})

    assert nuevas == []


def test_fila_nueva_deja_las_demas_columnas_en_blanco(tmp_path):
    ws = _ws_proyectos_con_filas(tmp_path, [])
    nuevas = af.crear_filas_proyectos_nuevos(ws, [], {"UMAG": "UMAG"})
    assert nuevas == [{"fila": 2, "tag": "UMAG", "nombre": "UMAG"}]
    # Cliente, Categoría, Estado, etc. deben quedar sin tocar (None) -- las
    # llena el resto del pipeline (autocompletado) o el usuario a mano.
    assert ws.cell(row=2, column=3).value is None
    assert ws.cell(row=2, column=5).value is None


def test_varios_prefijos_nuevos_se_crean_en_orden_alfabetico(tmp_path):
    ws = _ws_proyectos_con_filas(tmp_path, [])
    nuevas = af.crear_filas_proyectos_nuevos(
        ws, [], {"CVAL": "Caldera Valdivia", "BWIL": "Bomba Wilo Conchalí"},
    )
    assert [n["tag"] for n in nuevas] == ["BWIL", "CVAL"]
    assert ws.cell(row=2, column=1).value == "BWIL"
    assert ws.cell(row=3, column=1).value == "CVAL"
