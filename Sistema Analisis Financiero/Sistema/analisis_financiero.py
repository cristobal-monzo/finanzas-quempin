# -*- coding: utf-8 -*-
"""
analisis_financiero.py -- Consolidador de costos reales por proyecto para
QUEMPIN SpA. Lee Centro de Costos.xlsx (SOLO LECTURA, nunca lo escribe) y
mantiene Análisis de Proyectos.xlsx (3 hojas: Proyectos, Detalle Costos
Reales, Indicadores). Ver docs/superpowers/specs/2026-07-20-analisis-
financiero-design.md para el diseño completo.
"""

import json
import math
import shutil
import sys
import unicodedata
from collections import Counter
from datetime import datetime
from difflib import SequenceMatcher
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.styles.colors import Color
from openpyxl.utils import get_column_letter

# ── CONFIGURACIÓN ────────────────────────────────────────────────────────────

RAIZ = Path(__file__).resolve().parent
RAIZ_MODULO = RAIZ.parent
# Reorganizado 2026-07-21: el código/skill vive en "Sistema Analisis
# Financiero/" (esta carpeta), separado de "Análisis Financiero/" que
# contiene solo el Excel de trabajo -- ambas son carpetas hermanas bajo la
# raíz de Finanzas QUEMPIN/. RAIZ_DATOS apunta a la carpeta con el Excel.
RAIZ_DATOS = RAIZ_MODULO.parent / "Análisis Financiero"
RUTA_EXCEL = RAIZ_DATOS / "Análisis de Proyectos.xlsx"
RAIZ_RESPALDOS = RAIZ_MODULO / "Respaldos"
RUTA_CLIENTES_PENDIENTES = RAIZ / "clientes_pendientes.json"

RAIZ_CENTRO_COSTOS = RAIZ_MODULO.parent / "Centro de Costos"
RUTA_EXCEL_CENTRO_COSTOS = RAIZ_CENTRO_COSTOS / "Excel" / "Centro de Costos.xlsx"
RAIZ_FACTURAS_CENTRO_COSTOS = (
    RAIZ_CENTRO_COSTOS / "Sitio de comunicación - Centro de Costos 1" / "Facturas y Boletas"
)

RAIZ_VISUALIZADOR_WEB_AF = RAIZ_MODULO / "Visualizador Web"

HOJA_PROYECTOS = "Proyectos"
HOJA_DETALLE_COSTOS_REALES = "Detalle Costos Reales"
HOJA_INDICADORES = "Indicadores"
HOJA_CLIENTES = "Clientes"
HOJA_GLOSARIO_KPIS = "Glosario KPIs"

HEADERS_PROYECTOS = [
    "TAG proyecto", "Nombre del proyecto", "Cliente", "Categoría", "Estado",
    "Fecha de inicio", "Fecha de cierre", "Monto de Venta (sin IVA)",
    "Costos Materiales Proyectados", "Costos Equipos Proyectados",
    "Mano de Obra Proyectada", "Otros Costos Proyectados",
    "Costos Materiales Reales", "Costos Equipos Reales",
    "Otros Costos Reales", "Mano de Obra Real", "Total Proyectado",
    "Total Real", "Margen Proyectado", "Margen Real",
    "Desviación % (Real vs Proyectado)",
]
HEADERS_DETALLE_COSTOS_REALES = [
    "TAG proyecto", "Subcategoría", "Bucket", "Total sin IVA",
    "% del Total Real del proyecto",
]
# Playbook depurado 2026-07-28: se eliminaron "Rentabilidad sobre costo" y las
# 4 "Productividad" (Materiales/Equipos/MO/Otros) -- eran redundancias
# matemáticas exactas de "Margen neto %" y de "Costo % de venta" (ver
# CLAUDE.md, sección "Playbook de KPIs", y MEMORY.md 2026-07-28). Se agregaron
# 4 KPIs nuevos: Estructura % del costo real (mix, suma 100%), Desviación %
# Total (ya existía en "Proyectos", ahora también visible acá), y
# Ahorro/Sobrecosto neto en $ por categoría y total.
HEADERS_INDICADORES = [
    "TAG proyecto", "Nombre del proyecto", "Margen neto %",
    "Costo Materiales % de venta", "Costo Equipos % de venta",
    "Costo MO % de venta", "Costo Otros % de venta",
    "Estructura % Materiales", "Estructura % Equipos", "Estructura % MO",
    "Estructura % Otros",
    "Desviación % Materiales", "Desviación % Equipos", "Desviación % MO",
    "Desviación % Otros", "Desviación % Total",
    "Ahorro/Sobrecosto Materiales", "Ahorro/Sobrecosto Equipos",
    "Ahorro/Sobrecosto MO", "Ahorro/Sobrecosto Otros",
    "Ahorro/Sobrecosto Total",
    "Nota del Proyecto", "Evaluación",
    # 2 KPIs nuevos agregados 2026-07-28 (aprobados por el usuario; un
    # tercero propuesto, "Cumplimiento de plazo", quedó fuera a propósito
    # porque necesitaba un dato manual nuevo que el usuario decidió no
    # agregar por ahora -- ver MEMORY.md).
    "Peso del proyecto en la cartera de ventas (%)",
    "Margen por día de ejecución",
]
HEADERS_CLIENTES = [
    "Cliente", "AOV (Valor promedio de venta)",
    "Vida del cliente (n° de proyectos)", "Meses activo",
    "Frecuencia de compra (proyectos/año)", "Margen de utilidad %", "CLTV",
    "Clasificación",
]
HEADERS_GLOSARIO_KPIS = [
    "KPI", "Por qué importa", "Qué elementos usa", "Qué significa el resultado",
]

# Letra de columna de "Proyectos" por nombre de encabezado -- centraliza el
# mapeo para que las fórmulas nunca hardcodeen letras directamente (si el
# orden de HEADERS_PROYECTOS cambia, las fórmulas se recalculan solas).
LETRA_COL_PROYECTOS = {
    nombre: get_column_letter(idx) for idx, nombre in enumerate(HEADERS_PROYECTOS, start=1)
}
# Mismo patrón para "Indicadores" -- las fórmulas de Nota/Evaluación
# referencian columnas de esta misma hoja (no de "Proyectos") y no pueden
# hardcodear la letra, para no romperse si el playbook de KPIs cambia de
# nuevo (ya pasó una vez, 2026-07-28: se sacaron 5 columnas y se agregaron 9).
LETRA_COL_INDICADORES = {
    nombre: get_column_letter(idx) for idx, nombre in enumerate(HEADERS_INDICADORES, start=1)
}


# ── ESTILO VISUAL ────────────────────────────────────────────────────────────
# Formato que el usuario armó a mano en "Proyectos" (encabezado en negrita,
# centrado, con wrap, relleno por color de theme según grupo de columna,
# formato moneda/porcentaje) -- replicado acá para que se mantenga en cada
# corrida y se extienda a "Detalle Costos Reales" e "Indicadores", que hasta
# 2026-07-21 quedaban 100% regeneradas sin ningún estilo.

ALTURA_FILA_ENCABEZADO = 46.2
FUENTE_ENCABEZADO = Font(name="Calibri", size=11, bold=True)
ALINEACION_ENCABEZADO = Alignment(vertical="center", wrap_text=True)

FORMATO_MONEDA = '_ "$"* #,##0_ ;_ "$"* \\-#,##0_ ;_ "$"* "-"_ ;_ @_ '
FORMATO_PORCENTAJE = "0.0%"
FORMATO_RATIO = "0.00"
FORMATO_ENTERO = "0"
# Formato fijo DD-MM-AAAA para columnas de fecha -- pedido del usuario 2026-07-28,
# mismo estandar que Centro de Costos (ver DATE_FORMAT en auditor_centro_costos.py).
FORMATO_FECHA = "DD-MM-YYYY"


def _color_tema(theme: int, tint: float) -> Color:
    return Color(theme=theme, tint=tint)


# Mismos 4 colores que ya usaba "Proyectos" a mano, uno por grupo semántico
# de columna -- se reutilizan en las 3 hojas para que el libro se lea como
# un solo sistema visual.
COLOR_IDENTIFICACION = _color_tema(5, 0.3999755851924192)    # TAG/Nombre/Estado/fechas/venta
COLOR_COSTO_PROYECTADO = _color_tema(8, 0.3999755851924192)  # columnas "...Proyectado(s)"
COLOR_COSTO_REAL = _color_tema(9, 0.3999755851924192)        # columnas "...Real(es)"
COLOR_DERIVADO = _color_tema(3, 0.499984740745262)           # totales, márgenes, KPIs finales

# Por nombre de encabezado, no por letra -- se convierte a letra-keyed más
# abajo vía LETRA_COL_PROYECTOS. Evita el bug real que ya pasó el 2026-07-28
# al reordenar el playbook de Indicadores (ver asegurar_estructura_workbook):
# un dict hardcodeado por letra queda mal alineado apenas cambia el orden de
# HEADERS_PROYECTOS (ej. al mover "Categoría" junto a "Cliente").
ESTILO_COLUMNAS_PROYECTOS_POR_NOMBRE = {
    "TAG proyecto": (COLOR_IDENTIFICACION, None, 10),
    "Nombre del proyecto": (COLOR_IDENTIFICACION, None, 22),
    "Cliente": (COLOR_IDENTIFICACION, None, 22),
    "Categoría": (COLOR_IDENTIFICACION, None, 16),
    "Estado": (COLOR_IDENTIFICACION, None, 13),
    "Fecha de inicio": (COLOR_IDENTIFICACION, FORMATO_FECHA, 13),
    "Fecha de cierre": (COLOR_IDENTIFICACION, FORMATO_FECHA, 13),
    "Monto de Venta (sin IVA)": (COLOR_IDENTIFICACION, FORMATO_MONEDA, 16),
    "Costos Materiales Proyectados": (COLOR_COSTO_PROYECTADO, FORMATO_MONEDA, 14),
    "Costos Equipos Proyectados": (COLOR_COSTO_PROYECTADO, FORMATO_MONEDA, 14),
    "Mano de Obra Proyectada": (COLOR_COSTO_PROYECTADO, FORMATO_MONEDA, 14),
    "Otros Costos Proyectados": (COLOR_COSTO_PROYECTADO, FORMATO_MONEDA, 14),
    "Costos Materiales Reales": (COLOR_COSTO_REAL, FORMATO_MONEDA, 14),
    "Costos Equipos Reales": (COLOR_COSTO_REAL, FORMATO_MONEDA, 14),
    "Otros Costos Reales": (COLOR_COSTO_REAL, FORMATO_MONEDA, 14),
    "Mano de Obra Real": (COLOR_COSTO_REAL, FORMATO_MONEDA, 14),
    "Total Proyectado": (COLOR_DERIVADO, FORMATO_MONEDA, 14),
    "Total Real": (COLOR_DERIVADO, FORMATO_MONEDA, 14),
    "Margen Proyectado": (COLOR_DERIVADO, FORMATO_MONEDA, 14),
    "Margen Real": (COLOR_DERIVADO, FORMATO_MONEDA, 14),
    "Desviación % (Real vs Proyectado)": (COLOR_DERIVADO, FORMATO_PORCENTAJE, 16),
}
# columna (letra) -> (color de encabezado, formato numérico o None, ancho sugerido)
ESTILO_COLUMNAS_PROYECTOS = {
    LETRA_COL_PROYECTOS[nombre]: estilo
    for nombre, estilo in ESTILO_COLUMNAS_PROYECTOS_POR_NOMBRE.items()
}

ESTILO_COLUMNAS_DETALLE_COSTOS_REALES = {
    "A": (COLOR_IDENTIFICACION, None, 10),
    "B": (COLOR_COSTO_REAL, None, 22),
    "C": (COLOR_COSTO_REAL, None, 13),
    "D": (COLOR_COSTO_REAL, FORMATO_MONEDA, 14),
    "E": (COLOR_DERIVADO, FORMATO_PORCENTAJE, 18),
}

# Reordenado 2026-07-28 junto con HEADERS_INDICADORES: A/B identificación:
# C margen neto; D-G costo % de venta; H-K estructura % del costo real
# (mix, suma 100%); L-P desviación % (4 categorías + total); Q-U
# ahorro/sobrecosto neto en $ (4 categorías + total); V/W nota/evaluación.
ESTILO_COLUMNAS_INDICADORES = {
    "A": (COLOR_IDENTIFICACION, None, 10),
    "B": (COLOR_IDENTIFICACION, None, 22),
    "C": (COLOR_DERIVADO, FORMATO_PORCENTAJE, 16),
    "D": (COLOR_COSTO_REAL, FORMATO_PORCENTAJE, 14),
    "E": (COLOR_COSTO_REAL, FORMATO_PORCENTAJE, 14),
    "F": (COLOR_COSTO_REAL, FORMATO_PORCENTAJE, 14),
    "G": (COLOR_COSTO_REAL, FORMATO_PORCENTAJE, 14),
    "H": (COLOR_COSTO_REAL, FORMATO_PORCENTAJE, 14),
    "I": (COLOR_COSTO_REAL, FORMATO_PORCENTAJE, 14),
    "J": (COLOR_COSTO_REAL, FORMATO_PORCENTAJE, 14),
    "K": (COLOR_COSTO_REAL, FORMATO_PORCENTAJE, 14),
    "L": (COLOR_COSTO_PROYECTADO, FORMATO_PORCENTAJE, 14),
    "M": (COLOR_COSTO_PROYECTADO, FORMATO_PORCENTAJE, 14),
    "N": (COLOR_COSTO_PROYECTADO, FORMATO_PORCENTAJE, 14),
    "O": (COLOR_COSTO_PROYECTADO, FORMATO_PORCENTAJE, 14),
    "P": (COLOR_COSTO_PROYECTADO, FORMATO_PORCENTAJE, 16),
    "Q": (COLOR_DERIVADO, FORMATO_MONEDA, 16),
    "R": (COLOR_DERIVADO, FORMATO_MONEDA, 16),
    "S": (COLOR_DERIVADO, FORMATO_MONEDA, 16),
    "T": (COLOR_DERIVADO, FORMATO_MONEDA, 16),
    "U": (COLOR_DERIVADO, FORMATO_MONEDA, 16),
    "V": (COLOR_DERIVADO, FORMATO_ENTERO, 12),
    "W": (COLOR_DERIVADO, None, 20),
    "X": (COLOR_DERIVADO, FORMATO_PORCENTAJE, 18),
    "Y": (COLOR_DERIVADO, FORMATO_MONEDA, 18),
}

ESTILO_COLUMNAS_CLIENTES = {
    "A": (COLOR_IDENTIFICACION, None, 22),
    "B": (COLOR_DERIVADO, FORMATO_MONEDA, 18),
    "C": (COLOR_DERIVADO, FORMATO_ENTERO, 14),
    "D": (COLOR_DERIVADO, FORMATO_RATIO, 14),
    "E": (COLOR_DERIVADO, FORMATO_RATIO, 14),
    "F": (COLOR_DERIVADO, FORMATO_PORCENTAJE, 14),
    "G": (COLOR_DERIVADO, FORMATO_MONEDA, 18),
    "H": (COLOR_DERIVADO, None, 22),
}

ESTILO_COLUMNAS_GLOSARIO_KPIS = {
    "A": (COLOR_IDENTIFICACION, None, 30),
    "B": (COLOR_IDENTIFICACION, None, 50),
    "C": (COLOR_IDENTIFICACION, None, 40),
    "D": (COLOR_IDENTIFICACION, None, 50),
}


def aplicar_estilo_visual(wb) -> None:
    """Aplica a las 3 hojas el mismo lenguaje visual que el usuario armó a
    mano en 'Proyectos': encabezado en negrita, centrado y con wrap, alto de
    fila 46.2, relleno por color de theme según grupo de columna, y formato
    moneda/porcentaje/ratio en las columnas correspondientes. Solo toca
    estilo, nunca valores; el ancho de columna se deja intacto si el usuario
    ya lo fijó a mano (no se sobrescribe una vez seteado)."""
    for nombre_hoja, estilo_columnas in (
        (HOJA_PROYECTOS, ESTILO_COLUMNAS_PROYECTOS),
        (HOJA_DETALLE_COSTOS_REALES, ESTILO_COLUMNAS_DETALLE_COSTOS_REALES),
        (HOJA_INDICADORES, ESTILO_COLUMNAS_INDICADORES),
        (HOJA_CLIENTES, ESTILO_COLUMNAS_CLIENTES),
        (HOJA_GLOSARIO_KPIS, ESTILO_COLUMNAS_GLOSARIO_KPIS),
    ):
        ws = wb[nombre_hoja]
        ws.row_dimensions[1].height = ALTURA_FILA_ENCABEZADO
        for columna, (color, formato_numero, ancho) in estilo_columnas.items():
            # ws.column_dimensions[columna] autovivifica la entrada con un
            # width por defecto (13.0) apenas se accede -- por eso el "ya
            # tiene ancho manual" se revisa ANTES de tocar la columna, con
            # el operador "in" sobre el dict, no con ".width is None".
            ya_tenia_ancho_manual = columna in ws.column_dimensions
            celda_encabezado = ws[f"{columna}1"]
            celda_encabezado.font = FUENTE_ENCABEZADO
            celda_encabezado.alignment = ALINEACION_ENCABEZADO
            celda_encabezado.fill = PatternFill(fgColor=color, fill_type="solid")
            if formato_numero is not None:
                ws.column_dimensions[columna].number_format = formato_numero
            if not ya_tenia_ancho_manual:
                ws.column_dimensions[columna].width = ancho


# ── RESALTADO DE CELDAS DE INGRESO MANUAL ("Proyectos") ─────────────────────
# Pedido del usuario (2026-07-28): que las celdas que se llenan a mano nunca
# se confundan con las de fórmula/autocompletado. "Cliente" y "Categoría"
# quedan afuera aunque viven en el mismo bloque que nunca se reescribe entre
# corridas (asegurar_formulas_proyectos) -- se autocompletan solas
# (asegurar_columna_cliente / asegurar_categoria_proyectos) y ya tienen su
# propio lenguaje visual (rojo/azul marino). Las columnas L/M/N/P-T tampoco
# entran: son 100% fórmula (Materiales/Equipos/Otros Reales, Total
# Proyectado/Real, Margen Proyectado/Real, Desviación %).
NOMBRES_COLUMNAS_MANUALES_PROYECTOS = [
    "TAG proyecto", "Nombre del proyecto", "Estado", "Fecha de inicio",
    "Fecha de cierre", "Monto de Venta (sin IVA)",
    "Costos Materiales Proyectados", "Costos Equipos Proyectados",
    "Mano de Obra Proyectada", "Otros Costos Proyectados", "Mano de Obra Real",
]
COLOR_RESALTADO_MANUAL = "FFF2CC"  # amarillo pastel -- no se reutiliza en ningún otro relleno del libro
FUENTE_RESALTADO_MANUAL = Font(name="Calibri", size=11, italic=True)
FILAS_MINIMAS_RESALTADO_MANUAL = 60  # deja filas vacías pre-formateadas como plantilla aunque el libro no tenga proyectos cargados todavía
BUFFER_FILAS_RESALTADO_MANUAL = 20   # margen de filas nuevas pre-formateadas más allá de la última fila con datos reales
LEYENDA_RESALTADO_MANUAL = (
    "Relleno amarillo + cursiva = ingreso manual.\nSin relleno = fórmula o "
    "autocompletado (Cliente, Categoría) -- no editar a mano."
)


def migrar_formato_fecha_proyectos(wb) -> None:
    """Migracion de formato (idempotente, 2026-07-28, mismo patron que
    migrar_formato_fecha_corta de Centro de Costos): fuerza FORMATO_FECHA
    celda por celda en 'Fecha de inicio'/'Fecha de cierre' de TODAS las
    filas ya escritas de 'Proyectos'. Necesario porque el number_format a
    nivel de columna (aplicar_estilo_visual, via ws.column_dimensions) NO
    pisa el formato que Excel ya asigno a una celda cuando el usuario
    escribio la fecha a mano -- el formato por celda gana sobre el de
    columna para cualquier celda que ya tenga uno propio. Es formato, no
    contenido: no toca valores, misma excepcion a la regla de oro que el
    resaltado de celdas manuales de abajo."""
    ws = wb[HOJA_PROYECTOS]
    col_tag = LETRA_COL_PROYECTOS["TAG proyecto"]
    letras_fecha = [LETRA_COL_PROYECTOS["Fecha de inicio"], LETRA_COL_PROYECTOS["Fecha de cierre"]]

    for fila in range(2, ws.max_row + 1):
        if ws[f"{col_tag}{fila}"].value is None:
            continue
        for letra in letras_fecha:
            ws[f"{letra}{fila}"].number_format = FORMATO_FECHA


def aplicar_resaltado_celdas_manuales(wb) -> None:
    """Rellena de amarillo + cursiva las columnas de ingreso manual de
    'Proyectos' (ver NOMBRES_COLUMNAS_MANUALES_PROYECTOS), en un rango que
    siempre cubre al menos FILAS_MINIMAS_RESALTADO_MANUAL filas (para que el
    libro sirva de plantilla incluso sin proyectos cargados) y se extiende
    BUFFER_FILAS_RESALTADO_MANUAL más allá de la última fila con TAG real --
    calculado leyendo valores de la columna A, nunca ws.max_row a secas (que
    quedaría inflado por el relleno ya aplicado en corridas previas y
    crecería sin límite en cada corrida). Nunca toca valores, solo relleno/
    fuente -- se reaplica en cada corrida igual que aplicar_estilo_visual."""
    ws = wb[HOJA_PROYECTOS]
    col_tag = LETRA_COL_PROYECTOS["TAG proyecto"]

    ultima_fila_con_datos = 1
    for fila in range(2, ws.max_row + 1):
        if ws[f"{col_tag}{fila}"].value is not None:
            ultima_fila_con_datos = fila

    ultima_fila = max(
        FILAS_MINIMAS_RESALTADO_MANUAL, ultima_fila_con_datos + BUFFER_FILAS_RESALTADO_MANUAL
    )
    relleno = PatternFill(fgColor=COLOR_RESALTADO_MANUAL, fill_type="solid")
    letras = [LETRA_COL_PROYECTOS[nombre] for nombre in NOMBRES_COLUMNAS_MANUALES_PROYECTOS]

    for letra in letras:
        for fila in range(2, ultima_fila + 1):
            celda = ws[f"{letra}{fila}"]
            celda.fill = relleno
            celda.font = FUENTE_RESALTADO_MANUAL

    col_leyenda = len(HEADERS_PROYECTOS) + 1
    celda_leyenda = ws.cell(row=1, column=col_leyenda)
    celda_leyenda.value = LEYENDA_RESALTADO_MANUAL
    celda_leyenda.font = Font(name="Calibri", size=9, italic=True, color="808080")
    celda_leyenda.alignment = Alignment(wrap_text=True, vertical="center")
    letra_leyenda = get_column_letter(col_leyenda)
    if letra_leyenda not in ws.column_dimensions:
        ws.column_dimensions[letra_leyenda].width = 45


def asegurar_estructura_workbook(ruta_excel: Path) -> openpyxl.Workbook:
    """Abre ruta_excel si existe, o crea un libro nuevo. Garantiza que las 5
    hojas existan con encabezados en la fila 1.

    "Proyectos" es la única hoja con datos MANUALES del usuario: nunca se
    pisa un encabezado ya escrito ahí, solo se completan columnas nuevas al
    final del esquema (ej. "Cliente" en un archivo real creado antes de que
    esa columna existiera) -- regla de oro del módulo.

    Las otras 4 hojas ("Detalle Costos Reales"/"Indicadores"/"Clientes"/
    "Glosario KPIs") son 100% regeneradas cada corrida -- sus DATOS ya se
    reescriben completos en cada `ejecutar()` (ver regenerar_hoja_detalle_
    costos_reales / asegurar_hoja_indicadores / etc.), así que su fila de
    encabezados también se reescribe completa acá si no coincide con el
    esquema actual (incluye borrar columnas sobrantes si el esquema nuevo
    tiene menos columnas que el anterior). **Bug real encontrado y corregido
    2026-07-28**: antes esta función usaba el mismo criterio "solo llenar
    vacíos" para las 5 hojas -- al reordenar/depurar el playbook de KPIs de
    "Indicadores" (quitar 5 columnas, agregar 9, reordenar el resto), los
    encabezados viejos quedaron pisando columnas con fórmulas del esquema
    NUEVO (ej. columna C decía "Rentabilidad sobre costo" pero tenía la
    fórmula de "Margen neto %"), y aparecieron "Nota del Proyecto"/
    "Evaluación" duplicados en las columnas nuevas del final -- un archivo
    financiero real quedó con encabezados y datos desalineados. Se corrió
    contra el Excel real antes de detectarse (restaurado desde backup) --
    ver MEMORY.md 2026-07-28 para el detalle completo del incidente."""
    if ruta_excel.exists():
        wb = openpyxl.load_workbook(ruta_excel)
    else:
        wb = openpyxl.Workbook()

    for nombre_hoja, headers in (
        (HOJA_PROYECTOS, HEADERS_PROYECTOS),
        (HOJA_DETALLE_COSTOS_REALES, HEADERS_DETALLE_COSTOS_REALES),
        (HOJA_INDICADORES, HEADERS_INDICADORES),
        (HOJA_CLIENTES, HEADERS_CLIENTES),
        (HOJA_GLOSARIO_KPIS, HEADERS_GLOSARIO_KPIS),
    ):
        ws = wb[nombre_hoja] if nombre_hoja in wb.sheetnames else wb.create_sheet(nombre_hoja)
        if nombre_hoja == HOJA_PROYECTOS:
            for col, encabezado in enumerate(headers, start=1):
                if ws.cell(row=1, column=col).value is None:
                    ws.cell(row=1, column=col, value=encabezado)
            continue

        max_col_previo = ws.max_column if ws.max_row >= 1 else 0
        for col in range(1, max(max_col_previo, len(headers)) + 1):
            nuevo_valor = headers[col - 1] if col <= len(headers) else None
            celda = ws.cell(row=1, column=col)
            if celda.value != nuevo_valor:
                celda.value = nuevo_valor

    for nombre_default in ("Hoja1", "Sheet"):
        if nombre_default in wb.sheetnames:
            ws_default = wb[nombre_default]
            esta_vacia = all(
                celda.value is None
                for fila in ws_default.iter_rows()
                for celda in fila
            )
            if esta_vacia:
                del wb[nombre_default]

    return wb


# ── MAPEO DE CATEGORÍAS ──────────────────────────────────────────────────────

MAPEO_CATEGORIA_BUCKET = {
    "Materiales": "Materiales",
    "Consumibles": "Materiales",
    "Equipos-Herramientas": "Equipos",
}


def mapear_categoria_a_bucket(categoria_item: str | None) -> tuple[str, bool]:
    """Devuelve (bucket, es_mapeo_explicito). Cualquier categoria_item que no
    esté en MAPEO_CATEGORIA_BUCKET (incluyendo None) cae en "Otros" con
    es_mapeo_explicito=False, para poder avisar sin perder el monto."""
    if categoria_item in MAPEO_CATEGORIA_BUCKET:
        return MAPEO_CATEGORIA_BUCKET[categoria_item], True
    return "Otros", False


# ── CLIENTE: DERIVACIÓN Y EMPAREJAMIENTO FUZZY ──────────────────────────────
# "Cliente" no existe como dato separado en Centro de Costos ni en "Proyectos"
# -- se deriva del "Nombre del proyecto" (que a veces mezcla cliente +
# iteración/fecha, ej. "AGCID (I) FEBRERO") y se compara contra los clientes
# ya registrados para detectar recurrencia. Nunca pregunta en vivo (ver
# asegurar_columna_cliente): si hay duda, marca "pendiente" para revisión
# posterior vía confirmar_clientes_pendientes.

UMBRAL_SIMILITUD_CLIENTE = 0.6


def derivar_cliente(nombre_proyecto: str) -> str:
    """Corta el nombre del proyecto en el primer paréntesis (donde suele
    empezar la iteración/fecha, ej. "AGCID (I) FEBRERO" -> "AGCID"). Sin
    paréntesis, devuelve el nombre completo tal cual."""
    candidato = nombre_proyecto.split("(")[0].strip()
    return candidato if candidato else nombre_proyecto.strip()


def normalizar_texto(texto: str) -> str:
    """Mayúsculas, sin tildes, sin espacios repetidos -- para comparar
    nombres de cliente sin que un tilde o un espacio extra genere un falso
    "pendiente"."""
    sin_tildes = "".join(
        c for c in unicodedata.normalize("NFD", texto)
        if unicodedata.category(c) != "Mn"
    )
    return " ".join(sin_tildes.upper().split())


def emparejar_cliente(candidato: str, clientes_existentes: list[str]) -> dict:
    """Compara candidato contra clientes_existentes. Devuelve un dict con
    "cliente" (el valor final a escribir), "estado" ("exacto"/"pendiente"/
    "nuevo") y "similitud". Coincidencia exacta tras normalizar -> "exacto"
    (usa el nombre YA registrado, no el candidato, para no crear variantes de
    capitalización del mismo cliente). Similar pero no exacta (>= umbral)
    -> "pendiente", usa el existente como sugerencia. Sin parecido -> "nuevo",
    usa el candidato tal cual."""
    candidato_norm = normalizar_texto(candidato)
    mejor_match = None
    mejor_similitud = 0.0
    for existente in clientes_existentes:
        if normalizar_texto(existente) == candidato_norm:
            return {"cliente": existente, "estado": "exacto", "similitud": 1.0}
        similitud = SequenceMatcher(None, candidato_norm, normalizar_texto(existente)).ratio()
        if similitud > mejor_similitud:
            mejor_similitud = similitud
            mejor_match = existente
    if mejor_match is not None and mejor_similitud >= UMBRAL_SIMILITUD_CLIENTE:
        return {"cliente": mejor_match, "estado": "pendiente", "similitud": mejor_similitud}
    return {"cliente": candidato, "estado": "nuevo", "similitud": 0.0}


# Mismos hex que Centro de Costos (auditor_centro_costos.py: ROJO="C00000",
# NAVY_OSCURO="1F3864") -- mismo lenguaje visual "rojo = revisar, azul marino
# = corregido a mano" en todo Finanzas QUEMPIN.
FUENTE_PENDIENTE_REVISION_CLIENTE = Font(name="Calibri", size=11, color="C00000")
FUENTE_CONFIRMADO_CLIENTE = Font(name="Calibri", size=11, color="1F3864")


def leer_clientes_pendientes(ruta_pendientes: Path) -> list[dict]:
    """Lee clientes_pendientes.json. Devuelve [] si el archivo no existe
    todavía (primera corrida)."""
    if not ruta_pendientes.exists():
        return []
    with open(ruta_pendientes, "r", encoding="utf-8") as f:
        return json.load(f)


def asegurar_columna_cliente(ws_proyectos, filas_validas: list[dict], ruta_pendientes: Path) -> list[dict]:
    """Completa la columna 'Cliente' de las filas válidas que la tengan
    vacía: deriva un candidato del nombre del proyecto y lo empareja contra
    los clientes ya asignados (incluyendo los que se van asignando en esta
    misma corrida). Si el emparejamiento queda 'pendiente', pinta la celda de
    rojo y agrega una entrada a clientes_pendientes.json -- nunca pregunta en
    vivo (este módulo corre encadenado y no bloqueante al run de Centro de
    Costos). Devuelve solo las entradas pendientes NUEVAS de esta corrida."""
    col_cliente = HEADERS_PROYECTOS.index("Cliente") + 1

    clientes_existentes = []
    for fila_info in filas_validas:
        valor = ws_proyectos.cell(row=fila_info["fila"], column=col_cliente).value
        if valor:
            clientes_existentes.append(valor)

    pendientes_nuevos = []
    for fila_info in filas_validas:
        r = fila_info["fila"]
        celda = ws_proyectos.cell(row=r, column=col_cliente)
        if celda.value:
            continue

        candidato = derivar_cliente(fila_info["nombre"])
        resultado = emparejar_cliente(candidato, clientes_existentes)
        celda.value = resultado["cliente"]

        if resultado["estado"] == "pendiente":
            celda.font = FUENTE_PENDIENTE_REVISION_CLIENTE
            pendientes_nuevos.append({
                "tag": fila_info["tag"],
                "fila": r,
                "nombre_proyecto": fila_info["nombre"],
                "cliente_derivado": candidato,
                "cliente_sugerido": resultado["cliente"],
                "similitud": round(resultado["similitud"], 2),
                "estado": "Pendiente",
            })

        clientes_existentes.append(resultado["cliente"])

    if pendientes_nuevos:
        pendientes_totales = leer_clientes_pendientes(ruta_pendientes) + pendientes_nuevos
        with open(ruta_pendientes, "w", encoding="utf-8") as f:
            json.dump(pendientes_totales, f, ensure_ascii=False, indent=2)

    return pendientes_nuevos


def confirmar_clientes_pendientes(
    objetivo=None,
    ruta_excel: Path | None = None,
    ruta_pendientes: Path | None = None,
    ruta_respaldos: Path | None = None,
) -> list[dict]:
    """objetivo=None -> preview de solo lectura (no toca nada). objetivo=
    "TODOS" o lista de TAGs -> aplica: escribe cliente_sugerido en la celda,
    recolorea azul marino, marca "Confirmado" en clientes_pendientes.json.
    Mismo contrato que confirmar_correcciones de Centro de Costos."""
    ruta_excel = ruta_excel or RUTA_EXCEL
    ruta_pendientes = ruta_pendientes or RUTA_CLIENTES_PENDIENTES
    ruta_respaldos = ruta_respaldos or RAIZ_RESPALDOS

    pendientes = leer_clientes_pendientes(ruta_pendientes)
    seleccion_pendiente = [p for p in pendientes if p["estado"] == "Pendiente"]

    if objetivo is None:
        return seleccion_pendiente

    seleccion = (
        seleccion_pendiente if objetivo == "TODOS"
        else [p for p in seleccion_pendiente if p["tag"] in set(objetivo)]
    )
    if not seleccion:
        return []

    hacer_backup(ruta_excel, ruta_respaldos)
    wb = openpyxl.load_workbook(ruta_excel)
    ws_proyectos = wb[HOJA_PROYECTOS]
    col_cliente = HEADERS_PROYECTOS.index("Cliente") + 1

    for p in seleccion:
        celda = ws_proyectos.cell(row=p["fila"], column=col_cliente)
        celda.value = p["cliente_sugerido"]
        celda.font = FUENTE_CONFIRMADO_CLIENTE
        p["estado"] = "Confirmado"

    wb.save(ruta_excel)
    with open(ruta_pendientes, "w", encoding="utf-8") as f:
        json.dump(pendientes, f, ensure_ascii=False, indent=2)

    return seleccion


# ── LECTURA DE CENTRO DE COSTOS (SOLO LECTURA) ───────────────────────────────

def prefijo_de_n_ref(n_ref: str) -> str:
    """'UMAG-001' -> 'UMAG'. Mismo prefijo que PREFIJOS_PROYECTO en
    Centro de Costos/Sistema/auditor_centro_costos.py."""
    return n_ref.split("-")[0]


def leer_detalle_centro_costos(ruta_excel_cc: Path) -> list[dict]:
    """Lee la hoja 'Detalle' de Centro de Costos.xlsx -- SOLO LECTURA, este
    módulo nunca escribe ese archivo. Filas sin N° Ref. o sin Total sin IVA
    se ignoran (no se puede agrupar ni sumar sin esos dos datos)."""
    wb = openpyxl.load_workbook(ruta_excel_cc, data_only=True)
    ws = wb["Detalle"]
    encabezados = [celda.value for celda in ws[1]]
    col_n_ref = encabezados.index("N° Ref.") + 1
    col_categoria = encabezados.index("Categoría Ítem") + 1
    col_total_sin_iva = encabezados.index("Total sin IVA (CLP)") + 1

    items = []
    for fila in ws.iter_rows(min_row=2):
        n_ref = fila[col_n_ref - 1].value
        total = fila[col_total_sin_iva - 1].value
        if n_ref is None or total is None:
            continue
        categoria = fila[col_categoria - 1].value
        items.append({"n_ref": n_ref, "categoria_item": categoria, "total_sin_iva": float(total)})
    return items


def leer_tipo_proyecto_centro_costos(ruta_excel_cc: Path) -> dict[str, str]:
    """Lee la hoja 'Master' de Centro de Costos.xlsx (SOLO LECTURA) y devuelve,
    por prefijo de proyecto, el 'Tipo de Proyecto' más frecuente entre sus
    documentos. Filas sin N° Ref. o sin Tipo de Proyecto se ignoran. Si la
    hoja 'Master' no existe, devuelve dict vacío."""
    wb = openpyxl.load_workbook(ruta_excel_cc, data_only=True)
    if "Master" not in wb.sheetnames:
        return {}
    ws = wb["Master"]
    encabezados = [celda.value for celda in ws[1]]
    col_n_ref = encabezados.index("N° Ref.") + 1
    col_tipo = encabezados.index("Tipo de Proyecto") + 1

    tipos_por_prefijo: dict[str, Counter] = {}
    for fila in ws.iter_rows(min_row=2):
        n_ref = fila[col_n_ref - 1].value
        tipo = fila[col_tipo - 1].value
        if not n_ref or not tipo:
            continue
        prefijo = prefijo_de_n_ref(n_ref)
        tipos_por_prefijo.setdefault(prefijo, Counter())[tipo] += 1

    return {
        prefijo: contador.most_common(1)[0][0]
        for prefijo, contador in tipos_por_prefijo.items()
    }


def asegurar_categoria_proyectos(
    ws_proyectos, filas_validas: list[dict], categoria_por_prefijo: dict[str, str],
    columna: int,
) -> list[str]:
    """Escribe la 'Categoría' de cada fila válida (columna indicada por
    'columna' -- quien llama la calcula desde HEADERS_PROYECTOS, nunca
    hardcodeada acá) desde categoria_por_prefijo -- valor plano, no fórmula
    (no hay agregación posible, es un lookup 1 a 1). Si un proyecto no tiene
    ningún documento en Centro de Costos todavía, limpia la celda y avisa."""
    avisos = []
    for fila_info in filas_validas:
        prefijo = fila_info["tag"]
        categoria = categoria_por_prefijo.get(prefijo)
        cell = ws_proyectos.cell(row=fila_info["fila"], column=columna)
        if categoria is None:
            avisos.append(
                f"Proyecto '{fila_info['nombre']}' ({prefijo}) sin documentos en "
                f"Centro de Costos todavía -- Categoría queda vacía."
            )
            cell.value = None
            continue
        cell.value = categoria
    return avisos


def agrupar_por_proyecto_y_subcategoria(items_detalle: list[dict]) -> dict[tuple[str, str], float]:
    """Suma total_sin_iva agrupado por (prefijo de proyecto, categoria_item
    original -- sin colapsar a bucket todavía, eso lo hace la hoja 'Detalle
    Costos Reales' al escribir, para no perder la subcategoría real)."""
    agrupado: dict[tuple[str, str], float] = {}
    for item in items_detalle:
        prefijo = prefijo_de_n_ref(item["n_ref"])
        clave = (prefijo, item["categoria_item"])
        agrupado[clave] = agrupado.get(clave, 0.0) + item["total_sin_iva"]
    return agrupado


# ── LECTURA DE LA HOJA "PROYECTOS" ───────────────────────────────────────────

def leer_filas_proyectos(ws_proyectos) -> tuple[list[dict], list[str]]:
    """Recorre la hoja 'Proyectos' desde la fila 2. Filas sin TAG o sin
    Nombre se saltan con aviso. TAG duplicado: se queda con la primera
    fila, avisa de las siguientes."""
    filas_validas = []
    avisos = []
    tags_vistos = set()

    for fila_idx in range(2, ws_proyectos.max_row + 1):
        tag = ws_proyectos.cell(row=fila_idx, column=1).value
        nombre = ws_proyectos.cell(row=fila_idx, column=2).value

        if not tag or not nombre:
            if tag or nombre:
                avisos.append(f"Fila {fila_idx}: falta TAG o Nombre, se salta.")
            continue

        if tag in tags_vistos:
            avisos.append(f"Fila {fila_idx}: TAG '{tag}' duplicado, se usa la primera fila.")
            continue

        tags_vistos.add(tag)
        filas_validas.append({"fila": fila_idx, "tag": tag, "nombre": nombre})

    return filas_validas, avisos


# ── CARPETAS DE PROYECTO ─────────────────────────────────────────────────────

def asegurar_carpeta_proyecto(nombre_proyecto: str, raiz_facturas: Path) -> bool:
    """Crea raiz_facturas/<nombre_proyecto>/ si no existe. Devuelve True si
    la creó, False si ya existía. raiz_facturas debe ser la fuente REAL que
    lee Centro de Costos hoy (Sitio de comunicación - Centro de Costos 1/
    Facturas y Boletas/), nunca la carpeta legado."""
    carpeta = raiz_facturas / nombre_proyecto
    ya_existia = carpeta.exists()
    carpeta.mkdir(parents=True, exist_ok=True)
    return not ya_existia


def asegurar_carpetas_proyectos(filas_validas: list[dict], raiz_facturas: Path) -> list[str]:
    """Aplica asegurar_carpeta_proyecto a cada fila válida. Devuelve los
    nombres de las carpetas que se crearon (para el informe de consola)."""
    creadas = []
    for fila_info in filas_validas:
        if asegurar_carpeta_proyecto(fila_info["nombre"], raiz_facturas):
            creadas.append(fila_info["nombre"])
    return creadas


# ── BACKUP ────────────────────────────────────────────────────────────────

MESES_ES = {
    1: "Enero", 2: "Febrero", 3: "Marzo", 4: "Abril", 5: "Mayo", 6: "Junio",
    7: "Julio", 8: "Agosto", 9: "Septiembre", 10: "Octubre", 11: "Noviembre",
    12: "Diciembre",
}


def hacer_backup(ruta_excel: Path, raiz_respaldos: Path) -> Path | None:
    """Copia ruta_excel a raiz_respaldos/<Mes Año>/Análisis de Proyectos -
    backup <fecha> <hora>.xlsx antes de escribir -- mismo patrón que
    Centro de Costos. Devuelve None si ruta_excel todavía no existe (nada
    que respaldar)."""
    if not ruta_excel.exists():
        return None
    ahora = datetime.now()
    carpeta_mes = raiz_respaldos / f"{MESES_ES[ahora.month]} {ahora.year}"
    carpeta_mes.mkdir(parents=True, exist_ok=True)
    marca_tiempo = ahora.strftime("%Y-%m-%d %H%M%S")
    destino = carpeta_mes / f"Análisis de Proyectos - backup {marca_tiempo}.xlsx"
    shutil.copy2(ruta_excel, destino)
    return destino


# ── HOJA "DETALLE COSTOS REALES" (100% regenerada cada corrida) ─────────────

def regenerar_hoja_detalle_costos_reales(wb, agrupado: dict[tuple[str, str], float]) -> list[str]:
    """Borra todas las filas de datos (fila 2 en adelante) y las reescribe
    completas desde 'agrupado' -- mismo patrón que las hojas de proyecto de
    Centro de Costos: se recalcula entera, nunca se acumula a mano. Devuelve
    avisos de subcategorías sin mapeo explícito (caen en 'Otros').

    Columna E ("% del Total Real del proyecto", agregada 2026-07-28) = Total
    sin IVA de la fila / suma de TODAS las filas de ese mismo proyecto EN
    ESTA MISMA HOJA (no el Total Real de 'Proyectos', que además incluye
    Mano de Obra Real manual -- esa categoría no tiene detalle por
    subcategoría acá). Se calcula en Python, como el resto de esta hoja
    (nunca fórmula), para que sume ~100% de forma verificable sin depender
    de que Excel recalcule."""
    ws = wb[HOJA_DETALLE_COSTOS_REALES]
    if ws.max_row >= 2:
        ws.delete_rows(2, ws.max_row - 1)

    total_por_proyecto: dict[str, float] = {}
    for (tag, _), total in agrupado.items():
        total_por_proyecto[tag] = total_por_proyecto.get(tag, 0.0) + total

    avisos = []
    fila = 2
    for (tag, subcategoria), total in sorted(agrupado.items()):
        bucket, es_explicito = mapear_categoria_a_bucket(subcategoria)
        if not es_explicito:
            avisos.append(
                f"Categoría '{subcategoria}' (proyecto {tag}) sin mapeo explícito, va a 'Otros'."
            )
        total_proyecto = total_por_proyecto[tag]
        porcentaje = (total / total_proyecto) if total_proyecto else None
        ws.cell(row=fila, column=1, value=tag)
        ws.cell(row=fila, column=2, value=subcategoria)
        ws.cell(row=fila, column=3, value=bucket)
        ws.cell(row=fila, column=4, value=total)
        ws.cell(row=fila, column=5, value=porcentaje)
        fila += 1

    return avisos


# ── FÓRMULAS DE LA HOJA "PROYECTOS" ──────────────────────────────────────────

def asegurar_formulas_proyectos(ws_proyectos, filas_validas: list[dict]) -> None:
    """Escribe las columnas derivadas (Materiales/Equipos/Otros Reales =
    SUMIFS hacia 'Detalle Costos Reales'; Total Proyectado/Real, Margen
    Proyectado/Real, Desviación % = fórmulas sobre 'Proyectos') para cada
    fila válida -- todas resueltas por nombre vía LETRA_COL_PROYECTOS, nunca
    por letra fija, para no romperse si HEADERS_PROYECTOS cambia de orden.
    Nunca toca las columnas de ingreso manual (ver
    NOMBRES_COLUMNAS_MANUALES_PROYECTOS) ni Cliente/Categoría
    (autocompletadas)."""
    col = {nombre: HEADERS_PROYECTOS.index(nombre) + 1 for nombre in HEADERS_PROYECTOS}
    letra = LETRA_COL_PROYECTOS

    for fila_info in filas_validas:
        r = fila_info["fila"]
        tag_ref = f"$A{r}"

        for nombre, bucket in (
            ("Costos Materiales Reales", "Materiales"),
            ("Costos Equipos Reales", "Equipos"),
            ("Otros Costos Reales", "Otros"),
        ):
            ws_proyectos.cell(row=r, column=col[nombre], value=(
                f"=SUMIFS('{HOJA_DETALLE_COSTOS_REALES}'!$D:$D,"
                f"'{HOJA_DETALLE_COSTOS_REALES}'!$A:$A,{tag_ref},"
                f"'{HOJA_DETALLE_COSTOS_REALES}'!$C:$C,\"{bucket}\")"
            ))

        g, h, i, j = (letra[n] for n in (
            "Costos Materiales Proyectados", "Costos Equipos Proyectados",
            "Mano de Obra Proyectada", "Otros Costos Proyectados",
        ))
        l, m, n, o = (letra[n] for n in (
            "Costos Materiales Reales", "Costos Equipos Reales",
            "Otros Costos Reales", "Mano de Obra Real",
        ))
        venta = letra["Monto de Venta (sin IVA)"]
        total_proy = letra["Total Proyectado"]
        total_real = letra["Total Real"]

        ws_proyectos.cell(row=r, column=col["Total Proyectado"], value=f"={g}{r}+{h}{r}+{i}{r}+{j}{r}")
        ws_proyectos.cell(row=r, column=col["Total Real"], value=f"={l}{r}+{m}{r}+{n}{r}+{o}{r}")
        ws_proyectos.cell(row=r, column=col["Margen Proyectado"], value=f"={venta}{r}-{total_proy}{r}")
        ws_proyectos.cell(row=r, column=col["Margen Real"], value=f"={venta}{r}-{total_real}{r}")
        ws_proyectos.cell(
            row=r, column=col["Desviación % (Real vs Proyectado)"],
            value=f"={total_real}{r}/{total_proy}{r}-1",
        )


# ── NOTA DEL PROYECTO (hoja "Indicadores") ──────────────────────────────────
# 0-100, aprobatorio >=55. Rentabilidad domina el peso (70/30) -- decision
# del usuario en brainstorming, spec 2026-07-21 seccion 1. Constantes
# separadas de la formula para que recalibrar el benchmark no implique
# reescribir la logica, solo estos 3 valores.
MARGEN_OBJETIVO_NOTA = 0.25
PESO_RENTABILIDAD_NOTA = 0.7
PESO_DESVIACION_NOTA = 0.3

UMBRAL_EXCELENTE, UMBRAL_BUENO, UMBRAL_APROBADO = 85, 70, 55


# ── COMPLETITUD DE UN PROYECTO: UNA SOLA DEFINICION ─────────────────────────
# Un proyecto sin estos campos cargados a mano no recibe KPIs: ni reporte PDF
# ni fila en el dashboard (aparece como "pendiente de completar"). "Fecha de
# cierre" queda deliberadamente FUERA -- su ausencia marca al proyecto como
# "en desarrollo", no como incompleto. "Cliente"/"Categoría" tampoco cuentan:
# se autocompletan, no son carga del usuario.
#
# Hasta el 2026-07-28 esta regla estaba duplicada y los dos consumidores no
# coincidian: los reportes exigian estos 8 campos, el visualizador solo 6
# (sin "Estado" ni "Fecha de inicio") y ademas aceptaba la cadena vacia como
# valor cargado. Resultado: un proyecto podia salir con KPIs completos en el
# dashboard y a la vez ser rechazado con DatosIncompletosError al pedir su PDF.
CAMPOS_MANUALES_REQUERIDOS = [
    "Estado", "Fecha de inicio", "Monto de Venta (sin IVA)",
    "Costos Materiales Proyectados", "Costos Equipos Proyectados",
    "Mano de Obra Proyectada", "Otros Costos Proyectados", "Mano de Obra Real",
]


def tiene_datos_completos(valor_de_campo) -> bool:
    """True si el proyecto tiene los 8 campos manuales cargados.
    `valor_de_campo` es un callable que recibe el nombre de encabezado de
    "Proyectos" y devuelve su valor -- asi sirve igual para un dict keyed por
    encabezado (reportes) que para uno de claves cortas (visualizador), sin
    que ninguno de los dos tenga que reimplementar la regla.

    0 SI cuenta como cargado (un costo real en cero es un dato, no un vacio);
    None y la cadena vacia no."""
    return all(valor_de_campo(campo) not in (None, "") for campo in CAMPOS_MANUALES_REQUERIDOS)


# ── NOTA / EVALUACION: UNA SOLA DEFINICION, DOS RENDERIZADOS ────────────────
# La misma regla de negocio se expresa de dos formas: como formula de Excel
# (_formula_nota, para que el libro recalcule solo) y como calculo en Python
# (calcular_nota, para los consumidores que no pueden leer el resultado de una
# formula -- openpyxl nunca la evalua). Las dos viven pegadas A PROPOSITO:
# hasta el 2026-07-28 el calculo Python estaba duplicado en dos archivos
# distintos (Reportes/kpis_recalculados.py y Visualizador Web/
# build_visualizador.py) y se desincronizaron -- el visualizador siguio usando
# ABS() despues de que la regla cambio, y el MISMO proyecto mostraba nota 94 en
# el dashboard y 100 en el Excel/PDF. Si cambias una de las dos funciones de
# abajo, cambia la otra en el mismo commit; test_contrato_kpis.py falla si no.


def _redondear_excel(x: float) -> int:
    """ROUND() de Excel: 'half away from zero'. El round() nativo de Python
    usa banker's rounding (round(96.5) == 96), lo que puede dejar la Nota en
    el bucket de Evaluacion equivocado justo en un empate exacto en .5."""
    return math.floor(x + 0.5) if x >= 0 else math.ceil(x - 0.5)


def calcular_nota(margen_neto: float | None, desviacion_total: float | None) -> int | None:
    """Equivalente Python exacto de _formula_nota(). None si falta cualquiera
    de los dos insumos (mismo significado que una celda vacia en Excel).

    El componente de desviacion (30%) usa MAX(0, desviacion), NO ABS():
    penaliza solo el sobrecosto real (Real > Proyectado). Un proyecto en o
    bajo presupuesto obtiene el puntaje maximo de ese componente -- el
    beneficio de haber ahorrado ya esta capturado en el 70% de rentabilidad,
    asi que castigarlo aca ademas seria doble penalizacion (decision del
    usuario, 2026-07-28)."""
    if margen_neto is None or desviacion_total is None:
        return None
    score_margen = min(100, max(0, (margen_neto / MARGEN_OBJETIVO_NOTA) * 100))
    score_desviacion = min(100, max(0, 100 - max(0, desviacion_total) * 100))
    return _redondear_excel(
        PESO_RENTABILIDAD_NOTA * score_margen + PESO_DESVIACION_NOTA * score_desviacion
    )


def clasificar_evaluacion(nota: int | None) -> str | None:
    """Equivalente Python exacto de _formula_evaluacion()."""
    if nota is None:
        return None
    if nota >= UMBRAL_EXCELENTE:
        return "Excelente"
    if nota >= UMBRAL_BUENO:
        return "Bueno"
    if nota >= UMBRAL_APROBADO:
        return "Aprobado"
    return "Requiere atención"


def _formula_nota(fila_proyectos: int) -> str:
    """Corregido 2026-07-28 (aprobado por el usuario): el componente de
    control de desviación (30%) ya NO usa ABS() -- antes penalizaba gastar
    de menos igual que gastar de más, pese a que ahorrar ya sube el margen
    (capturado en el 70% de rentabilidad), así que era un doble castigo
    disfrazado de doble premio. Ahora MAX(0, desviación) anula el término
    para cualquier proyecto en o bajo presupuesto (desviación <= 0): ese
    componente da el puntaje máximo (100), sin restar ni sumar de más. Solo
    resta puntos cuando Real > Proyectado (sobrecosto real, desviación
    positiva). Ver MEMORY.md 2026-07-28 para la verificación a mano contra
    UMAG (Nota pasó de 91 a 100)."""
    r = fila_proyectos
    margen_real = LETRA_COL_PROYECTOS["Margen Real"]
    venta = LETRA_COL_PROYECTOS["Monto de Venta (sin IVA)"]
    desviacion = LETRA_COL_PROYECTOS["Desviación % (Real vs Proyectado)"]
    score_margen = (
        f"MIN(100,MAX(0,(Proyectos!{margen_real}{r}/Proyectos!{venta}{r})"
        f"/{MARGEN_OBJETIVO_NOTA}*100))"
    )
    score_desviacion = (
        f"MIN(100,MAX(0,100-MAX(0,Proyectos!{desviacion}{r})*100))"
    )
    return f"=ROUND({PESO_RENTABILIDAD_NOTA}*{score_margen}+{PESO_DESVIACION_NOTA}*{score_desviacion},0)"


def _formula_evaluacion(fila_destino: int) -> str:
    # Referencia a la columna "Nota del Proyecto" de esta misma hoja
    # ("Indicadores"), calculada desde LETRA_COL_INDICADORES -- nunca
    # hardcodeada, para no romperse si el playbook de KPIs vuelve a cambiar
    # de orden/cantidad de columnas.
    col_nota = LETRA_COL_INDICADORES["Nota del Proyecto"]
    ref = f"{col_nota}{fila_destino}"
    return (
        f'=IF({ref}>=85,"Excelente",IF({ref}>=70,"Bueno",'
        f'IF({ref}>=55,"Aprobado","Requiere atención")))'
    )


# ── FÓRMULAS DE LA HOJA "INDICADORES" (100% regenerada cada corrida) ────────

def asegurar_hoja_indicadores(wb, filas_validas: list[dict]) -> None:
    """Regenera 'Indicadores' completa: una fila compacta por proyecto
    válido (sin huecos), pero cada fórmula referencia la fila REAL del
    proyecto en 'Proyectos' (que sí puede tener huecos).

    Orden de columnas (reordenado 2026-07-28, ver HEADERS_INDICADORES):
    margen neto; costo % de venta (4); estructura % del costo real / mix,
    suma 100% (4); desviación % por categoría (4) + desviación % total;
    ahorro/sobrecosto neto en $ por categoría (4) + total; nota; evaluación."""
    ws = wb[HOJA_INDICADORES]
    if ws.max_row >= 2:
        ws.delete_rows(2, ws.max_row - 1)

    lp = LETRA_COL_PROYECTOS
    tag, nombre = lp["TAG proyecto"], lp["Nombre del proyecto"]
    venta = lp["Monto de Venta (sin IVA)"]
    margen_real, total_real, total_proy = (
        lp["Margen Real"], lp["Total Real"], lp["Total Proyectado"],
    )
    desviacion_total = lp["Desviación % (Real vs Proyectado)"]
    mat_r, eq_r, mo_r, otros_r = (
        lp["Costos Materiales Reales"], lp["Costos Equipos Reales"],
        lp["Mano de Obra Real"], lp["Otros Costos Reales"],
    )
    mat_p, eq_p, mo_p, otros_p = (
        lp["Costos Materiales Proyectados"], lp["Costos Equipos Proyectados"],
        lp["Mano de Obra Proyectada"], lp["Otros Costos Proyectados"],
    )
    fecha_inicio, fecha_cierre = lp["Fecha de inicio"], lp["Fecha de cierre"]

    fila_destino = 2
    for fila_info in filas_validas:
        r = fila_info["fila"]
        f = fila_destino
        ws.cell(row=f, column=1, value=f"=Proyectos!{tag}{r}")
        ws.cell(row=f, column=2, value=f"=Proyectos!{nombre}{r}")
        ws.cell(row=f, column=3, value=f"=Proyectos!{margen_real}{r}/Proyectos!{venta}{r}")
        ws.cell(row=f, column=4, value=f"=Proyectos!{mat_r}{r}/Proyectos!{venta}{r}")
        ws.cell(row=f, column=5, value=f"=Proyectos!{eq_r}{r}/Proyectos!{venta}{r}")
        ws.cell(row=f, column=6, value=f"=Proyectos!{mo_r}{r}/Proyectos!{venta}{r}")
        ws.cell(row=f, column=7, value=f"=Proyectos!{otros_r}{r}/Proyectos!{venta}{r}")
        ws.cell(row=f, column=8, value=f"=Proyectos!{mat_r}{r}/Proyectos!{total_real}{r}")
        ws.cell(row=f, column=9, value=f"=Proyectos!{eq_r}{r}/Proyectos!{total_real}{r}")
        ws.cell(row=f, column=10, value=f"=Proyectos!{mo_r}{r}/Proyectos!{total_real}{r}")
        ws.cell(row=f, column=11, value=f"=Proyectos!{otros_r}{r}/Proyectos!{total_real}{r}")
        ws.cell(row=f, column=12, value=f"=Proyectos!{mat_r}{r}/Proyectos!{mat_p}{r}-1")
        ws.cell(row=f, column=13, value=f"=Proyectos!{eq_r}{r}/Proyectos!{eq_p}{r}-1")
        ws.cell(row=f, column=14, value=f"=Proyectos!{mo_r}{r}/Proyectos!{mo_p}{r}-1")
        ws.cell(row=f, column=15, value=f"=Proyectos!{otros_r}{r}/Proyectos!{otros_p}{r}-1")
        ws.cell(row=f, column=16, value=f"=Proyectos!{desviacion_total}{r}")
        ws.cell(row=f, column=17, value=f"=Proyectos!{mat_p}{r}-Proyectos!{mat_r}{r}")
        ws.cell(row=f, column=18, value=f"=Proyectos!{eq_p}{r}-Proyectos!{eq_r}{r}")
        ws.cell(row=f, column=19, value=f"=Proyectos!{mo_p}{r}-Proyectos!{mo_r}{r}")
        ws.cell(row=f, column=20, value=f"=Proyectos!{otros_p}{r}-Proyectos!{otros_r}{r}")
        ws.cell(row=f, column=21, value=f"=Proyectos!{total_proy}{r}-Proyectos!{total_real}{r}")
        ws.cell(row=f, column=22, value=_formula_nota(r))
        ws.cell(row=f, column=23, value=_formula_evaluacion(f))
        # X: Peso del proyecto en la cartera de ventas (%) -- venta del
        # proyecto sobre la suma de TODA la columna "Monto de Venta (sin
        # IVA)" de "Proyectos" (no solo la propia fila), incluyendo
        # proyectos que no están "Terminado" -- SUM ignora celdas vacías/
        # texto, así que un proyecto sin venta cargada no distorsiona el
        # denominador solo, no hace falta filtrar por Estado.
        ws.cell(row=f, column=24, value=(
            f"=Proyectos!{venta}{r}/SUM(Proyectos!${venta}:${venta})"
        ))
        # Y: Margen por día de ejecución -- IF(Fecha de cierre vacía, "",
        # ...) para que un proyecto "en desarrollo" (sin Fecha de cierre)
        # quede vacío en vez de una fórmula con error/número sin sentido;
        # el guard vive en la fórmula (se recalcula solo si el usuario
        # completa la fecha después, sin correr el script de nuevo).
        # MAX(1, días) evita #DIV/0! si Fecha de cierre = Fecha de inicio.
        ws.cell(row=f, column=25, value=(
            f'=IF(Proyectos!{fecha_cierre}{r}="","",'
            f"Proyectos!{margen_real}{r}/MAX(1,Proyectos!{fecha_cierre}{r}-Proyectos!{fecha_inicio}{r}))"
        ))
        fila_destino += 1


# ── HOJA "CLIENTES" (CLTV, 100% regenerada cada corrida) ────────────────────

def asegurar_hoja_clientes(wb, filas_validas: list[dict], ws_proyectos) -> None:
    """Regenera 'Clientes' completa: una fila por valor único de la columna
    'Cliente' de 'Proyectos' (filas sin Cliente asignado se ignoran). Todas
    las columnas son fórmulas que agregan sobre 'Proyectos' filtrando por
    Cliente -- nunca valores calculados en Python -- para que un proyecto
    nuevo del mismo cliente se sume solo la próxima vez que Excel recalcule."""
    ws = wb[HOJA_CLIENTES]
    if ws.max_row >= 2:
        ws.delete_rows(2, ws.max_row - 1)

    col_cliente = HEADERS_PROYECTOS.index("Cliente") + 1
    clientes_unicos = []
    vistos = set()
    for fila_info in filas_validas:
        cliente = ws_proyectos.cell(row=fila_info["fila"], column=col_cliente).value
        if cliente and cliente not in vistos:
            vistos.add(cliente)
            clientes_unicos.append(cliente)

    cliente_col = LETRA_COL_PROYECTOS["Cliente"]
    venta_col = LETRA_COL_PROYECTOS["Monto de Venta (sin IVA)"]
    fecha_inicio_col = LETRA_COL_PROYECTOS["Fecha de inicio"]
    margen_real_col = LETRA_COL_PROYECTOS["Margen Real"]

    for i, cliente in enumerate(sorted(clientes_unicos), start=2):
        ws.cell(row=i, column=1, value=cliente)
        ws.cell(row=i, column=2, value=(
            f"=AVERAGEIF(Proyectos!${cliente_col}:${cliente_col},$A{i},"
            f"Proyectos!${venta_col}:${venta_col})"
        ))
        ws.cell(row=i, column=3, value=f"=COUNTIF(Proyectos!${cliente_col}:${cliente_col},$A{i})")
        ws.cell(row=i, column=4, value=(
            f"=MAX(1,(_xlfn.MAXIFS(Proyectos!${fecha_inicio_col}:${fecha_inicio_col},"
            f"Proyectos!${cliente_col}:${cliente_col},$A{i})"
            f"-_xlfn.MINIFS(Proyectos!${fecha_inicio_col}:${fecha_inicio_col},"
            f"Proyectos!${cliente_col}:${cliente_col},$A{i}))/30)"
        ))
        ws.cell(row=i, column=5, value=f"=C{i}/(D{i}/12)")
        ws.cell(row=i, column=6, value=(
            f"=SUMIF(Proyectos!${cliente_col}:${cliente_col},$A{i},Proyectos!${margen_real_col}:${margen_real_col})"
            f"/SUMIF(Proyectos!${cliente_col}:${cliente_col},$A{i},Proyectos!${venta_col}:${venta_col})"
        ))
        ws.cell(row=i, column=7, value=f"=B{i}*E{i}*C{i}*F{i}")
        ws.cell(row=i, column=8, value=(
            f'=IF(G{i}>=PERCENTILE(Clientes!$G:$G,0.67),"Clientes estratégicos",'
            f'IF(G{i}>=PERCENTILE(Clientes!$G:$G,0.33),"Clientes potenciales","Clientes de oportunidad"))'
        ))


# ── HOJA "GLOSARIO KPIS" (texto estatico, 100% regenerado cada corrida) ─────
# Contenido fijo -- no depende de datos del usuario. Cubre el playbook
# original (spec 2026-07-20) mas los KPIs de este spec (2026-07-21): Nota,
# Evaluacion, y los 6 de la hoja Clientes.

GLOSARIO_KPIS: list[tuple[str, str, str, str]] = [
    (
        "Margen neto %",
        "El indicador de rentabilidad más directo y comparable entre proyectos de distinto tamaño.",
        "Margen Real, Monto de Venta",
        "20% significa que de cada $100 vendidos quedan $20 de utilidad tras cubrir todos los costos reales.",
    ),
    (
        "Costo % de venta (por categoría)",
        "Muestra qué parte de cada peso VENDIDO se va en esa categoría de costo — mide el impacto de la categoría sobre el precio de venta.",
        "Costo Real de la categoría, Monto de Venta",
        "35% en Costo MO % de venta → un tercio de cada venta se destina a mano de obra; detecta categorías que consumen desproporcionadamente el margen.",
    ),
    (
        "Estructura % del costo real (mix, por categoría)",
        "Muestra cómo se reparte el GASTO real entre categorías (no contra la venta, sino entre sí) — las 4 categorías suman 100%, a diferencia de 'Costo % de venta' que no suma 100%.",
        "Costo Real de la categoría, Costos Totales Real",
        "60% en Estructura % Otros → 6 de cada $10 gastados en el proyecto fueron en 'Otros'; útil para ver dónde se concentra el gasto real sin importar cuán rentable resultó el proyecto.",
    ),
    (
        "Desviación % (por categoría, Real vs Proyectado)",
        "Mide qué tan preciso fue el presupuesto original para esa categoría — clave para mejorar futuras cotizaciones.",
        "Costo Real, Costo Proyectado de la categoría",
        "+15% = se gastó 15% más de lo presupuestado; negativo = se gastó menos de lo previsto.",
    ),
    (
        "Desviación % Total",
        "Mismo concepto que la desviación por categoría, pero a nivel de todo el proyecto — el número que resume si el presupuesto completo se cumplió.",
        "Costos Totales Real, Costos Totales Proyectado (columna ya existente en 'Proyectos', traída acá como columna visible)",
        "-10% = el proyecto costó 10% menos que lo presupuestado en total; +10% = costó 10% más.",
    ),
    (
        "Ahorro/Sobrecosto neto en $ (por categoría y total)",
        "Traduce la desviación % a pesos concretos — más accionable para decisiones de gestión que un porcentaje solo, sobre todo en categorías con montos grandes.",
        "Costo Proyectado − Costo Real, de cada categoría y del total",
        "Positivo = ahorro (se gastó menos de lo presupuestado); negativo = sobrecosto (se gastó más). $800.000 en Ahorro MO → la mano de obra costó $800.000 menos que lo presupuestado.",
    ),
    (
        "Nota del Proyecto",
        "Resume rentabilidad y control de presupuesto en un solo número comparable entre proyectos, para priorizar dónde poner atención de gestión.",
        "Margen neto % (70%, contra objetivo de 25%) y Desviación % Total, solo penalizando sobrecosto (30%)",
        "≥55 = proyecto en rango aceptable; <55 = requiere revisión (rentabilidad baja y/o descontrol presupuestario). Un proyecto que ahorra (Real ≤ Proyectado) obtiene el puntaje máximo del componente de control — no se penaliza gastar de menos, ese beneficio ya se refleja en el margen.",
    ),
    (
        "Evaluación",
        "Traduce la nota a una etiqueta rápida de lectura para revisiones ejecutivas.",
        "Nota del Proyecto",
        "Excelente / Bueno / Aprobado / Requiere atención.",
    ),
    (
        "% del Total Real del proyecto (Detalle Costos Reales)",
        "Muestra el peso de cada subcategoría real de Centro de Costos (ej. Consumibles, Equipos-Herramientas, Combustible, Servicios) dentro del gasto detallado del proyecto — más granular que 'Estructura %' de 'Indicadores', que solo agrupa en 4 buckets.",
        "Total sin IVA de la fila, suma de todas las filas de ese proyecto en 'Detalle Costos Reales'",
        "Suma ~100% por proyecto. No incluye Mano de Obra Real (es manual, sin detalle por subcategoría en Centro de Costos) -- el 100% es sobre el gasto SÍ trazado a documentos, no sobre el Total Real completo del proyecto.",
    ),
    (
        "Peso del proyecto en la cartera de ventas (%)",
        "Mide qué tan concentrado está el ingreso de la empresa en un solo proyecto -- riesgo de dependencia si un proyecto grande se cae o se retrasa.",
        "Monto de Venta del proyecto, suma de Monto de Venta de todos los proyectos con venta cargada",
        "25% significa que un cuarto de todo el ingreso de la cartera actual depende de ese único proyecto; valores altos ameritan revisar el riesgo de concentración.",
    ),
    (
        "Margen por día de ejecución",
        "Mide cuánto margen genera el proyecto por unidad de tiempo -- útil para priorizar proyectos que compiten por la misma capacidad de equipo/tiempo, no solo por margen total.",
        "Margen Real, Fecha de cierre − Fecha de inicio (en días)",
        "$50.000/día = el proyecto generó en promedio $50.000 de margen por cada día que duró su ejecución. Queda vacío si el proyecto todavía no tiene Fecha de cierre (en desarrollo) -- no se calcula sobre una duración que aún no terminó.",
    ),
    (
        "AOV (Clientes)",
        "Mide el tamaño promedio de una venta a ese cliente.",
        "Monto de Venta de sus proyectos",
        "AOV alto = cliente que trae proyectos grandes por transacción.",
    ),
    (
        "Vida del cliente",
        "Mide cuántas veces ha comprado el cliente en total — la base para saber si es recurrente.",
        "Conteo de proyectos del cliente",
        "Vida=1 → cliente de una sola compra hasta ahora; vida>1 → recurrente.",
    ),
    (
        "Meses activo",
        "Mide cuánto tiempo lleva comprando el cliente — el denominador para anualizar la frecuencia.",
        "Fecha más antigua y más reciente entre sus proyectos",
        "Meses activo alto + vida baja → cliente esporádico; meses activo bajo + vida alta → cliente muy activo recientemente.",
    ),
    (
        "Frecuencia de compra (Clientes)",
        "Mide qué tan seguido vuelve a comprar el cliente, anualizado — clave para proyectar ingresos futuros de ese cliente.",
        "Vida del cliente, Meses activo",
        "Frecuencia=2 → el cliente compra en promedio 2 veces al año.",
    ),
    (
        "Margen de utilidad % (Clientes)",
        "Mide qué tan rentable es la relación completa con ese cliente, ponderado por tamaño de proyecto.",
        "Suma de Margen Real y de Monto de Venta de todos sus proyectos",
        "Mismo significado que Margen neto % pero a nivel cliente.",
    ),
    (
        "CLTV",
        "Estima el valor total que el cliente representa para QUEMPIN a lo largo de su relación completa — la métrica central para decidir dónde invertir esfuerzo comercial.",
        "AOV × Frecuencia de compra × Vida del cliente × Margen de utilidad %",
        "CLTV alto = cliente que ha generado y probablemente seguirá generando mucho valor; prioridad para retención.",
    ),
    (
        "Clasificación (Clientes)",
        "Traduce el CLTV a un tier accionable, relativo a la cartera actual de QUEMPIN, no a un corte fijo en pesos que quede obsoleto con el crecimiento de la empresa.",
        "Percentil del CLTV entre todos los clientes registrados",
        "'Clientes estratégicos' (top 33%) → atención prioritaria; 'Clientes de oportunidad' (bottom 33%) → candidatos a desarrollar o repensar la relación.",
    ),
]


def asegurar_hoja_glosario_kpis(wb) -> None:
    """Reescribe 'Glosario KPIs' completa desde la constante GLOSARIO_KPIS --
    texto estático, no depende de datos del usuario ni de fórmulas."""
    ws = wb[HOJA_GLOSARIO_KPIS]
    if ws.max_row >= 2:
        ws.delete_rows(2, ws.max_row - 1)

    for i, fila in enumerate(GLOSARIO_KPIS, start=2):
        for col, valor in enumerate(fila, start=1):
            ws.cell(row=i, column=col, value=valor)


# ── ORQUESTADOR ───────────────────────────────────────────────────────────

def actualizar_visualizador_af() -> bool:
    """Regenera el visualizador web (Visualizador Web/build/index.html) a
    partir del Excel recien guardado -- mismo patron que actualizar_
    visualizador() en Centro de Costos: corre al final de ejecutar(), solo
    lee el Excel (no lo modifica), y si falla no aborta el run, solo
    advierte -- el Excel ya quedo guardado igual. Ver Visualizador Web/
    CLAUDE.md para la arquitectura del build."""
    ruta_build_script = RAIZ_VISUALIZADOR_WEB_AF / "build_visualizador.py"
    if not ruta_build_script.exists():
        return False
    ya_en_path = str(RAIZ_VISUALIZADOR_WEB_AF) in sys.path
    if not ya_en_path:
        sys.path.insert(0, str(RAIZ_VISUALIZADOR_WEB_AF))
    try:
        sys.modules.pop("build_visualizador", None)
        import build_visualizador as bv
        return bv.build() == 0
    finally:
        if not ya_en_path and str(RAIZ_VISUALIZADOR_WEB_AF) in sys.path:
            sys.path.remove(str(RAIZ_VISUALIZADOR_WEB_AF))


def ejecutar(
    ruta_excel_af: Path = RUTA_EXCEL,
    ruta_excel_cc: Path = RUTA_EXCEL_CENTRO_COSTOS,
    raiz_facturas_cc: Path = RAIZ_FACTURAS_CENTRO_COSTOS,
    raiz_respaldos: Path = RAIZ_RESPALDOS,
    ruta_clientes_pendientes: Path = RUTA_CLIENTES_PENDIENTES,
    dry_run: bool = False,
) -> dict:
    """Orquesta todo el flujo. Con dry_run=True no escribe nada -- ni backup,
    ni carpetas, ni el Excel -- solo reporta qué pasaría (usado por el
    comando 'status' del skill). Captura PermissionError/OSError de operaciones
    de archivo (backup, carpetas, save); excepciones de lectura de datos
    propagarán hacia afuera."""
    resumen = {
        "avisos": [], "carpetas_creadas": [], "categorias_no_mapeadas": [],
        "clientes_pendientes": [], "error": None,
    }

    wb = asegurar_estructura_workbook(ruta_excel_af)
    ws_proyectos = wb[HOJA_PROYECTOS]
    filas_validas, avisos_lectura = leer_filas_proyectos(ws_proyectos)
    resumen["avisos"].extend(avisos_lectura)

    if not ruta_excel_cc.exists():
        resumen["avisos"].append(
            f"No se encontró {ruta_excel_cc}, no se actualizan costos reales."
        )
        return resumen

    items_detalle = leer_detalle_centro_costos(ruta_excel_cc)
    agrupado = agrupar_por_proyecto_y_subcategoria(items_detalle)

    if dry_run:
        for fila_info in filas_validas:
            if not (raiz_facturas_cc / fila_info["nombre"]).exists():
                resumen["carpetas_creadas"].append(fila_info["nombre"])
        categorias_no_mapeadas = set()
        for _, subcategoria in agrupado:
            _, es_explicito = mapear_categoria_a_bucket(subcategoria)
            if not es_explicito:
                categorias_no_mapeadas.add(subcategoria)
        resumen["categorias_no_mapeadas"] = sorted(categorias_no_mapeadas)
        return resumen

    try:
        hacer_backup(ruta_excel_af, raiz_respaldos)
    except PermissionError as exc:
        resumen["avisos"].append(f"No se pudo respaldar (¿archivo abierto?): {exc}")

    try:
        resumen["carpetas_creadas"] = asegurar_carpetas_proyectos(filas_validas, raiz_facturas_cc)
    except OSError as exc:
        resumen["avisos"].append(f"No se pudieron crear una o más carpetas de proyecto: {exc}")

    avisos_detalle = regenerar_hoja_detalle_costos_reales(wb, agrupado)
    resumen["avisos"].extend(avisos_detalle)
    categorias_no_mapeadas = set()
    for _, subcategoria in agrupado:
        _, es_explicito = mapear_categoria_a_bucket(subcategoria)
        if not es_explicito:
            categorias_no_mapeadas.add(subcategoria)
    resumen["categorias_no_mapeadas"] = sorted(categorias_no_mapeadas)

    resumen["clientes_pendientes"] = asegurar_columna_cliente(
        ws_proyectos, filas_validas, ruta_clientes_pendientes
    )

    asegurar_formulas_proyectos(ws_proyectos, filas_validas)
    tipos_por_prefijo = leer_tipo_proyecto_centro_costos(ruta_excel_cc)
    col_categoria = HEADERS_PROYECTOS.index("Categoría") + 1
    resumen["avisos"].extend(
        asegurar_categoria_proyectos(ws_proyectos, filas_validas, tipos_por_prefijo, col_categoria)
    )
    asegurar_hoja_indicadores(wb, filas_validas)
    asegurar_hoja_clientes(wb, filas_validas, ws_proyectos)
    asegurar_hoja_glosario_kpis(wb)
    aplicar_estilo_visual(wb)
    migrar_formato_fecha_proyectos(wb)
    aplicar_resaltado_celdas_manuales(wb)

    try:
        wb.save(ruta_excel_af)
    except PermissionError as exc:
        resumen["error"] = f"No se pudo guardar {ruta_excel_af} (¿archivo abierto?): {exc}"

    try:
        if not actualizar_visualizador_af():
            resumen["avisos"].append(
                "No se pudo actualizar el visualizador web -- correr manualmente "
                "'python driver.py visualizador' despues."
            )
    except Exception as exc:
        resumen["avisos"].append(f"No se pudo actualizar el visualizador web: {exc}")

    return resumen


def main() -> None:
    resumen = ejecutar()
    print("=== Análisis Financiero ===")
    if resumen["carpetas_creadas"]:
        print(f"Carpetas de proyecto creadas: {', '.join(resumen['carpetas_creadas'])}")
    if resumen["categorias_no_mapeadas"]:
        print(
            "Categorías sin mapeo explícito (van a 'Otros'): "
            + ", ".join(resumen["categorias_no_mapeadas"])
        )
    if resumen["clientes_pendientes"]:
        print(
            f"[AVISO] {len(resumen['clientes_pendientes'])} cliente(s) nuevo(s) parecido(s) "
            "a uno existente -- revisar con 'python driver.py confirmar-cliente'."
        )
    for aviso in resumen["avisos"]:
        print(f"[AVISO] {aviso}")
    if resumen["error"]:
        print(f"[ERROR] {resumen['error']}")


if __name__ == "__main__":
    import sys
    sys.stdout.reconfigure(encoding="utf-8", errors="backslashreplace")
    sys.stderr.reconfigure(encoding="utf-8", errors="backslashreplace")
    main()
