# -*- coding: utf-8 -*-
"""
driver.py -- Comandos status/run del skill Reportes_Analisis_Financiero.
status: solo lectura, calcula que reportes quedaron pendientes/desactualizados.
run: igual que status, mas contexto para que el AGENTE (no este script) redacte
y renderice cada uno -- este driver nunca genera contenido de reporte.
"""

import sys
from pathlib import Path

RAIZ_SKILL = Path(__file__).resolve().parent
RAIZ_MODULO = RAIZ_SKILL.parent.parent.parent
RAIZ_REPORTES = RAIZ_MODULO / "Reportes"
RAIZ_SISTEMA = RAIZ_MODULO / "Sistema"
for raiz in (RAIZ_REPORTES, RAIZ_SISTEMA):
    if str(raiz) not in sys.path:
        sys.path.insert(0, str(raiz))

import datos_reportes as dr  # noqa: E402
import estado_reportes as er  # noqa: E402
from analisis_financiero import HOJA_PROYECTOS, RUTA_EXCEL  # noqa: E402

import openpyxl  # noqa: E402

RUTA_ESTADO_REPORTES = RAIZ_REPORTES / "estado_reportes.json"


def listar_entidades(ruta_excel: Path = RUTA_EXCEL) -> dict[str, tuple[str, str]]:
    """Recorre 'Proyectos' y arma la clave->tipo/identificador de cada
    proyecto, cliente unico, y categoria unica presentes hoy. Excluye por
    completo los proyectos sin datos manuales completos (spec §6) -- ni
    generan su propia entrada 'proyecto:TAG', ni aportan a 'cliente:'/
    'categoria:' salvo que OTRO proyecto completo ya la haya registrado."""
    wb = openpyxl.load_workbook(ruta_excel, data_only=True)
    ws = wb[HOJA_PROYECTOS]
    mapa = {celda.value: idx + 1 for idx, celda in enumerate(ws[1]) if celda.value}
    col_tag = mapa.get("TAG proyecto")

    entidades: dict[str, tuple[str, str]] = {}
    clientes_vistos = set()
    categorias_vistas = set()
    for fila_idx in range(2, ws.max_row + 1):
        tag = ws.cell(row=fila_idx, column=col_tag).value if col_tag else None
        if not tag:
            continue
        fila = {h: ws.cell(row=fila_idx, column=c).value for h, c in mapa.items()}
        if not dr.proyecto_tiene_datos_completos(fila):
            continue

        entidades[f"proyecto:{tag}"] = ("proyecto", tag)

        cliente = fila.get("Cliente")
        if cliente and cliente not in clientes_vistos:
            clientes_vistos.add(cliente)
            entidades[f"cliente:{cliente}"] = ("cliente", cliente)

        categoria = fila.get("Categoría")
        if categoria and categoria not in categorias_vistas:
            categorias_vistas.add(categoria)
            entidades[f"categoria:{categoria}"] = ("categoria", categoria)

    return entidades


_FUNCIONES_POR_TIPO = {
    "proyecto": dr.paquete_datos_proyecto,
    "cliente": dr.paquete_datos_cliente,
    "categoria": dr.paquete_datos_categoria,
}


def calcular_reportes_pendientes(
    ruta_excel: Path = RUTA_EXCEL, ruta_estado: Path = RUTA_ESTADO_REPORTES,
) -> list[str]:
    """Claves de entidades cuyo reporte PDF no existe o quedo desactualizado."""
    entidades = listar_entidades(ruta_excel)
    paquetes_actuales = {
        clave: _FUNCIONES_POR_TIPO[tipo](ruta_excel, identificador)
        for clave, (tipo, identificador) in entidades.items()
    }
    estado = er.cargar_estado(ruta_estado)
    return er.detectar_desactualizados(paquetes_actuales, estado)


def status() -> None:
    pendientes = calcular_reportes_pendientes()
    print("=== Reportes Analisis Financiero -- status ===")
    if not pendientes:
        print("Todos los reportes estan al dia.")
        return
    print(f"{len(pendientes)} reporte(s) pendiente(s)/desactualizado(s):")
    for clave in pendientes:
        print(f"  - {clave}")


def run() -> None:
    pendientes = calcular_reportes_pendientes()
    print("=== Reportes Analisis Financiero -- run ===")
    if not pendientes:
        print("Todos los reportes estan al dia. Nada que generar.")
        return
    print(
        f"{len(pendientes)} reporte(s) pendiente(s) -- este comando NO los genera "
        f"solo (requiere redaccion del agente). Pidele al agente que los redacte "
        f"y renderice uno por uno usando datos_reportes/brand/graficos/motor_reportes."
    )
    for clave in pendientes:
        print(f"  - {clave}")


if __name__ == "__main__":
    import sys as _sys
    _sys.stdout.reconfigure(encoding="utf-8", errors="backslashreplace")
    comando = _sys.argv[1] if len(_sys.argv) > 1 else "status"
    {"status": status, "run": run}.get(comando, status)()
