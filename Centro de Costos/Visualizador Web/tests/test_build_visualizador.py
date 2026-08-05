import importlib.util
import sys
from pathlib import Path

import openpyxl

# Los 3 modulos financieros tienen su propio "build_visualizador.py". Como
# sys.modules cachea por NOMBRE, un "import build_visualizador" le entrega a
# los 3 el que se haya importado primero al correr la suite completa (y los
# tests de los otros 2 fallan). Se carga por ruta bajo un nombre unico --
# mismo patron que Sistema/tests/test_driver_preview_renombrados.py.
_RUTA_BV = Path(__file__).resolve().parent.parent / "build_visualizador.py"
_spec = importlib.util.spec_from_file_location("build_visualizador_cc", _RUTA_BV)
bv = importlib.util.module_from_spec(_spec)
sys.modules["build_visualizador_cc"] = bv
_spec.loader.exec_module(bv)

HEADERS_MASTER = [
    "N° Ref.", "Proyecto", "Tipo de Proyecto", "Fecha", "N° Documento",
    "Tipo Documento", "Proveedor", "Proveedor (Razón Social)", "Categoría",
    "Resumen Ítems", "Total sin IVA (CLP)", "IVA 19% (CLP)",
    "Total con IVA (CLP)", "Estado", "Archivo origen", "Fecha modificación",
]
HEADERS_DETALLE = [
    "N° Ref.", "Proyecto", "Tipo de Proyecto", "N° Documento", "Nombre Ítem",
    "Descripción", "Categoría Ítem", "Cantidad", "P. Unitario sin IVA",
    "Total sin IVA (CLP)", "Total con IVA (CLP)",
]


def _wb_con_un_documento(tmp_path, archivo_origen):
    wb = openpyxl.Workbook()
    ws_master = wb.active
    ws_master.title = "Master"
    for c, h in enumerate(HEADERS_MASTER, 1):
        ws_master.cell(row=1, column=c, value=h)
    fila = {
        "N° Ref.": "TEST-001", "Proyecto": "UMAG", "Tipo de Proyecto": "I+D+i",
        "Fecha": None, "N° Documento": "123", "Tipo Documento": "Factura",
        "Proveedor": "Proveedor", "Proveedor (Razón Social)": "Proveedor SpA",
        "Categoría": "Materiales", "Resumen Ítems": "Item",
        "Total sin IVA (CLP)": 1000, "IVA 19% (CLP)": 190,
        "Total con IVA (CLP)": 1190, "Estado": "Pagado",
        "Archivo origen": archivo_origen, "Fecha modificación": None,
    }
    for c, h in enumerate(HEADERS_MASTER, 1):
        ws_master.cell(row=2, column=c, value=fila[h])

    ws_detalle = wb.create_sheet("Detalle")
    for c, h in enumerate(HEADERS_DETALLE, 1):
        ws_detalle.cell(row=1, column=c, value=h)

    ruta = tmp_path / "Centro de Costos.xlsx"
    wb.save(str(ruta))
    return ruta


def test_extraer_datos_saneados_incluye_archivo_origen(tmp_path):
    ruta = _wb_con_un_documento(tmp_path, "TEST-001_Proveedor_2026-07-01.jpg")
    data = bv.extraer_datos_saneados(ruta)
    assert data["documentos"][0]["archivo_origen"] == "TEST-001_Proveedor_2026-07-01.jpg"


def test_extraer_datos_saneados_archivo_origen_ausente_es_none(tmp_path):
    ruta = _wb_con_un_documento(tmp_path, None)
    data = bv.extraer_datos_saneados(ruta)
    assert data["documentos"][0]["archivo_origen"] is None


def test_pendiente_coincide_con_el_color_real_de_auditor_centro_costos(tmp_path):
    """La deteccion de "celda roja" de este modulo es una reimplementacion
    independiente de _celda_es_roja (Sistema/auditor_centro_costos.py) --
    nada comparaba que las dos coincidieran sobre los mismos colores reales.
    Encontro un bug real: el match exacto anterior nunca reconocia el ARGB
    que openpyxl realmente devuelve ("00C00000", no "FFC00000") -- el
    dashboard nunca marcaba ningun documento como "pendiente de revision".
    Usa los Font() reales del modulo que sí escribe el Excel, no strings de
    color adivinados, y siempre relee desde disco (como hacen ambos caminos
    en producción) para que openpyxl normalice el ARGB igual que en un
    archivo real."""
    raiz_sistema = Path(__file__).resolve().parents[2] / "Sistema"
    if str(raiz_sistema) not in sys.path:
        sys.path.insert(0, str(raiz_sistema))
    import auditor_centro_costos as acc

    col_n_documento = HEADERS_MASTER.index("N° Documento") + 1
    casos = (
        (acc.ROJO_FONT, True),
        (acc.AZUL_MARINO_FONT, False),
        (acc.NORMAL_FONT, False),
    )
    for font, debe_marcar in casos:
        ruta = _wb_con_un_documento(tmp_path, "TEST-001_Proveedor_2026-07-01.jpg")
        wb = openpyxl.load_workbook(str(ruta))
        wb["Master"].cell(row=2, column=col_n_documento).font = font
        wb.save(str(ruta))

        wb_verificacion = openpyxl.load_workbook(str(ruta))
        cell_recargada = wb_verificacion["Master"].cell(row=2, column=col_n_documento)
        assert acc._celda_es_roja(cell_recargada) is debe_marcar

        data = bv.extraer_datos_saneados(ruta)
        assert data["documentos"][0]["pendiente_revision"] is debe_marcar, (
            f"build_visualizador y auditor_centro_costos discreparon para {font.color.rgb}"
        )
