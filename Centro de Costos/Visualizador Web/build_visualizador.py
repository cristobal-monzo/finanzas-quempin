# -*- coding: utf-8 -*-
"""
build_visualizador.py — genera el visualizador web de Centro de Costos.

No lee `Centro de Costos.xlsx` desde el HTML: este script hace el "paso de
export saneado" (ver ../../Visualizador Web/CLAUDE.md, doc maestro) y luego
incrusta ese export en `template.html` (versionado, sin datos) para producir
un `build/index.html` autocontenido — un solo archivo, sin servidor, listo
para abrir en el navegador o subir como Claude Artifact.

Salidas (ambas gitignoradas, se regeneran completas en cada corrida — nunca
edites nada dentro de `data/` o `build/` a mano):
  data/centro-de-costos.json   — snapshot saneado intermedio (auditable)
  build/index.html             — visualizador final con los datos incrustados

`template.html` (estructura/CSS/JS, sin datos reales) sí está versionado:
las próximas corridas, con más facturas ya registradas, solo necesitan
volver a correr este script — no tocar la plantilla.

Uso:
  python build_visualizador.py
  (o, desde el driver de la skill: python driver.py visualizador)
"""

import base64
import io
import json
import re
import sys
from datetime import datetime
from pathlib import Path

import openpyxl

RAIZ = Path(__file__).resolve().parent          # Centro de Costos/Visualizador Web/
RAIZ_MODULO = RAIZ.parent                        # Centro de Costos/
RUTA_EXCEL = RAIZ_MODULO / "Excel" / "Centro de Costos.xlsx"
RUTA_TEMPLATE = RAIZ / "template.html"
RUTA_DATA_JSON = RAIZ / "data" / "centro-de-costos.json"
RUTA_BUILD_HTML = RAIZ / "build" / "index.html"

REF_RE = re.compile(r"^[A-Z]+-\d+$")
# Debe coincidir con ROJO ("C00000") en Sistema/auditor_centro_costos.py, el
# unico lugar que realmente escribe esta fuente (ROJO_FONT) para marcar una
# celda "requiere revision" -- si ese valor cambia, actualizar tambien aca.
# Bug real encontrado por test_pendiente_coincide_con_el_color_real_de_
# auditor_centro_costos: openpyxl NO siempre antepone "FF" de alpha al
# releer un .xlsx guardado -- Font(color="C00000") vuelve como "00C00000",
# no "FFC00000". El match exacto contra un set fijo (antes {"FFFF0000",
# "FFC00000"}, ninguno de los dos el real) nunca detectaba una celda roja
# de verdad. Mismo criterio "endswith" que ya usa _celda_es_roja en
# auditor_centro_costos.py, que por eso nunca tuvo este bug.
ROJO_SUFIJO = "C00000"


def extraer_datos_saneados(ruta_excel=RUTA_EXCEL):
    """Lee Master + Detalle y arma el snapshot saneado (sin razón social
    en el nivel de tabla — esa sí se incluye porque el visualizador
    de Detalles la muestra al expandir cada fila, resuelto con el usuario
    2026-07-19: tag corto en la tabla, razón social completa en el detalle).

    `ruta_excel` es parametrizable para poder testear contra un workbook
    temporal en vez del Excel real (que tiene datos financieros reales)."""
    wb = openpyxl.load_workbook(str(ruta_excel), data_only=True)
    ws_master = wb["Master"]
    ws_detalle = wb["Detalle"]

    headers_m = [c.value for c in ws_master[1]]
    docs = []
    for row in ws_master.iter_rows(min_row=2):
        ref_val = row[0].value
        if not isinstance(ref_val, str) or not REF_RE.match(ref_val):
            continue
        d = {headers_m[i]: row[i].value for i in range(len(headers_m))}

        pendiente = False
        for cell in row:
            color = cell.font.color.rgb if (cell.font and cell.font.color) else None
            if isinstance(color, str) and color.upper().endswith(ROJO_SUFIJO):
                pendiente = True
                break

        fecha = d.get("Fecha")
        fecha_iso = fecha.strftime("%Y-%m-%d") if isinstance(fecha, datetime) else None
        fecha_mod = d.get("Fecha modificación")

        docs.append({
            "ref": d.get("N° Ref."),
            "proyecto": d.get("Proyecto"),
            "tipo_proyecto": d.get("Tipo de Proyecto"),
            "fecha": fecha_iso,
            "fecha_modificacion": str(fecha_mod) if fecha_mod is not None else None,
            "n_documento": str(d.get("N° Documento")) if d.get("N° Documento") is not None else None,
            "tipo_documento": d.get("Tipo Documento"),
            "proveedor_tag": d.get("Proveedor"),
            "proveedor_razon_social": d.get("Proveedor (Razón Social)"),
            "categoria": d.get("Categoría"),
            "resumen_items": d.get("Resumen Ítems"),
            "total_sin_iva": d.get("Total sin IVA (CLP)"),
            "iva": d.get("IVA 19% (CLP)"),
            "total_con_iva": d.get("Total con IVA (CLP)"),
            "estado": d.get("Estado"),
            "pendiente_revision": pendiente,
            "archivo_origen": d.get("Archivo origen"),
            "items": [],
        })

    by_ref = {d["ref"]: d for d in docs}
    headers_d = [c.value for c in ws_detalle[1]]
    for row in ws_detalle.iter_rows(min_row=2, values_only=True):
        if not isinstance(row[0], str) or not REF_RE.match(row[0]):
            continue
        rd = {headers_d[i]: row[i] for i in range(len(headers_d))}
        doc = by_ref.get(rd.get("N° Ref."))
        if doc is not None:
            doc["items"].append({
                "nombre_item": rd.get("Nombre Ítem"),
                "descripcion": rd.get("Descripción"),
                "categoria_item": rd.get("Categoría Ítem"),
                "cantidad": rd.get("Cantidad"),
                "p_unitario_sin_iva": rd.get("P. Unitario sin IVA"),
                "total_sin_iva": rd.get("Total sin IVA (CLP)"),
                "total_con_iva": rd.get("Total con IVA (CLP)"),
            })

    # "Total sin IVA" y "Total con IVA" en Master son formulas de Excel
    # (SUMIF / K+L) -- openpyxl nunca las calcula, asi que su valor cacheado
    # queda vacio justo despues de que este mismo script (via 'run') vuelva a
    # guardar el Excel, hasta que alguien lo abra en Excel de verdad. Se
    # recalculan sumando los items de Detalle (valores Python, no formula,
    # siempre confiables) en vez de confiar en la celda de Master -- mismo
    # numero que la formula calcularia, solo que sin depender de si Excel ya
    # recalculo. Encontrado 2026-07-19 corriendo el build justo despues de
    # incorporarlo a 'run' (PASO 12c): el primer intento mostro $0 de gasto
    # total pese a que Detalle tenia los montos correctos.
    for d in docs:
        if d["items"]:
            d["total_sin_iva"] = sum(it["total_sin_iva"] or 0 for it in d["items"])
            d["total_con_iva"] = sum(it["total_con_iva"] or 0 for it in d["items"])
        else:
            d["total_sin_iva"] = d["total_sin_iva"] or 0
            d["total_con_iva"] = d["total_con_iva"] or 0

    fechas_mod = [d["fecha_modificacion"] for d in docs if d["fecha_modificacion"]]
    output = {
        "generado": datetime.now().strftime("%Y-%m-%d %H:%M"),
        "ultima_actualizacion": max(fechas_mod) if fechas_mod else None,
        "kpis": {
            "total_sin_iva": sum(d["total_sin_iva"] or 0 for d in docs),
            "total_iva": sum(d["iva"] or 0 for d in docs),
            "total_con_iva": sum(d["total_con_iva"] or 0 for d in docs),
            "n_documentos": len(docs),
            "n_pendientes": sum(1 for d in docs if d["pendiente_revision"]),
        },
        "documentos": docs,
    }
    return output


def build():
    try:
        sys.stdout.reconfigure(encoding="utf-8", errors="backslashreplace")
    except Exception:
        pass
    if not RUTA_EXCEL.exists():
        print(f"[ERROR] No existe el Excel: {RUTA_EXCEL}")
        return 1
    if not RUTA_TEMPLATE.exists():
        print(f"[ERROR] No existe la plantilla: {RUTA_TEMPLATE}")
        return 1

    data = extraer_datos_saneados()

    RUTA_DATA_JSON.parent.mkdir(parents=True, exist_ok=True)
    with io.open(RUTA_DATA_JSON, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)

    data_json_text = json.dumps(data, ensure_ascii=False)
    data_b64 = base64.b64encode(data_json_text.encode("utf-8")).decode("ascii")

    with io.open(RUTA_TEMPLATE, "r", encoding="utf-8") as f:
        template = f.read()
    if "__CC_DATA_B64__" not in template:
        print("[ERROR] template.html no tiene el placeholder __CC_DATA_B64__")
        return 1
    html = template.replace("__CC_DATA_B64__", data_b64)

    RUTA_BUILD_HTML.parent.mkdir(parents=True, exist_ok=True)
    with io.open(RUTA_BUILD_HTML, "w", encoding="utf-8") as f:
        f.write(html)

    print(f"OK — {len(data['documentos'])} documentos, "
          f"{sum(len(d['items']) for d in data['documentos'])} items, "
          f"gasto total c/IVA ${data['kpis']['total_con_iva']:,}".replace(",", "."))
    print(f"Última actualización de los datos: {data['ultima_actualizacion']}")
    print(f"Snapshot: {RUTA_DATA_JSON}")
    print(f"Visualizador: {RUTA_BUILD_HTML}")
    print("Para verlo: publícalo como Claude Artifact (mismo link de siempre, "
          "ver MEMORY.md de este skill) o ábrelo directo en el navegador.")
    return 0


if __name__ == "__main__":
    sys.stdout.reconfigure(encoding="utf-8", errors="backslashreplace")
    raise SystemExit(build())
