# -*- coding: utf-8 -*-
"""
auditor_centro_costos.py — Registro incremental de facturas/boletas para QUEMPIN SpA.

Estructura "rica" (reconstruida el 2026-07-16 a partir de Centro de Costos.xlsx,
ver reconciliacion_archivos.json para el detalle):
- Detalle: hoja de edicion, una fila por ITEM de linea de cada documento.
- Master: una fila por DOCUMENTO (N Ref), con formulas SUMIF hacia Detalle.
- Una hoja de solo lectura por proyecto, 100% formulas hacia Master.

La hoja _Claude (registro interno de una version anterior del pipeline, ya
perdida) se deja intacta si existe, pero esta version no escribe en ella: no
lo necesita porque nunca vuelve a tocar una fila de datos ya creada (ver
regla siguiente), asi que no hay nada que la corrija a mano.

Reglas:
- El CONTENIDO de una fila de datos ya escrita (items en Detalle, documentos en
  Master) NUNCA se vuelve a tocar una vez creado -- es historial editable a mano
  por el usuario. Su POSICION si se reordena en cada corrida (ver
  reordenar_por_fecha): fila 2 = documento con la fecha mas reciente, fondo de la
  tabla = el mas antiguo. Reordenar mueve valores/formato tal cual, nunca los
  reescribe con datos distintos.
- Las filas de pie (TOTAL GENERAL + leyenda) y las hojas de proyecto SI se regeneran
  en cada corrida, porque son 100% derivadas.
- Backup siempre antes de escribir.
"""

import json
import re
import shutil
import zipfile
from datetime import datetime
from pathlib import Path
from xml.etree import ElementTree as ET

import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.styles.numbers import BUILTIN_FORMATS
from openpyxl.utils import get_column_letter, column_index_from_string

# ── CONFIGURACIÓN ────────────────────────────────────────────────────────────

RAIZ = Path(__file__).resolve().parent
RAIZ_MODULO = RAIZ.parent
# Desde 2026-07-17 la fuente oficial de documentos es el acceso directo de
# OneDrive "Sitio de comunicacion - Centro de Costos 1" (carpeta compartida
# donde suben facturas/boletas los colegas), no la carpeta local "Facturas y
# Boletas/" -- ver CLAUDE.md seccion "Sitio de comunicacion".
RAIZ_SITIO_COMUNICACION = RAIZ_MODULO / "Sitio de comunicación - Centro de Costos 1"
RAIZ_DOCS = RAIZ_SITIO_COMUNICACION / "Facturas y Boletas"
RUTA_EXCEL = RAIZ_MODULO / "Excel" / "Centro de Costos.xlsx"
RUTA_EXCEL_SITIO_COMUNICACION = RAIZ_SITIO_COMUNICACION / "Centro de Costos.xlsx"
RUTA_JSON = RAIZ / "datos_extraidos.json"
RUTA_RECONCILIACION = RAIZ / "reconciliacion_archivos.json"
RUTA_BACKUPS = RAIZ_MODULO / "Excel" / "Respaldos"
RUTA_CORRECCIONES = RAIZ / "correcciones_manuales.json"
RUTA_ERRORES_MD = RAIZ_MODULO / ".claude" / "skills" / "Registro_Centro_de_Costos" / "ERRORES.md"
RAIZ_VISUALIZADOR_WEB = RAIZ_MODULO / "Visualizador Web"
RAIZ_ANALISIS_FINANCIERO = RAIZ_MODULO.parent / "Análisis Financiero"

PREFIJOS_PROYECTO = {
    "UMAG": "UMAG",
    "Cesfam Limache": "CFLI",
    "Cesfam Constitución": "CCON",
    "Gastos Generales": "GGEN",
    "Microturbina LER": "MLER",
}

PALETA = [
    "FFB7CE", "FFBE9D", "E9CF87", "B6DFA0", "86E6D3", "89DFFF",
    "B9CFFF", "E9BFFC",
]

# Tag corto (mas facil de reconocer visualmente) por razon social completa del
# proveedor. La razon social completa siempre se conserva en la columna oculta
# "Proveedor (Razon Social)" -- este diccionario solo controla que se MUESTRA
# en la columna "Proveedor". Curado a mano para los proveedores ya conocidos;
# para uno nuevo que no este aqui, generar_tag_proveedor() aplica una
# heuristica automatica (ver esa funcion) -- si el resultado no queda
# representativo, agregar la entrada correcta aqui.
TAGS_PROVEEDOR_CURADOS = {
    "Comercial Beckman SpA": "Beckman",
    "Easy Retail S.A.": "Easy",
    "Air Express Cargo SpA": "Air Express",
    "Sociedad Comercial Patagonica SpA": "Patagónica",
    "Soc. Com. El Estuche Ltda.": "El Estuche",
    "Comercial e Inversiones Crosur Ltda.": "Crosur",
    "Esteban Guic y Cia. Ltda. (RECASUR)": "RECASUR",
    "Ferreteria El Aguila SpA": "El Águila",
    "Sebastian Prado y Compania Ltda. (BOLT)": "BOLT",
    "Empresas Tur Bus": "Tur Bus",
    "Estaciones de Servicios Fandos Ltda. (Shell Ruta 68)": "Shell",
    "ACO S.A.": "ACO",
    "Danus Conexiones SpA": "Danus",
    "Undurraga Tecnica y Comercial S.A. (UTECSA)": "UTECSA",
    "LATAM Airlines Group S.A.": "LATAM",
    "Ortuzar SpA (Dezar Rent a Car)": "Dezar",
    "Engas Chile SpA": "Engas",
    "Antonio Ruiz Ch. e Hijos Ltda.": "Antonio Ruiz",
    "Comercial Anwo S.A.": "Anwo",
}

_SUFIJOS_LEGALES_RE = re.compile(
    r"\b(S\.?A\.?|SpA|Ltda\.?|E\.?I\.?R\.?L\.?|Group)\b\.?", re.IGNORECASE
)
_PALABRAS_GENERICAS_TAG = {
    "comercial", "sociedad", "soc", "com", "empresas", "compañía", "compania",
    "cia", "inversiones", "e", "y", "de", "del", "la", "el", "los", "las",
    "servicios", "grupo", "group",
}


def generar_tag_proveedor(razon_social):
    """Deriva un tag corto y representativo de una razon social completa.
    1) Si esta en TAGS_PROVEEDOR_CURADOS, se usa ese (fuente de verdad manual).
    2) Si el nombre trae una marca entre parentesis (ej. "... (Shell Ruta 68)"),
       se usa la primera palabra de ese parentesis (la marca, sin el
       descriptor que suele acompañarla).
    3) Si no, se limpia la razon social de sufijos legales (SpA, S.A., Ltda.,
       etc.) y palabras genericas (Comercial, Sociedad, Inversiones, ...) y se
       toman las 1-2 palabras significativas que queden.
    Es un fallback heuristico para proveedores nuevos -- si el resultado no es
    representativo, agregar la entrada correcta a TAGS_PROVEEDOR_CURADOS."""
    if razon_social in TAGS_PROVEEDOR_CURADOS:
        return TAGS_PROVEEDOR_CURADOS[razon_social]

    m = re.search(r"\(([^)]+)\)", razon_social)
    if m:
        contenido = m.group(1).strip()
        if contenido:
            return contenido.split()[0]

    base = _SUFIJOS_LEGALES_RE.sub("", razon_social)
    palabras = [
        p for p in re.split(r"\s+", base.strip(" .,"))
        if p and p.strip(".,").lower() not in _PALABRAS_GENERICAS_TAG
    ]
    tag = " ".join(palabras[:2]).strip(" .,")
    return tag or razon_social

EXTENSIONES_VALIDAS = {".png", ".jpg", ".jpeg", ".heic", ".pdf"}
EXTENSIONES_IGNORAR = {".html", ".txt", ".ini", ".tmp"}

ENCABEZADOS_MASTER = [
    "N° Ref.", "Proyecto", "Tipo de Proyecto", "Fecha", "N° Documento",
    "Tipo Documento", "Proveedor", "Proveedor (Razón Social)", "Categoría",
    "Resumen Ítems", "Total sin IVA (CLP)", "IVA 19% (CLP)",
    "Total con IVA (CLP)", "Estado", "Archivo origen", "Fecha modificación",
]
ENCABEZADOS_DETALLE = [
    "N° Ref.", "Proyecto", "Tipo de Proyecto", "N° Documento", "Nombre Ítem",
    "Descripción", "Categoría Ítem", "Cantidad", "P. Unitario sin IVA",
    "Total sin IVA (CLP)", "Total con IVA (CLP)",
]
ENCABEZADOS_PROYECTO = [
    "N° Ref.", "Proyecto", "Tipo de Proyecto", "Fecha", "N° Documento",
    "Tipo Documento", "Proveedor", "Proveedor (Razón Social)", "Categoría",
    "Resumen Ítems", "Total sin IVA (CLP)", "Total con IVA (CLP)", "Estado",
]

# Columna (1-indexada) de "Proveedor (Razón Social)" en Master y en las hojas
# de proyecto -- se usa para ocultarla y para el mapeo de columnas al migrar.
COL_PROVEEDOR_TAG_MASTER = 7
COL_PROVEEDOR_RAZON_SOCIAL_MASTER = 8

LEYENDA_MASTER = [
    "Cursiva = celda editable a mano",
    "Rojo = valor que requiere revisión (pasa el cursor por la celda para ver el motivo)",
    "Azul marino = valor corregido a mano por ti (Claude lo respeta y no lo sobreescribe)",
    "Fondo de color = proyecto (se pinta solo según la columna 'Proyecto'; si la cambias, "
    "el color se actualiza) · la foto de cada documento lleva su N° como nombre de archivo",
    "⚠️ 'Total sin IVA' y 'Total con IVA' se calculan desde Detalle. 'Archivo origen' y "
    "'Fecha modificación' son el registro de lo ya procesado: no editar.",
]
LEYENDA_DETALLE = [
    "Cursiva = celda editable a mano",
    "Rojo = valor que requiere revisión (pasa el cursor por la celda para ver el motivo)",
    "Azul marino = valor corregido a mano por ti (Claude lo respeta y no lo sobreescribe)",
    "Fondo de color = proyecto (se pinta solo según la columna 'Proyecto'; si la cambias, "
    "el color se actualiza) · la foto de cada documento lleva su N° como nombre de archivo",
    "✔ Ésta es la hoja de edición: Master y las hojas de proyecto se calculan desde aquí.",
]
LEYENDA_PROYECTO = [
    "Cursiva = celda editable a mano",
    "Rojo = valor que requiere revisión (pasa el cursor por la celda para ver el motivo)",
    "Azul marino = valor corregido a mano por ti (Claude lo respeta y no lo sobreescribe)",
    "Fondo de color = proyecto (se pinta solo según la columna 'Proyecto'; si la cambias, "
    "el color se actualiza) · la foto de cada documento lleva su N° como nombre de archivo",
    "⚠️ Vista de sólo lectura: ninguna celda es editable. Editar los ítems en la hoja Detalle.",
]

PATRON_NREF = re.compile(r"^[A-Za-zÁÉÍÓÚÑ]+-\d+$")

NAVY = "1F4E79"
NAVY_OSCURO = "1F3864"
ROJO = "C00000"
BLANCO = "FFFFFF"
NEGRO = "000000"

HEADER_FILL = PatternFill("solid", fgColor=NAVY)
HEADER_FONT = Font(name="Calibri", bold=True, size=11, color=BLANCO)
NORMAL_FONT = Font(name="Calibri", size=11, color=NEGRO)
ROJO_FONT = Font(name="Calibri", size=11, color=ROJO)
AZUL_MARINO_FONT = Font(name="Calibri", size=11, color=NAVY_OSCURO)

# Columnas de Master que se repiten en Detalle -- si una correccion manual
# confirmada toca una de estas columnas en Master, se propaga tambien a la
# columna correspondiente de Detalle (una fila por cada item del N Ref).
CAMPOS_PROPAGADOS_A_DETALLE = {5: 4}  # N Documento: Master col E -> Detalle col D
MONEY_FORMAT = '"$"#,##0'
# numFmtId 14 = "Fecha corta" nativo de Excel (Formato de celdas > Fecha > Fecha
# corta): se adapta a la configuracion regional de quien abre el archivo, en vez
# de fijar un orden de dia/mes/año como hacia el string "DD/MM/YYYY" anterior
# -- pedido del usuario 2026-07-17.
DATE_FORMAT = BUILTIN_FORMATS[14]
THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)


# ── UTILIDADES DE FORMATO ────────────────────────────────────────────────────

def formato_encabezado(ws, ncols):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=1, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER


def pintar_fila(ws, row, ncols, hex_color):
    fill = PatternFill("solid", fgColor=hex_color)
    for c in range(1, ncols + 1):
        ws.cell(row=row, column=c).fill = fill


def ajustar_anchos(ws):
    """Autoajusta el ancho SOLO de columnas que todavia no tienen un ancho fijado.
    Si la columna ya tiene width (fijado por el script en una corrida anterior, o a
    mano por el usuario en Excel), se respeta y no se vuelve a tocar -- asi el
    usuario puede angostar/ensanchar columnas en Excel sin que el proximo run se lo
    pise."""
    for col in ws.columns:
        column_letter = get_column_letter(col[0].column)
        dim = ws.column_dimensions[column_letter]
        if dim.width:
            continue
        max_length = 0
        for cell in col:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except TypeError:
                pass
        dim.width = min(max(max_length + 2, 10), 40)


def escribir_leyenda(ws, fila, textos, ncols):
    colores = [NAVY, ROJO, NAVY_OSCURO, NAVY, NAVY]
    for i, texto in enumerate(textos):
        cell = ws.cell(row=fila + i, column=1, value=texto)
        cell.font = Font(name="Calibri", size=9, italic=(i == 0), color=colores[i % len(colores)])
    return fila + len(textos)


# ── BACKUP ───────────────────────────────────────────────────────────────────

MESES_ES = {
    1: "Enero", 2: "Febrero", 3: "Marzo", 4: "Abril", 5: "Mayo", 6: "Junio",
    7: "Julio", 8: "Agosto", 9: "Septiembre", 10: "Octubre", 11: "Noviembre", 12: "Diciembre",
}


def carpeta_mes(ruta_backups, fecha):
    """Subcarpeta del mes correspondiente (ej. 'Julio 2026') dentro de ruta_backups,
    creandola si todavia no existe."""
    ruta = ruta_backups / f"{MESES_ES[fecha.month]} {fecha.year}"
    ruta.mkdir(parents=True, exist_ok=True)
    return ruta


def hacer_backup(ruta_excel, ruta_backups=RUTA_BACKUPS):
    if not ruta_excel.exists():
        print(f"[WARN] El archivo {ruta_excel} no existe, se creara desde cero.")
        return None
    ahora_dt = datetime.now()
    ruta_carpeta_mes = carpeta_mes(ruta_backups, ahora_dt)
    ahora = ahora_dt.strftime("%Y-%m-%d %H%M")
    ruta_backup = ruta_carpeta_mes / f"Centro de Costos - backup {ahora}.xlsx"
    shutil.copy2(str(ruta_excel), str(ruta_backup))
    print(f"[OK] Backup creado: Respaldos/{ruta_carpeta_mes.name}/{ruta_backup.name}")
    return ruta_backup


# ── LECTURA DE ESTADO EXISTENTE ─────────────────────────────────────────────

def ultima_fila_datos(ws):
    """Ultima fila (desde la 2) cuya columna A calza con el patron N Ref (PROYECTO-NNN)."""
    fila = 1
    r = 2
    while True:
        v = ws.cell(row=r, column=1).value
        if isinstance(v, str) and PATRON_NREF.match(v):
            fila = r
            r += 1
        else:
            break
    return fila


def leer_master(ws_master):
    """
    Devuelve:
    - filas: lista de dicts por documento ya registrado (fila, n_ref, proyecto, ...)
    - max_seq: {prefijo: numero_max_usado}
    - docs_registrados: set de N Documento (tal cual, mas version sin ceros a la izquierda)
    """
    filas = []
    max_seq = {}
    docs_registrados = set()

    ultima = ultima_fila_datos(ws_master)
    for r in range(2, ultima + 1):
        n_ref = ws_master.cell(row=r, column=1).value
        if not isinstance(n_ref, str) or not PATRON_NREF.match(n_ref):
            continue
        prefijo, seq = n_ref.rsplit("-", 1)
        max_seq[prefijo] = max(max_seq.get(prefijo, 0), int(seq))

        proyecto = ws_master.cell(row=r, column=2).value
        archivo_origen = ws_master.cell(row=r, column=15).value
        n_doc = ws_master.cell(row=r, column=5).value
        if n_doc:
            docs_registrados.add(str(n_doc))
            docs_registrados.add(normalizar_n_documento(n_doc))

        filas.append({
            "fila": r, "n_ref": n_ref, "proyecto": proyecto,
            "archivo_origen": archivo_origen,
            "proveedor_tag": ws_master.cell(row=r, column=7).value,
            "fecha": ws_master.cell(row=r, column=4).value,
        })

    return filas, max_seq, docs_registrados


def cargar_reconciliacion():
    if not RUTA_RECONCILIACION.exists():
        return {}
    with open(RUTA_RECONCILIACION, "r", encoding="utf-8") as f:
        return json.load(f).get("mapeo", {})


# ── CORRECCIONES MANUALES (celda roja -> azul marino, con confirmacion) ────
#
# Flujo (ver Sistema/tests/test_correcciones_manuales.py y
# .claude/skills/Registro_Centro_de_Costos/ERRORES.md):
#   1. 'run' detecta (detectar_correcciones_manuales) comparando el backup mas
#      reciente que YA existia contra el estado actual, y las deja anotadas
#      como "Pendiente" (registrar_correcciones_pendientes) -- no toca el
#      Excel todavia.
#   2. Un comando aparte, 'confirmar', aplica: recolorea azul marino oscuro y
#      propaga a Detalle (confirmar_correcciones). Nunca se aplica solo,
#      sin que alguien lo pida explicitamente.

def backup_mas_reciente(ruta_backups=RUTA_BACKUPS):
    """El backup mas reciente que YA existia (por mtime), antes de que esta
    corrida cree el suyo. Es el punto de comparacion para detectar ediciones
    manuales hechas desde la corrida anterior."""
    if not ruta_backups.exists():
        return None
    archivos = list(ruta_backups.rglob("*.xlsx"))
    if not archivos:
        return None
    return max(archivos, key=lambda p: p.stat().st_mtime)


def _mapa_filas_por_n_ref(ws):
    mapa = {}
    for r in range(2, ultima_fila_datos(ws) + 1):
        n_ref = ws.cell(row=r, column=1).value
        if isinstance(n_ref, str) and PATRON_NREF.match(n_ref):
            mapa[n_ref] = r
    return mapa


def _celda_es_roja(cell):
    font = cell.font
    color = font.color.rgb if font and font.color else None
    return isinstance(color, str) and color.upper().endswith(ROJO)


def detectar_correcciones_manuales(wb_anterior, wb_actual):
    """Compara la hoja Master de wb_anterior (backup mas reciente antes de esta
    corrida) contra wb_actual (estado actual, antes de escribir nada nuevo) y
    devuelve las celdas que estaban en rojo en wb_anterior y cambiaron de
    valor -- candidatas a correccion manual. Empareja filas por N Ref, no por
    numero de fila (reordenar_por_fecha puede haber movido las filas)."""
    if "Master" not in wb_anterior.sheetnames or "Master" not in wb_actual.sheetnames:
        return []
    ws_ant = wb_anterior["Master"]
    ws_act = wb_actual["Master"]
    filas_ant = _mapa_filas_por_n_ref(ws_ant)
    filas_act = _mapa_filas_por_n_ref(ws_act)

    detectadas = []
    for n_ref, fila_ant in filas_ant.items():
        fila_act = filas_act.get(n_ref)
        if fila_act is None:
            continue
        for col in range(1, len(ENCABEZADOS_MASTER) + 1):
            cell_ant = ws_ant.cell(row=fila_ant, column=col)
            if not _celda_es_roja(cell_ant):
                continue
            valor_anterior = cell_ant.value
            valor_actual = ws_act.cell(row=fila_act, column=col).value
            if valor_actual in (None, "") or valor_actual == valor_anterior:
                continue
            detectadas.append({
                "n_ref": n_ref, "hoja": "Master", "columna": col,
                "campo": ENCABEZADOS_MASTER[col - 1],
                "valor_anterior": valor_anterior, "valor_corregido": valor_actual,
            })
    return detectadas


def cargar_correcciones_manuales(ruta=RUTA_CORRECCIONES):
    if not ruta.exists():
        return []
    with open(ruta, "r", encoding="utf-8") as f:
        return json.load(f)


def guardar_correcciones_manuales(correcciones, ruta=RUTA_CORRECCIONES):
    with open(ruta, "w", encoding="utf-8") as f:
        json.dump(correcciones, f, ensure_ascii=False, indent=2)


ENCABEZADO_TABLA_CORRECCIONES = "## Correcciones manuales pendientes de recolorear"


def regenerar_tabla_errores_md(correcciones, ruta_errores=RUTA_ERRORES_MD):
    """Reescribe SOLO la tabla bajo ENCABEZADO_TABLA_CORRECCIONES en
    ERRORES.md a partir de correcciones_manuales.json -- es 100% derivada
    (mismo espiritu que los pies de tabla de Master/Detalle), el resto del
    archivo (historial de errores en prosa) no se toca."""
    if not ruta_errores.exists():
        return
    texto = ruta_errores.read_text(encoding="utf-8")
    inicio = texto.find(ENCABEZADO_TABLA_CORRECCIONES)
    if inicio == -1:
        return
    resto = texto[inicio + len(ENCABEZADO_TABLA_CORRECCIONES):]
    m = re.search(r"\n## ", resto)
    fin = inicio + len(ENCABEZADO_TABLA_CORRECCIONES) + (m.start() if m else len(resto))

    filas = sorted(correcciones, key=lambda c: (c["fecha_detectado"], c["n_ref"]))
    lineas = [
        "| Fecha | Hoja | N° Ref. | Campo / Columna | Valor anterior (rojo) | Valor corregido | Estado |",
        "|---|---|---|---|---|---|---|",
    ]
    if filas:
        for c in filas:
            estado = c["estado"] if c["estado"] == "Pendiente" else f"Aplicado ({c.get('fecha_aplicado') or ''})"
            lineas.append(
                f"| {c['fecha_detectado']} | {c['hoja']} | {c['n_ref']} | {c['campo']} | "
                f"{c['valor_anterior']} | {c['valor_corregido']} | {estado} |"
            )
    else:
        lineas.append("| *(sin entradas todavía)* | | | | | | |")

    nuevo_bloque = ENCABEZADO_TABLA_CORRECCIONES + "\n\n" + "\n".join(lineas) + "\n"
    ruta_errores.write_text(texto[:inicio] + nuevo_bloque + texto[fin:], encoding="utf-8")


def registrar_correcciones_pendientes(detectadas, escribir=True,
                                       ruta_correcciones=None, ruta_errores=None):
    """Fusiona las candidatas detectadas con correcciones_manuales.json:
    - Si no habia entrada para (n_ref, columna): la agrega como "Pendiente".
    - Si ya habia una "Pendiente": actualiza su valor_corregido/fecha (el
      usuario volvio a editar la celda antes de confirmar) -- no duplica.
    - Si ya habia una "Aplicado" con el mismo valor_corregido: la ignora (ya
      se confirmo en una corrida anterior; el backup "anterior" de esta
      corrida todavia puede mostrar la celda en rojo, asi que sin este
      chequeo se re-detectaria como nueva indefinidamente).
    Con escribir=False (usado por 'status', de solo lectura) no toca
    correcciones_manuales.json ni ERRORES.md -- solo devuelve el preview."""
    ruta_correcciones = ruta_correcciones or RUTA_CORRECCIONES
    ruta_errores = ruta_errores or RUTA_ERRORES_MD
    hoy = datetime.now().strftime("%Y-%m-%d")
    correcciones = cargar_correcciones_manuales(ruta_correcciones)
    indice = {(c["n_ref"], c["columna"]): c for c in correcciones}

    pendientes_para_reportar = []
    for d in detectadas:
        clave = (d["n_ref"], d["columna"])
        existente = indice.get(clave)
        if existente is None:
            nueva = {
                "n_ref": d["n_ref"], "hoja": d["hoja"], "columna": d["columna"],
                "campo": d["campo"], "valor_anterior": d["valor_anterior"],
                "valor_corregido": d["valor_corregido"], "estado": "Pendiente",
                "fecha_detectado": hoy, "fecha_aplicado": None,
            }
            correcciones.append(nueva)
            indice[clave] = nueva
            pendientes_para_reportar.append(nueva)
        elif existente["estado"] == "Pendiente":
            if existente["valor_corregido"] != d["valor_corregido"]:
                existente["valor_corregido"] = d["valor_corregido"]
                existente["fecha_detectado"] = hoy
            pendientes_para_reportar.append(existente)
        elif existente["estado"] == "Aplicado" and existente["valor_corregido"] == d["valor_corregido"]:
            continue  # ya confirmada en una corrida anterior, no es nueva

    if escribir:
        guardar_correcciones_manuales(correcciones, ruta_correcciones)
        regenerar_tabla_errores_md(correcciones, ruta_errores)

    return pendientes_para_reportar


def _n_ref_a_filas_detalle(ws_detalle, n_ref):
    return [
        r for r in range(2, ultima_fila_datos(ws_detalle) + 1)
        if ws_detalle.cell(row=r, column=1).value == n_ref
    ]


def confirmar_correcciones(objetivo=None, ruta_excel=None, ruta_correcciones=None,
                            ruta_errores=None, ruta_backups=None):
    """objetivo=None -> preview de solo lectura (no toca el Excel).
    objetivo="TODOS" o lista de N Ref -> aplica: recolorea azul marino
    oscuro en Master (y en Detalle si el campo esta en
    CAMPOS_PROPAGADOS_A_DETALLE) y marca la correccion como "Aplicado"."""
    ruta_excel = ruta_excel or RUTA_EXCEL
    ruta_correcciones = ruta_correcciones or RUTA_CORRECCIONES
    ruta_errores = ruta_errores or RUTA_ERRORES_MD
    ruta_backups = ruta_backups or RUTA_BACKUPS
    correcciones = cargar_correcciones_manuales(ruta_correcciones)
    pendientes = [c for c in correcciones if c["estado"] == "Pendiente"]

    if objetivo is None:
        print(f"\nCorrecciones manuales pendientes de confirmar: {len(pendientes)}")
        for c in pendientes:
            print(f"  - {c['n_ref']} / {c['campo']}: '{c['valor_anterior']}' -> '{c['valor_corregido']}'")
        if pendientes:
            print("\nPara aplicar: python driver.py confirmar --todos"
                  " (o 'python driver.py confirmar <N_REF> ...' para solo algunas)")
        return pendientes

    seleccion = pendientes if objetivo == "TODOS" else [c for c in pendientes if c["n_ref"] in set(objetivo)]
    if not seleccion:
        print("\nNo hay correcciones pendientes que coincidan con lo pedido.")
        return []

    if excel_esta_bloqueado(ruta_excel):
        print("\n[ERROR] El archivo esta abierto en Excel (o bloqueado). Cierralo antes de confirmar.")
        return []

    hacer_backup(ruta_excel, ruta_backups)
    wb = openpyxl.load_workbook(str(ruta_excel), data_only=False)
    ws_master = wb["Master"]
    ws_detalle = wb["Detalle"] if "Detalle" in wb.sheetnames else None
    filas_master = _mapa_filas_por_n_ref(ws_master)

    aplicadas = []
    for c in seleccion:
        fila_m = filas_master.get(c["n_ref"])
        if fila_m is None:
            print(f"  [WARN] {c['n_ref']}: ya no existe en Master, se salta.")
            continue
        cell = ws_master.cell(row=fila_m, column=c["columna"])
        if cell.value != c["valor_corregido"] and str(cell.value) != str(c["valor_corregido"]):
            print(f"  [WARN] {c['n_ref']} / {c['campo']}: el valor cambio de nuevo desde que se "
                  f"detecto ('{cell.value}' != '{c['valor_corregido']}'), se salta -- correr 'run' "
                  f"de nuevo para re-detectarlo.")
            continue

        valor_aplicar = c["valor_corregido"]
        if c["columna"] == 5 and isinstance(valor_aplicar, str):
            # N Documento tampoco puede quedar con ceros a la izquierda cuando se
            # corrige a mano (ver MEMORY.md "Reglas de negocio", pedido 2026-07-17):
            # se valida contra lo que el usuario escribio tal cual (arriba), pero se
            # escribe la version normalizada.
            valor_aplicar = normalizar_n_documento(valor_aplicar)
        cell.value = valor_aplicar
        cell.font = AZUL_MARINO_FONT

        col_detalle = CAMPOS_PROPAGADOS_A_DETALLE.get(c["columna"])
        if col_detalle and ws_detalle is not None:
            for fila_d in _n_ref_a_filas_detalle(ws_detalle, c["n_ref"]):
                celda_d = ws_detalle.cell(row=fila_d, column=col_detalle, value=valor_aplicar)
                celda_d.font = AZUL_MARINO_FONT

        c["estado"] = "Aplicado"
        c["fecha_aplicado"] = datetime.now().strftime("%Y-%m-%d")
        aplicadas.append(c)
        propagado = ", propagado a Detalle" if col_detalle else ""
        print(f"  [OK] {c['n_ref']} / {c['campo']} -> '{c['valor_corregido']}' (azul marino{propagado}).")

    if aplicadas:
        try:
            wb.save(str(ruta_excel))
            suprimir_aviso_numero_texto(ruta_excel, _hojas_columna_n_documento(wb))
        except PermissionError:
            print("\n[ERROR] El archivo esta abierto en Excel. Cierralo y vuelve a confirmar.")
            return []
        guardar_correcciones_manuales(correcciones, ruta_correcciones)
        regenerar_tabla_errores_md(correcciones, ruta_errores)

    print(f"\n{len(aplicadas)} correccion(es) aplicada(s).")
    return aplicadas


# ── REVISIÓN GUIADA DE ERRORES (skill Revision_de_Errores) ─────────────────
#
# Camino distinto al de arriba: en vez de que el usuario edite la celda a
# mano en Excel y 'run' la detecte comparando contra el backup anterior, el
# agente recorre las celdas rojas UNA A UNA en la conversacion (mostrando el
# documento origen), pregunta el valor correcto, y lo aplica de inmediato con
# corregir_valor_manual -- mismo resultado final (celda azul marino, propagada
# a Detalle, registrada como "Aplicado" en correcciones_manuales.json y
# ERRORES.md) sin pasar por el estado "Pendiente".

# Unicas columnas de Master que el script pinta de rojo hoy (ver
# escribir_fila_master): N Documento ilegible/"S/N" y IVA que no cuadra con
# el 19% del Neto. Si en el futuro se pinta de rojo otra columna, agregarla
# aqui para que listar_celdas_rojas() la detecte.
COLUMNAS_REVISABLES = (5, 12)


def listar_celdas_rojas(ws_master):
    """Todas las celdas de Master actualmente en rojo (requieren revision),
    sin importar si ya fueron detectadas comparando backups -- es la fuente
    del recorrido guiado de la skill Revision_de_Errores. Una entrada por
    celda roja: {n_ref, fila, columna, campo, valor_actual, proyecto,
    archivo_origen}."""
    encontradas = []
    for r in range(2, ultima_fila_datos(ws_master) + 1):
        n_ref = ws_master.cell(row=r, column=1).value
        if not isinstance(n_ref, str) or not PATRON_NREF.match(n_ref):
            continue
        for col in COLUMNAS_REVISABLES:
            cell = ws_master.cell(row=r, column=col)
            if _celda_es_roja(cell):
                encontradas.append({
                    "n_ref": n_ref, "fila": r, "columna": col,
                    "campo": ENCABEZADOS_MASTER[col - 1],
                    "valor_actual": cell.value,
                    "proyecto": ws_master.cell(row=r, column=2).value,
                    "archivo_origen": ws_master.cell(row=r, column=15).value,
                })
    return encontradas


# Convencion ya usada en el modulo desde CCON-004 (ver MEMORY.md / ERRORES.md):
# cuando una parte de la factura es fisicamente ilegible, esas lineas se
# agrupan en 1 item aparte cuyo nombre incluye la palabra "varios" (ej.
# "Materiales varios"). No hay celda pintada de ningun color para este caso
# (a diferencia de listar_celdas_rojas) -- se detecta por el nombre del item.
PATRON_ITEM_AGRUPADO = re.compile(r"\bvarios\b", re.IGNORECASE)


def listar_items_agrupados(ws_detalle):
    """Filas de Detalle que agrupan en 1 sola linea una parte de la compra
    que no se pudo desglosar item por item (ver PATRON_ITEM_AGRUPADO) -- es
    la fuente de la parte "items agrupados" del recorrido de
    Revision_de_Errores, complementaria a listar_celdas_rojas. Una entrada
    por fila agrupada: {n_ref, fila, proyecto, nombre_item, descripcion,
    cantidad, p_unitario_sin_iva, total_sin_iva}."""
    encontradas = []
    for r in range(2, ultima_fila_datos(ws_detalle) + 1):
        n_ref = ws_detalle.cell(row=r, column=1).value
        if not isinstance(n_ref, str) or not PATRON_NREF.match(n_ref):
            continue
        nombre_item = ws_detalle.cell(row=r, column=5).value
        if isinstance(nombre_item, str) and PATRON_ITEM_AGRUPADO.search(nombre_item):
            encontradas.append({
                "n_ref": n_ref, "fila": r,
                "proyecto": ws_detalle.cell(row=r, column=2).value,
                "nombre_item": nombre_item,
                "descripcion": ws_detalle.cell(row=r, column=6).value,
                "cantidad": ws_detalle.cell(row=r, column=8).value,
                "p_unitario_sin_iva": ws_detalle.cell(row=r, column=9).value,
                "total_sin_iva": ws_detalle.cell(row=r, column=10).value,
            })
    return encontradas


def corregir_valor_manual(n_ref, columna, valor_nuevo, ruta_excel=None,
                           ruta_correcciones=None, ruta_errores=None,
                           ruta_backups=None):
    """Aplica UNA correccion manual directa (valor ya confirmado por el
    usuario en la conversacion, no detectado comparando backups): backup,
    escribe valor_nuevo en Master (fila de n_ref, columna dada), recolorea la
    fuente a azul marino oscuro, propaga a Detalle si la columna esta en
    CAMPOS_PROPAGADOS_A_DETALLE, y deja el cambio como "Aplicado" en
    correcciones_manuales.json + ERRORES.md -- mismo destino final que
    confirmar_correcciones(), pero en un solo paso.

    Solo opera sobre celdas actualmente en rojo: si la celda de
    (n_ref, columna) no esta en rojo, no toca nada y devuelve None (evita
    pisar un dato ya bueno por un N Ref o columna equivocados)."""
    ruta_excel = ruta_excel or RUTA_EXCEL
    ruta_correcciones = ruta_correcciones or RUTA_CORRECCIONES
    ruta_errores = ruta_errores or RUTA_ERRORES_MD
    ruta_backups = ruta_backups or RUTA_BACKUPS

    if excel_esta_bloqueado(ruta_excel):
        print("\n[ERROR] El archivo esta abierto en Excel (o bloqueado). Cierralo antes de corregir.")
        return None

    wb = openpyxl.load_workbook(str(ruta_excel), data_only=False)
    if "Master" not in wb.sheetnames:
        print("\n[ERROR] No existe la hoja Master.")
        return None
    ws_master = wb["Master"]
    fila_m = _mapa_filas_por_n_ref(ws_master).get(n_ref)
    if fila_m is None:
        print(f"\n[ERROR] {n_ref} no existe en Master.")
        return None

    cell = ws_master.cell(row=fila_m, column=columna)
    if not _celda_es_roja(cell):
        print(f"\n[WARN] {n_ref} / columna {columna} no esta en rojo (no requiere revision); no se toca.")
        return None

    if columna == 5 and isinstance(valor_nuevo, str):
        # N Documento tampoco puede quedar con ceros a la izquierda cuando se
        # corrige a mano (ver MEMORY.md "Reglas de negocio", pedido 2026-07-17).
        valor_nuevo = normalizar_n_documento(valor_nuevo)

    valor_anterior = cell.value
    hacer_backup(ruta_excel, ruta_backups)
    cell.value = valor_nuevo
    cell.font = AZUL_MARINO_FONT

    ws_detalle = wb["Detalle"] if "Detalle" in wb.sheetnames else None
    col_detalle = CAMPOS_PROPAGADOS_A_DETALLE.get(columna)
    if col_detalle and ws_detalle is not None:
        for fila_d in _n_ref_a_filas_detalle(ws_detalle, n_ref):
            ws_detalle.cell(row=fila_d, column=col_detalle, value=valor_nuevo).font = AZUL_MARINO_FONT

    try:
        wb.save(str(ruta_excel))
        suprimir_aviso_numero_texto(ruta_excel, _hojas_columna_n_documento(wb))
    except PermissionError:
        print("\n[ERROR] El archivo esta abierto en Excel. Cierralo y vuelve a intentar.")
        return None

    hoy = datetime.now().strftime("%Y-%m-%d")
    campo = ENCABEZADOS_MASTER[columna - 1]
    correcciones = cargar_correcciones_manuales(ruta_correcciones)
    entrada = next((c for c in correcciones if c["n_ref"] == n_ref and c["columna"] == columna), None)
    if entrada is None:
        entrada = {"n_ref": n_ref, "hoja": "Master", "columna": columna, "campo": campo}
        correcciones.append(entrada)
    entrada.update({
        "valor_anterior": valor_anterior, "valor_corregido": valor_nuevo,
        "estado": "Aplicado", "fecha_detectado": entrada.get("fecha_detectado", hoy),
        "fecha_aplicado": hoy,
    })
    guardar_correcciones_manuales(correcciones, ruta_correcciones)
    regenerar_tabla_errores_md(correcciones, ruta_errores)

    propagado = ", propagado a Detalle" if col_detalle else ""
    print(f"  [OK] {n_ref} / {campo}: '{valor_anterior}' -> '{valor_nuevo}' (azul marino{propagado}).")
    return entrada


def desglosar_item_agrupado(n_ref, items_nuevos, ruta_excel=None, ruta_correcciones=None,
                             ruta_errores=None, ruta_backups=None):
    """Reemplaza la UNICA fila de Detalle 'agrupada' de n_ref (ver
    listar_items_agrupados) por una fila nueva por cada item de
    items_nuevos -- cada uno un dict {"nombre_item", "descripcion",
    "categoria_item", "cantidad", "p_unitario_sin_iva"}, ya confirmado por el
    usuario en la conversacion (misma logica que corregir_valor_manual, pero
    para Detalle en vez de una celda de Master).

    Recalcula "Total sin IVA"/"Total con IVA" de cada fila nueva con la
    misma formula que escribir_items_detalle (tasa real del documento = IVA
    de Master, que no se toca, sobre el Total sin IVA recalculado del
    documento con el desglose nuevo), reconstruye "Resumen Item" en Master
    leyendo Detalle ya actualizado, y regenera el pie de Detalle (la cantidad
    de filas de la hoja cambia salvo que items_nuevos traiga exactamente 1
    elemento). Las filas nuevas quedan en azul marino (valor corregido a
    mano), heredando el relleno de color de proyecto de la fila que
    reemplazan. Excepcion deliberada a la regla de oro de "nunca tocar una
    fila ya escrita" -- mismo tipo de excepcion que corregir_valor_manual.

    Solo actua si encuentra EXACTAMENTE una fila agrupada para n_ref -- si no
    encuentra ninguna, o mas de una, no toca nada y devuelve None (evita
    adivinar cual fila reemplazar; un documento con mas de 1 fila agrupada
    hoy no esta soportado por este comando, se resuelve a mano)."""
    from copy import copy as _copy

    ruta_excel = ruta_excel or RUTA_EXCEL
    ruta_correcciones = ruta_correcciones or RUTA_CORRECCIONES
    ruta_errores = ruta_errores or RUTA_ERRORES_MD
    ruta_backups = ruta_backups or RUTA_BACKUPS

    if excel_esta_bloqueado(ruta_excel):
        print("\n[ERROR] El archivo esta abierto en Excel (o bloqueado). Cierralo antes de desglosar.")
        return None

    wb = openpyxl.load_workbook(str(ruta_excel), data_only=False)
    if "Master" not in wb.sheetnames or "Detalle" not in wb.sheetnames:
        print("\n[ERROR] Faltan hojas Master/Detalle.")
        return None
    ws_master = wb["Master"]
    ws_detalle = wb["Detalle"]

    fila_master = _mapa_filas_por_n_ref(ws_master).get(n_ref)
    if fila_master is None:
        print(f"\n[ERROR] {n_ref} no existe en Master.")
        return None

    agrupados = [f for f in listar_items_agrupados(ws_detalle) if f["n_ref"] == n_ref]
    if len(agrupados) != 1:
        print(f"\n[ERROR] Se esperaba exactamente 1 fila agrupada para {n_ref} en Detalle, "
              f"se encontraron {len(agrupados)}.")
        return None
    fila_vieja = agrupados[0]["fila"]
    nombre_item_anterior = agrupados[0]["nombre_item"]
    descripcion_anterior = agrupados[0]["descripcion"]

    ncols_detalle = len(ENCABEZADOS_DETALLE)
    ultima_detalle = ultima_fila_datos(ws_detalle)
    snapshot_vieja = capturar_fila(ws_detalle, fila_vieja, ncols_detalle)
    fill_proyecto = snapshot_vieja[0]["fill"]  # relleno de color por proyecto, igual en toda la fila
    n_doc_str = ws_detalle.cell(row=fila_vieja, column=4).value
    tipo_proyecto = ws_detalle.cell(row=fila_vieja, column=3).value

    # Total sin IVA del documento CON el desglose nuevo (todas las demas filas
    # de este n_ref, mas los items nuevos) -- para recalcular la tasa real de
    # IVA del documento, igual que escribir_items_detalle.
    total_otros_items = sum(
        (ws_detalle.cell(row=r, column=10).value or 0)
        for r in range(2, ultima_detalle + 1)
        if ws_detalle.cell(row=r, column=1).value == n_ref and r != fila_vieja
    )
    total_items_nuevos = sum(it["cantidad"] * it["p_unitario_sin_iva"] for it in items_nuevos)
    total_sin_iva_doc = total_otros_items + total_items_nuevos

    iva_doc = ws_master.cell(row=fila_master, column=12).value or 0
    tasa_iva_doc = (iva_doc / total_sin_iva_doc) if total_sin_iva_doc else 0

    hacer_backup(ruta_excel, ruta_backups)

    snapshots_todas = [capturar_fila(ws_detalle, r, ncols_detalle) for r in range(2, ultima_detalle + 1)]
    indice_vieja = fila_vieja - 2

    nuevas_filas_snapshot = []
    for item in items_nuevos:
        total_item = item["cantidad"] * item["p_unitario_sin_iva"]
        total_item_con_iva = round(total_item * (1 + tasa_iva_doc))
        valores = [
            n_ref, agrupados[0]["proyecto"], tipo_proyecto, n_doc_str,
            item["nombre_item"], item.get("descripcion", ""), item.get("categoria_item", ""),
            item["cantidad"], item["p_unitario_sin_iva"], total_item, total_item_con_iva,
        ]
        nuevas_filas_snapshot.append([
            {
                "value": v,
                "font": _copy(AZUL_MARINO_FONT),
                "fill": _copy(fill_proyecto),
                "border": _copy(THIN_BORDER),
                "alignment": _copy(snapshot_vieja[c - 1]["alignment"]),
                "number_format": MONEY_FORMAT if c in (9, 10, 11) else "General",
            }
            for c, v in enumerate(valores, 1)
        ])

    filas_finales = (
        snapshots_todas[:indice_vieja] + nuevas_filas_snapshot + snapshots_todas[indice_vieja + 1:]
    )
    for i, snap in enumerate(filas_finales):
        escribir_snapshot_fila(ws_detalle, 2 + i, snap)

    primera_fila_libre_detalle = 2 + len(filas_finales)
    limpiar_pie(ws_detalle)
    regenerar_pie(ws_detalle, ncols_detalle, [10, 11], "TOTAL GENERAL", primera_fila_libre_detalle, LEYENDA_DETALLE)

    # "Resumen Item" en Master: reconstruido leyendo Detalle ya actualizado
    # (no reemplazo de texto) para no depender de que el nombre viejo sea unico.
    nombres_finales = [
        ws_detalle.cell(row=r, column=5).value
        for r in range(2, primera_fila_libre_detalle)
        if ws_detalle.cell(row=r, column=1).value == n_ref
    ]
    ws_master.cell(row=fila_master, column=10, value="; ".join(nombres_finales))

    try:
        wb.save(str(ruta_excel))
        suprimir_aviso_numero_texto(ruta_excel, _hojas_columna_n_documento(wb))
    except PermissionError:
        print("\n[ERROR] El archivo esta abierto en Excel. Cierralo y vuelve a intentar.")
        return None

    hoy = datetime.now().strftime("%Y-%m-%d")
    resumen_nuevo = "; ".join(it["nombre_item"] for it in items_nuevos)
    valor_anterior = f"{nombre_item_anterior} ({descripcion_anterior})" if descripcion_anterior else nombre_item_anterior
    correcciones = cargar_correcciones_manuales(ruta_correcciones)
    entrada = {
        "n_ref": n_ref, "hoja": "Detalle", "columna": 5, "campo": "Ítems agrupados (desglose)",
        "valor_anterior": valor_anterior, "valor_corregido": resumen_nuevo,
        "estado": "Aplicado", "fecha_detectado": hoy, "fecha_aplicado": hoy,
    }
    correcciones.append(entrada)
    guardar_correcciones_manuales(correcciones, ruta_correcciones)
    regenerar_tabla_errores_md(correcciones, ruta_errores)

    print(f"  [OK] {n_ref} / Ítems agrupados: '{nombre_item_anterior}' -> "
          f"{len(items_nuevos)} ítem(s) ({resumen_nuevo}) (azul marino).")
    return entrada


# ── MIGRACIÓN: COLUMNA "PROVEEDOR (RAZÓN SOCIAL)" (2026-07-16) ─────────────

def _desplazar_rango_columna(rango_str, punto_insercion):
    """Desplaza en +1 cualquier letra de columna >= punto_insercion dentro de
    un rango tipo 'H2:H225'. Se usa para corregir validaciones de datos y
    formato condicional despues de insertar una columna con ws.insert_cols
    (openpyxl no ajusta esos rangos solo -- ver "Formato Centro de Costos.md" #8/#9/#14)."""
    def desplazar_letra(letra):
        idx = column_index_from_string(letra)
        return get_column_letter(idx + 1) if idx >= punto_insercion else letra

    m = re.match(r"^([A-Z]+)(\d+):([A-Z]+)(\d+)$", rango_str)
    if not m:
        return rango_str
    c1, r1, c2, r2 = m.groups()
    return f"{desplazar_letra(c1)}{r1}:{desplazar_letra(c2)}{r2}"


def _migrar_rangos_columna(ws, punto_insercion):
    for dv in ws.data_validations.dataValidation:
        dv.sqref = " ".join(
            _desplazar_rango_columna(p, punto_insercion) for p in str(dv.sqref).split()
        )

    # ConditionalFormattingList indexa sus reglas en un dict cuya clave es el
    # propio objeto ConditionalFormatting, con hash derivado de su sqref --
    # mutar cf.sqref en el lugar corrompe ese dict (el hash cambia pero la
    # entrada sigue en el bucket viejo). Hay que reconstruir el dict interno
    # con objetos ConditionalFormatting nuevos en vez de mutar los existentes.
    from openpyxl.formatting.formatting import ConditionalFormatting
    from collections import OrderedDict

    cf_list = ws.conditional_formatting
    nuevas_reglas = OrderedDict()
    for cf, reglas in cf_list._cf_rules.items():
        nuevo_sqref = " ".join(
            _desplazar_rango_columna(p, punto_insercion) for p in str(cf.sqref).split()
        )
        nuevo_cf = ConditionalFormatting(sqref=nuevo_sqref)
        nuevas_reglas[nuevo_cf] = reglas
    cf_list._cf_rules = nuevas_reglas


def migrar_columna_proveedor(wb):
    """Migracion unica (idempotente): inserta 'Proveedor (Razón Social)' junto
    a 'Proveedor' en Master, mueve ahi la razon social completa de cada fila
    ya registrada, y deja en 'Proveedor' el tag corto (TAGS_PROVEEDOR_CURADOS
    o generar_tag_proveedor() como fallback). Es una excepcion deliberada a
    "nunca tocar una fila de datos ya escrita" -- decision del usuario
    2026-07-16, porque la columna 'Proveedor' vieja no se toca como dato
    financiero, solo se resume a un tag visual (la razon social completa
    igual queda preservada, en la columna oculta).

    Debe correr ANTES de leer_master/inventariar, apenas se abre el libro,
    porque insertar una columna desplaza las formulas K/M, las validaciones
    de datos y el formato condicional heredado -- ver "Formato Centro de Costos.md" #14."""
    if "Master" not in wb.sheetnames:
        return
    ws = wb["Master"]
    ya_migrado = ws.cell(row=1, column=COL_PROVEEDOR_RAZON_SOCIAL_MASTER).value == (
        ENCABEZADOS_MASTER[COL_PROVEEDOR_RAZON_SOCIAL_MASTER - 1]
    )
    if ya_migrado:
        return

    print("  Migrando Master: agregando columna oculta 'Proveedor (Razón Social)'...")
    ultima = ultima_fila_datos(ws)

    ws.insert_cols(COL_PROVEEDOR_RAZON_SOCIAL_MASTER)
    _migrar_rangos_columna(ws, COL_PROVEEDOR_RAZON_SOCIAL_MASTER)

    header_cell = ws.cell(
        row=1, column=COL_PROVEEDOR_RAZON_SOCIAL_MASTER,
        value=ENCABEZADOS_MASTER[COL_PROVEEDOR_RAZON_SOCIAL_MASTER - 1],
    )
    header_cell.fill = HEADER_FILL
    header_cell.font = HEADER_FONT
    header_cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    header_cell.border = THIN_BORDER

    migradas = 0
    for r in range(2, ultima + 1):
        celda_proveedor = ws.cell(row=r, column=COL_PROVEEDOR_TAG_MASTER)
        razon_social = celda_proveedor.value
        if not razon_social:
            continue
        celda_razon = ws.cell(row=r, column=COL_PROVEEDOR_RAZON_SOCIAL_MASTER, value=razon_social)
        celda_razon.font = NORMAL_FONT
        celda_razon.border = THIN_BORDER
        celda_proveedor.value = generar_tag_proveedor(razon_social)
        escribir_formulas_master(ws, r)
        migradas += 1

    ws.column_dimensions[get_column_letter(COL_PROVEEDOR_RAZON_SOCIAL_MASTER)].hidden = True
    print(f"  [OK] {migradas} fila(s) migrada(s) (tag en 'Proveedor', razón social movida a columna oculta).")


# Reasignacion de colores por proyecto para la migracion de paleta del
# 2026-07-17 (ver migrar_paleta_colores). Elegida a mano para preservar la
# identidad visual que cada proyecto ya tenia (ej. UMAG se seguia viendo
# rosado antes de la migracion, aunque tapado por el formato condicional
# heredado -- ver Formato Centro de Costos.md SS4/SS9), en vez de reasignar
# en orden alfabetico como hace asignar_colores_proyectos() para proyectos
# nuevos.
REASIGNACION_COLORES_2026_07_17 = {
    "UMAG": "FFB7CE",
    "Cesfam Limache": "89DFFF",
    "Cesfam Constitución": "FFBE9D",
    "Gastos Generales": "B6DFA0",
    "Microturbina LER": "E9CF87",
}


def migrar_paleta_colores(wb, ws_master, ws_detalle):
    """Migracion unica e idempotente (2026-07-17): corrige dos problemas de
    color heredados del pipeline perdido, reportados por el usuario --
    "las hojas no coinciden con el color de las filas, ademas hay proyectos
    con colores demasiado parecidos o los mismos colores":

    1. Master/Detalle tenian 3 reglas de formato condicional heredadas
       (Proyecto = UMAG/Cesfam Limache/Gastos Generales) que pintaban esas
       filas con colores fijos, con prioridad visual sobre el relleno
       directo de celda que sí controla el script -- por eso el color de la
       fila no coincidia con el tabColor de la hoja de ese proyecto. Se
       eliminan: de aqui en adelante el relleno directo (pintar_fila) es la
       unica fuente de color, para cualquier cantidad de proyectos.
    2. La PALETA vieja de 12 pasteles tenia colores casi indistinguibles
       entre si (ej. FCE4D6 vs FBE5D6 vs FDE9D9 -- diferencia de 1 unidad en
       un canal RGB). La PALETA nueva usa 8 tonos espaciados uniformemente
       en el circulo de matices (mismo metodo de color categorico del
       skill dataviz) para que sean distinguibles a simple vista.

    Se detecta si ya corrio revisando si sigue existiendo alguna de las 3
    reglas de formato condicional legado; si no estan, no hace nada."""
    formulas_legado = ('$B2="UMAG"', '$B2="Cesfam Limache"', '$B2="Gastos Generales"')

    def tiene_formato_legado(ws):
        for reglas in ws.conditional_formatting._cf_rules.values():
            for regla in reglas:
                if any(f in str(regla.formula) for f in formulas_legado):
                    return True
        return False

    if not (tiene_formato_legado(ws_master) or tiene_formato_legado(ws_detalle)):
        return

    print("  Migrando colores: quitando formato condicional heredado y repintando con la paleta nueva...")
    from collections import OrderedDict
    ws_master.conditional_formatting._cf_rules = OrderedDict()
    ws_detalle.conditional_formatting._cf_rules = OrderedDict()

    colores = dict(REASIGNACION_COLORES_2026_07_17)
    libres = [c for c in PALETA if c not in colores.values()]

    def color_de(proyecto):
        if proyecto not in colores:
            colores[proyecto] = libres.pop(0) if libres else PALETA[hash(proyecto) % len(PALETA)]
        return colores[proyecto]

    repintadas = 0
    for ws, ncols in ((ws_master, len(ENCABEZADOS_MASTER)), (ws_detalle, len(ENCABEZADOS_DETALLE))):
        ultima = ultima_fila_datos(ws)
        for r in range(2, ultima + 1):
            proyecto = ws.cell(row=r, column=2).value
            if proyecto:
                pintar_fila(ws, r, ncols, color_de(proyecto))
                repintadas += 1

    print(f"  [OK] {repintadas} fila(s) repintada(s) entre Master y Detalle con la paleta nueva.")


def migrar_formato_fecha_corta(ws_master):
    """Migracion de formato (idempotente, 2026-07-17): normaliza la columna Fecha
    (D) de TODAS las filas ya escritas de Master al formato nativo 'Fecha corta'
    de Excel (DATE_FORMAT = numFmtId 14), no solo las filas nuevas -- pedido del
    usuario. Es formato, no contenido: no viola la regla de oro de no tocar filas
    de datos ya escritas (mismo principio que reordenar_por_fecha, que tambien
    toca solo donde/como se ve una fila, nunca su valor). Las hojas de proyecto
    no necesitan esta migracion porque se regeneran completas en cada corrida
    (ver regenerar_hoja_proyecto)."""
    ultima = ultima_fila_datos(ws_master)
    for r in range(2, ultima + 1):
        cell = ws_master.cell(row=r, column=4)
        if cell.number_format != DATE_FORMAT:
            cell.number_format = DATE_FORMAT


def migrar_n_documento_sin_ceros(ws_master, ws_detalle):
    """Migracion idempotente (2026-07-17): la regla 'n_documento no puede empezar
    con 0' (ver MEMORY.md) se agrego a las reglas de negocio del skill pero nunca
    se implemento en el codigo -- documentos como CFLI-001 ('0000130842') y
    UMAG-017 ('0000125796') quedaron registrados con los ceros. Recorre TODAS las
    filas ya escritas de Master (columna E) y Detalle (columna D) y les quita los
    ceros a la izquierda con normalizar_n_documento(), sin tocar font/fill/border
    (incluidas celdas ya corregidas a mano en azul marino -- decision explicita
    del usuario 2026-07-17, es una excepcion deliberada a la regla de oro, igual
    que migrar_columna_proveedor()). Las hojas de proyecto no necesitan tocarse:
    son formulas hacia Master!E, se actualizan solas."""
    corregidas = 0
    ultima_m = ultima_fila_datos(ws_master)
    for r in range(2, ultima_m + 1):
        cell = ws_master.cell(row=r, column=5)
        actual = cell.value
        if isinstance(actual, str):
            nuevo = normalizar_n_documento(actual)
            if nuevo != actual:
                cell.value = nuevo
                corregidas += 1

    ultima_d = ultima_fila_datos(ws_detalle)
    for r in range(2, ultima_d + 1):
        cell = ws_detalle.cell(row=r, column=4)
        actual = cell.value
        if isinstance(actual, str):
            nuevo = normalizar_n_documento(actual)
            if nuevo != actual:
                cell.value = nuevo
                corregidas += 1

    if corregidas:
        print(f"  [OK] {corregidas} celda(s) de N Documento corregida(s): ceros a la izquierda eliminados.")


def migrar_columna_total_con_iva_detalle(ws_master, ws_detalle):
    """Migracion idempotente (2026-07-17): agrega 'Total con IVA (CLP)' como
    ultima columna de Detalle (pedido del usuario) y la rellena para TODAS las
    filas ya escritas -- no solo las nuevas. Se agrega AL FINAL de la hoja
    (columna 11), no en medio, asi que no hace falta desplazar columnas ni
    reescribir formulas existentes (a diferencia de migrar_columna_proveedor).
    Cada fila usa la tasa REAL del documento (IVA de Master columna L / Neto
    sumado desde Detalle columna J), no 19% fijo, para que sirva tambien en
    documentos exentos (pasajes de bus) y de Zona Franca -- misma logica que
    escribir_items_detalle() aplica a filas nuevas. No se lee Master columna K
    (Neto) porque es una formula: openpyxl con data_only=False devuelve el
    texto de la formula, no el numero calculado por Excel; sumar Detalle!J
    directo da el mismo resultado sin depender de que el archivo se haya
    abierto en Excel antes."""
    if ws_detalle.cell(row=1, column=11).value == "Total con IVA (CLP)":
        return

    print("  Migrando Detalle: agregando columna 'Total con IVA (CLP)'...")
    header_cell = ws_detalle.cell(row=1, column=11, value="Total con IVA (CLP)")
    header_cell.fill = HEADER_FILL
    header_cell.font = HEADER_FONT
    header_cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    header_cell.border = THIN_BORDER

    ultima_m = ultima_fila_datos(ws_master)
    iva_por_nref = {
        ws_master.cell(row=r, column=1).value: ws_master.cell(row=r, column=12).value or 0
        for r in range(2, ultima_m + 1)
    }

    ultima_d = ultima_fila_datos(ws_detalle)
    from collections import defaultdict
    from copy import copy as _copy
    neto_por_nref = defaultdict(float)
    for r in range(2, ultima_d + 1):
        n_ref = ws_detalle.cell(row=r, column=1).value
        j = ws_detalle.cell(row=r, column=10).value
        if isinstance(j, (int, float)):
            neto_por_nref[n_ref] += j

    rellenadas = 0
    for r in range(2, ultima_d + 1):
        n_ref = ws_detalle.cell(row=r, column=1).value
        j = ws_detalle.cell(row=r, column=10).value or 0
        neto_doc = neto_por_nref.get(n_ref, 0)
        iva_doc = iva_por_nref.get(n_ref, 0)
        ratio = (iva_doc / neto_doc) if neto_doc else 0
        cell = ws_detalle.cell(row=r, column=11, value=round(j * (1 + ratio)))
        cell.number_format = MONEY_FORMAT
        cell.border = _copy(ws_detalle.cell(row=r, column=10).border)
        cell.font = _copy(ws_detalle.cell(row=r, column=10).font)
        origen_fill = ws_detalle.cell(row=r, column=10).fill
        if origen_fill and origen_fill.patternType:
            cell.fill = _copy(origen_fill)
        rellenadas += 1

    print(f"  [OK] {rellenadas} fila(s) de Detalle con 'Total con IVA (CLP)' calculado.")


# ── SUPRESION DEL AVISO "NUMERO ALMACENADO COMO TEXTO" (N Documento) ────────
# N Documento se guarda siempre como texto, no como numero (aunque ya no lleve
# ceros a la izquierda desde normalizar_n_documento() -- puede seguir teniendo
# formato "S/N (archivo)" o letras), lo que hace que Excel marque toda la
# columna con el triangulo verde "Numero almacenado como texto". openpyxl no expone
# <ignoredErrors> en su API publica (la clase existe en
# openpyxl.worksheet.errors pero no esta conectada al lector/escritor de
# hojas), asi que hay que inyectar ese XML directo en el .xlsx ya guardado.
# Se debe reaplicar despues de CADA wb.save(): openpyxl descarta este
# elemento (no lo modela) en cualquier ciclo de carga+guardado posterior.

_NS_MAIN = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
_NS_REL = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"
# Elementos de CT_Worksheet que, segun el esquema OOXML, deben ir DESPUES de
# ignoredErrors -- se inserta el bloque justo antes del primero que aparezca
# (o antes de </worksheet> si no hay ninguno) para no corromper el orden.
_TAGS_DESPUES_DE_IGNORED_ERRORS = (
    "<drawing", "<legacyDrawing", "<legacyDrawingHF", "<picture",
    "<oleObjects", "<controls", "<webPublishItems", "<tableParts", "<extLst",
)


def _hojas_columna_n_documento(wb):
    """N Documento vive en E (col 5) en Master y en las hojas de proyecto, y en
    D (col 4) en Detalle."""
    hojas = {}
    if "Master" in wb.sheetnames:
        hojas["Master"] = "E"
    if "Detalle" in wb.sheetnames:
        hojas["Detalle"] = "D"
    for nombre in wb.sheetnames:
        if nombre not in ("Master", "Detalle", "_Claude"):
            hojas[nombre] = "E"
    return hojas


def _insertar_ignored_errors(xml_hoja, bloque):
    posiciones = [xml_hoja.find(tag) for tag in _TAGS_DESPUES_DE_IGNORED_ERRORS]
    posiciones = [p for p in posiciones if p != -1]
    punto = min(posiciones) if posiciones else xml_hoja.rfind("</worksheet>")
    return xml_hoja[:punto] + bloque + xml_hoja[punto:]


def suprimir_aviso_numero_texto(ruta_excel, hojas_columnas):
    """Post-procesa ruta_excel (ya guardado por openpyxl) para que Excel ignore
    el aviso 'Numero almacenado como texto' en toda la columna de N Documento
    de cada hoja en hojas_columnas (ej. {'Master': 'E', 'Detalle': 'D'})."""
    if not hojas_columnas:
        return

    with zipfile.ZipFile(ruta_excel, "r") as zf:
        infos = zf.infolist()
        contenidos = {n: zf.read(n) for n in zf.namelist()}

    wb_root = ET.fromstring(contenidos["xl/workbook.xml"])
    rels_root = ET.fromstring(contenidos["xl/_rels/workbook.xml.rels"])
    rid_a_target = {rel.get("Id"): rel.get("Target") for rel in rels_root}

    ruta_por_hoja = {}
    for sheet_el in wb_root.find(f"{{{_NS_MAIN}}}sheets"):
        nombre = sheet_el.get("name")
        rid = sheet_el.get(f"{{{_NS_REL}}}id")
        target = rid_a_target.get(rid, "").lstrip("/")
        if target:
            ruta_por_hoja[nombre] = target if target.startswith("xl/") else f"xl/{target}"

    cambiados = {}
    for hoja, col in hojas_columnas.items():
        ruta_hoja = ruta_por_hoja.get(hoja)
        if not ruta_hoja or ruta_hoja not in contenidos:
            continue
        xml_hoja = contenidos[ruta_hoja].decode("utf-8")
        xml_hoja = re.sub(r"<ignoredErrors>.*?</ignoredErrors>", "", xml_hoja, flags=re.DOTALL)
        bloque = (
            f'<ignoredErrors><ignoredError sqref="{col}2:{col}1048576" '
            f'numberStoredAsText="1"/></ignoredErrors>'
        )
        cambiados[ruta_hoja] = _insertar_ignored_errors(xml_hoja, bloque).encode("utf-8")

    if not cambiados:
        return

    ruta_tmp = str(ruta_excel) + ".tmp"
    with zipfile.ZipFile(ruta_tmp, "w", zipfile.ZIP_DEFLATED) as zout:
        for info in infos:
            zout.writestr(info, cambiados.get(info.filename, contenidos[info.filename]))
    shutil.move(ruta_tmp, str(ruta_excel))


def prefijo_para_proyecto(proyecto):
    if proyecto in PREFIJOS_PROYECTO:
        return PREFIJOS_PROYECTO[proyecto]
    letras = re.sub(r"[^A-Za-zÁÉÍÓÚÑ]", "", proyecto).upper()
    derivado = (letras[:4] or "PROY")
    print(f"  [WARN] Proyecto '{proyecto}' no tiene prefijo de N Ref definido en "
          f"PREFIJOS_PROYECTO. Usando '{derivado}' derivado automaticamente -- "
          f"agregalo a PREFIJOS_PROYECTO si no es el que quieres.")
    return derivado


def siguiente_n_ref(proyecto, max_seq):
    prefijo = prefijo_para_proyecto(proyecto)
    siguiente = max_seq.get(prefijo, 0) + 1
    max_seq[prefijo] = siguiente
    return f"{prefijo}-{siguiente:03d}"


# ── INVENTARIO DE ARCHIVOS ──────────────────────────────────────────────────

def inventariar_archivos(raiz, archivos_registrados):
    """archivos_registrados: set de 'Proyecto\\archivo.ext' ya cubiertos (Master + reconciliacion)."""
    pendientes = []
    omitidos = []

    for subdir in sorted(raiz.iterdir()):
        if not subdir.is_dir() or subdir.name.startswith(("_", ".")):
            continue
        proyecto = subdir.name

        for archivo in sorted(subdir.iterdir()):
            if not archivo.is_file():
                continue
            ext = archivo.suffix.lower()
            if ext in EXTENSIONES_IGNORAR or archivo.name == "desktop.ini":
                continue
            if ext not in EXTENSIONES_VALIDAS:
                continue

            ruta_rel = f"{proyecto}\\{archivo.name}"
            stat = archivo.stat()
            info = {
                "archivo": archivo.name,
                "proyecto": proyecto,
                "ruta_relativa": ruta_rel,
                "ruta_absoluta": str(archivo),
                "fecha_mod": datetime.fromtimestamp(stat.st_mtime).strftime("%Y-%m-%d %H:%M:%S"),
            }
            if ruta_rel in archivos_registrados:
                omitidos.append(info)
            else:
                pendientes.append(info)

    return pendientes, omitidos


# ── DATOS EXTRAÍDOS (JSON) ──────────────────────────────────────────────────

def cargar_datos_json(ruta_json):
    if not ruta_json.exists():
        return []
    with open(ruta_json, "r", encoding="utf-8") as f:
        return json.load(f)


def buscar_dato_por_archivo(datos, proyecto, archivo):
    for d in datos:
        if d.get("archivo") == archivo and d.get("proyecto") == proyecto:
            return d
    return None


def total_sin_iva_items(items):
    return sum(it["cantidad"] * it["p_unitario_sin_iva"] for it in items)


def normalizar_n_documento(valor):
    """Quita los ceros a la izquierda de un N Documento (pedido 2026-07-17,
    ver MEMORY.md 'Reglas de negocio'): '0000130020' -> '130020'. No toca
    valores que no empiecen con un digito '0' (ej. 'S/N (IMG_1234)'). Si el
    valor es solo ceros (caso extremo, no deberia pasar en la practica), deja
    un solo '0' en vez de vaciar la celda."""
    s = str(valor)
    despojado = s.lstrip("0")
    return despojado if despojado else (s[-1:] if s else s)


def calcular_iva_documento(dato, total_sin_iva):
    """Logica compartida para el IVA a nivel documento: usa el valor explicito
    del JSON si viene, si no calcula 19% del Neto para Factura/Guia de
    Despacho (0 para el resto). La usan tanto escribir_fila_master (columna L
    de Master) como escribir_items_detalle (para prorratear 'Total con IVA'
    por item en Detalle), asi ambas hojas quedan siempre consistentes entre si."""
    iva = dato.get("iva")
    if iva is None:
        iva = round(total_sin_iva * 0.19) if dato.get("tipo_documento") in ("Factura", "Guía de Despacho") else 0
    return iva


# ── ESCRITURA: DOCUMENTOS NUEVOS ────────────────────────────────────────────

def escribir_items_detalle(ws_detalle, fila_inicio, n_ref, dato, color):
    n_doc_str = normalizar_n_documento(dato["n_documento"])
    total_sin_iva_doc = total_sin_iva_items(dato["items"])
    iva_doc = calcular_iva_documento(dato, total_sin_iva_doc)
    # Tasa real del documento (IVA / Neto), no 19% fijo -- para que sirva tambien en
    # documentos exentos (pasajes de bus) y de Zona Franca (0%), y para que la suma de
    # "Total con IVA" de Detalle por N Ref coincida siempre con Master (pedido 2026-07-17).
    tasa_iva_doc = (iva_doc / total_sin_iva_doc) if total_sin_iva_doc else 0

    fila = fila_inicio
    for item in dato["items"]:
        total_item = item["cantidad"] * item["p_unitario_sin_iva"]
        total_item_con_iva = round(total_item * (1 + tasa_iva_doc))
        valores = [
            n_ref, dato["proyecto"], dato.get("tipo_proyecto", ""), n_doc_str,
            item["nombre_item"], item.get("descripcion", ""), item.get("categoria_item", ""),
            item["cantidad"], item["p_unitario_sin_iva"], total_item, total_item_con_iva,
        ]
        for c, v in enumerate(valores, 1):
            cell = ws_detalle.cell(row=fila, column=c, value=v)
            cell.font = NORMAL_FONT
            cell.border = THIN_BORDER
        ws_detalle.cell(row=fila, column=9).number_format = MONEY_FORMAT
        ws_detalle.cell(row=fila, column=10).number_format = MONEY_FORMAT
        ws_detalle.cell(row=fila, column=11).number_format = MONEY_FORMAT
        if color:
            pintar_fila(ws_detalle, fila, len(ENCABEZADOS_DETALLE), color)
        fila += 1
    return fila


def celda_requiere_revision(valor):
    return isinstance(valor, str) and ("S/N" in valor.upper() or "SIN_NUMERO" in valor.upper())


def escribir_fila_master(ws_master, fila, n_ref, dato, info_archivo, color):
    resumen_items = "; ".join(it["nombre_item"] for it in dato["items"])
    total_sin_iva = total_sin_iva_items(dato["items"])
    iva = calcular_iva_documento(dato, total_sin_iva)

    try:
        fecha_val = datetime.strptime(dato["fecha"], "%d/%m/%Y")
    except (ValueError, KeyError):
        fecha_val = dato.get("fecha", "")

    for c in range(1, len(ENCABEZADOS_MASTER) + 1):
        cell = ws_master.cell(row=fila, column=c)
        cell.font = NORMAL_FONT
        cell.border = THIN_BORDER

    proveedor_completo = dato.get("proveedor", "")
    valores = {
        1: n_ref, 2: dato["proyecto"], 3: dato.get("tipo_proyecto", ""), 4: fecha_val,
        6: dato.get("tipo_documento", ""),
        7: generar_tag_proveedor(proveedor_completo) if proveedor_completo else "",
        8: proveedor_completo,
        9: dato.get("categoria", ""), 10: resumen_items, 12: iva,
        14: dato.get("estado", "Pendiente"), 15: info_archivo["ruta_relativa"],
        16: info_archivo["fecha_mod"],
    }
    for c, v in valores.items():
        ws_master.cell(row=fila, column=c, value=v)

    n_doc_str = normalizar_n_documento(dato["n_documento"])
    ws_master.cell(row=fila, column=5, value=n_doc_str)
    if celda_requiere_revision(n_doc_str):
        ws_master.cell(row=fila, column=5).font = ROJO_FONT

    if isinstance(fecha_val, datetime):
        ws_master.cell(row=fila, column=4).number_format = DATE_FORMAT

    escribir_formulas_master(ws_master, fila)

    iva_cell = ws_master.cell(row=fila, column=12)
    iva_cell.number_format = MONEY_FORMAT
    if dato.get("tipo_documento") in ("Factura", "Guía de Despacho") and total_sin_iva > 0:
        esperado = round(total_sin_iva * 0.19)
        if abs(iva - esperado) > 1:
            iva_cell.font = ROJO_FONT

    if color:
        pintar_fila(ws_master, fila, len(ENCABEZADOS_MASTER), color)


def escribir_formulas_master(ws_master, fila):
    """(Re)escribe K y M como formulas -- son siempre derivadas, nunca se editan a mano.
    Si la celda ya tiene un valor NO-formula (corregida a mano), se respeta y no se toca.
    (Hasta antes de la columna "Proveedor (Razón Social)" estas eran J y L --
    se corrieron una posición al agregar esa columna, ver migrar_columna_proveedor.)"""
    k_actual = ws_master.cell(row=fila, column=11).value
    if k_actual is None or (isinstance(k_actual, str) and k_actual.startswith("=")):
        ws_master.cell(row=fila, column=11, value=f"=SUMIF(Detalle!$A:$A,$A{fila},Detalle!$J:$J)")
    ws_master.cell(row=fila, column=11).number_format = MONEY_FORMAT

    m_actual = ws_master.cell(row=fila, column=13).value
    if m_actual is None or (isinstance(m_actual, str) and m_actual.startswith("=")):
        ws_master.cell(row=fila, column=13, value=f"=K{fila}+L{fila}")
    ws_master.cell(row=fila, column=13).number_format = MONEY_FORMAT


# ── REORDENAR POR FECHA (mas reciente arriba) ───────────────────────────────
# Master/Detalle se muestran con el documento mas reciente en la fila 2 y el
# mas antiguo al fondo de la tabla (pedido del usuario 2026-07-17). Es una
# excepcion deliberada a "nunca tocar una fila de datos ya escrita" -- igual
# que migrar_columna_proveedor()/aplicar_renombrados(): el CONTENIDO de cada
# fila (valores, colores, fuentes, correcciones a mano) no cambia, solo su
# POSICION. Corre en cada run (incluso sin documentos nuevos), porque agregar
# hoy un documento con fecha antigua debe intercalarlo en su lugar, no solo
# agregarlo al final.

def capturar_fila(ws, fila, ncols):
    """Snapshot de una fila completa (valores + estilo) para poder reubicarla
    a otra fila sin perder formato manual (fuente, color, borde, formato).
    cell.font/fill/border/alignment devuelven un StyleProxy de solo lectura
    (no un Font/PatternFill real) -- copy.copy() lo desenvuelve a un objeto
    de estilo real que si se puede reasignar a otra celda."""
    from copy import copy as _copy
    return [
        {
            "value": ws.cell(row=fila, column=c).value,
            "number_format": ws.cell(row=fila, column=c).number_format,
            "font": _copy(ws.cell(row=fila, column=c).font),
            "fill": _copy(ws.cell(row=fila, column=c).fill),
            "border": _copy(ws.cell(row=fila, column=c).border),
            "alignment": _copy(ws.cell(row=fila, column=c).alignment),
        }
        for c in range(1, ncols + 1)
    ]


def escribir_snapshot_fila(ws, fila, snapshot):
    for c, datos in enumerate(snapshot, 1):
        cell = ws.cell(row=fila, column=c)
        cell.value = datos["value"]
        cell.number_format = datos["number_format"]
        cell.font = datos["font"]
        cell.fill = datos["fill"]
        cell.border = datos["border"]
        cell.alignment = datos["alignment"]


def _clave_orden_fecha(info):
    """info = (indice_original, fila, n_ref, fecha). Mas reciente primero;
    fechas no interpretables (string/vacio) van al final, en su orden original."""
    idx, _fila, _n_ref, fecha = info
    if isinstance(fecha, datetime):
        return (0, -fecha.timestamp(), idx)
    return (1, 0, idx)


def reordenar_por_fecha(ws_master, ws_detalle, primera_fila_libre_master, primera_fila_libre_detalle):
    """Reordena las filas de datos de Master (fila 2 = fecha mas reciente,
    fondo = mas antigua) y reagrupa los bloques de items de Detalle que le
    corresponden a cada N Ref siguiendo ese mismo orden. No modifica
    valor/formato de ninguna celda, solo la fila donde vive -- ver nota de la
    seccion. Las formulas K/M de Master (que referencian su propia fila, ej.
    $A15) se regeneran para apuntar a la fila nueva; los valores fijos
    (corregidos a mano) se preservan tal cual."""
    ultima_master = primera_fila_libre_master - 1
    if ultima_master < 2:
        return

    ncols_master = len(ENCABEZADOS_MASTER)
    info_filas = [
        (idx, r, ws_master.cell(row=r, column=1).value, ws_master.cell(row=r, column=4).value)
        for idx, r in enumerate(range(2, ultima_master + 1))
    ]
    orden = sorted(info_filas, key=_clave_orden_fecha)
    filas_origen = [f[1] for f in orden]

    if filas_origen == list(range(2, ultima_master + 1)):
        return  # ya esta en el orden correcto, no tocar nada

    snapshots_master = [capturar_fila(ws_master, r, ncols_master) for r in filas_origen]
    for i, snap in enumerate(snapshots_master):
        escribir_snapshot_fila(ws_master, 2 + i, snap)
        escribir_formulas_master(ws_master, 2 + i)  # ajusta $A<fila> de K/M a la nueva posicion

    n_ref_orden = [ws_master.cell(row=2 + i, column=1).value for i in range(len(snapshots_master))]

    ultima_detalle = primera_fila_libre_detalle - 1
    if ultima_detalle < 2:
        return

    ncols_detalle = len(ENCABEZADOS_DETALLE)
    grupos = {}
    orden_original_grupos = []
    for r in range(2, ultima_detalle + 1):
        n_ref_d = ws_detalle.cell(row=r, column=1).value
        if n_ref_d not in grupos:
            grupos[n_ref_d] = []
            orden_original_grupos.append(n_ref_d)
        grupos[n_ref_d].append(capturar_fila(ws_detalle, r, ncols_detalle))

    fila = 2
    vistos = set()
    for n_ref in n_ref_orden + orden_original_grupos:
        if n_ref in vistos:
            continue
        vistos.add(n_ref)
        for snap in grupos.get(n_ref, []):
            escribir_snapshot_fila(ws_detalle, fila, snap)
            fila += 1


# ── PIES DE TABLA (regenerados en cada corrida) ─────────────────────────────

def regenerar_pie(ws, ncols, col_totales, etiqueta_total, primera_fila_libre, leyenda):
    fila = primera_fila_libre
    if primera_fila_libre > 2:
        fila += 1  # fila en blanco separadora
        etiqueta_fila = fila
        ws.cell(row=etiqueta_fila, column=col_totales[0] - 1, value=etiqueta_total)
        for c in col_totales:
            letra = get_column_letter(c)
            ws.cell(row=etiqueta_fila, column=c, value=f"=SUM({letra}2:{letra}{primera_fila_libre - 1})")
            ws.cell(row=etiqueta_fila, column=c).number_format = MONEY_FORMAT
        for c in range(1, ncols + 1):
            cell = ws.cell(row=etiqueta_fila, column=c)
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
        fila += 2
    return escribir_leyenda(ws, fila, leyenda, ncols)


def limpiar_pie(ws):
    ultima = ultima_fila_datos(ws)
    if ws.max_row > ultima:
        ws.delete_rows(ultima + 1, ws.max_row - ultima)
    return ultima


# ── HOJAS DE PROYECTO (100% derivadas, se regeneran completas) ─────────────

def regenerar_hoja_proyecto(wb, proyecto, filas_master, color):
    """Regenera el CONTENIDO (filas 2 en adelante: datos + pie + leyenda) de la hoja
    de proyecto. El nombre/pestaña de la hoja es el PREFIJO del proyecto (ej. "Cesfam
    Limache" -> "CFLI"), no el nombre completo -- pedido del usuario 2026-07-17, para
    que las pestañas del Excel sean compactas. El nombre completo solo se usa para
    contenido dentro de la hoja (ej. el rótulo "TOTAL <proyecto>" del pie). Si la hoja
    ya existe, se reutiliza el mismo objeto de hoja en vez de borrarla y recrearla --
    borrar+recrear tiraba tambien columnas ocultas, anchos manuales, autofiltro,
    freeze panes y validaciones de datos que el usuario haya dejado en esa hoja. Solo
    se borran las filas de datos/pie/leyenda (fila 2 en adelante); el encabezado (fila
    1) y todo el formato a nivel de columna/hoja quedan intactos."""
    nombre_hoja = prefijo_para_proyecto(proyecto)
    if nombre_hoja in wb.sheetnames:
        ws = wb[nombre_hoja]
        if ws.max_row > 1:
            ws.delete_rows(2, ws.max_row - 1)
    else:
        ws = wb.create_sheet(title=nombre_hoja, index=len(wb.sheetnames))

    for c, h in enumerate(ENCABEZADOS_PROYECTO, 1):
        ws.cell(row=1, column=c, value=h)
    formato_encabezado(ws, len(ENCABEZADOS_PROYECTO))
    ws.column_dimensions[get_column_letter(COL_PROVEEDOR_RAZON_SOCIAL_MASTER)].hidden = True

    # Columnas de Master que se muestran en la hoja de proyecto (se salta L =
    # IVA, que no se repite aca). Refleja el layout de ENCABEZADOS_MASTER
    # despues de agregar "Proveedor (Razón Social)" en H.
    columnas_master = ["A", "B", "C", "D", "E", "F", "G", "H", "I", "J", "K", "M", "N"]
    fila = 2
    for fila_m in filas_master:
        for c, col_m in enumerate(columnas_master, 1):
            ws.cell(row=fila, column=c, value=f"=Master!{col_m}{fila_m}")
        for c in range(1, len(ENCABEZADOS_PROYECTO) + 1):
            ws.cell(row=fila, column=c).font = NORMAL_FONT
            ws.cell(row=fila, column=c).border = THIN_BORDER
        ws.cell(row=fila, column=4).number_format = DATE_FORMAT
        ws.cell(row=fila, column=11).number_format = MONEY_FORMAT
        ws.cell(row=fila, column=12).number_format = MONEY_FORMAT
        if color:
            pintar_fila(ws, fila, len(ENCABEZADOS_PROYECTO), color)
        fila += 1

    if fila > 2:
        fila += 1
        etiqueta = f"TOTAL {proyecto.upper()}"
        ws.cell(row=fila, column=10, value=etiqueta)
        ws.cell(row=fila, column=11, value=f"=SUM(K2:K{fila - 2})")
        ws.cell(row=fila, column=12, value=f"=SUM(L2:L{fila - 2})")
        for c in range(1, len(ENCABEZADOS_PROYECTO) + 1):
            cell = ws.cell(row=fila, column=c)
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
        ws.cell(row=fila, column=11).number_format = MONEY_FORMAT
        ws.cell(row=fila, column=12).number_format = MONEY_FORMAT
        fila += 2

    escribir_leyenda(ws, fila, LEYENDA_PROYECTO, len(ENCABEZADOS_PROYECTO))
    if color:
        ws.sheet_properties.tabColor = color
    ajustar_anchos(ws)
    return ws


def asignar_colores_proyectos(wb, proyectos):
    colores = {}
    if "Master" in wb.sheetnames:
        ws_master = wb["Master"]
        ultima = ultima_fila_datos(ws_master)
        for r in range(2, ultima + 1):
            proy = ws_master.cell(row=r, column=2).value
            fill = ws_master.cell(row=r, column=2).fill
            if proy and fill and fill.fgColor and fill.fgColor.rgb and fill.fgColor.rgb not in ("00000000", None):
                rgb = fill.fgColor.rgb
                colores[proy] = rgb[2:] if len(rgb) == 8 else rgb

    usados = set(colores.values())
    libres = [c for c in PALETA if c not in usados]
    for proy in sorted(proyectos):
        if proy not in colores:
            colores[proy] = libres.pop(0) if libres else PALETA[hash(proy) % len(PALETA)]
    return colores


# ── VERIFICACIONES ──────────────────────────────────────────────────────────

def verificar_aritmetica(datos):
    inconsistencias = []
    for d in datos:
        total_sin_iva = total_sin_iva_items(d["items"])
        iva = d.get("iva")
        if iva is None:
            continue
        if d.get("tipo_documento") in ("Factura", "Guía de Despacho") and total_sin_iva > 0:
            esperado = round(total_sin_iva * 0.19)
            if abs(iva - esperado) > 1:
                inconsistencias.append({
                    "archivo": d["archivo"], "n_documento": d["n_documento"],
                    "neto": total_sin_iva, "iva": iva, "iva_esperado": esperado,
                    "nota": d.get("notas", ""),
                })
    return inconsistencias


# ── RENOMBRADO Y CONVERSIÓN DE ARCHIVOS ─────────────────────────────────────
# Renombra cada foto/PDF a "<N Ref>_<TagProveedor>_<Fecha ISO>.<ext>" y convierte
# HEIC->JPG. Cubre documentos nuevos y, retroactivamente, los ya registrados en
# Master (incluidos los del bootstrap via reconciliacion_archivos.json), con el
# mismo mecanismo idempotente: se compara el nombre fisico actual contra el
# esperado y solo se actua si difieren.

CARACTERES_INVALIDOS_ARCHIVO = re.compile(r'[\\/:*?"<>|]')


def sanitizar_nombre(texto):
    """Reemplaza espacios y caracteres invalidos en nombres de archivo Windows
    (\\ / : * ? " < > |) por '_'."""
    limpio = CARACTERES_INVALIDOS_ARCHIVO.sub("_", texto)
    limpio = re.sub(r"\s+", "_", limpio.strip())
    return re.sub(r"_+", "_", limpio)


def fecha_iso_desde_valor(valor):
    """Convierte Master.Fecha (datetime/date, o string 'dd/mm/yyyy') a 'yyyy-mm-dd'.
    Si no se puede interpretar, sanitiza el valor tal cual para no romper el nombre."""
    if isinstance(valor, datetime):
        return valor.strftime("%Y-%m-%d")
    if hasattr(valor, "strftime"):
        return valor.strftime("%Y-%m-%d")
    if isinstance(valor, str):
        try:
            return datetime.strptime(valor, "%d/%m/%Y").strftime("%Y-%m-%d")
        except ValueError:
            return sanitizar_nombre(valor)
    return "sin-fecha"


def nombre_esperado_archivo(n_ref, proveedor_tag, fecha_valor, extension):
    """Nombre de archivo esperado: '<N Ref>_<TagProveedor>_<Fecha ISO><ext>'.
    Los .heic/.HEIC pasan a .jpg (se convierten); el resto conserva su extension."""
    tag = sanitizar_nombre(proveedor_tag or "SinProveedor")
    fecha = fecha_iso_desde_valor(fecha_valor)
    ext = extension.lower()
    if ext == ".heic":
        ext = ".jpg"
    return f"{n_ref}_{tag}_{fecha}{ext}"


def construir_reconciliacion_inversa(reconciliacion):
    """{n_ref: ruta_relativa} a partir del mapeo 'ruta_relativa -> n_ref' de
    reconciliacion_archivos.json."""
    return {n_ref: ruta for ruta, n_ref in reconciliacion.items()}


def resolver_ruta_actual(fila_dict, reconciliacion_inversa):
    """Ruta relativa ('Proyecto\\archivo.ext') del archivo fisico actual para
    una fila de Master, o None si no se puede determinar.

    OJO: la ubicacion fisica puede NO coincidir con Master.Proyecto (ver nota
    de UMAG-002 en reconciliacion_archivos.json) -- por eso se usa el Proyecto
    embebido en la propia ruta relativa (de Archivo origen o de la
    reconciliacion), nunca fila_dict['proyecto'].

    Prueba 'Archivo origen' primero, pero si esa ruta no existe en disco (caso
    de los 24 documentos del bootstrap, cuyo 'Archivo origen' quedo con el
    nombre que les daba un pipeline anterior ya perdido -- ver
    reconciliacion_archivos.json) cae a la ruta de la reconciliacion, que
    apunta al archivo fisico real."""
    candidatos = []
    if fila_dict.get("archivo_origen"):
        candidatos.append(str(fila_dict["archivo_origen"]))
    ruta_reconciliacion = reconciliacion_inversa.get(fila_dict["n_ref"])
    if ruta_reconciliacion and ruta_reconciliacion not in candidatos:
        candidatos.append(ruta_reconciliacion)

    for candidato in candidatos:
        if "\\" not in candidato:
            continue
        proyecto_fisico, nombre_archivo = candidato.split("\\", 1)
        if (RAIZ_DOCS / proyecto_fisico / nombre_archivo).exists():
            return candidato

    return candidatos[0] if candidatos else None


def planificar_renombrado_fila(fila_dict, reconciliacion_inversa):
    """Calcula (sin tocar disco) que accion corresponde para una fila de Master:
    {'n_ref', 'fila', 'accion', 'ruta_actual', 'ruta_nueva', 'nombre_nuevo'}.
    'accion' es uno de: 'ya_correcto', 'renombrar', 'convertir_heic',
    'archivo_no_encontrado'. Se usa tanto para el preview de status como,
    antes de ejecutar, para el run real."""
    base = {
        "n_ref": fila_dict["n_ref"], "fila": fila_dict["fila"],
        "ruta_actual": None, "ruta_nueva": None, "nombre_nuevo": None,
    }

    ruta_relativa = resolver_ruta_actual(fila_dict, reconciliacion_inversa)
    if not ruta_relativa:
        return {**base, "accion": "archivo_no_encontrado"}
    if "\\" not in ruta_relativa:
        return {**base, "accion": "archivo_no_encontrado"}

    proyecto_fisico, nombre_archivo = ruta_relativa.split("\\", 1)
    ruta_actual = RAIZ_DOCS / proyecto_fisico / nombre_archivo
    if not ruta_actual.exists():
        return {**base, "accion": "archivo_no_encontrado", "ruta_actual": ruta_actual}

    extension_actual = ruta_actual.suffix
    nombre_nuevo = nombre_esperado_archivo(
        fila_dict["n_ref"], fila_dict["proveedor_tag"], fila_dict["fecha"], extension_actual
    )
    ruta_nueva = ruta_actual.parent / nombre_nuevo

    if ruta_actual.name == nombre_nuevo:
        return {**base, "accion": "ya_correcto", "ruta_actual": ruta_actual,
                "ruta_nueva": ruta_actual, "nombre_nuevo": nombre_nuevo}

    accion = "convertir_heic" if extension_actual.lower() == ".heic" else "renombrar"
    return {**base, "accion": accion, "ruta_actual": ruta_actual,
            "ruta_nueva": ruta_nueva, "nombre_nuevo": nombre_nuevo}


def planificar_renombrados(filas_master, reconciliacion_inversa):
    """planificar_renombrado_fila() aplicado a cada fila de Master. No toca disco."""
    return [planificar_renombrado_fila(fm, reconciliacion_inversa) for fm in filas_master]


def convertir_heic_a_jpg(ruta_origen, ruta_destino):
    """Decodifica un HEIC y lo guarda como JPG, respetando la orientacion EXIF
    (las fotos de celular la traen y sin esto quedarian rotadas). Calidad 90,
    sin redimensionar -- son documentos tributarios que pueden necesitar zoom."""
    from PIL import Image, ImageOps
    import pillow_heif

    pillow_heif.register_heif_opener()
    with Image.open(ruta_origen) as imagen:
        imagen = ImageOps.exif_transpose(imagen)
        imagen.convert("RGB").save(str(ruta_destino), "JPEG", quality=90)


def ejecutar_plan_renombrado(item):
    """Ejecuta en disco la accion de un item de planificar_renombrado_fila
    ('renombrar' o 'convertir_heic'; 'ya_correcto'/'archivo_no_encontrado' no
    hacen nada). Devuelve (ok, error) -- no toca Master, eso lo hace el
    llamador (Task 5)."""
    if item["accion"] == "renombrar":
        try:
            item["ruta_actual"].rename(item["ruta_nueva"])
        except Exception as e:
            return False, str(e)
        return True, None

    if item["accion"] == "convertir_heic":
        try:
            convertir_heic_a_jpg(item["ruta_actual"], item["ruta_nueva"])
        except Exception as e:
            return False, str(e)
        item["ruta_actual"].unlink()
        return True, None

    return True, None


def excel_esta_bloqueado(ruta_excel):
    """True si ruta_excel existe y no se puede abrir para escritura (ej. porque
    esta abierto en Excel). Se usa como chequeo previo a PASO 9 -- ese paso
    renombra/borra archivos reales en disco sin poder deshacerse, asi que hay
    que confirmar que el Excel es escribible ANTES de tocar ningun archivo,
    no despues (si se descubriera recien al guardar en PASO 11, ya seria
    tarde: las fotos ya estarian renombradas/borradas y el Master actualizado
    solo en memoria, nunca persistido)."""
    if not ruta_excel.exists():
        return False
    try:
        with open(ruta_excel, "r+b"):
            pass
        return False
    except PermissionError:
        return True


def aplicar_renombrados(ws_master, filas_master, reconciliacion_inversa):
    """Recorre filas_master, ejecuta los renombrados/conversiones pendientes y
    actualiza Master.Archivo origen (col 15) y Master.Fecha modificacion
    (col 16, con el mtime real del archivo resultante) en las filas afectadas.
    Excepcion deliberada a la regla de oro de "nunca tocar una fila ya
    escrita" -- mismo tipo de excepcion que migrar_columna_proveedor().
    Devuelve (cantidad_renombrados, advertencias)."""
    renombrados = 0
    advertencias = []

    for item in planificar_renombrados(filas_master, reconciliacion_inversa):
        if item["accion"] == "archivo_no_encontrado":
            advertencias.append({
                "n_ref": item["n_ref"],
                "detalle": "Archivo fisico no encontrado para renombrar/convertir.",
            })
            continue
        if item["accion"] == "ya_correcto":
            continue

        ok, error = ejecutar_plan_renombrado(item)
        if not ok:
            accion_desc = "la conversion HEIC" if item["accion"] == "convertir_heic" else "el renombrado"
            advertencias.append({
                "n_ref": item["n_ref"],
                "detalle": f"Fallo {accion_desc}: {error}",
            })
            continue

        proyecto_fisico = item["ruta_nueva"].parent.name
        ws_master.cell(row=item["fila"], column=15, value=f"{proyecto_fisico}\\{item['nombre_nuevo']}")
        mtime = datetime.fromtimestamp(item["ruta_nueva"].stat().st_mtime).strftime("%Y-%m-%d %H:%M:%S")
        ws_master.cell(row=item["fila"], column=16, value=mtime)
        renombrados += 1

    return renombrados, advertencias


def reflejar_a_sitio_comunicacion(ruta_excel=None, ruta_sitio=None):
    """Copia (shutil.copy2) el Excel local encima de la copia de solo lectura
    en 'Sitio de comunicacion - Centro de Costos 1/' -- mismo paso que hace
    'run' (PASO 12b) al final de cada corrida. Factorizado para que tambien
    lo puedan llamar comandos que no pasan por main(), como el driver de la
    skill Revision_de_Errores (corregir/desglosar solo tocan el Excel local;
    sin este paso aparte, la copia compartida queda desactualizada). No falla
    si el destino esta bloqueado, solo advierte -- igual que en 'run'.
    Devuelve True si la copia se actualizo, False si no se pudo."""
    ruta_excel = ruta_excel or RUTA_EXCEL
    ruta_sitio = ruta_sitio or RUTA_EXCEL_SITIO_COMUNICACION
    try:
        shutil.copy2(ruta_excel, ruta_sitio)
        print(f"  [OK] Copia actualizada: {ruta_sitio}")
        return True
    except (PermissionError, OSError) as e:
        print(f"  [WARN] No se pudo actualizar la copia en Sitio de comunicacion ({e}).")
        print("         El Excel local si quedo guardado; reintenta mas tarde (ej. si esta abierto).")
        return False


def actualizar_visualizador():
    """Regenera el visualizador web (Visualizador Web/build/index.html) a partir
    del Excel recien guardado -- mismo patron que reflejar_a_sitio_comunicacion:
    corre al final de cada 'run' (PASO 12c), solo lee el Excel (no lo modifica),
    y si falla no aborta el run, solo advierte -- el Excel ya quedo guardado
    igual. Pedido del usuario 2026-07-19: que actualizar el Centro de Costos
    actualice el visualizador automaticamente, sin paso manual aparte. Ver
    Visualizador Web/CLAUDE.md para la arquitectura del build."""
    ruta_build_script = RAIZ_VISUALIZADOR_WEB / "build_visualizador.py"
    if not ruta_build_script.exists():
        print(f"  [WARN] No existe {ruta_build_script}, se omite este paso.")
        return False
    import sys
    ya_en_path = str(RAIZ_VISUALIZADOR_WEB) in sys.path
    if not ya_en_path:
        sys.path.insert(0, str(RAIZ_VISUALIZADOR_WEB))
    try:
        sys.modules.pop("build_visualizador", None)
        import build_visualizador as bv
        return bv.build() == 0
    except Exception as e:
        print(f"  [WARN] No se pudo actualizar el visualizador web ({e}).")
        print("         El Excel si quedo guardado; correr manualmente "
              "'python driver.py visualizador' despues.")
        return False
    finally:
        if not ya_en_path and str(RAIZ_VISUALIZADOR_WEB) in sys.path:
            sys.path.remove(str(RAIZ_VISUALIZADOR_WEB))


def actualizar_analisis_financiero():
    """Actualiza Análisis Financiero/Análisis de Proyectos.xlsx (costos
    reales por proyecto/categoría + hoja Indicadores) a partir del Excel
    recien guardado -- mismo patron que actualizar_visualizador: corre al
    final de cada 'run' (PASO 12d), solo lee Centro de Costos.xlsx (no lo
    modifica), y si falla no aborta el run, solo advierte."""
    ruta_script = RAIZ_ANALISIS_FINANCIERO / "Sistema" / "analisis_financiero.py"
    if not ruta_script.exists():
        print(f"  [WARN] No existe {ruta_script}, se omite este paso.")
        return False
    import sys
    raiz_sistema_af = ruta_script.parent
    ya_en_path = str(raiz_sistema_af) in sys.path
    if not ya_en_path:
        sys.path.insert(0, str(raiz_sistema_af))
    try:
        sys.modules.pop("analisis_financiero", None)
        import analisis_financiero as af
        resumen = af.ejecutar()
        if resumen["error"]:
            print(f"  [WARN] Análisis Financiero terminó con error: {resumen['error']}")
            return False
        return True
    except Exception as e:
        print(f"  [WARN] No se pudo actualizar Análisis Financiero ({e}).")
        print("         El Excel de Centro de Costos si quedo guardado; correr manualmente "
              "'python driver.py run' en Análisis Financiero despues.")
        return False


# --- MAIN ---------------------------------------------------------------------

def main():
    import sys
    sys.stdout.reconfigure(encoding="utf-8", errors="backslashreplace")
    sys.stderr.reconfigure(encoding="utf-8", errors="backslashreplace")

    print("=" * 70)
    print("  REGISTRO CENTRO DE COSTOS - QUEMPIN SpA")
    print(f"  Fecha: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    print("=" * 70)

    if not RAIZ_DOCS.exists():
        print(f"ERROR: No existe la carpeta raiz: {RAIZ_DOCS}")
        return
    if not RUTA_JSON.exists():
        print(f"ERROR: No existe el JSON de datos: {RUTA_JSON}")
        return

    print("\n--- PASO 1: Backup ---")
    ruta_backup_anterior = backup_mas_reciente()
    hacer_backup(RUTA_EXCEL)

    print("\n--- PASO 2: Abrir Excel ---")
    if RUTA_EXCEL.exists():
        wb = openpyxl.load_workbook(str(RUTA_EXCEL), data_only=False)
        print(f"  Excel abierto. Hojas existentes: {wb.sheetnames}")
    else:
        wb = openpyxl.Workbook()
        wb.remove(wb.active)
        print("  Excel creado desde cero.")

    migrar_columna_proveedor(wb)

    if "Master" not in wb.sheetnames:
        ws_master = wb.create_sheet("Master", 0)
        for c, h in enumerate(ENCABEZADOS_MASTER, 1):
            ws_master.cell(row=1, column=c, value=h)
        formato_encabezado(ws_master, len(ENCABEZADOS_MASTER))
        ws_master.column_dimensions[get_column_letter(COL_PROVEEDOR_RAZON_SOCIAL_MASTER)].hidden = True
    else:
        ws_master = wb["Master"]

    if "Detalle" not in wb.sheetnames:
        ws_detalle = wb.create_sheet("Detalle", 1)
        for c, h in enumerate(ENCABEZADOS_DETALLE, 1):
            ws_detalle.cell(row=1, column=c, value=h)
        formato_encabezado(ws_detalle, len(ENCABEZADOS_DETALLE))
    else:
        ws_detalle = wb["Detalle"]

    migrar_paleta_colores(wb, ws_master, ws_detalle)
    migrar_formato_fecha_corta(ws_master)
    migrar_n_documento_sin_ceros(ws_master, ws_detalle)
    migrar_columna_total_con_iva_detalle(ws_master, ws_detalle)

    print("\n--- PASO 2B: Detectar correcciones manuales (celdas rojas editadas a mano) ---")
    correcciones_pendientes = []
    if ruta_backup_anterior is not None:
        wb_anterior = openpyxl.load_workbook(str(ruta_backup_anterior), data_only=False)
        detectadas = detectar_correcciones_manuales(wb_anterior, wb)
        correcciones_pendientes = registrar_correcciones_pendientes(detectadas)
        if correcciones_pendientes:
            print(f"  [INFO] {len(correcciones_pendientes)} correccion(es) manual(es) detectada(s), "
                  f"pendiente(s) de confirmar:")
            for c in correcciones_pendientes:
                print(f"    - {c['n_ref']} / {c['campo']}: '{c['valor_anterior']}' -> '{c['valor_corregido']}'")
            print("    Registradas en ERRORES.md. Para aplicarlas (recolorear azul marino + propagar "
                  "a Detalle): python driver.py confirmar --todos")
        else:
            print("  Sin correcciones manuales nuevas.")
    else:
        print("  [INFO] No hay backup anterior contra el cual comparar (primera corrida).")

    print("\n--- PASO 3: Leer registros existentes ---")
    filas_master, max_seq, docs_registrados = leer_master(ws_master)
    reconciliacion = cargar_reconciliacion()
    archivos_registrados = set(reconciliacion.keys())
    for fm in filas_master:
        if fm["archivo_origen"]:
            archivos_registrados.add(str(fm["archivo_origen"]))
    print(f"  Documentos ya en Master: {len(filas_master)}")
    print(f"  Archivos ya cubiertos (Master + reconciliacion): {len(archivos_registrados)}")

    print("\n--- PASO 4: Inventariar archivos ---")
    pendientes, omitidos = inventariar_archivos(RAIZ_DOCS, archivos_registrados)
    print(f"  Pendientes: {len(pendientes)}")
    print(f"  Omitidos (ya registrados): {len(omitidos)}")

    print("\n--- PASO 5: Cargar datos extraidos ---")
    datos_json = cargar_datos_json(RUTA_JSON)
    print(f"  Entradas en {RUTA_JSON.name}: {len(datos_json)}")

    limpiar_pie(ws_detalle)
    limpiar_pie(ws_master)
    fila_detalle = ultima_fila_datos(ws_detalle) + 1
    fila_master = ultima_fila_datos(ws_master) + 1

    proyectos_tocados = set(fm["proyecto"] for fm in filas_master if fm["proyecto"])
    colores = asignar_colores_proyectos(wb, proyectos_tocados | {p["proyecto"] for p in pendientes})

    registrados_ok = 0
    limitaciones = []
    alertas_legibilidad = []
    posibles_duplicados = []

    print("\n--- PASO 6: Escribir documentos nuevos ---")
    for info in pendientes:
        dato = buscar_dato_por_archivo(datos_json, info["proyecto"], info["archivo"])
        if not dato:
            limitaciones.append({
                "archivo": info["archivo"], "proyecto": info["proyecto"],
                "detalle": "No se encontraron datos extraidos en el JSON para este archivo.",
                "accion": "Agregar manualmente los datos (con items) al JSON y re-ejecutar.",
            })
            print(f"  [WARN] Sin datos para: {info['proyecto']}\\{info['archivo']}")
            continue
        if not dato.get("items"):
            limitaciones.append({
                "archivo": info["archivo"], "proyecto": info["proyecto"],
                "detalle": "La entrada del JSON no tiene 'items' (lista de lineas).",
                "accion": "Agregar al menos un item con nombre_item/cantidad/p_unitario_sin_iva.",
            })
            continue

        n_doc_str = str(dato["n_documento"])
        n_doc_norm = normalizar_n_documento(n_doc_str)
        if n_doc_str in docs_registrados or n_doc_norm in docs_registrados:
            posibles_duplicados.append({
                "archivo": info["archivo"], "proyecto": info["proyecto"], "n_documento": n_doc_str,
            })

        n_ref = siguiente_n_ref(dato["proyecto"], max_seq)
        color = colores.get(dato["proyecto"])

        fila_detalle = escribir_items_detalle(ws_detalle, fila_detalle, n_ref, dato, color)
        escribir_fila_master(ws_master, fila_master, n_ref, dato, info, color)
        docs_registrados.add(n_doc_str)
        docs_registrados.add(n_doc_norm)

        proyectos_tocados.add(dato["proyecto"])
        registrados_ok += 1
        print(f"  [OK] {info['proyecto']}\\{info['archivo']} -> {n_ref} "
              f"(doc {dato['n_documento']}, {len(dato['items'])} item(s))")

        if celda_requiere_revision(n_doc_str):
            alertas_legibilidad.append({
                "archivo": info["archivo"], "proyecto": dato["proyecto"],
                "detalle": "N Documento requiere revision manual.",
            })
        fila_master += 1

    print("\n--- PASO 7: Reordenar por fecha (mas reciente arriba) ---")
    reordenar_por_fecha(ws_master, ws_detalle, fila_master, fila_detalle)

    print("\n--- PASO 8: Regenerar pies de Detalle/Master ---")
    regenerar_pie(ws_detalle, len(ENCABEZADOS_DETALLE), [10, 11], "TOTAL GENERAL", fila_detalle, LEYENDA_DETALLE)
    regenerar_pie(ws_master, len(ENCABEZADOS_MASTER), [11, 12, 13], "TOTAL GENERAL", fila_master, LEYENDA_MASTER)

    print("\n--- PASO 9: Regenerar hojas de proyecto ---")
    ultima_master = fila_master - 1
    for proyecto in sorted(proyectos_tocados):
        filas_de_este_proyecto = [
            r for r in range(2, ultima_master + 1)
            if ws_master.cell(row=r, column=2).value == proyecto
        ]
        if not filas_de_este_proyecto:
            continue
        regenerar_hoja_proyecto(wb, proyecto, filas_de_este_proyecto, colores.get(proyecto))
        print(f"  [OK] Hoja '{prefijo_para_proyecto(proyecto)}' ({proyecto}) regenerada "
              f"({len(filas_de_este_proyecto)} documento(s))")

    if excel_esta_bloqueado(RUTA_EXCEL):
        print("\n[ERROR] El archivo esta abierto en Excel (o bloqueado). Cierralo antes de continuar.")
        print("        No se renombro ni convirtio ningun archivo; no se guardaron cambios.")
        return

    print("\n--- PASO 10: Renombrar y convertir archivos ---")
    reconciliacion_inversa = construir_reconciliacion_inversa(reconciliacion)
    filas_master_actual, _, _ = leer_master(ws_master)
    renombrados, advertencias_renombrado = aplicar_renombrados(
        ws_master, filas_master_actual, reconciliacion_inversa
    )
    print(f"  [OK] {renombrados} archivo(s) renombrado(s)/convertido(s).")
    for adv in advertencias_renombrado:
        print(f"  [WARN] {adv['n_ref']}: {adv['detalle']}")

    print("\n--- PASO 11: Formato final ---")
    ajustar_anchos(ws_master)
    ajustar_anchos(ws_detalle)
    orden_deseado = ["Master", "Detalle"] + [prefijo_para_proyecto(p) for p in sorted(proyectos_tocados)]
    for i, nombre in enumerate(orden_deseado):
        if nombre in wb.sheetnames:
            wb.move_sheet(nombre, offset=i - wb.sheetnames.index(nombre))
    print(f"  [OK] Hojas ordenadas: {wb.sheetnames}")

    print("\n--- PASO 12: Guardar ---")
    try:
        wb.save(str(RUTA_EXCEL))
        suprimir_aviso_numero_texto(RUTA_EXCEL, _hojas_columna_n_documento(wb))
        print(f"  [OK] Excel guardado: {RUTA_EXCEL.name}")
    except PermissionError:
        print("  ERROR: El archivo esta abierto en Excel. Cierralo y vuelve a ejecutar.")
        return

    print("\n--- PASO 12b: Reflejar en Sitio de comunicacion ---")
    reflejar_a_sitio_comunicacion()

    print("\n--- PASO 12c: Actualizar visualizador web ---")
    actualizar_visualizador()

    print("\n--- PASO 12d: Actualizar Análisis Financiero ---")
    actualizar_analisis_financiero()

    print("\n--- PASO 13: Verificaciones aritmeticas (sobre todo el JSON) ---")
    inconsistencias = verificar_aritmetica(datos_json)

    print("\n" + "=" * 70)
    print("  INFORME DE AUDITORIA")
    print("=" * 70)

    print("\n1. ALERTAS DE LEGIBILIDAD")
    if alertas_legibilidad:
        for a in alertas_legibilidad:
            print(f"   * {a['archivo']} | Proyecto: {a['proyecto']} | {a['detalle']}")
    else:
        print("   Sin hallazgos.")

    print("\n2. INCONSISTENCIAS ARITMETICAS (Neto vs IVA 19%)")
    if inconsistencias:
        for inc in inconsistencias:
            print(f"   * Doc {inc['n_documento']} ({inc['archivo']}): "
                  f"Neto={inc['neto']:,} | IVA registrado={inc['iva']:,} vs esperado={inc['iva_esperado']:,}")
            if inc["nota"]:
                print(f"     Nota: {inc['nota'][:120]}")
    else:
        print("   Sin hallazgos.")

    print("\n3. POSIBLES DUPLICADOS (mismo N Documento que uno ya registrado)")
    if posibles_duplicados:
        for dup in posibles_duplicados:
            print(f"   * {dup['proyecto']}\\{dup['archivo']} | N Documento: {dup['n_documento']}")
    else:
        print("   Sin hallazgos.")

    print("\n4. LIMITACIONES DE REGISTRO")
    if limitaciones:
        for lim in limitaciones:
            print(f"   * {lim['archivo']} | Proyecto: {lim['proyecto']}")
            print(f"     {lim['detalle']} Accion: {lim['accion']}")
    else:
        print("   Sin hallazgos.")

    print("\n5. RENOMBRADO/CONVERSION DE ARCHIVOS")
    if advertencias_renombrado:
        for adv in advertencias_renombrado:
            print(f"   * {adv['n_ref']}: {adv['detalle']}")
    else:
        print("   Sin hallazgos.")

    print("\n6. CAMBIOS MANUALES PENDIENTES DE CONFIRMAR")
    if correcciones_pendientes:
        for c in correcciones_pendientes:
            print(f"   * {c['n_ref']} / {c['campo']}: '{c['valor_anterior']}' -> '{c['valor_corregido']}'")
        print("   Aplicar con: python driver.py confirmar --todos")
    else:
        print("   Sin hallazgos.")

    print("\n" + "-" * 70)
    print("  RESUMEN FINAL")
    print("-" * 70)
    print(f"  {'Documentos nuevos registrados:':<40} {registrados_ok}")
    print(f"  {'Documentos omitidos (ya registrados):':<40} {len(omitidos)}")
    print(f"  {'Posibles duplicados:':<40} {len(posibles_duplicados)}")
    print(f"  {'Limitaciones (faltan datos en JSON):':<40} {len(limitaciones)}")
    print(f"  {'Archivos renombrados/convertidos:':<40} {renombrados}")
    print(f"  {'Correcciones manuales pendientes de confirmar:':<40} {len(correcciones_pendientes)}")
    print("\n" + "=" * 70)


if __name__ == "__main__":
    import sys as _sys
    if len(_sys.argv) > 1 and _sys.argv[1] == "--confirmar":
        resto = _sys.argv[2:]
        if not resto:
            confirmar_correcciones(None)
        elif resto == ["--todos"]:
            confirmar_correcciones("TODOS")
        else:
            confirmar_correcciones(resto)
    else:
        main()
