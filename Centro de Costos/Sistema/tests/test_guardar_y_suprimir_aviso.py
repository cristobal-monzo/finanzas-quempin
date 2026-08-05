from unittest.mock import patch

import openpyxl

import auditor_centro_costos as acc


def _wb_minimo():
    wb = openpyxl.Workbook()
    wb.active.title = "Master"
    wb.create_sheet("Detalle")
    return wb


def test_guarda_y_suprime_aviso_en_flujo_normal(tmp_path):
    ruta = tmp_path / "Centro de Costos.xlsx"
    acc._guardar_y_suprimir_aviso(_wb_minimo(), ruta)
    assert ruta.exists()
    assert openpyxl.load_workbook(str(ruta)).sheetnames == ["Master", "Detalle"]


def test_permission_error_de_wb_save_se_propaga(tmp_path):
    """El caso real que todos los llamadores capturan: el archivo esta
    abierto en Excel. Debe seguir abortando -- no es el caso best-effort."""
    ruta = tmp_path / "Centro de Costos.xlsx"
    wb = _wb_minimo()
    with patch.object(wb, "save", side_effect=PermissionError("abierto en Excel")):
        try:
            acc._guardar_y_suprimir_aviso(wb, ruta)
            assert False, "deberia haber propagado PermissionError"
        except PermissionError:
            pass


def test_falla_de_suprimir_aviso_no_se_propaga_y_el_excel_ya_quedo_guardado(tmp_path, capsys):
    """Bug real que esto reemplaza: antes, cualquier falla de
    suprimir_aviso_numero_texto (cirugia de zip/XML best-effort, no la
    escritura real) quedaba sin capturar y tumbaba la corrida completa --
    aunque el Excel ya se hubiera guardado bien un instante antes."""
    ruta = tmp_path / "Centro de Costos.xlsx"
    with patch.object(acc, "suprimir_aviso_numero_texto", side_effect=RuntimeError("zip corrupto")):
        acc._guardar_y_suprimir_aviso(_wb_minimo(), ruta)

    assert ruta.exists()
    assert openpyxl.load_workbook(str(ruta)).sheetnames == ["Master", "Detalle"]
    salida = capsys.readouterr().out
    assert "zip corrupto" in salida
    assert "quedo guardado correctamente" in salida
