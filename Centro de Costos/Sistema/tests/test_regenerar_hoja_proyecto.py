import openpyxl

import auditor_centro_costos as acc


def _wb():
    wb = openpyxl.Workbook()
    wb.active.title = "Master"
    return wb


def test_crea_la_hoja_con_el_prefijo_del_proyecto_no_el_nombre_completo():
    wb = _wb()
    acc.regenerar_hoja_proyecto(wb, "Cesfam Limache", filas_master=[2], color="89DFFF")
    assert "CFLI" in wb.sheetnames
    assert "Cesfam Limache" not in wb.sheetnames


def test_cada_fila_es_una_formula_hacia_master_de_la_fila_correspondiente():
    wb = _wb()
    ws = acc.regenerar_hoja_proyecto(wb, "UMAG", filas_master=[5, 9], color="FFB7CE")

    assert ws.cell(row=2, column=1).value == "=Master!A5"
    assert ws.cell(row=2, column=4).value == "=Master!D5"
    assert ws.cell(row=3, column=1).value == "=Master!A9"
    assert ws.cell(row=3, column=4).value == "=Master!D9"


def test_fila_total_suma_solo_las_filas_de_datos_escritas():
    wb = _wb()
    ws = acc.regenerar_hoja_proyecto(wb, "UMAG", filas_master=[5, 9, 12], color="FFB7CE")
    # 3 filas de datos (2,3,4) -> el total va en la fila 6 (con 1 en blanco de por medio).
    assert ws.cell(row=6, column=11).value == "=SUM(K2:K4)"
    assert ws.cell(row=6, column=10).value == "TOTAL UMAG"


def test_tabcolor_se_aplica_a_la_hoja():
    wb = _wb()
    ws = acc.regenerar_hoja_proyecto(wb, "UMAG", filas_master=[5], color="FFB7CE")
    assert ws.sheet_properties.tabColor.rgb.endswith("FFB7CE")


def test_reutiliza_la_hoja_existente_y_no_deja_filas_viejas(monkeypatch):
    """regenerar_hoja_proyecto no borra+recrea la hoja (perderia formato de
    columna/autofiltro/freeze panes dejado a mano) -- solo limpia filas 2 en
    adelante y reescribe. Si se corre con MENOS filas que la vez anterior,
    no deben sobrevivir filas viejas de mas."""
    wb = _wb()
    acc.regenerar_hoja_proyecto(wb, "UMAG", filas_master=[5, 9, 12, 20, 25], color="FFB7CE")
    assert wb.sheetnames.count("UMAG") == 1

    ws = acc.regenerar_hoja_proyecto(wb, "UMAG", filas_master=[5], color="FFB7CE")

    assert wb.sheetnames.count("UMAG") == 1
    assert ws.cell(row=2, column=1).value == "=Master!A5"
    # Fila 3 en adelante ya no debe tener el contenido de la corrida anterior
    # (ni datos de otro N Ref, ni el pie/leyenda viejos en una posicion que
    # ya no corresponde).
    assert ws.cell(row=3, column=1).value in (None, "", "TOTAL UMAG")
