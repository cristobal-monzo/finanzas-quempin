import openpyxl

import auditor_centro_costos as acc


# ── normalizar_n_documento ───────────────────────────────────────────────

def test_normalizar_n_documento_quita_ceros_a_la_izquierda():
    assert acc.normalizar_n_documento("0000130020") == "130020"
    assert acc.normalizar_n_documento("0456") == "456"


def test_normalizar_n_documento_no_toca_valores_sin_ceros():
    assert acc.normalizar_n_documento("130020") == "130020"


def test_normalizar_n_documento_no_toca_sn_ilegible():
    assert acc.normalizar_n_documento("S/N (IMG_7533)") == "S/N (IMG_7533)"


def test_normalizar_n_documento_solo_ceros_deja_un_cero():
    assert acc.normalizar_n_documento("000") == "0"


# ── escribir_items_detalle: Total con IVA por item ──────────────────────

def _ws_detalle():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Detalle"
    for c, h in enumerate(acc.ENCABEZADOS_DETALLE, 1):
        ws.cell(row=1, column=c, value=h)
    return ws


def test_escribir_items_detalle_normaliza_n_documento_y_calcula_total_con_iva_19():
    ws = _ws_detalle()
    dato = {
        "proyecto": "UMAG", "tipo_proyecto": "I+D+i", "n_documento": "0000130020",
        "tipo_documento": "Factura",
        "items": [
            {"nombre_item": "Taladro", "cantidad": 1, "p_unitario_sin_iva": 100000},
        ],
    }
    acc.escribir_items_detalle(ws, 2, "UMAG-001", dato, color=None)

    assert ws.cell(row=2, column=4).value == "130020"  # N Documento sin ceros
    assert ws.cell(row=2, column=10).value == 100000  # Total sin IVA
    assert ws.cell(row=2, column=11).value == 119000  # Total con IVA (19%)


def test_escribir_items_detalle_documento_exento_no_agrega_iva():
    ws = _ws_detalle()
    dato = {
        "proyecto": "UMAG", "tipo_proyecto": "I+D+i", "n_documento": "162354",
        "tipo_documento": "Boleta", "iva": 0,
        "items": [
            {"nombre_item": "Pasaje bus", "cantidad": 1, "p_unitario_sin_iva": 5000},
        ],
    }
    acc.escribir_items_detalle(ws, 2, "UMAG-009", dato, color=None)

    assert ws.cell(row=2, column=10).value == 5000
    assert ws.cell(row=2, column=11).value == 5000  # sin IVA agregado


def test_escribir_items_detalle_prorratea_iva_custom_entre_varios_items():
    ws = _ws_detalle()
    dato = {
        "proyecto": "UMAG", "tipo_proyecto": "I+D+i", "n_documento": "1",
        "tipo_documento": "Factura", "iva": 10878,
        "items": [
            {"nombre_item": "A", "cantidad": 1, "p_unitario_sin_iva": 24859},
            {"nombre_item": "B", "cantidad": 1, "p_unitario_sin_iva": 32395},
        ],
    }
    acc.escribir_items_detalle(ws, 2, "UMAG-002", dato, color=None)

    total_sin_iva = 24859 + 32395
    ratio = 10878 / total_sin_iva
    assert ws.cell(row=2, column=11).value == round(24859 * (1 + ratio))
    assert ws.cell(row=3, column=11).value == round(32395 * (1 + ratio))


# ── migrar_n_documento_sin_ceros ─────────────────────────────────────────

def test_migrar_n_documento_sin_ceros_corrige_master_y_detalle():
    wb = openpyxl.Workbook()
    ws_m = wb.active
    ws_m.title = "Master"
    for c, h in enumerate(acc.ENCABEZADOS_MASTER, 1):
        ws_m.cell(row=1, column=c, value=h)
    ws_m.cell(row=2, column=1, value="CFLI-001")
    ws_m.cell(row=2, column=5, value="0000130842")

    ws_d = wb.create_sheet("Detalle")
    for c, h in enumerate(acc.ENCABEZADOS_DETALLE, 1):
        ws_d.cell(row=1, column=c, value=h)
    ws_d.cell(row=2, column=1, value="CFLI-001")
    ws_d.cell(row=2, column=4, value="0000130842")

    acc.migrar_n_documento_sin_ceros(ws_m, ws_d)

    assert ws_m.cell(row=2, column=5).value == "130842"
    assert ws_d.cell(row=2, column=4).value == "130842"


def test_migrar_n_documento_sin_ceros_es_idempotente():
    wb = openpyxl.Workbook()
    ws_m = wb.active
    ws_m.title = "Master"
    for c, h in enumerate(acc.ENCABEZADOS_MASTER, 1):
        ws_m.cell(row=1, column=c, value=h)
    ws_m.cell(row=2, column=1, value="UMAG-001")
    ws_m.cell(row=2, column=5, value="164")

    ws_d = wb.create_sheet("Detalle")
    for c, h in enumerate(acc.ENCABEZADOS_DETALLE, 1):
        ws_d.cell(row=1, column=c, value=h)

    acc.migrar_n_documento_sin_ceros(ws_m, ws_d)

    assert ws_m.cell(row=2, column=5).value == "164"


# ── migrar_columna_total_con_iva_detalle ─────────────────────────────────

def test_migrar_columna_total_con_iva_detalle_rellena_filas_existentes():
    wb = openpyxl.Workbook()
    ws_m = wb.active
    ws_m.title = "Master"
    for c, h in enumerate(acc.ENCABEZADOS_MASTER, 1):
        ws_m.cell(row=1, column=c, value=h)
    ws_m.cell(row=2, column=1, value="UMAG-001")
    ws_m.cell(row=2, column=12, value=19000)  # IVA

    encabezados_detalle_viejos = acc.ENCABEZADOS_DETALLE[:-1]  # sin "Total con IVA"
    ws_d = wb.create_sheet("Detalle")
    for c, h in enumerate(encabezados_detalle_viejos, 1):
        ws_d.cell(row=1, column=c, value=h)
    ws_d.cell(row=2, column=1, value="UMAG-001")
    ws_d.cell(row=2, column=10, value=100000)  # Total sin IVA

    acc.migrar_columna_total_con_iva_detalle(ws_m, ws_d)

    assert ws_d.cell(row=1, column=11).value == "Total con IVA (CLP)"
    assert ws_d.cell(row=2, column=11).value == 119000


def test_migrar_columna_total_con_iva_detalle_es_idempotente():
    wb = openpyxl.Workbook()
    ws_m = wb.active
    ws_m.title = "Master"
    ws_d = wb.create_sheet("Detalle")
    for c, h in enumerate(acc.ENCABEZADOS_DETALLE, 1):
        ws_d.cell(row=1, column=c, value=h)
    ws_d.cell(row=2, column=11, value=12345)

    acc.migrar_columna_total_con_iva_detalle(ws_m, ws_d)

    assert ws_d.cell(row=2, column=11).value == 12345  # no se toco de nuevo
