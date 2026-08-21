from pathlib import Path

import openpyxl

import auditor_centro_costos as acc


def test_main_pe_corre_sin_errores_sobre_arbol_vacio(tmp_path, monkeypatch, capsys):
    """Simula un arbol de carpetas Peru completamente vacio (0 proyectos, 0
    documentos, JSON vacio) y verifica que main(pais='PE') lo procesa sin
    lanzar excepciones, guarda un Excel con los encabezados de Peru, y no
    intenta actualizar el sitio de comunicacion (Peru no tiene uno).

    IMPORTANTE: main(pais="PE") llama configurar_pais("PE") como su PRIMERA
    linea -- eso pisaria cualquier monkeypatch.setattr(acc, "RUTA_EXCEL", ...)
    hecho ANTES de esta llamada. Por eso este test parchea el diccionario
    PAISES["PE"] (via monkeypatch.setitem, que se revierte solo al terminar
    el test) en vez de los globals directamente -- asi configurar_pais("PE")
    resuelve exactamente a estas rutas de tmp_path cuando main() la invoque."""
    raiz_docs = tmp_path / "Facturas" / "Perú"
    raiz_docs.mkdir(parents=True)
    (tmp_path / "Excel").mkdir()  # wb.save() no crea el directorio padre solo
    ruta_excel = tmp_path / "Excel" / "Centro de Costos Perú.xlsx"
    ruta_backups = tmp_path / "Excel" / "Respaldos"
    ruta_json = tmp_path / "datos_extraidos_peru.json"
    ruta_json.write_text("[]", encoding="utf-8")
    ruta_logs = tmp_path / "logs"

    pe_cfg = dict(acc.PAISES["PE"])
    pe_cfg["ruta_docs"] = raiz_docs
    pe_cfg["ruta_excel"] = ruta_excel
    pe_cfg["ruta_backups"] = ruta_backups
    pe_cfg["ruta_json"] = ruta_json
    pe_cfg["ruta_reconciliacion"] = tmp_path / "reconciliacion_archivos_peru.json"
    pe_cfg["ruta_logs"] = ruta_logs
    pe_cfg["ruta_excel_sitio_comunicacion"] = None
    pe_cfg["ruta_visualizador_web"] = tmp_path / "Visualizador Web"
    monkeypatch.setitem(acc.PAISES, "PE", pe_cfg)

    acc.main(pais="PE")

    salida = capsys.readouterr().out
    assert "No se pudo actualizar la copia en Sitio de comunicacion" not in salida
    assert ruta_excel.exists()

    wb = openpyxl.load_workbook(str(ruta_excel))
    assert wb["Master"].cell(row=1, column=11).value == "Total sin IGV (PEN)"
    assert wb["Master"].cell(row=1, column=12).value == "IGV 18% (PEN)"


def test_main_sin_argumentos_sigue_siendo_chile(monkeypatch):
    """main() sin argumentos (la firma que usa hoy driver.py 'run') debe
    seguir apuntando a Chile -- no debe requerir pasar pais explicitamente."""
    llamadas = []
    monkeypatch.setattr(acc, "configurar_pais", lambda pais="CL": llamadas.append(pais))
    # Forzamos que falle temprano (carpeta no existe) para no ejecutar main()
    # completo -- solo nos interesa que configurar_pais("CL") se haya llamado.
    monkeypatch.setattr(acc, "RAIZ_DOCS", Path("Z:/no-existe-de-verdad"))
    acc.main()
    assert llamadas == ["CL"]
