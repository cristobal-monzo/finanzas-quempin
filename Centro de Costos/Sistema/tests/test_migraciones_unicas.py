"""Tests de las migraciones unicas/idempotentes (migrar_columna_proveedor,
migrar_paleta_colores) que ya corrieron una vez contra el Excel real
(2026-07-16 y 2026-07-17). Cubren sobre todo el guard de idempotencia: que
NO vuelvan a tocar nada si el libro ya esta migrado -- son destructivas por
diseno (insertan columnas / borran formato condicional), asi que si el guard
se rompe alguna vez, correrian de nuevo contra el Excel real."""

import openpyxl

import auditor_centro_costos as acc


# ── migrar_columna_proveedor ─────────────────────────────────────────────

def _wb_master_sin_migrar():
    """Master en el layout ANTERIOR a la migracion: columna 8 = 'Categoría'
    (no 'Proveedor (Razón Social)' todavia)."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Master"
    headers_viejos = [
        "N° Ref.", "Proyecto", "Tipo de Proyecto", "Fecha", "N° Documento",
        "Tipo Documento", "Proveedor", "Categoría",
    ]
    for c, h in enumerate(headers_viejos, 1):
        ws.cell(row=1, column=c, value=h)
    ws.cell(row=2, column=1, value="UMAG-001")
    ws.cell(row=2, column=7, value="Comercial Beckman SpA")
    ws.cell(row=2, column=8, value="Materiales")
    return wb, ws


def test_migrar_columna_proveedor_inserta_razon_social_y_deja_tag_corto():
    wb, ws = _wb_master_sin_migrar()

    acc.migrar_columna_proveedor(wb)

    assert ws.cell(row=1, column=8).value == "Proveedor (Razón Social)"
    assert ws.cell(row=2, column=8).value == "Comercial Beckman SpA"
    assert ws.cell(row=2, column=7).value == "Beckman"
    # La columna que antes era "Categoría" (col 8) se corrio una posicion (col 9).
    assert ws.cell(row=1, column=9).value == "Categoría"
    assert ws.cell(row=2, column=9).value == "Materiales"


def test_migrar_columna_proveedor_es_no_op_si_ya_esta_migrado():
    wb, ws = _wb_master_sin_migrar()
    acc.migrar_columna_proveedor(wb)
    estado_tras_primera_corrida = [ws.cell(row=2, column=c).value for c in range(1, 10)]

    acc.migrar_columna_proveedor(wb)  # segunda corrida: no deberia tocar nada

    assert [ws.cell(row=2, column=c).value for c in range(1, 10)] == estado_tras_primera_corrida
    assert ws.cell(row=1, column=8).value == "Proveedor (Razón Social)"  # no se inserto una 2a vez


# ── migrar_paleta_colores ────────────────────────────────────────────────

def test_migrar_paleta_colores_no_op_sin_formato_condicional_heredado():
    """Un libro ya migrado (o creado desde cero) no tiene ninguna de las 3
    reglas de formato condicional heredadas del pipeline perdido -- debe
    ser un no-op total, sin excepciones, sin tocar el relleno de ninguna fila."""
    wb = openpyxl.Workbook()
    ws_master = wb.active
    ws_master.title = "Master"
    ws_master.cell(row=2, column=1, value="UMAG-001")
    ws_master.cell(row=2, column=2, value="UMAG")
    ws_detalle = wb.create_sheet("Detalle")

    relleno_previo_rgb = ws_master.cell(row=2, column=2).fill.fgColor.rgb

    acc.migrar_paleta_colores(wb, ws_master, ws_detalle)

    assert ws_master.cell(row=2, column=2).fill.fgColor.rgb == relleno_previo_rgb


# ── migrar_columna_total_con_iva_detalle ─────────────────────────────────

def test_migrar_total_con_iva_es_no_op_si_encabezado_ya_es_el_del_pais_activo(tmp_path):
    acc.configurar_pais("PE")
    try:
        wb = openpyxl.Workbook()
        ws_master = wb.active
        ws_master.title = "Master"
        ws_detalle = wb.create_sheet("Detalle")
        for c, h in enumerate(acc.ENCABEZADOS_DETALLE, 1):
            ws_detalle.cell(row=1, column=c, value=h)

        acc.migrar_columna_total_con_iva_detalle(ws_master, ws_detalle)

        # Header sigue siendo el de Peru -- no se piso con el literal chileno.
        assert ws_detalle.cell(row=1, column=11).value == "Total con IGV (PEN)"
    finally:
        acc.configurar_pais("CL")
