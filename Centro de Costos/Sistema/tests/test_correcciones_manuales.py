import json
import time

import openpyxl

import auditor_centro_costos as acc


def _wb_master_con_fila(n_ref, valores, filas_previas=0):
    """Crea un Workbook con una hoja Master minima: fila 1 encabezados,
    `filas_previas` filas de relleno (con N Ref valido, para que
    ultima_fila_datos/PATRON_NREF las recorra sin cortar antes de tiempo) y
    luego la fila con `n_ref` y los valores (dict columna->(valor, es_rojo))
    indicados en `valores`. `filas_previas` sirve para simular que
    reordenar_por_fecha movio la fila de posicion entre corridas."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Master"
    for c, h in enumerate(acc.ENCABEZADOS_MASTER, 1):
        ws.cell(row=1, column=c, value=h)

    for i in range(filas_previas):
        ws.cell(row=2 + i, column=1, value=f"REL-{i + 1:03d}")

    fila = 2 + filas_previas
    ws.cell(row=fila, column=1, value=n_ref)
    for col, (valor, es_rojo) in valores.items():
        cell = ws.cell(row=fila, column=col, value=valor)
        cell.font = acc.ROJO_FONT if es_rojo else acc.NORMAL_FONT
    return wb


# ── backup_mas_reciente ──────────────────────────────────────────────────

def test_backup_mas_reciente_devuelve_none_si_no_existe_carpeta(tmp_path):
    assert acc.backup_mas_reciente(tmp_path / "NoExiste") is None


def test_backup_mas_reciente_devuelve_none_si_carpeta_vacia(tmp_path):
    assert acc.backup_mas_reciente(tmp_path) is None


def test_backup_mas_reciente_elige_el_de_mtime_mas_reciente(tmp_path):
    viejo = tmp_path / "Centro de Costos - backup 2026-07-16 1000.xlsx"
    nuevo = tmp_path / "Centro de Costos - backup 2026-07-17 1000.xlsx"
    viejo.write_bytes(b"x")
    time.sleep(0.05)
    nuevo.write_bytes(b"x")

    assert acc.backup_mas_reciente(tmp_path) == nuevo


# ── detectar_correcciones_manuales ───────────────────────────────────────

def test_detecta_celda_roja_cuyo_valor_cambio():
    wb_anterior = _wb_master_con_fila("UMAG-014", {5: ("S/N (IMG_7533)", True)})
    wb_actual = _wb_master_con_fila("UMAG-014", {5: ("12345", False)})

    detectadas = acc.detectar_correcciones_manuales(wb_anterior, wb_actual)

    assert detectadas == [{
        "n_ref": "UMAG-014", "hoja": "Master", "columna": 5,
        "campo": "N° Documento",
        "valor_anterior": "S/N (IMG_7533)", "valor_corregido": "12345",
    }]


def test_ignora_celda_que_no_estaba_en_rojo():
    wb_anterior = _wb_master_con_fila("UMAG-014", {5: ("ALGO", False)})
    wb_actual = _wb_master_con_fila("UMAG-014", {5: ("OTRA_COSA", False)})

    assert acc.detectar_correcciones_manuales(wb_anterior, wb_actual) == []


def test_ignora_celda_roja_cuyo_valor_no_cambio():
    wb_anterior = _wb_master_con_fila("UMAG-014", {5: ("S/N (IMG_7533)", True)})
    wb_actual = _wb_master_con_fila("UMAG-014", {5: ("S/N (IMG_7533)", True)})

    assert acc.detectar_correcciones_manuales(wb_anterior, wb_actual) == []


def test_empareja_por_n_ref_no_por_numero_de_fila():
    # En wb_anterior la fila esta en la posicion 2; en wb_actual (tras
    # reordenar_por_fecha) la misma fila quedo en la posicion 6.
    wb_anterior = _wb_master_con_fila("UMAG-014", {5: ("S/N (IMG_7533)", True)}, filas_previas=0)
    wb_actual = _wb_master_con_fila("UMAG-014", {5: ("12345", False)}, filas_previas=4)

    detectadas = acc.detectar_correcciones_manuales(wb_anterior, wb_actual)

    assert len(detectadas) == 1
    assert detectadas[0]["n_ref"] == "UMAG-014"
    assert detectadas[0]["valor_corregido"] == "12345"


def test_ignora_n_ref_que_no_existe_en_wb_anterior():
    wb_anterior = _wb_master_con_fila("UMAG-001", {5: ("S/N (x)", True)})
    wb_actual = _wb_master_con_fila("UMAG-002", {5: ("999", False)})

    assert acc.detectar_correcciones_manuales(wb_anterior, wb_actual) == []


def test_detecta_correccion_de_iva_en_columna_12():
    wb_anterior = _wb_master_con_fila("UMAG-014", {12: (100, True)})
    wb_actual = _wb_master_con_fila("UMAG-014", {12: (190, False)})

    detectadas = acc.detectar_correcciones_manuales(wb_anterior, wb_actual)

    assert len(detectadas) == 1
    assert detectadas[0]["campo"] == "IVA 19% (CLP)"
    assert detectadas[0]["valor_corregido"] == 190


# ── registrar_correcciones_pendientes ────────────────────────────────────

def _correccion(n_ref="UMAG-014", columna=5, valor_corregido="12345", **overrides):
    d = {
        "n_ref": n_ref, "hoja": "Master", "columna": columna,
        "campo": "N° Documento", "valor_anterior": "S/N (x)",
        "valor_corregido": valor_corregido,
    }
    d.update(overrides)
    return d


def _errores_md_de_prueba(tmp_path):
    ruta = tmp_path / "ERRORES.md"
    ruta.write_text(
        "# Errores\n\n"
        "## Correcciones manuales pendientes de recolorear\n\n"
        "| Fecha | Hoja | N° Ref. | Campo / Columna | Valor anterior (rojo) | Valor corregido | Estado |\n"
        "|---|---|---|---|---|---|---|\n"
        "| *(sin entradas todavía)* | | | | | | |\n\n"
        "## Historial de errores detectados\n\n"
        "- contenido de prueba que no debe tocarse\n",
        encoding="utf-8",
    )
    return ruta


def test_registrar_agrega_nueva_pendiente(tmp_path):
    ruta_json = tmp_path / "correcciones_manuales.json"
    ruta_md = _errores_md_de_prueba(tmp_path)

    pendientes = acc.registrar_correcciones_pendientes(
        [_correccion()], ruta_correcciones=ruta_json, ruta_errores=ruta_md,
    )

    assert len(pendientes) == 1
    assert pendientes[0]["estado"] == "Pendiente"
    guardadas = json.loads(ruta_json.read_text(encoding="utf-8"))
    assert len(guardadas) == 1
    assert guardadas[0]["n_ref"] == "UMAG-014"

    texto_md = ruta_md.read_text(encoding="utf-8")
    assert "UMAG-014" in texto_md
    assert "contenido de prueba que no debe tocarse" in texto_md


def test_registrar_no_duplica_pendiente_actualiza_valor(tmp_path):
    ruta_json = tmp_path / "correcciones_manuales.json"
    ruta_md = _errores_md_de_prueba(tmp_path)

    acc.registrar_correcciones_pendientes(
        [_correccion(valor_corregido="12345")], ruta_correcciones=ruta_json, ruta_errores=ruta_md,
    )
    acc.registrar_correcciones_pendientes(
        [_correccion(valor_corregido="99999")], ruta_correcciones=ruta_json, ruta_errores=ruta_md,
    )

    guardadas = json.loads(ruta_json.read_text(encoding="utf-8"))
    assert len(guardadas) == 1
    assert guardadas[0]["valor_corregido"] == "99999"


def test_registrar_ignora_ya_aplicado_con_mismo_valor(tmp_path):
    ruta_json = tmp_path / "correcciones_manuales.json"
    ruta_md = _errores_md_de_prueba(tmp_path)
    ruta_json.write_text(json.dumps([{
        "n_ref": "UMAG-014", "hoja": "Master", "columna": 5, "campo": "N° Documento",
        "valor_anterior": "S/N (x)", "valor_corregido": "12345",
        "estado": "Aplicado", "fecha_detectado": "2026-07-16", "fecha_aplicado": "2026-07-16",
    }]), encoding="utf-8")

    pendientes = acc.registrar_correcciones_pendientes(
        [_correccion(valor_corregido="12345")], ruta_correcciones=ruta_json, ruta_errores=ruta_md,
    )

    assert pendientes == []
    guardadas = json.loads(ruta_json.read_text(encoding="utf-8"))
    assert len(guardadas) == 1
    assert guardadas[0]["estado"] == "Aplicado"


def test_registrar_con_escribir_false_no_toca_archivos(tmp_path):
    ruta_json = tmp_path / "correcciones_manuales.json"
    ruta_md = _errores_md_de_prueba(tmp_path)
    texto_original = ruta_md.read_text(encoding="utf-8")

    pendientes = acc.registrar_correcciones_pendientes(
        [_correccion()], escribir=False, ruta_correcciones=ruta_json, ruta_errores=ruta_md,
    )

    assert len(pendientes) == 1
    assert not ruta_json.exists()
    assert ruta_md.read_text(encoding="utf-8") == texto_original


# ── regenerar_tabla_errores_md ───────────────────────────────────────────

def test_regenerar_tabla_reemplaza_solo_el_bloque_de_la_tabla(tmp_path):
    ruta_md = _errores_md_de_prueba(tmp_path)
    correcciones = [{
        "n_ref": "UMAG-014", "hoja": "Master", "campo": "N° Documento",
        "valor_anterior": "S/N (x)", "valor_corregido": "12345",
        "estado": "Aplicado", "fecha_detectado": "2026-07-17", "fecha_aplicado": "2026-07-17",
    }]

    acc.regenerar_tabla_errores_md(correcciones, ruta_errores=ruta_md)

    texto = ruta_md.read_text(encoding="utf-8")
    assert "| 2026-07-17 | Master | UMAG-014 | N° Documento | S/N (x) | 12345 | Aplicado (2026-07-17) |" in texto
    assert "contenido de prueba que no debe tocarse" in texto
    assert "*(sin entradas todavía)*" not in texto


# ── confirmar_correcciones ────────────────────────────────────────────────

def _excel_master_detalle(tmp_path, n_ref="UMAG-014", valor_master="S/N (IMG_7533)"):
    ruta_excel = tmp_path / "Centro de Costos.xlsx"
    wb = openpyxl.Workbook()
    ws_master = wb.active
    ws_master.title = "Master"
    for c, h in enumerate(acc.ENCABEZADOS_MASTER, 1):
        ws_master.cell(row=1, column=c, value=h)
    ws_master.cell(row=2, column=1, value=n_ref)
    ws_master.cell(row=2, column=5, value=valor_master).font = acc.ROJO_FONT

    ws_detalle = wb.create_sheet("Detalle")
    for c, h in enumerate(acc.ENCABEZADOS_DETALLE, 1):
        ws_detalle.cell(row=1, column=c, value=h)
    for fila in (2, 3):
        ws_detalle.cell(row=fila, column=1, value=n_ref)
        ws_detalle.cell(row=fila, column=4, value=valor_master).font = acc.ROJO_FONT

    wb.save(str(ruta_excel))
    return ruta_excel


def test_confirmar_preview_no_toca_el_excel(tmp_path):
    ruta_json = tmp_path / "correcciones_manuales.json"
    acc.guardar_correcciones_manuales([_correccion(estado="Pendiente", fecha_detectado="2026-07-17",
                                                    fecha_aplicado=None)], ruta=ruta_json)

    pendientes = acc.confirmar_correcciones(None, ruta_correcciones=ruta_json)

    assert len(pendientes) == 1


def test_confirmar_todos_recolorea_y_propaga_a_detalle(tmp_path):
    # El Excel ya tiene el valor corregido a mano (font todavia en rojo,
    # pendiente de confirmar) -- confirmar_correcciones no cambia el VALOR,
    # solo lo recolorea y lo propaga a Detalle.
    ruta_excel = _excel_master_detalle(tmp_path, valor_master="12345")
    ruta_json = tmp_path / "correcciones_manuales.json"
    ruta_md = _errores_md_de_prueba(tmp_path)
    ruta_backups = tmp_path / "Respaldos"
    acc.guardar_correcciones_manuales([_correccion(
        valor_corregido="12345", estado="Pendiente", fecha_detectado="2026-07-17", fecha_aplicado=None,
    )], ruta=ruta_json)

    aplicadas = acc.confirmar_correcciones(
        "TODOS", ruta_excel=ruta_excel, ruta_correcciones=ruta_json,
        ruta_errores=ruta_md, ruta_backups=ruta_backups,
    )

    assert len(aplicadas) == 1
    assert aplicadas[0]["estado"] == "Aplicado"

    wb = openpyxl.load_workbook(str(ruta_excel))
    ws_master = wb["Master"]
    assert ws_master.cell(row=2, column=5).value == "12345"
    assert ws_master.cell(row=2, column=5).font.color.rgb.upper().endswith(acc.NAVY_OSCURO)

    ws_detalle = wb["Detalle"]
    for fila in (2, 3):
        assert ws_detalle.cell(row=fila, column=4).value == "12345"
        assert ws_detalle.cell(row=fila, column=4).font.color.rgb.upper().endswith(acc.NAVY_OSCURO)

    guardadas = json.loads(ruta_json.read_text(encoding="utf-8"))
    assert guardadas[0]["estado"] == "Aplicado"
    assert list(ruta_backups.rglob("*.xlsx")), "confirmar debe respaldar antes de escribir"


def test_confirmar_salta_si_el_valor_cambio_de_nuevo(tmp_path):
    ruta_excel = _excel_master_detalle(tmp_path, valor_master="S/N (IMG_7533)")
    ruta_json = tmp_path / "correcciones_manuales.json"
    ruta_md = _errores_md_de_prueba(tmp_path)
    # El JSON quedo con un valor_corregido que ya no coincide con la celda real
    # (alguien la volvio a editar despues de la deteccion).
    acc.guardar_correcciones_manuales([_correccion(
        valor_corregido="99999", estado="Pendiente", fecha_detectado="2026-07-17", fecha_aplicado=None,
    )], ruta=ruta_json)

    aplicadas = acc.confirmar_correcciones(
        "TODOS", ruta_excel=ruta_excel, ruta_correcciones=ruta_json,
        ruta_errores=ruta_md, ruta_backups=tmp_path / "Respaldos",
    )

    assert aplicadas == []
    guardadas = json.loads(ruta_json.read_text(encoding="utf-8"))
    assert guardadas[0]["estado"] == "Pendiente"


def test_confirmar_por_n_ref_puntual_no_afecta_otras_pendientes(tmp_path):
    ruta_excel = _excel_master_detalle(tmp_path, n_ref="UMAG-014", valor_master="12345")
    ruta_json = tmp_path / "correcciones_manuales.json"
    ruta_md = _errores_md_de_prueba(tmp_path)
    acc.guardar_correcciones_manuales([
        _correccion(n_ref="UMAG-014", valor_corregido="12345", estado="Pendiente",
                    fecha_detectado="2026-07-17", fecha_aplicado=None),
        _correccion(n_ref="CFLI-002", valor_corregido="777", estado="Pendiente",
                    fecha_detectado="2026-07-17", fecha_aplicado=None),
    ], ruta=ruta_json)

    aplicadas = acc.confirmar_correcciones(
        ["UMAG-014"], ruta_excel=ruta_excel, ruta_correcciones=ruta_json,
        ruta_errores=ruta_md, ruta_backups=tmp_path / "Respaldos",
    )

    assert [a["n_ref"] for a in aplicadas] == ["UMAG-014"]
    guardadas = {g["n_ref"]: g["estado"] for g in json.loads(ruta_json.read_text(encoding="utf-8"))}
    assert guardadas == {"UMAG-014": "Aplicado", "CFLI-002": "Pendiente"}
