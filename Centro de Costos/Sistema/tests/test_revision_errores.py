import json

import openpyxl

import auditor_centro_costos as acc


# ── listar_celdas_rojas ──────────────────────────────────────────────────

def _wb_master(filas):
    """filas: lista de (n_ref, proyecto, {columna: (valor, es_rojo)})."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Master"
    for c, h in enumerate(acc.ENCABEZADOS_MASTER, 1):
        ws.cell(row=1, column=c, value=h)

    for i, (n_ref, proyecto, valores) in enumerate(filas):
        fila = 2 + i
        ws.cell(row=fila, column=1, value=n_ref)
        ws.cell(row=fila, column=2, value=proyecto)
        for col, (valor, es_rojo) in valores.items():
            cell = ws.cell(row=fila, column=col, value=valor)
            cell.font = acc.ROJO_FONT if es_rojo else acc.NORMAL_FONT
    return ws


def test_listar_celdas_rojas_encuentra_n_documento_y_iva():
    ws = _wb_master([
        ("UMAG-014", "UMAG", {5: ("S/N (IMG_7533)", True)}),
        ("CFLI-002", "Cesfam Limache", {12: (100, True)}),
    ])

    encontradas = acc.listar_celdas_rojas(ws)

    assert len(encontradas) == 2
    assert encontradas[0] == {
        "n_ref": "UMAG-014", "fila": 2, "columna": 5, "campo": "N° Documento",
        "valor_actual": "S/N (IMG_7533)", "proyecto": "UMAG", "archivo_origen": None,
    }
    assert encontradas[1]["n_ref"] == "CFLI-002"
    assert encontradas[1]["campo"] == "IVA 19% (CLP)"


def test_listar_celdas_rojas_ignora_celdas_no_rojas():
    ws = _wb_master([("UMAG-014", "UMAG", {5: ("12345", False)})])

    assert acc.listar_celdas_rojas(ws) == []


def test_listar_celdas_rojas_ignora_columnas_no_revisables():
    # Columna 9 (Categoria) en rojo no cuenta -- el script nunca pinta esa
    # columna de rojo, listar_celdas_rojas solo mira COLUMNAS_REVISABLES.
    ws = _wb_master([("UMAG-014", "UMAG", {9: ("Materiales", True)})])

    assert acc.listar_celdas_rojas(ws) == []


def test_listar_celdas_rojas_vacio_sin_filas_de_datos():
    ws = _wb_master([])

    assert acc.listar_celdas_rojas(ws) == []


# ── corregir_valor_manual ────────────────────────────────────────────────

def _excel_master_detalle_rojo(tmp_path, n_ref="UMAG-014", columna=5, valor="S/N (IMG_7533)"):
    ruta_excel = tmp_path / "Centro de Costos.xlsx"
    wb = openpyxl.Workbook()
    ws_master = wb.active
    ws_master.title = "Master"
    for c, h in enumerate(acc.ENCABEZADOS_MASTER, 1):
        ws_master.cell(row=1, column=c, value=h)
    ws_master.cell(row=2, column=1, value=n_ref)
    ws_master.cell(row=2, column=columna, value=valor).font = acc.ROJO_FONT

    ws_detalle = wb.create_sheet("Detalle")
    for c, h in enumerate(acc.ENCABEZADOS_DETALLE, 1):
        ws_detalle.cell(row=1, column=c, value=h)
    for fila in (2, 3):
        ws_detalle.cell(row=fila, column=1, value=n_ref)
        ws_detalle.cell(row=fila, column=4, value=valor).font = acc.ROJO_FONT

    wb.save(str(ruta_excel))
    return ruta_excel


def _errores_md_de_prueba(tmp_path):
    ruta = tmp_path / "ERRORES.md"
    ruta.write_text(
        "# Errores\n\n"
        "## Correcciones manuales pendientes de recolorear\n\n"
        "| Fecha | Hoja | N° Ref. | Campo / Columna | Valor anterior (rojo) | Valor corregido | Estado |\n"
        "|---|---|---|---|---|---|---|\n"
        "| *(sin entradas todavía)* | | | | | | |\n\n"
        "## Historial de errores detectados\n\n"
        "- contenido de prueba que no debe tocarse\n",
        encoding="utf-8",
    )
    return ruta


def test_corregir_valor_manual_recolorea_y_propaga_n_documento(tmp_path):
    ruta_excel = _excel_master_detalle_rojo(tmp_path, valor="S/N (IMG_7533)")
    ruta_json = tmp_path / "correcciones_manuales.json"
    ruta_md = _errores_md_de_prueba(tmp_path)
    ruta_backups = tmp_path / "Respaldos"

    resultado = acc.corregir_valor_manual(
        "UMAG-014", 5, "12345", ruta_excel=ruta_excel, ruta_correcciones=ruta_json,
        ruta_errores=ruta_md, ruta_backups=ruta_backups,
    )

    assert resultado["estado"] == "Aplicado"
    assert resultado["valor_anterior"] == "S/N (IMG_7533)"
    assert resultado["valor_corregido"] == "12345"

    wb = openpyxl.load_workbook(str(ruta_excel))
    ws_master = wb["Master"]
    assert ws_master.cell(row=2, column=5).value == "12345"
    assert ws_master.cell(row=2, column=5).font.color.rgb.upper().endswith(acc.NAVY_OSCURO)

    ws_detalle = wb["Detalle"]
    for fila in (2, 3):
        assert ws_detalle.cell(row=fila, column=4).value == "12345"
        assert ws_detalle.cell(row=fila, column=4).font.color.rgb.upper().endswith(acc.NAVY_OSCURO)

    guardadas = json.loads(ruta_json.read_text(encoding="utf-8"))
    assert guardadas[0]["estado"] == "Aplicado"
    assert "12345" in ruta_md.read_text(encoding="utf-8")
    assert list(ruta_backups.rglob("*.xlsx")), "corregir_valor_manual debe respaldar antes de escribir"


def test_corregir_valor_manual_no_propaga_iva_a_detalle(tmp_path):
    ruta_excel = _excel_master_detalle_rojo(tmp_path, columna=12, valor=100)
    ruta_json = tmp_path / "correcciones_manuales.json"
    ruta_md = _errores_md_de_prueba(tmp_path)

    resultado = acc.corregir_valor_manual(
        "UMAG-014", 12, 190, ruta_excel=ruta_excel, ruta_correcciones=ruta_json,
        ruta_errores=ruta_md, ruta_backups=tmp_path / "Respaldos",
    )

    assert resultado["campo"] == "IVA 19% (CLP)"
    wb = openpyxl.load_workbook(str(ruta_excel))
    assert wb["Master"].cell(row=2, column=12).value == 190
    # Detalle no tiene columna de IVA -- la fila 2 de Detalle no debe alterarse
    assert wb["Detalle"].cell(row=2, column=4).value == 100


def test_corregir_valor_manual_no_toca_celda_que_no_esta_roja(tmp_path):
    ruta_excel = tmp_path / "Centro de Costos.xlsx"
    wb = openpyxl.Workbook()
    ws_master = wb.active
    ws_master.title = "Master"
    for c, h in enumerate(acc.ENCABEZADOS_MASTER, 1):
        ws_master.cell(row=1, column=c, value=h)
    ws_master.cell(row=2, column=1, value="UMAG-014")
    ws_master.cell(row=2, column=5, value="12345").font = acc.AZUL_MARINO_FONT
    wb.save(str(ruta_excel))

    resultado = acc.corregir_valor_manual(
        "UMAG-014", 5, "99999", ruta_excel=ruta_excel,
        ruta_correcciones=tmp_path / "correcciones_manuales.json",
        ruta_errores=tmp_path / "ERRORES.md", ruta_backups=tmp_path / "Respaldos",
    )

    assert resultado is None
    wb2 = openpyxl.load_workbook(str(ruta_excel))
    assert wb2["Master"].cell(row=2, column=5).value == "12345"


def test_corregir_valor_manual_devuelve_none_si_n_ref_no_existe(tmp_path):
    ruta_excel = _excel_master_detalle_rojo(tmp_path, n_ref="UMAG-014")

    resultado = acc.corregir_valor_manual(
        "UMAG-999", 5, "12345", ruta_excel=ruta_excel,
        ruta_correcciones=tmp_path / "correcciones_manuales.json",
        ruta_errores=tmp_path / "ERRORES.md", ruta_backups=tmp_path / "Respaldos",
    )

    assert resultado is None


# ── listar_items_agrupados / desglosar_item_agrupado ────────────────────

def _excel_con_item_agrupado(tmp_path, n_ref="CCON-004", proyecto="Cesfam Constitución",
                              iva=33082, nombre_agrupado="Materiales varios"):
    """Master con 1 documento + Detalle con 2 items legibles y 1 agrupado
    (mismo patron que CCON-004: precedente real del modulo, ver ERRORES.md)."""
    ruta_excel = tmp_path / "Centro de Costos.xlsx"
    wb = openpyxl.Workbook()
    ws_master = wb.active
    ws_master.title = "Master"
    for c, h in enumerate(acc.ENCABEZADOS_MASTER, 1):
        ws_master.cell(row=1, column=c, value=h)
    ws_master.cell(row=2, column=1, value=n_ref)
    ws_master.cell(row=2, column=2, value=proyecto)
    ws_master.cell(row=2, column=10, value="Codo bronce; Terminal bronce; " + nombre_agrupado)
    ws_master.cell(row=2, column=12, value=iva)

    ws_detalle = wb.create_sheet("Detalle")
    for c, h in enumerate(acc.ENCABEZADOS_DETALLE, 1):
        ws_detalle.cell(row=1, column=c, value=h)

    filas = [
        (n_ref, proyecto, "Mantenimiento", "130020", "Codo bronce", "Codo bronce 1 1/2", "Materiales", 3, 6546, 19638, 23369),
        (n_ref, proyecto, "Mantenimiento", "130020", "Terminal bronce", "Terminal bronce 1 1/2", "Materiales", 2, 7105, 14210, 16910),
        (n_ref, proyecto, "Mantenimiento", "130020", nombre_agrupado, "Resto de la factura no legible", "Materiales", 1, 115206, 115206, 137095),
    ]
    for i, fila_valores in enumerate(filas):
        fila = 2 + i
        for c, v in enumerate(fila_valores, 1):
            cell = ws_detalle.cell(row=fila, column=c, value=v)
            cell.font = acc.NORMAL_FONT
        ws_detalle.cell(row=fila, column=9).number_format = acc.MONEY_FORMAT
        ws_detalle.cell(row=fila, column=10).number_format = acc.MONEY_FORMAT
        ws_detalle.cell(row=fila, column=11).number_format = acc.MONEY_FORMAT

    acc.regenerar_pie(ws_detalle, len(acc.ENCABEZADOS_DETALLE), [10, 11], "TOTAL GENERAL", 5, acc.LEYENDA_DETALLE)

    wb.save(str(ruta_excel))
    return ruta_excel


def test_listar_items_agrupados_encuentra_fila_con_varios(tmp_path):
    ruta_excel = _excel_con_item_agrupado(tmp_path)
    wb = openpyxl.load_workbook(str(ruta_excel))

    encontrados = acc.listar_items_agrupados(wb["Detalle"])

    assert len(encontrados) == 1
    assert encontrados[0]["n_ref"] == "CCON-004"
    assert encontrados[0]["nombre_item"] == "Materiales varios"
    assert encontrados[0]["fila"] == 4


def test_listar_items_agrupados_ignora_items_normales(tmp_path):
    ruta_excel = _excel_con_item_agrupado(tmp_path, nombre_agrupado="Varios")
    wb = openpyxl.load_workbook(str(ruta_excel))
    ws = wb["Detalle"]
    ws.cell(row=2, column=5, value="Variador de frecuencia")  # no debe matchear "varios"

    encontrados = acc.listar_items_agrupados(ws)

    assert len(encontrados) == 1
    assert encontrados[0]["fila"] == 4


def test_desglosar_item_agrupado_reemplaza_fila_y_recalcula_totales(tmp_path):
    ruta_excel = _excel_con_item_agrupado(tmp_path)
    ruta_json = tmp_path / "correcciones_manuales.json"
    ruta_md = _errores_md_de_prueba(tmp_path)

    items_nuevos = [
        {"nombre_item": "Cañería cobre", "descripcion": "Cañería cobre 1 1/4", "categoria_item": "Materiales", "cantidad": 2, "p_unitario_sin_iva": 30000},
        {"nombre_item": "Soldadura plata", "descripcion": "Soldadura al 6%", "categoria_item": "Materiales", "cantidad": 1, "p_unitario_sin_iva": 55206},
    ]

    resultado = acc.desglosar_item_agrupado(
        "CCON-004", items_nuevos, ruta_excel=ruta_excel, ruta_correcciones=ruta_json,
        ruta_errores=ruta_md, ruta_backups=tmp_path / "Respaldos",
    )

    assert resultado["estado"] == "Aplicado"
    assert resultado["hoja"] == "Detalle"

    wb = openpyxl.load_workbook(str(ruta_excel))
    ws_detalle = wb["Detalle"]

    assert acc.listar_items_agrupados(ws_detalle) == []
    assert acc.ultima_fila_datos(ws_detalle) == 5  # 2 legibles + 2 nuevos, la agrupada se fue

    assert ws_detalle.cell(row=4, column=5).value == "Cañería cobre"
    assert ws_detalle.cell(row=4, column=10).value == 60000
    assert ws_detalle.cell(row=4, column=4).value == "130020"  # N Documento heredado
    assert ws_detalle.cell(row=4, column=4).font.color.rgb.upper().endswith(acc.NAVY_OSCURO)

    assert ws_detalle.cell(row=5, column=5).value == "Soldadura plata"
    assert ws_detalle.cell(row=5, column=10).value == 55206

    # Pie regenerado en la fila correcta (2 legibles + 2 nuevos = filas 2-5, pie en 7)
    assert ws_detalle.cell(row=7, column=9).value == "TOTAL GENERAL"

    wb_backups = list((tmp_path / "Respaldos").rglob("*.xlsx"))
    assert wb_backups, "desglosar_item_agrupado debe respaldar antes de escribir"


def test_desglosar_item_agrupado_actualiza_resumen_items_en_master(tmp_path):
    ruta_excel = _excel_con_item_agrupado(tmp_path)
    items_nuevos = [
        {"nombre_item": "Cañería cobre", "categoria_item": "Materiales", "cantidad": 2, "p_unitario_sin_iva": 30000},
    ]

    acc.desglosar_item_agrupado(
        "CCON-004", items_nuevos, ruta_excel=ruta_excel,
        ruta_correcciones=tmp_path / "correcciones_manuales.json",
        ruta_errores=_errores_md_de_prueba(tmp_path), ruta_backups=tmp_path / "Respaldos",
    )

    wb = openpyxl.load_workbook(str(ruta_excel))
    resumen = wb["Master"].cell(row=2, column=10).value
    assert resumen == "Codo bronce; Terminal bronce; Cañería cobre"


def test_desglosar_item_agrupado_registra_correccion_en_json_y_md(tmp_path):
    ruta_excel = _excel_con_item_agrupado(tmp_path)
    ruta_json = tmp_path / "correcciones_manuales.json"
    ruta_md = _errores_md_de_prueba(tmp_path)
    items_nuevos = [
        {"nombre_item": "Cañería cobre", "categoria_item": "Materiales", "cantidad": 2, "p_unitario_sin_iva": 30000},
    ]

    acc.desglosar_item_agrupado(
        "CCON-004", items_nuevos, ruta_excel=ruta_excel, ruta_correcciones=ruta_json,
        ruta_errores=ruta_md, ruta_backups=tmp_path / "Respaldos",
    )

    guardadas = json.loads(ruta_json.read_text(encoding="utf-8"))
    assert guardadas[0]["hoja"] == "Detalle"
    assert guardadas[0]["estado"] == "Aplicado"
    assert "Cañería cobre" in guardadas[0]["valor_corregido"]
    assert "Cañería cobre" in ruta_md.read_text(encoding="utf-8")


def test_desglosar_item_agrupado_falla_si_no_hay_fila_agrupada(tmp_path):
    ruta_excel = _excel_con_item_agrupado(tmp_path, nombre_agrupado="Repuesto especifico")

    resultado = acc.desglosar_item_agrupado(
        "CCON-004", [{"nombre_item": "X", "categoria_item": "Materiales", "cantidad": 1, "p_unitario_sin_iva": 100}],
        ruta_excel=ruta_excel, ruta_correcciones=tmp_path / "correcciones_manuales.json",
        ruta_errores=tmp_path / "ERRORES.md", ruta_backups=tmp_path / "Respaldos",
    )

    assert resultado is None


def test_desglosar_item_agrupado_falla_si_hay_mas_de_una_fila_agrupada(tmp_path):
    ruta_excel = _excel_con_item_agrupado(tmp_path)
    wb = openpyxl.load_workbook(str(ruta_excel))
    ws_detalle = wb["Detalle"]
    ws_detalle.cell(row=2, column=5, value="Insumos varios")  # segunda fila agrupada
    wb.save(str(ruta_excel))

    resultado = acc.desglosar_item_agrupado(
        "CCON-004", [{"nombre_item": "X", "categoria_item": "Materiales", "cantidad": 1, "p_unitario_sin_iva": 100}],
        ruta_excel=ruta_excel, ruta_correcciones=tmp_path / "correcciones_manuales.json",
        ruta_errores=tmp_path / "ERRORES.md", ruta_backups=tmp_path / "Respaldos",
    )

    assert resultado is None


def test_corregir_valor_manual_actualiza_entrada_existente_sin_duplicar(tmp_path):
    ruta_excel = _excel_master_detalle_rojo(tmp_path, valor="S/N (IMG_7533)")
    ruta_json = tmp_path / "correcciones_manuales.json"
    ruta_md = _errores_md_de_prueba(tmp_path)
    acc.guardar_correcciones_manuales([{
        "n_ref": "UMAG-014", "hoja": "Master", "columna": 5, "campo": "N° Documento",
        "valor_anterior": "S/N (algo viejo)", "valor_corregido": "00000",
        "estado": "Aplicado", "fecha_detectado": "2026-07-10", "fecha_aplicado": "2026-07-10",
    }], ruta=ruta_json)

    acc.corregir_valor_manual(
        "UMAG-014", 5, "12345", ruta_excel=ruta_excel, ruta_correcciones=ruta_json,
        ruta_errores=ruta_md, ruta_backups=tmp_path / "Respaldos",
    )

    guardadas = json.loads(ruta_json.read_text(encoding="utf-8"))
    assert len(guardadas) == 1
    assert guardadas[0]["valor_corregido"] == "12345"
    assert guardadas[0]["valor_anterior"] == "S/N (IMG_7533)"
