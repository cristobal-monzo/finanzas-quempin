from datetime import datetime

import openpyxl

import auditor_centro_costos as acc


def _wb_master_detalle(filas_master, filas_detalle):
    """filas_master: lista de (n_ref, fecha, marca) -- 'marca' vive en
    'Resumen Ítems' (columna 10), solo para verificar que el contenido viaja
    intacto junto con la fila. filas_detalle: lista de (n_ref, marca_item)
    en el orden FISICO original (no agrupado), 'marca_item' en 'Nombre
    Ítem' (columna 5)."""
    wb = openpyxl.Workbook()
    ws_master = wb.active
    ws_master.title = "Master"
    for c, h in enumerate(acc.ENCABEZADOS_MASTER, 1):
        ws_master.cell(row=1, column=c, value=h)
    for i, (n_ref, fecha, marca) in enumerate(filas_master):
        fila = 2 + i
        ws_master.cell(row=fila, column=1, value=n_ref)
        ws_master.cell(row=fila, column=4, value=fecha)
        ws_master.cell(row=fila, column=10, value=marca)

    ws_detalle = wb.create_sheet("Detalle")
    for c, h in enumerate(acc.ENCABEZADOS_DETALLE, 1):
        ws_detalle.cell(row=1, column=c, value=h)
    for i, (n_ref, marca_item) in enumerate(filas_detalle):
        fila = 2 + i
        ws_detalle.cell(row=fila, column=1, value=n_ref)
        ws_detalle.cell(row=fila, column=5, value=marca_item)

    return ws_master, ws_detalle


def test_reordena_master_mas_reciente_primero_preservando_el_contenido():
    ws_master, ws_detalle = _wb_master_detalle(
        filas_master=[
            ("A-001", datetime(2026, 1, 1), "vieja"),
            ("A-002", datetime(2026, 3, 1), "nueva"),
            ("A-003", datetime(2026, 2, 1), "media"),
        ],
        filas_detalle=[],
    )

    acc.reordenar_por_fecha(ws_master, ws_detalle, primera_fila_libre_master=5, primera_fila_libre_detalle=2)

    assert [ws_master.cell(row=r, column=1).value for r in (2, 3, 4)] == ["A-002", "A-003", "A-001"]
    assert [ws_master.cell(row=r, column=10).value for r in (2, 3, 4)] == ["nueva", "media", "vieja"]


def test_fechas_no_interpretables_quedan_al_final_en_su_orden_original():
    ws_master, ws_detalle = _wb_master_detalle(
        filas_master=[
            ("A-001", "fecha ilegible", "sin fecha 1"),
            ("A-002", datetime(2026, 3, 1), "nueva"),
            ("A-003", "otra ilegible", "sin fecha 2"),
        ],
        filas_detalle=[],
    )

    acc.reordenar_por_fecha(ws_master, ws_detalle, primera_fila_libre_master=5, primera_fila_libre_detalle=2)

    assert [ws_master.cell(row=r, column=1).value for r in (2, 3, 4)] == ["A-002", "A-001", "A-003"]


def test_reagrupa_detalle_siguiendo_el_nuevo_orden_de_master():
    ws_master, ws_detalle = _wb_master_detalle(
        filas_master=[
            ("A-001", datetime(2026, 1, 1), "vieja"),
            ("A-002", datetime(2026, 3, 1), "nueva"),
        ],
        filas_detalle=[
            ("A-001", "item de A-001"),
            ("A-002", "item 1 de A-002"),
            ("A-002", "item 2 de A-002"),
        ],
    )

    acc.reordenar_por_fecha(ws_master, ws_detalle, primera_fila_libre_master=4, primera_fila_libre_detalle=5)

    # A-002 (mas reciente) ahora encabeza Master, y sus 2 items de Detalle
    # deben quedar agrupados primero tambien, en su orden relativo original.
    assert [ws_detalle.cell(row=r, column=1).value for r in (2, 3, 4)] == ["A-002", "A-002", "A-001"]
    assert [ws_detalle.cell(row=r, column=5).value for r in (2, 3, 4)] == [
        "item 1 de A-002", "item 2 de A-002", "item de A-001",
    ]


def test_no_toca_nada_si_ya_esta_en_el_orden_correcto():
    ws_master, ws_detalle = _wb_master_detalle(
        filas_master=[
            ("A-002", datetime(2026, 3, 1), "nueva"),
            ("A-001", datetime(2026, 1, 1), "vieja"),
        ],
        filas_detalle=[],
    )
    # Centinela: si reordenar_por_fecha regenerara las formulas K/M (solo
    # deberia pasar cuando SI reordena), este valor fijo se pisaria por una
    # formula "=SUMIF(...)".
    ws_master.cell(row=2, column=11, value=999)

    acc.reordenar_por_fecha(ws_master, ws_detalle, primera_fila_libre_master=4, primera_fila_libre_detalle=2)

    assert [ws_master.cell(row=r, column=1).value for r in (2, 3)] == ["A-002", "A-001"]
    assert ws_master.cell(row=2, column=11).value == 999
