# -*- coding: utf-8 -*-
"""
driver.py — arnés de ejecución para la skill Revision_de_Errores.

No reimplementa nada: importa auditor_centro_costos.py desde Sistema/ (igual
que el driver de Registro_Centro_de_Costos) y expone dos comandos:

  errores    → SOLO LECTURA. Lista todas las celdas de Master actualmente en
               rojo (requieren revisión) con su N° Ref, proyecto, campo, valor
               actual y la ruta relativa a la foto del documento original.

  corregir   → Aplica UNA corrección directa: recolorea la celda a azul marino
               oscuro, propaga a Detalle si corresponde, y la registra como
               "Aplicado" en correcciones_manuales.json + ERRORES.md. Solo
               opera sobre celdas que están en rojo (si no, no hace nada).
               Acepta un flag opcional --nota "<texto>" (en cualquier
               posición tras el VALOR) que queda como comentario de Excel
               sobre la celda corregida, además de en la tabla de
               ERRORES.md -- para desgloses que no caben en un solo valor
               (ej. IVA de combustible que incluye IEF/IEV-FEPP).

  agrupados  → SOLO LECTURA. Lista las filas de Detalle que agrupan en 1 sola
               línea una parte de la compra que no se pudo desglosar ítem por
               ítem (nombre con la palabra "varios", ej. "Materiales varios" —
               ver precedente CCON-004 en ERRORES.md), con su N° Ref,
               proyecto, descripción del grupo y la ruta a la foto.

  desglosar  → Reemplaza la fila agrupada de un N° Ref por N filas nuevas (una
               por ítem ya identificado por el usuario), recalcula los totales
               y "Resumen Ítems" de Master, y registra el cambio como
               "Aplicado" en correcciones_manuales.json + ERRORES.md. Solo
               opera si encuentra EXACTAMENTE 1 fila agrupada para ese N° Ref.

  reflejar   → Copia el Excel local encima de la copia de "Sitio de
               comunicación - Centro de Costos 1/" (mismo paso que hace 'run'
               al final de cada corrida). Correr siempre al terminar un
               recorrido de esta skill, haya o no correcciones aplicadas.

Uso:
  python driver.py errores
  python driver.py corregir <N_REF> <CAMPO_O_COLUMNA> <VALOR>
  python driver.py agrupados
  python driver.py desglosar <N_REF> '<ITEMS_JSON>'
  python driver.py reflejar

  Ejemplos:
  python driver.py corregir UMAG-014 "N° Documento" 12345
  python driver.py corregir UMAG-014 5 12345
  python driver.py corregir CFLI-002 iva 190
  python driver.py corregir CVAL-018 iva 8799 --nota "IVA: $7188 / IEF: $4111 / IEV/FEPP: $-2500"
  python driver.py desglosar CCON-004 "[{\"nombre_item\": \"Cañería cobre\", \"descripcion\": \"1 1/4 pulg\", \"categoria_item\": \"Materiales\", \"cantidad\": 2, \"p_unitario_sin_iva\": 30000}]"
"""

import json
import sys
from pathlib import Path

ROOT = Path(__file__).resolve().parents[3]
sys.path.insert(0, str(ROOT / "Sistema"))

sys.dont_write_bytecode = True
import auditor_centro_costos as acc  # noqa: E402


def resolver_columna(campo_o_num):
    """Acepta un número de columna o un fragmento del nombre del campo
    (insensible a mayúsculas/acentos simples) y devuelve el número de
    columna -- restringido a acc.COLUMNAS_REVISABLES, así este comando nunca
    puede tocar una columna que el script no pinta de rojo."""
    if campo_o_num.isdigit():
        col = int(campo_o_num)
        if col in acc.COLUMNAS_REVISABLES:
            return col
        return None

    objetivo = campo_o_num.strip().lower()
    for col in acc.COLUMNAS_REVISABLES:
        campo = acc.ENCABEZADOS_MASTER[col - 1].lower()
        if objetivo in campo or campo.startswith(objetivo):
            return col
    return None


def ruta_foto_para_fila(fila_dict, reconciliacion_inversa):
    """Ruta absoluta a la foto del documento (o None si no se puede
    resolver) -- reusa resolver_ruta_actual, el mismo mecanismo que usa el
    renombrado automático para encontrar el archivo físico real."""
    ruta_relativa = acc.resolver_ruta_actual(fila_dict, reconciliacion_inversa)
    if not ruta_relativa or "\\" not in ruta_relativa:
        return None
    proyecto_fisico, nombre_archivo = ruta_relativa.split("\\", 1)
    ruta_absoluta = acc.RAIZ_DOCS / proyecto_fisico / nombre_archivo
    return ruta_absoluta if ruta_absoluta.exists() else None


def cmd_errores():
    sys.stdout.reconfigure(encoding="utf-8", errors="backslashreplace")

    print("=" * 70)
    print("  ERRORES PENDIENTES DE REVISIÓN (solo lectura, no escribe nada)")
    print("=" * 70)

    if not acc.RUTA_EXCEL.exists():
        print("\n[ERROR] No existe el Excel todavía.")
        return 1

    import openpyxl
    wb = openpyxl.load_workbook(str(acc.RUTA_EXCEL), data_only=False)
    if "Master" not in wb.sheetnames:
        print("\n[ERROR] No existe la hoja Master.")
        return 1

    celdas = acc.listar_celdas_rojas(wb["Master"])
    print(f"\nCeldas en rojo encontradas: {len(celdas)}")
    if not celdas:
        print("\nNo hay errores pendientes de revisión.")
        return 0

    reconciliacion_inversa = acc.construir_reconciliacion_inversa(acc.cargar_reconciliacion())
    filas_master, _, _ = acc.leer_master(wb["Master"])
    filas_por_n_ref = {f["n_ref"]: f for f in filas_master}

    for c in celdas:
        print(f"\n- {c['n_ref']} ({c['proyecto']}) / {c['campo']}: '{c['valor_actual']}'")
        fila_dict = filas_por_n_ref.get(c["n_ref"])
        ruta_foto = ruta_foto_para_fila(fila_dict, reconciliacion_inversa) if fila_dict else None
        if ruta_foto:
            print(f"  Foto: {ruta_foto}")
        else:
            print(f"  [WARN] No se encontró la foto en disco (Archivo origen: {c['archivo_origen']}).")

    print("\n" + "=" * 70)
    print("  Nada fue escrito. Para aplicar una corrección: python driver.py corregir <N_REF> <CAMPO> <VALOR>")
    print("=" * 70)
    return 0


def cmd_agrupados():
    sys.stdout.reconfigure(encoding="utf-8", errors="backslashreplace")

    print("=" * 70)
    print("  ÍTEMS AGRUPADOS PENDIENTES DE DESGLOSAR (solo lectura, no escribe nada)")
    print("=" * 70)

    if not acc.RUTA_EXCEL.exists():
        print("\n[ERROR] No existe el Excel todavía.")
        return 1

    import openpyxl
    wb = openpyxl.load_workbook(str(acc.RUTA_EXCEL), data_only=False)
    if "Detalle" not in wb.sheetnames:
        print("\n[ERROR] No existe la hoja Detalle.")
        return 1

    agrupados = acc.listar_items_agrupados(wb["Detalle"])
    print(f"\nÍtems agrupados encontrados: {len(agrupados)}")
    if not agrupados:
        print("\nNo hay ítems agrupados pendientes de desglosar.")
        return 0

    reconciliacion_inversa = acc.construir_reconciliacion_inversa(acc.cargar_reconciliacion())
    filas_master, _, _ = acc.leer_master(wb["Master"])
    filas_por_n_ref = {f["n_ref"]: f for f in filas_master}

    for a in agrupados:
        print(f"\n- {a['n_ref']} ({a['proyecto']}) / '{a['nombre_item']}' "
              f"(fila {a['fila']} de Detalle, monto ${a['total_sin_iva']}):")
        print(f"  {a['descripcion']}")
        fila_dict = filas_por_n_ref.get(a["n_ref"])
        ruta_foto = ruta_foto_para_fila(fila_dict, reconciliacion_inversa) if fila_dict else None
        if ruta_foto:
            print(f"  Foto: {ruta_foto}")
        else:
            archivo_origen = fila_dict["archivo_origen"] if fila_dict else None
            print(f"  [WARN] No se encontró la foto en disco (Archivo origen: {archivo_origen}).")

    print("\n" + "=" * 70)
    print("  Nada fue escrito. Para desglosar: python driver.py desglosar <N_REF> '<ITEMS_JSON>'")
    print("=" * 70)
    return 0


def cmd_desglosar(args):
    if len(args) < 2:
        print("Uso: python driver.py desglosar <N_REF> '<ITEMS_JSON>'")
        return 2
    n_ref, items_json = args[0], " ".join(args[1:])

    try:
        items_nuevos = json.loads(items_json)
    except json.JSONDecodeError as e:
        print(f"[ERROR] ITEMS_JSON no es JSON válido: {e}")
        return 1
    if not isinstance(items_nuevos, list) or not items_nuevos:
        print("[ERROR] ITEMS_JSON debe ser una lista no vacía de ítems.")
        return 1

    resultado = acc.desglosar_item_agrupado(n_ref, items_nuevos)
    return 0 if resultado is not None else 1


def cmd_reflejar():
    ok = acc.reflejar_a_sitio_comunicacion()
    return 0 if ok else 1


def cmd_corregir(args):
    nota = None
    if "--nota" in args:
        idx = args.index("--nota")
        if idx + 1 >= len(args):
            print('[ERROR] --nota requiere un valor: --nota "<texto>"')
            return 2
        nota = args[idx + 1]
        args = args[:idx] + args[idx + 2:]

    if len(args) < 3:
        print('Uso: python driver.py corregir <N_REF> <CAMPO_O_COLUMNA> <VALOR> [--nota "<texto>"]')
        return 2
    n_ref, campo_o_num, valor_nuevo = args[0], args[1], " ".join(args[2:])

    columna = resolver_columna(campo_o_num)
    if columna is None:
        print(f"[ERROR] '{campo_o_num}' no es una columna revisable "
              f"({[acc.ENCABEZADOS_MASTER[c - 1] for c in acc.COLUMNAS_REVISABLES]}).")
        return 1

    if columna == 12:  # IVA 19% (CLP) es numérico
        try:
            valor_nuevo = int(valor_nuevo)
        except ValueError:
            try:
                valor_nuevo = float(valor_nuevo)
            except ValueError:
                pass

    resultado = acc.corregir_valor_manual(n_ref, columna, valor_nuevo, nota=nota)
    return 0 if resultado is not None else 1


COMANDOS = ("errores", "corregir", "agrupados", "desglosar", "reflejar")


def main():
    if len(sys.argv) < 2 or sys.argv[1] not in COMANDOS:
        print(f"Uso: python driver.py [{'|'.join(COMANDOS)}] ...")
        return 2

    comando = sys.argv[1]
    if comando == "errores":
        return cmd_errores()
    if comando == "agrupados":
        return cmd_agrupados()
    if comando == "desglosar":
        return cmd_desglosar(sys.argv[2:])
    if comando == "reflejar":
        return cmd_reflejar()
    return cmd_corregir(sys.argv[2:])


if __name__ == "__main__":
    raise SystemExit(main())
