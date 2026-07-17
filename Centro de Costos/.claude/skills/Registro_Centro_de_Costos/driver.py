# -*- coding: utf-8 -*-
"""
driver.py — arnés de ejecución para la skill Registro_Centro_de_Costos.

No reimplementa el registrador: importa auditor_centro_costos.py desde la raíz
del proyecto y expone dos comandos seguros de invocar desde un agente:

  status  → SOLO LECTURA. Inventaria Documentos Centro de Costos/, dice qué
            se registraría (pendientes/omitidos), qué archivos no tienen
            datos (con items) en datos_extraidos.json, y corre la
            verificación aritmética sobre ese JSON. No crea backup ni
            escribe el Excel.

  run     → Ejecución real: equivalente a `python auditor_centro_costos.py`
            (backup + escritura). Idempotente: correrlo varias veces no
            duplica filas (las filas de datos ya escritas nunca se tocan;
            solo se regeneran los pies de tabla y las hojas de proyecto,
            que son 100% derivadas).

Uso:
  python driver.py status
  python driver.py run
"""

import sys
from pathlib import Path

ROOT = Path(__file__).resolve().parents[3]
sys.path.insert(0, str(ROOT))

sys.dont_write_bytecode = True
import auditor_centro_costos as acc  # noqa: E402


def mostrar_preview_renombrados(filas_master, reconciliacion):
    """Preview de status (solo lectura): que archivos se renombrarian/
    convertirian si se corriera 'run', sin tocar disco."""
    reconciliacion_inversa = acc.construir_reconciliacion_inversa(reconciliacion)
    planes = acc.planificar_renombrados(filas_master, reconciliacion_inversa)
    a_renombrar = [p for p in planes if p["accion"] in ("renombrar", "convertir_heic")]
    no_encontrados = [p for p in planes if p["accion"] == "archivo_no_encontrado"]

    print(f"\nArchivos que se renombrarian/convertirian si corres 'run': {len(a_renombrar)}")
    for p in a_renombrar:
        print(f"  - {p['n_ref']}: {p['ruta_actual'].name} -> {p['nombre_nuevo']} ({p['accion']})")

    if no_encontrados:
        print(f"\n[WARN] {len(no_encontrados)} fila(s) sin archivo fisico encontrado para renombrar:")
        for p in no_encontrados:
            print(f"  - {p['n_ref']}")


def cmd_status():
    sys.stdout.reconfigure(encoding="utf-8", errors="backslashreplace")

    print("=" * 70)
    print("  ESTADO CENTRO DE COSTOS (solo lectura, no escribe nada)")
    print("=" * 70)

    print(f"\nRaíz documentos: {acc.RAIZ_DOCS}")
    print(f"  Existe: {acc.RAIZ_DOCS.exists()}")
    print(f"Excel: {acc.RUTA_EXCEL.name}")
    print(f"  Existe: {acc.RUTA_EXCEL.exists()}")
    print(f"JSON datos: {acc.RUTA_JSON.name}")
    print(f"  Existe: {acc.RUTA_JSON.exists()}")
    print(f"Reconciliación (bootstrap de documentos preexistentes): {acc.RUTA_RECONCILIACION.name}")
    print(f"  Existe: {acc.RUTA_RECONCILIACION.exists()}")

    if not acc.RAIZ_DOCS.exists() or not acc.RUTA_JSON.exists():
        print("\n[ERROR] Falta la carpeta de documentos o el JSON. Abortando status.")
        return 1

    import openpyxl

    ws_master = None
    if acc.RUTA_EXCEL.exists():
        wb = openpyxl.load_workbook(str(acc.RUTA_EXCEL), data_only=False)
        print(f"\nHojas existentes: {wb.sheetnames}")
        if "Master" in wb.sheetnames:
            ws_master = wb["Master"]
    else:
        print("\n[INFO] El Excel aún no existe, se crearía desde cero en un 'run'.")

    if ws_master is not None:
        filas_master, max_seq, docs_registrados = acc.leer_master(ws_master)
    else:
        filas_master, max_seq, docs_registrados = [], {}, set()

    reconciliacion = acc.cargar_reconciliacion()
    archivos_registrados = set(reconciliacion.keys())
    for fm in filas_master:
        if fm["archivo_origen"]:
            archivos_registrados.add(str(fm["archivo_origen"]))

    print(f"\nDocumentos ya en Master: {len(filas_master)}")
    print(f"N° Documento distintos ya registrados: {len(docs_registrados)}")
    print(f"Archivos ya cubiertos (Master + reconciliación): {len(archivos_registrados)}")

    pendientes, omitidos = acc.inventariar_archivos(acc.RAIZ_DOCS, archivos_registrados)
    print("\nInventario de Documentos Centro de Costos/:")
    print(f"  Pendientes (no registrados):            {len(pendientes)}")
    print(f"  Omitidos (ya registrados):               {len(omitidos)}")

    proyectos = sorted({p.name for p in acc.RAIZ_DOCS.iterdir() if p.is_dir()})
    print(f"\nProyectos detectados ({len(proyectos)}): {', '.join(proyectos)}")

    datos_json = acc.cargar_datos_json(acc.RUTA_JSON)
    print(f"\nEntradas en datos_extraidos.json: {len(datos_json)}")

    sin_datos = []
    for info in pendientes:
        dato = acc.buscar_dato_por_archivo(datos_json, info["proyecto"], info["archivo"])
        if not dato or not dato.get("items"):
            sin_datos.append(info)

    print(f"\nPendientes SIN datos (o sin items) en el JSON (bloquean el registro): {len(sin_datos)}")
    for info in sin_datos:
        print(f"  - [{info['proyecto']}] {info['archivo']}")

    escribibles = len(pendientes) - len(sin_datos)
    print(f"\nSi corres 'run' ahora se registrarían: {escribibles} documento(s).")

    print("\nVerificación aritmética sobre TODO datos_extraidos.json (Neto vs IVA 19%):")
    inconsistencias = acc.verificar_aritmetica(datos_json)
    if inconsistencias:
        for inc in inconsistencias:
            print(f"  * Doc {inc['n_documento']} ({inc['archivo']}): "
                  f"Neto={inc['neto']:,} IVA={inc['iva']:,} esperado={inc['iva_esperado']:,}")
    else:
        print("  Sin inconsistencias.")

    mostrar_preview_renombrados(filas_master, reconciliacion)

    print("\n" + "=" * 70)
    print("  Nada fue escrito. Para ejecutar de verdad: python driver.py run")
    print("=" * 70)
    return 0


def cmd_run():
    acc.main()
    return 0


def main():
    if len(sys.argv) != 2 or sys.argv[1] not in ("status", "run"):
        print("Uso: python driver.py [status|run]")
        return 2

    if sys.argv[1] == "status":
        return cmd_status()
    return cmd_run()


if __name__ == "__main__":
    raise SystemExit(main())
