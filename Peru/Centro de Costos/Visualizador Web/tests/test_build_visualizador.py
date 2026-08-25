import importlib.util
import sys
from pathlib import Path

import openpyxl

# Mismo patron que los otros 3 modulos (ver Centro de Costos/Visualizador
# Web/tests/test_build_visualizador.py): "importlib" import-mode evita que
# pytest choque los 4 archivos de test por basename, pero el MODULO FUENTE
# igual hay que cargarlo por ruta bajo un nombre unico para que
# sys.modules no le entregue a este test el build_visualizador.py de otro
# pais/modulo.
_RUTA_BV = Path(__file__).resolve().parent.parent / "build_visualizador.py"
_spec = importlib.util.spec_from_file_location("build_visualizador_cc_pe", _RUTA_BV)
bv = importlib.util.module_from_spec(_spec)
sys.modules["build_visualizador_cc_pe"] = bv
_spec.loader.exec_module(bv)

HEADERS_MASTER = [
    "N° Ref.", "Proyecto", "Tipo de Proyecto", "Fecha", "N° Documento",
    "Tipo Documento", "Proveedor", "Proveedor (Razón Social)", "Categoría",
    "Resumen Ítems", "Total sin IGV (PEN)", "IGV 18% (PEN)",
    "Total con IGV (PEN)", "Estado", "Archivo origen", "Fecha modificación",
]
HEADERS_DETALLE = [
    "N° Ref.", "Proyecto", "Tipo de Proyecto", "N° Documento", "Nombre Ítem",
    "Descripción", "Categoría Ítem", "Cantidad", "P. Unitario sin IGV",
    "Total sin IGV (PEN)", "Total con IGV (PEN)",
]


def _wb_con_un_documento(tmp_path, archivo_origen):
    wb = openpyxl.Workbook()
    ws_master = wb.active
    ws_master.title = "Master"
    for c, h in enumerate(HEADERS_MASTER, 1):
        ws_master.cell(row=1, column=c, value=h)
    fila = {
        "N° Ref.": "TEST-001", "Proyecto": "Lima Proyecto", "Tipo de Proyecto": "I+D+i",
        "Fecha": None, "N° Documento": "123", "Tipo Documento": "Factura",
        "Proveedor": "Proveedor", "Proveedor (Razón Social)": "Proveedor SAC",
        "Categoría": "Materiales", "Resumen Ítems": "Item",
        "Total sin IGV (PEN)": 1000, "IGV 18% (PEN)": 180,
        "Total con IGV (PEN)": 1180, "Estado": "Pagado",
        "Archivo origen": archivo_origen, "Fecha modificación": None,
    }
    for c, h in enumerate(HEADERS_MASTER, 1):
        ws_master.cell(row=2, column=c, value=fila[h])

    ws_detalle = wb.create_sheet("Detalle")
    for c, h in enumerate(HEADERS_DETALLE, 1):
        ws_detalle.cell(row=1, column=c, value=h)

    ruta = tmp_path / "Centro de Costos Perú.xlsx"
    wb.save(str(ruta))
    return ruta


def test_extraer_datos_saneados_incluye_archivo_origen(tmp_path):
    ruta = _wb_con_un_documento(tmp_path, "TEST-001_Proveedor_2026-08-01.jpg")
    data = bv.extraer_datos_saneados(ruta)
    assert data["documentos"][0]["archivo_origen"] == "TEST-001_Proveedor_2026-08-01.jpg"


def test_extraer_datos_saneados_lee_columnas_igv_pen(tmp_path):
    ruta = _wb_con_un_documento(tmp_path, None)
    data = bv.extraer_datos_saneados(ruta)
    doc = data["documentos"][0]
    assert doc["total_sin_iva"] == 1000
    assert doc["iva"] == 180
    assert doc["total_con_iva"] == 1180


def test_extraer_datos_saneados_sin_documentos_no_falla(tmp_path):
    wb = openpyxl.Workbook()
    ws_master = wb.active
    ws_master.title = "Master"
    for c, h in enumerate(HEADERS_MASTER, 1):
        ws_master.cell(row=1, column=c, value=h)
    ws_detalle = wb.create_sheet("Detalle")
    for c, h in enumerate(HEADERS_DETALLE, 1):
        ws_detalle.cell(row=1, column=c, value=h)
    ruta = tmp_path / "Centro de Costos Perú.xlsx"
    wb.save(str(ruta))

    data = bv.extraer_datos_saneados(ruta)
    assert data["documentos"] == []
    assert data["kpis"]["n_documentos"] == 0
    assert data["kpis"]["total_con_iva"] == 0


def test_pendiente_coincide_con_el_color_real_de_auditor_centro_costos(tmp_path):
    """Misma verificacion cruzada que Centro de Costos/Visualizador Web/tests/
    test_build_visualizador.py -- la deteccion de "celda roja" es una
    reimplementacion independiente de _celda_es_roja (Sistema/
    auditor_centro_costos.py), y el color de "requiere revision" no cambia
    entre paises (configurar_pais() no lo toca), asi que el mismo caso de
    prueba aplica sin adaptar nada mas que la ruta de import."""
    raiz_sistema = Path(__file__).resolve().parents[4] / "Centro de Costos" / "Sistema"
    if str(raiz_sistema) not in sys.path:
        sys.path.insert(0, str(raiz_sistema))
    import auditor_centro_costos as acc

    col_n_documento = HEADERS_MASTER.index("N° Documento") + 1
    casos = (
        (acc.ROJO_FONT, True),
        (acc.AZUL_MARINO_FONT, False),
        (acc.NORMAL_FONT, False),
    )
    for font, debe_marcar in casos:
        ruta = _wb_con_un_documento(tmp_path, "TEST-001_Proveedor_2026-08-01.jpg")
        wb = openpyxl.load_workbook(str(ruta))
        wb["Master"].cell(row=2, column=col_n_documento).font = font
        wb.save(str(ruta))

        data = bv.extraer_datos_saneados(ruta)
        assert data["documentos"][0]["pendiente_revision"] is debe_marcar, (
            f"build_visualizador (Peru) discrepo para {font.color.rgb}"
        )
