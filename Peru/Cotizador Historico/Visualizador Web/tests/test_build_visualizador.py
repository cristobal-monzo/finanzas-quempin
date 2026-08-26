import importlib.util
import sys
from pathlib import Path

import openpyxl

# Mismo patron que los otros pares CL/PE de este repo (ver Centro de
# Costos/Visualizador Web/tests/test_build_visualizador.py): "importlib"
# import-mode evita colisiones de test por basename, pero el MODULO FUENTE
# se carga por ruta bajo un nombre unico para que sys.modules no le
# entregue a este test el build_visualizador.py de otro pais/modulo.
_RUTA_BV = Path(__file__).resolve().parent.parent / "build_visualizador.py"
_spec = importlib.util.spec_from_file_location("build_visualizador_ch_pe", _RUTA_BV)
bv = importlib.util.module_from_spec(_spec)
sys.modules["build_visualizador_ch_pe"] = bv
_spec.loader.exec_module(bv)

HEADERS_DETALLE = [
    "N° Ref.", "Nombre Ítem", "Descripción", "P. Unitario sin IGV",
    "Total sin IGV (PEN)", "Total con IGV (PEN)",
]


def _wb_con_un_item(tmp_path):
    wb = openpyxl.Workbook()
    ws_detalle = wb.active
    ws_detalle.title = "Detalle"
    for c, h in enumerate(HEADERS_DETALLE, 1):
        ws_detalle.cell(row=1, column=c, value=h)
    fila = ("LIMA-001", "Taladro", "Taladro percutor 20V", 300, 300, 354)
    for c, v in enumerate(fila, 1):
        ws_detalle.cell(row=2, column=c, value=v)

    ws_master = wb.create_sheet("Master")
    for c, h in enumerate(["N° Ref.", "Fecha"], 1):
        ws_master.cell(row=1, column=c, value=h)
    from datetime import datetime
    ws_master.cell(row=2, column=1, value="LIMA-001")
    ws_master.cell(row=2, column=2, value=datetime(2026, 1, 15))

    ruta = tmp_path / "Centro de Costos Perú.xlsx"
    wb.save(str(ruta))
    return ruta


def test_extraer_indice_saneado_no_tiene_campos_de_uf(tmp_path):
    ruta = _wb_con_un_item(tmp_path)
    data = bv.extraer_indice_saneado(ruta)
    assert "uf_hoy" not in data
    assert "uf_fecha" not in data
    assert "uf_fuente" not in data


def test_extraer_indice_saneado_precio_es_nominal(tmp_path):
    ruta = _wb_con_un_item(tmp_path)
    data = bv.extraer_indice_saneado(ruta)
    assert len(data["items"]) == 1
    item = data["items"][0]
    assert item["precio_reajustado_hoy"] == 300
    assert item["precio_reajustado_hoy_con_iva"] == 354


def test_extraer_indice_saneado_sin_items_no_falla(tmp_path):
    wb = openpyxl.Workbook()
    ws_detalle = wb.active
    ws_detalle.title = "Detalle"
    for c, h in enumerate(HEADERS_DETALLE, 1):
        ws_detalle.cell(row=1, column=c, value=h)
    ws_master = wb.create_sheet("Master")
    for c, h in enumerate(["N° Ref.", "Fecha"], 1):
        ws_master.cell(row=1, column=c, value=h)
    ruta = tmp_path / "Centro de Costos Perú.xlsx"
    wb.save(str(ruta))

    data = bv.extraer_indice_saneado(ruta)
    assert data["items"] == []
    assert data["excluidos_count"] == 0
    assert data["sin_uf_count"] == 0
