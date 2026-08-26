# -*- coding: utf-8 -*-
"""
build_visualizador.py -- genera el visualizador web de Análisis Financiero Perú.

Copia de Sistema Analisis Financiero/Visualizador Web/build_visualizador.py
-- ninguna de sus funciones depende de nombres de columna que varíen entre
países (a diferencia de Cotizador Histórico), así que el único cambio real
es de dónde lee el Excel. Ver ese archivo para el detalle de cada decisión
de recomputo (por qué nunca lee celdas de fórmula de Indicadores/Clientes).

Ver docs/superpowers/specs/2026-07-23-analisis-financiero-visualizador-web-
design.md para el diseno completo (Chile), y docs/superpowers/specs/
2026-08-21-peru-expansion-design.md para el contexto de Perú.
"""

import base64
import io
import json
import sys
from datetime import datetime
from pathlib import Path

import openpyxl

sys.path.insert(0, str(Path(__file__).resolve().parents[3] / "Sistema Analisis Financiero" / "Sistema"))
import analisis_financiero as af  # noqa: E402

RAIZ = Path(__file__).resolve().parent  # Peru/Análisis Financiero/Visualizador Web/
RUTA_EXCEL = af.PAISES["PE"]["ruta_excel_af"]
RUTA_TEMPLATE = RAIZ / "template.html"
RUTA_DATA_JSON = RAIZ / "data" / "analisis-financiero-peru.json"
RUTA_BUILD_HTML = RAIZ / "build" / "index.html"
RAIZ_REPORTES = RAIZ.parent / "Reportes"

# Perú no tiene todavía un link de SharePoint real para su planilla (no
# existe fuera de este repo) -- el mensaje de "pendientes" queda sin link
# hasta que se publique una. Con 0 proyectos hoy, "pendientes" siempre está
# vacío igual, así que esto no tiene efecto visible por ahora.
URL_PLANILLA_PENDIENTE = None

CLAVE_POR_ENCABEZADO = {
    "Estado": "estado",
    "Fecha de inicio": "fecha_inicio",
    "Monto de Venta (sin IVA)": "monto_venta",
    "Costos Materiales Proyectados": "materiales_proy",
    "Costos Equipos Proyectados": "equipos_proy",
    "Mano de Obra Proyectada": "mo_proy",
    "Otros Costos Proyectados": "otros_proy",
    "Mano de Obra Real": "mo_real",
}


def _valor_columna(ws, fila, nombre_columna):
    col = af.HEADERS_PROYECTOS.index(nombre_columna) + 1
    return ws.cell(row=fila, column=col).value


def leer_proyectos(ws_proyectos) -> list[dict]:
    proyectos = []
    for fila in range(2, ws_proyectos.max_row + 1):
        tag = _valor_columna(ws_proyectos, fila, "TAG proyecto")
        nombre = _valor_columna(ws_proyectos, fila, "Nombre del proyecto")
        if not tag or not nombre:
            continue
        proyectos.append({
            "fila": fila,
            "tag": tag,
            "nombre": nombre,
            "cliente": _valor_columna(ws_proyectos, fila, "Cliente"),
            "estado": _valor_columna(ws_proyectos, fila, "Estado"),
            "fecha_inicio": _valor_columna(ws_proyectos, fila, "Fecha de inicio"),
            "fecha_cierre": _valor_columna(ws_proyectos, fila, "Fecha de cierre"),
            "categoria": _valor_columna(ws_proyectos, fila, "Categoría"),
            "monto_venta": _valor_columna(ws_proyectos, fila, "Monto de Venta (sin IVA)"),
            "materiales_proy": _valor_columna(ws_proyectos, fila, "Costos Materiales Proyectados"),
            "equipos_proy": _valor_columna(ws_proyectos, fila, "Costos Equipos Proyectados"),
            "mo_proy": _valor_columna(ws_proyectos, fila, "Mano de Obra Proyectada"),
            "otros_proy": _valor_columna(ws_proyectos, fila, "Otros Costos Proyectados"),
            "mo_real": _valor_columna(ws_proyectos, fila, "Mano de Obra Real"),
        })
    return proyectos


def es_proyecto_completo(p: dict) -> bool:
    return af.tiene_datos_completos(lambda campo: p.get(CLAVE_POR_ENCABEZADO[campo]))


def sumar_costos_reales_por_bucket(ws_detalle, tag: str) -> dict:
    sumas = {"Materiales": 0.0, "Equipos": 0.0, "Otros": 0.0}
    for fila in range(2, ws_detalle.max_row + 1):
        fila_tag = ws_detalle.cell(row=fila, column=1).value
        bucket = ws_detalle.cell(row=fila, column=3).value
        total = ws_detalle.cell(row=fila, column=4).value
        if fila_tag == tag and bucket in sumas and total is not None:
            sumas[bucket] += total
    return sumas


def leer_detalle_subcategorias(ws_detalle) -> dict[str, list[dict]]:
    filas_por_tag: dict[str, list[dict]] = {}
    for fila in range(2, ws_detalle.max_row + 1):
        tag = ws_detalle.cell(row=fila, column=1).value
        total = ws_detalle.cell(row=fila, column=4).value
        if not tag or total is None:
            continue
        filas_por_tag.setdefault(tag, []).append({
            "subcategoria": ws_detalle.cell(row=fila, column=2).value,
            "bucket": ws_detalle.cell(row=fila, column=3).value,
            "total": total,
        })

    resultado: dict[str, list[dict]] = {}
    for tag, filas in filas_por_tag.items():
        total_proyecto = sum(f["total"] for f in filas)
        resultado[tag] = [
            dict(f, pct=(f["total"] / total_proyecto) if total_proyecto else None)
            for f in filas
        ]
    return resultado


def calcular_peso_cartera(proyectos: list[dict]) -> dict[str, float]:
    total_venta = sum(p["monto_venta"] for p in proyectos if p["monto_venta"] is not None)
    return {
        p["tag"]: (p["monto_venta"] / total_venta if total_venta and p["monto_venta"] is not None else 0.0)
        for p in proyectos
    }


def _fecha_str(valor):
    if valor is None:
        return None
    if hasattr(valor, "strftime"):
        return valor.strftime("%d-%m-%Y")
    return str(valor)


def _kpis_por_categoria(p: dict, costos_reales: dict, total_real: float) -> dict:
    reales = {
        "materiales": costos_reales["Materiales"], "equipos": costos_reales["Equipos"],
        "mo": p["mo_real"], "otros": costos_reales["Otros"],
    }
    proyectados = {
        "materiales": p["materiales_proy"], "equipos": p["equipos_proy"],
        "mo": p["mo_proy"], "otros": p["otros_proy"],
    }
    venta = p["monto_venta"]

    costo_pct_venta = {k: (v / venta if venta else 0.0) for k, v in reales.items()}
    estructura_pct = {k: (v / total_real if total_real else 0.0) for k, v in reales.items()}
    desviacion_pct_categoria = {
        k: (reales[k] / proyectados[k] - 1) if proyectados[k] else 0.0 for k in reales
    }
    ahorro_sobrecosto = {k: proyectados[k] - reales[k] for k in reales}

    return {
        "costo_pct_venta": costo_pct_venta,
        "estructura_pct": estructura_pct,
        "desviacion_pct_categoria": desviacion_pct_categoria,
        "ahorro_sobrecosto": ahorro_sobrecosto,
    }


def _margen_por_dia(p: dict, margen_real: float):
    fecha_inicio, fecha_cierre = p["fecha_inicio"], p["fecha_cierre"]
    if fecha_inicio is None or fecha_cierre is None:
        return None
    dias = (fecha_cierre - fecha_inicio).days
    return margen_real / max(1, dias)


def calcular_kpis_proyecto(p: dict, costos_reales: dict) -> dict:
    total_proyectado = p["materiales_proy"] + p["equipos_proy"] + p["mo_proy"] + p["otros_proy"]
    total_real = costos_reales["Materiales"] + costos_reales["Equipos"] + costos_reales["Otros"] + p["mo_real"]
    margen_real = p["monto_venta"] - total_real
    desviacion_pct = (total_real / total_proyectado - 1) if total_proyectado else 0.0

    margen_neto = (margen_real / p["monto_venta"]) if p["monto_venta"] else 0.0
    nota = af.calcular_nota(margen_neto, desviacion_pct)
    evaluacion = af.clasificar_evaluacion(nota)

    resultado = {
        "tag": p["tag"], "nombre": p["nombre"], "cliente": p["cliente"], "estado": p["estado"],
        "fecha_inicio": _fecha_str(p["fecha_inicio"]), "fecha_cierre": _fecha_str(p["fecha_cierre"]),
        "categoria": p["categoria"],
        "monto_venta": p["monto_venta"], "total_proyectado": total_proyectado,
        "total_real": total_real, "margen_real": margen_real, "desviacion_pct": desviacion_pct,
        "nota": nota, "evaluacion": evaluacion,
        "costos_proyectados": {
            "materiales": p["materiales_proy"], "equipos": p["equipos_proy"],
            "mo": p["mo_proy"], "otros": p["otros_proy"],
        },
        "costos_reales": {
            "materiales": costos_reales["Materiales"], "equipos": costos_reales["Equipos"],
            "mo": p["mo_real"], "otros": costos_reales["Otros"],
        },
        "ahorro_sobrecosto_total": total_proyectado - total_real,
        "margen_por_dia": _margen_por_dia(p, margen_real),
    }
    resultado.update(_kpis_por_categoria(p, costos_reales, total_real))
    return resultado


def percentil_inclusivo(valores: list[float], p: float) -> float:
    ordenados = sorted(valores)
    n = len(ordenados)
    if n == 1:
        return ordenados[0]
    idx = p * (n - 1)
    lo = int(idx)
    hi = min(lo + 1, n - 1)
    frac = idx - lo
    return ordenados[lo] + (ordenados[hi] - ordenados[lo]) * frac


def calcular_clientes(kpis_proyectos_completos: list[dict], proyectos_por_tag: dict) -> list[dict]:
    por_cliente: dict[str, list[dict]] = {}
    for kpi in kpis_proyectos_completos:
        cliente = kpi["cliente"]
        if not cliente:
            continue
        por_cliente.setdefault(cliente, []).append(kpi)

    filas = []
    for cliente, kpis in por_cliente.items():
        n = len(kpis)
        aov = sum(k["monto_venta"] for k in kpis) / n
        vida = n
        fechas = [proyectos_por_tag[k["tag"]]["fecha_inicio"] for k in kpis]
        fechas = [f for f in fechas if f is not None]
        if len(fechas) >= 2:
            meses_activo = max(12.0, (max(fechas) - min(fechas)).days / 30)
        else:
            meses_activo = 12.0
        frecuencia = vida / (meses_activo / 12)
        suma_venta = sum(k["monto_venta"] for k in kpis)
        suma_margen = sum(k["margen_real"] for k in kpis)
        margen_pct = suma_margen / suma_venta if suma_venta else 0.0
        cltv = aov * frecuencia * vida * margen_pct
        filas.append({
            "cliente": cliente, "aov": aov, "vida": vida, "meses_activo": meses_activo,
            "frecuencia": frecuencia, "margen_pct": margen_pct, "cltv": cltv,
        })

    cltvs = [f["cltv"] for f in filas]
    for f in filas:
        p67 = percentil_inclusivo(cltvs, 0.67)
        p33 = percentil_inclusivo(cltvs, 0.33)
        if f["cltv"] >= p67:
            f["clasificacion"] = "Clientes estratégicos"
        elif f["cltv"] >= p33:
            f["clasificacion"] = "Clientes potenciales"
        else:
            f["clasificacion"] = "Clientes de oportunidad"

    return filas


def calcular_categorias(kpis_proyectos_completos: list[dict]) -> list[dict]:
    por_categoria: dict[str, list[dict]] = {}
    for kpi in kpis_proyectos_completos:
        categoria = kpi["categoria"] or "Sin categoría"
        por_categoria.setdefault(categoria, []).append(kpi)

    filas = []
    for categoria, kpis in por_categoria.items():
        n = len(kpis)
        filas.append({
            "categoria": categoria,
            "n_proyectos": n,
            "margen_real_total": sum(k["margen_real"] for k in kpis),
            "nota_promedio": sum(k["nota"] for k in kpis) / n,
            "tags_proyectos": [k["tag"] for k in kpis],
        })
    return filas


def embeber_reportes_pdf(proyectos: list[dict], categorias: list[dict]) -> dict[str, str]:
    reportes: dict[str, str] = {}
    if not RAIZ_REPORTES.exists():
        return reportes

    for p in proyectos:
        ruta = RAIZ_REPORTES / "Proyectos" / f"{p['tag']}.pdf"
        if ruta.exists():
            reportes[f"proyecto:{p['tag']}"] = base64.b64encode(ruta.read_bytes()).decode("ascii")

    for c in categorias:
        ruta = RAIZ_REPORTES / "Categorías" / f"{c['categoria']}.pdf"
        if ruta.exists():
            reportes[f"categoria:{c['categoria']}"] = base64.b64encode(ruta.read_bytes()).decode("ascii")

    return reportes


def extraer_datos_saneados(ruta_excel=RUTA_EXCEL) -> dict:
    wb = openpyxl.load_workbook(str(ruta_excel), data_only=True)
    ws_proyectos = wb[af.HOJA_PROYECTOS]
    ws_detalle = wb[af.HOJA_DETALLE_COSTOS_REALES]

    proyectos = leer_proyectos(ws_proyectos)
    proyectos_por_tag = {p["tag"]: p for p in proyectos}
    peso_cartera_por_tag = calcular_peso_cartera(proyectos)
    detalle_subcategorias_por_tag = leer_detalle_subcategorias(ws_detalle)

    completos = []
    pendientes = []
    for p in proyectos:
        if es_proyecto_completo(p):
            costos_reales = sumar_costos_reales_por_bucket(ws_detalle, p["tag"])
            kpi = calcular_kpis_proyecto(p, costos_reales)
            kpi["peso_cartera_pct"] = peso_cartera_por_tag.get(p["tag"], 0.0)
            kpi["detalle_subcategorias"] = detalle_subcategorias_por_tag.get(p["tag"], [])
            completos.append(kpi)
        else:
            pendientes.append({
                "tag": p["tag"],
                "nombre": p["nombre"],
                "mensaje": f"{p['nombre']} — Falta ingresar información en 'Análisis de Proyectos'",
                "link": URL_PLANILLA_PENDIENTE,
            })

    clientes = calcular_clientes(completos, proyectos_por_tag)
    categorias = calcular_categorias(completos)
    reportes_pdf = embeber_reportes_pdf(completos, categorias)

    pendientes_por_cliente: dict[str, int] = {}
    for p in proyectos:
        if not es_proyecto_completo(p) and p["cliente"]:
            pendientes_por_cliente[p["cliente"]] = pendientes_por_cliente.get(p["cliente"], 0) + 1
    for c in clientes:
        c["proyectos_pendientes"] = pendientes_por_cliente.get(c["cliente"], 0)

    n_completos = len(completos)
    return {
        "generado": datetime.now().strftime("%d-%m-%Y %H:%M"),
        "kpis_proyectos": {
            "n_completos": n_completos,
            "margen_real_total": sum(k["margen_real"] for k in completos),
            "monto_venta_total": sum(k["monto_venta"] for k in completos),
            "total_real_total": sum(k["total_real"] for k in completos),
            "nota_promedio": (sum(k["nota"] for k in completos) / n_completos) if n_completos else 0,
            "n_requiere_atencion": sum(1 for k in completos if k["evaluacion"] == "Requiere atención"),
        },
        "proyectos": completos,
        "clientes": clientes,
        "categorias": categorias,
        "pendientes": pendientes,
        "reportes_pdf": reportes_pdf,
    }


def build() -> int:
    try:
        sys.stdout.reconfigure(encoding="utf-8", errors="backslashreplace")
    except Exception:
        pass
    if not RUTA_EXCEL.exists():
        print(f"[ERROR] No existe el Excel: {RUTA_EXCEL}")
        return 1
    if not RUTA_TEMPLATE.exists():
        print(f"[ERROR] No existe la plantilla: {RUTA_TEMPLATE}")
        return 1

    data = extraer_datos_saneados(RUTA_EXCEL)

    RUTA_DATA_JSON.parent.mkdir(parents=True, exist_ok=True)
    with io.open(RUTA_DATA_JSON, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)

    data_json_text = json.dumps(data, ensure_ascii=False)
    data_b64 = base64.b64encode(data_json_text.encode("utf-8")).decode("ascii")

    with io.open(RUTA_TEMPLATE, "r", encoding="utf-8") as f:
        template = f.read()
    if "__AF_DATA_B64__" not in template:
        print("[ERROR] template.html no tiene el placeholder __AF_DATA_B64__")
        return 1
    html = template.replace("__AF_DATA_B64__", data_b64)

    RUTA_BUILD_HTML.parent.mkdir(parents=True, exist_ok=True)
    with io.open(RUTA_BUILD_HTML, "w", encoding="utf-8") as f:
        f.write(html)

    print(f"OK — {len(data['proyectos'])} proyecto(s) completo(s), "
          f"{len(data['pendientes'])} pendiente(s), {len(data['clientes'])} cliente(s)")
    print(f"Snapshot: {RUTA_DATA_JSON}")
    print(f"Visualizador: {RUTA_BUILD_HTML}")
    return 0


if __name__ == "__main__":
    sys.stdout.reconfigure(encoding="utf-8", errors="backslashreplace")
    raise SystemExit(build())
