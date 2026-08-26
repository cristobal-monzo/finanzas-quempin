import importlib.util
import sys
from pathlib import Path

import openpyxl

_RUTA_BV = Path(__file__).resolve().parent.parent / "build_visualizador.py"
_spec = importlib.util.spec_from_file_location("build_visualizador_af_pe", _RUTA_BV)
bv = importlib.util.module_from_spec(_spec)
sys.modules["build_visualizador_af_pe"] = bv
_spec.loader.exec_module(bv)


def test_extraer_datos_saneados_sin_proyectos_no_falla(tmp_path):
    wb = openpyxl.Workbook()
    ws_p = wb.active
    ws_p.title = "Proyectos"
    for c, h in enumerate(bv.af.HEADERS_PROYECTOS, 1):
        ws_p.cell(row=1, column=c, value=h)
    ws_d = wb.create_sheet("Detalle Costos Reales")
    for c, h in enumerate(bv.af.HEADERS_DETALLE_COSTOS_REALES, 1):
        ws_d.cell(row=1, column=c, value=h)
    ruta = tmp_path / "Análisis de Proyectos Perú.xlsx"
    wb.save(str(ruta))

    data = bv.extraer_datos_saneados(ruta)
    assert data["proyectos"] == []
    assert data["clientes"] == []
    assert data["pendientes"] == []
    assert data["kpis_proyectos"]["n_completos"] == 0
    assert data["kpis_proyectos"]["nota_promedio"] == 0
